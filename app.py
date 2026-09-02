#!/usr/bin/env python3
"""
GRATZ Logistics Company
Transport Fleet & Finance Management System
"""

import ctypes
import os
import sys
import csv
import difflib
import hashlib
import hmac
import io
import json
import logging
import re
import secrets
import shutil
import socket
import threading
import time
import uuid
import webbrowser
import zipfile
from logging.handlers import RotatingFileHandler
from datetime import datetime, date, timedelta, timezone
from functools import wraps
from types import SimpleNamespace

import openpyxl
import requests
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, PageBreak, SimpleDocTemplate, Spacer, Table, TableStyle
from dotenv import load_dotenv
from flask import (Flask, render_template, render_template_string, request, redirect, url_for,
                   flash, jsonify, make_response, session, send_from_directory, send_file, abort)
from flask_sqlalchemy import SQLAlchemy
from flask_login import (LoginManager, UserMixin, login_user, logout_user,
                         login_required, current_user)
from flask_wtf.csrf import CSRFProtect, generate_csrf
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from werkzeug.exceptions import HTTPException
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from sqlalchemy import event, func
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import with_loader_criteria

# A PyInstaller-frozen spoke .exe needs its .env and default SQLite file
# next to the .exe itself (sys.executable — stable across runs), not
# resolved the normal way: the module's own __file__/Flask's instance_path
# both point inside the frozen bundle's extraction dir, which for --onefile
# is a fresh temp folder every single run. Defaulting the DB there would
# silently lose all local data on every restart. Untouched for the normal
# case (dev machine, Render) — same load_dotenv()/default URI as before.
FROZEN = getattr(sys, 'frozen', False)
if FROZEN:
    _exe_dir = os.path.dirname(sys.executable)
    _env_path = os.path.join(_exe_dir, '.env')
    load_dotenv(_env_path)
    _default_db_uri = 'sqlite:///' + os.path.join(_exe_dir, 'transport_erp.db').replace('\\', '/')
else:
    _env_path = os.path.join(os.getcwd(), '.env')
    load_dotenv()
    _default_db_uri = 'sqlite:///transport_erp.db'

# Same "next to the .exe, not the frozen bundle's temp extraction dir"
# reasoning as _env_path above — used for this spoke's own VERSION marker
# (see check_for_spoke_update) and as the default landing spot for
# published release downloads on the hub.
BASE_DIR = _exe_dir if FROZEN else os.getcwd()
_version_path = os.path.join(BASE_DIR, 'VERSION')
APP_VERSION = open(_version_path, encoding='utf-8').read().strip() if os.path.exists(_version_path) else 'dev'


def persist_env_updates(updates):
    """Rewrite/append KEY=value lines in the local .env file (creating it
    if absent) and mirror them into os.environ for the current process.
    Used by first-run setup (SECRET_KEY, /setup's hub enrollment) so a
    shipped spoke .exe never requires hand-editing .env — settings chosen
    once in the browser survive a restart the same as if a person had
    typed them in themselves."""
    lines = []
    if os.path.exists(_env_path):
        with open(_env_path, encoding='utf-8') as f:
            lines = f.readlines()
    seen = set()
    for i, line in enumerate(lines):
        stripped = line.strip()
        if not stripped or stripped.startswith('#') or '=' not in stripped:
            continue
        key = stripped.split('=', 1)[0]
        if key in updates:
            lines[i] = f'{key}={updates[key]}\n'
            seen.add(key)
    for key, value in updates.items():
        if key not in seen:
            lines.append(f'{key}={value}\n')
    with open(_env_path, 'w', encoding='utf-8') as f:
        f.writelines(lines)
    for key, value in updates.items():
        os.environ[key] = value


# ─────────────────────────────────────────────────────────────
# App Setup
# ─────────────────────────────────────────────────────────────
# A shipped spoke .exe must always run production-hardened regardless of
# whether FLASK_ENV got set — there's no dev machine on the other end of
# it to deliberately opt into debug mode.
IS_PRODUCTION = True if FROZEN else os.environ.get('FLASK_ENV', 'production').lower() == 'production'

secret_key = os.environ.get('SECRET_KEY')
if not secret_key and FROZEN:
    # First launch of a shipped .exe with no SECRET_KEY yet — generate one
    # and persist it next to the .exe so sessions survive a restart.
    # (Never regenerate silently once one exists: that would invalidate
    # every open session on every restart.)
    secret_key = secrets.token_hex(32)
    persist_env_updates({'SECRET_KEY': secret_key})
if not secret_key:
    if IS_PRODUCTION:
        raise RuntimeError(
            'SECRET_KEY environment variable must be set in production. '
            'Set FLASK_ENV=development for local work with an auto-generated key.'
        )
    secret_key = secrets.token_hex(32)
    print('WARNING: SECRET_KEY not set — using a random development key '
          '(sessions will not persist across restarts).')

app = Flask(__name__)
app.config.update(
    SECRET_KEY=secret_key,
    SQLALCHEMY_DATABASE_URI=os.environ.get('DATABASE_URL', _default_db_uri),
    SQLALCHEMY_TRACK_MODIFICATIONS=False,
    COMMISSION_DRIVER_RATE=0.15,
    COMMISSION_CONDUCTOR_RATE=0.08,
    VEHICLE_USEFUL_LIFE_YEARS=5,
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE='Lax',
    SESSION_COOKIE_SECURE=IS_PRODUCTION,
    PERMANENT_SESSION_LIFETIME=timedelta(hours=12),
    SITE_ID=os.environ.get('SITE_ID', 'central'),
    SYNC_ENABLED=os.environ.get('SYNC_ENABLED', 'false').lower() == 'true',
    SYNC_HUB_URL=os.environ.get('SYNC_HUB_URL', '').rstrip('/'),
    SYNC_API_KEY=os.environ.get('SYNC_API_KEY', ''),
    SYNC_INTERVAL_SECONDS=int(os.environ.get('SYNC_INTERVAL_SECONDS', '60')),
    APP_VERSION=APP_VERSION,
    # Hub-side storage for published spoke .exe builds (see SpokeRelease) —
    # a spoke never sets this, only reads releases over the API. Defaults
    # next to the app for local/dev; Render points it at the mounted
    # persistent disk (see render.yaml) so a release survives a redeploy.
    SPOKE_RELEASES_DIR=os.environ.get('SPOKE_RELEASES_DIR', os.path.join(BASE_DIR, 'spoke_releases')),
    # How often a spoke checks the hub for a new published version — much
    # coarser than SYNC_INTERVAL_SECONDS above, since this is a one-shot
    # metadata check, not per-record data sync, and there's no urgency:
    # the update only ever applies on the next natural restart anyway.
    SPOKE_UPDATE_CHECK_SECONDS=int(os.environ.get('SPOKE_UPDATE_CHECK_SECONDS', str(6 * 3600))),
    WHATSAPP_TOKEN=os.environ.get('WHATSAPP_TOKEN', ''),
    WHATSAPP_PHONE_NUMBER_ID=os.environ.get('WHATSAPP_PHONE_NUMBER_ID', ''),
    WHATSAPP_VERIFY_TOKEN=os.environ.get('WHATSAPP_VERIFY_TOKEN', ''),
    WHATSAPP_APP_SECRET=os.environ.get('WHATSAPP_APP_SECRET', ''),
)

db = SQLAlchemy(app)
login_manager = LoginManager(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Please log in to access this page.'
login_manager.login_message_category = 'warning'

csrf = CSRFProtect(app)
limiter = Limiter(get_remote_address, app=app, storage_uri='memory://',
                  default_limits=[])

# Durable error/warning log, next to the .exe for a spoke install (same
# BASE_DIR as .env/VERSION/spoke_releases — see FROZEN handling above) or
# next to app.py for a dev checkout/Render. Without this, app.logger calls
# (sync failures, WhatsApp errors, the unhandled-exception handler below)
# go to stderr only — which for a windowed spoke .exe with no console
# (see _disable_console_quick_edit) and nothing redirecting its output is
# nowhere at all: a spoke could fail silently for weeks with zero trace on
# that machine. Rotated so an unattended install can't fill its disk.
_log_dir = os.path.join(BASE_DIR, 'logs')
os.makedirs(_log_dir, exist_ok=True)
_file_handler = RotatingFileHandler(
    os.path.join(_log_dir, 'app.log'), maxBytes=2_000_000, backupCount=5, encoding='utf-8')
_file_handler.setFormatter(logging.Formatter('%(asctime)s %(levelname)s: %(message)s'))
_file_handler.setLevel(logging.INFO)
app.logger.addHandler(_file_handler)
app.logger.setLevel(logging.INFO)

_ERROR_PAGE_HTML = """<!doctype html>
<html><head><meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1">
<title>Something went wrong</title>
<style>
  body{font-family:system-ui,-apple-system,Segoe UI,sans-serif;background:#f4f5f3;color:#20241f;
       display:flex;align-items:center;justify-content:center;min-height:100vh;margin:0;padding:24px;}
  .box{max-width:420px;text-align:center;}
  h1{font-size:1.3rem;margin:0 0 8px;}
  p{color:#5a625a;line-height:1.5;}
  a{display:inline-block;margin-top:16px;padding:10px 20px;background:#20241f;color:#fff;
    text-decoration:none;border-radius:6px;}
</style></head>
<body><div class="box">
  <h1>Something went wrong on our end</h1>
  <p>The error has been logged. Try again in a moment — if it keeps happening, tell an admin what you were doing.</p>
  <a href="{{ url_for('dashboard') }}">Back to dashboard</a>
</div></body></html>"""


@app.errorhandler(Exception)
def handle_unexpected_error(e):
    """Last-resort catch for anything handle_form_errors doesn't (a POST's
    KeyError/ValueError, or literally any exception on a GET) — without
    this, an unexpected error was a raw unstyled 500 with no record of what
    happened anywhere (see _file_handler above). HTTPException (404, 403,
    the abort()s scattered through the app, etc.) is deliberately passed
    through unchanged so Flask's normal handling of those still applies."""
    if isinstance(e, HTTPException):
        return e
    db.session.rollback()
    app.logger.exception('Unhandled exception on %s %s', request.method, request.path)
    return render_template_string(_ERROR_PAGE_HTML), 500


@app.template_global()
def asset_v(rel_path):
    """Cache-busting query string for a static asset, e.g. `{{ url_for('static',
    filename='js/app.js') }}?v={{ asset_v('js/app.js') }}` — keyed off the file's
    own mtime so a JS/CSS fix takes effect on next load instead of sitting behind
    a stale browser cache (Flask's static handler sends a 12h max-age by default)."""
    try:
        return int(os.path.getmtime(os.path.join(app.static_folder, rel_path)))
    except OSError:
        return 0


# ─────────────────────────────────────────────────────────────
# Models
# ─────────────────────────────────────────────────────────────
class User(UserMixin, db.Model):
    __tablename__ = 'users'
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password_hash = db.Column(db.String(256), nullable=False)
    role = db.Column(db.String(20), nullable=False, default='manager')
    is_active = db.Column(db.Boolean, default=True)
    permissions = db.Column(db.Text, default='[]')
    driver_id = db.Column(db.Integer, db.ForeignKey('drivers.id'), nullable=True)
    # Digits-only WhatsApp number (country code + number, no '+' or spaces —
    # e.g. "263771234567") this user is reachable/authenticated as on the
    # WhatsApp report bot. Null means the user hasn't been linked yet, and
    # the bot won't respond to their number. See whatsapp_dispatch below.
    whatsapp_phone = db.Column(db.String(20), unique=True, nullable=True)
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))

    linked_driver = db.relationship('Driver', foreign_keys=[driver_id])

    def set_password(self, pw):
        self.password_hash = generate_password_hash(pw)

    def check_password(self, pw):
        return check_password_hash(self.password_hash, pw)

    def has_permission(self, perm):
        if self.role == 'admin':
            return True
        try:
            return perm in json.loads(self.permissions or '[]')
        except (json.JSONDecodeError, TypeError):
            return False

    def get_permissions(self):
        try:
            return json.loads(self.permissions or '[]')
        except (json.JSONDecodeError, TypeError):
            return []


class Vehicle(db.Model):
    __tablename__ = 'vehicles'
    id = db.Column(db.Integer, primary_key=True)
    registration = db.Column(db.String(20), unique=True, nullable=False)
    make = db.Column(db.String(50), nullable=False)
    model = db.Column(db.String(50), nullable=False)
    year = db.Column(db.Integer, nullable=False)
    acquisition_cost = db.Column(db.Float, default=0.0)
    status = db.Column(db.String(20), default='active')
    fuel_type = db.Column(db.String(10), default='diesel')
    # Expected daily fare for this vehicle, set by the admin. Null means no
    # target is tracked for this vehicle — it's excluded from shortfall
    # flagging (see report_shortfalls / DailyLog.garnish).
    daily_target = db.Column(db.Float, nullable=True)
    # Insurance is tracked as its own first-class field (not a generic
    # VehicleDocument) since every vehicle needs exactly one current policy
    # and it's the compliance item admins check most often — it gets its
    # own alerting the same way documents do (see insurance_status below).
    insurance_provider = db.Column(db.String(100))
    insurance_policy_number = db.Column(db.String(100))
    # Classification of cover held (Comprehensive, Third Party Only, etc.) —
    # a free-text-backed field like VehicleDocument.doc_type rather than an
    # enum, so a class specific to one insurer/market can still be entered.
    insurance_type = db.Column(db.String(50))
    insurance_expiry = db.Column(db.Date, nullable=True)
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    updated_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    sync_uuid = db.Column(db.String(36), unique=True, index=True)
    pending_push = db.Column(db.Boolean, default=False)
    last_modified_site = db.Column(db.String(50))
    deleted_at = db.Column(db.DateTime, nullable=True)
    last_synced_updated_at = db.Column(db.DateTime, nullable=True)
    server_touched_at = db.Column(db.DateTime, nullable=True)

    documents = db.relationship('VehicleDocument', backref='vehicle',
                                lazy=True, cascade='all, delete-orphan')
    daily_logs = db.relationship('DailyLog', backref='vehicle', lazy=True)
    fuel_logs = db.relationship('FuelLog', backref='vehicle', lazy=True)
    maintenance_logs = db.relationship('MaintenanceLog', backref='vehicle', lazy=True)

    @property
    def total_revenue(self):
        # SQL-side sum, not sum(l.x for l in self.daily_logs) — that loaded
        # this vehicle's entire lifetime of logs into Python on every call,
        # which is fine for one vehicle here on the detail page but was ruinous
        # when called per-vehicle in a fleet list (see vehicles() below, which
        # now precomputes these in bulk instead of touching this property at all).
        return db.session.query(func.sum(DailyLog.gross_revenue)).filter(
            DailyLog.vehicle_id == self.id).scalar() or 0

    @property
    def total_fuel_liters(self):
        return db.session.query(func.sum(FuelLog.liters)).filter(
            FuelLog.vehicle_id == self.id).scalar() or 0

    @property
    def total_maintenance_cost(self):
        return db.session.query(func.sum(MaintenanceLog.total_cost)).filter(
            MaintenanceLog.vehicle_id == self.id).scalar() or 0

    @property
    def insurance_days_to_expiry(self):
        return (self.insurance_expiry - date.today()).days if self.insurance_expiry else None

    @property
    def insurance_status(self):
        if not self.insurance_expiry:
            return 'none'
        d = self.insurance_days_to_expiry
        if d < 0:
            return 'expired'
        if d <= 30:
            return 'warning'
        return 'valid'


class VehicleDocument(db.Model):
    __tablename__ = 'vehicle_documents'
    id = db.Column(db.Integer, primary_key=True)
    vehicle_id = db.Column(db.Integer, db.ForeignKey('vehicles.id'), nullable=False)
    doc_type = db.Column(db.String(50), nullable=False)
    reference_number = db.Column(db.String(100))
    issue_date = db.Column(db.Date)
    expiry_date = db.Column(db.Date, nullable=False)
    notes = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    updated_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    sync_uuid = db.Column(db.String(36), unique=True, index=True)
    pending_push = db.Column(db.Boolean, default=False)
    last_modified_site = db.Column(db.String(50))
    deleted_at = db.Column(db.DateTime, nullable=True)
    last_synced_updated_at = db.Column(db.DateTime, nullable=True)
    server_touched_at = db.Column(db.DateTime, nullable=True)

    @property
    def days_to_expiry(self):
        return (self.expiry_date - date.today()).days

    @property
    def status(self):
        d = self.days_to_expiry
        if d < 0:
            return 'expired'
        if d <= 30:
            return 'warning'
        return 'valid'


class Driver(db.Model):
    __tablename__ = 'drivers'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    # Only drivers need a license — conductors don't drive, so this is
    # optional and only enforced as required at the form level for role='driver'.
    license_number = db.Column(db.String(50), unique=True, nullable=True)
    # National ID / passport number — not DB-unique (existing SQLite files
    # can't gain a UNIQUE constraint via ALTER TABLE), so duplicates are
    # caught with the same check_unique() friendly-error path as license_number.
    id_number = db.Column(db.String(30))
    phone = db.Column(db.String(20))
    role = db.Column(db.String(20), default='driver')
    commission_rate = db.Column(db.Float)
    status = db.Column(db.String(20), default='active')
    # For a conductor, the driver they normally work under — informational,
    # and used to auto-select the conductor when logging a trip for that driver.
    paired_driver_id = db.Column(db.Integer, db.ForeignKey('drivers.id'), nullable=True)
    # The vehicle this driver/conductor is normally assigned to — informational,
    # and used to auto-select the driver/conductor when logging a trip for that vehicle.
    assigned_vehicle_id = db.Column(db.Integer, db.ForeignKey('vehicles.id'), nullable=True)
    next_of_kin_name = db.Column(db.String(100))
    next_of_kin_phone = db.Column(db.String(20))
    next_of_kin_relationship = db.Column(db.String(50))
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    updated_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    sync_uuid = db.Column(db.String(36), unique=True, index=True)
    pending_push = db.Column(db.Boolean, default=False)
    last_modified_site = db.Column(db.String(50))
    deleted_at = db.Column(db.DateTime, nullable=True)
    last_synced_updated_at = db.Column(db.DateTime, nullable=True)
    server_touched_at = db.Column(db.DateTime, nullable=True)

    paired_driver = db.relationship('Driver', remote_side=[id], backref='paired_conductors')
    assigned_vehicle = db.relationship('Vehicle', backref='assigned_crew')
    driven_logs = db.relationship('DailyLog', foreign_keys='DailyLog.driver_id',
                                  backref='driver', lazy=True)
    conducted_logs = db.relationship('DailyLog', foreign_keys='DailyLog.conductor_id',
                                     backref='conductor', lazy=True)


class Route(db.Model):
    __tablename__ = 'routes'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    start_point = db.Column(db.String(100), nullable=False)
    end_point = db.Column(db.String(100), nullable=False)
    distance_km = db.Column(db.Float)
    fare_rate = db.Column(db.Float, nullable=False)
    status = db.Column(db.String(20), default='active')
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    updated_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    sync_uuid = db.Column(db.String(36), unique=True, index=True)
    pending_push = db.Column(db.Boolean, default=False)
    last_modified_site = db.Column(db.String(50))
    deleted_at = db.Column(db.DateTime, nullable=True)
    last_synced_updated_at = db.Column(db.DateTime, nullable=True)
    server_touched_at = db.Column(db.DateTime, nullable=True)

    logs = db.relationship('DailyLog', backref='route', lazy=True)


class DailyLog(db.Model):
    __tablename__ = 'daily_logs'
    __table_args__ = (
        db.Index('ix_daily_logs_vehicle_date', 'vehicle_id', 'log_date'),
        db.Index('ix_daily_logs_date', 'log_date'),
    )
    id = db.Column(db.Integer, primary_key=True)
    vehicle_id = db.Column(db.Integer, db.ForeignKey('vehicles.id'), nullable=False)
    # Nullable so an activity/fare can be logged before a driver is known or
    # assigned — see driver_ledger_add. Garnish still requires a driver since
    # it nets against a specific person's commission.
    driver_id = db.Column(db.Integer, db.ForeignKey('drivers.id'), nullable=True)
    conductor_id = db.Column(db.Integer, db.ForeignKey('drivers.id'))
    # Nullable because Daily Transactions (the vehicle ledger) doesn't
    # collect a route per entry — it mirrors a plain date/driver/fare/
    # diesel/mileage logbook. Left over from a route-per-trip form that
    # has since been merged into this page; route_profitability reports
    # simply see None for rows entered here.
    route_id = db.Column(db.Integer, db.ForeignKey('routes.id'), nullable=True)
    log_date = db.Column(db.Date, nullable=False, default=date.today)
    trips_completed = db.Column(db.Integer, default=0)
    gross_revenue = db.Column(db.Float, nullable=False, default=0.0)
    # Garnished from the driver/conductor's commission for this day — typically
    # because they fell short of the admin-set revenue target and the admin
    # decided not to pay out the full percentage. See reason_for_shortfall.
    # Netted against commission in the payroll report.
    garnish = db.Column(db.Float, nullable=False, default=0.0)
    reason_for_shortfall = db.Column(db.Text)
    notes = db.Column(db.Text)
    created_by = db.Column(db.Integer, db.ForeignKey('users.id'))
    updated_by = db.Column(db.Integer, db.ForeignKey('users.id'))
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    updated_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    sync_uuid = db.Column(db.String(36), unique=True, index=True)
    pending_push = db.Column(db.Boolean, default=False)
    last_modified_site = db.Column(db.String(50))
    deleted_at = db.Column(db.DateTime, nullable=True)
    last_synced_updated_at = db.Column(db.DateTime, nullable=True)
    server_touched_at = db.Column(db.DateTime, nullable=True)

    creator = db.relationship('User', foreign_keys=[created_by])
    updater = db.relationship('User', foreign_keys=[updated_by])


class FuelLog(db.Model):
    __tablename__ = 'fuel_logs'
    __table_args__ = (
        db.Index('ix_fuel_logs_vehicle_date', 'vehicle_id', 'log_date'),
        db.Index('ix_fuel_logs_date', 'log_date'),
    )
    id = db.Column(db.Integer, primary_key=True)
    vehicle_id = db.Column(db.Integer, db.ForeignKey('vehicles.id'), nullable=False)
    log_date = db.Column(db.Date, nullable=False, default=date.today)
    liters = db.Column(db.Float, nullable=False)
    # Populated automatically from FuelPrice (see fuel_price_for) at the
    # vehicle's fuel_type — either derived from a cash amount entered in the
    # Daily Transactions ledger (cost known, liters computed) or from liters
    # entered directly here (liters known, cost computed). Stays 0 if no
    # price has been configured yet, same as before this was wired up.
    cost_per_liter = db.Column(db.Float, nullable=False, default=0.0)
    total_cost = db.Column(db.Float, nullable=False, default=0.0)
    odometer = db.Column(db.Float)
    supplier = db.Column(db.String(100))
    notes = db.Column(db.Text)
    created_by = db.Column(db.Integer, db.ForeignKey('users.id'))
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    updated_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    sync_uuid = db.Column(db.String(36), unique=True, index=True)
    pending_push = db.Column(db.Boolean, default=False)
    last_modified_site = db.Column(db.String(50))
    deleted_at = db.Column(db.DateTime, nullable=True)
    last_synced_updated_at = db.Column(db.DateTime, nullable=True)
    server_touched_at = db.Column(db.DateTime, nullable=True)

    creator = db.relationship('User', foreign_keys=[created_by])


class FuelPrice(db.Model):
    """Current pump price per litre for diesel and petrol, set by an admin
    on this site. Deliberately local/unsynced (like SpokeUpdateState) rather
    than propagated hub-to-spoke — fuel prices are set per site/region and
    change often, so each site's own current price is what should apply to
    fuel logged there. Used to convert between a cash fuel spend and liters
    (see fuel_price_for) so fuel efficiency can be computed automatically
    without crew having to read and enter a liter figure themselves.
    Singleton: one row (id=1), same convention as SpokeUpdateState."""
    __tablename__ = 'fuel_prices'
    id = db.Column(db.Integer, primary_key=True)
    diesel_price = db.Column(db.Float, nullable=True)
    petrol_price = db.Column(db.Float, nullable=True)
    updated_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    updated_by = db.Column(db.Integer, db.ForeignKey('users.id'))

    updater = db.relationship('User', foreign_keys=[updated_by])


class MaintenanceLog(db.Model):
    __tablename__ = 'maintenance_logs'
    __table_args__ = (
        db.Index('ix_maintenance_logs_vehicle_date', 'vehicle_id', 'log_date'),
        db.Index('ix_maintenance_logs_date', 'log_date'),
    )
    id = db.Column(db.Integer, primary_key=True)
    vehicle_id = db.Column(db.Integer, db.ForeignKey('vehicles.id'), nullable=False)
    log_date = db.Column(db.Date, nullable=False, default=date.today)
    description = db.Column(db.Text, nullable=False)
    parts_cost = db.Column(db.Float, default=0.0)
    labor_cost = db.Column(db.Float, default=0.0)
    total_cost = db.Column(db.Float, nullable=False)
    mechanic = db.Column(db.String(100))
    notes = db.Column(db.Text)
    created_by = db.Column(db.Integer, db.ForeignKey('users.id'))
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    updated_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    sync_uuid = db.Column(db.String(36), unique=True, index=True)
    pending_push = db.Column(db.Boolean, default=False)
    last_modified_site = db.Column(db.String(50))
    deleted_at = db.Column(db.DateTime, nullable=True)
    last_synced_updated_at = db.Column(db.DateTime, nullable=True)
    server_touched_at = db.Column(db.DateTime, nullable=True)

    creator = db.relationship('User', foreign_keys=[created_by])


class AuditLog(db.Model):
    __tablename__ = 'audit_logs'
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'))
    action = db.Column(db.String(20), nullable=False)
    table_name = db.Column(db.String(50))
    record_id = db.Column(db.Integer)
    description = db.Column(db.Text)
    ip_address = db.Column(db.String(50))
    timestamp = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))

    user = db.relationship('User')


class OfflineSyncLog(db.Model):
    """One row per successfully-synced offline form submission, keyed by the
    client-generated UUID that offline.js attaches to every queued POST — lets
    already_synced() reject a replayed submission (e.g. the device retries
    before seeing the first response) without creating a duplicate record."""
    __tablename__ = 'offline_sync_log'
    id = db.Column(db.Integer, primary_key=True)
    client_id = db.Column(db.String(36), unique=True, nullable=False)
    endpoint = db.Column(db.String(100), nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'))
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))


class SyncSite(db.Model):
    """One row per remote spoke (a local-server PC at a site) authorized to
    call /api/sync/push and /api/sync/pull. Per-site keys — rather than one
    shared secret — so a lost/decommissioned field PC can be revoked on its
    own without rotating a key every other site depends on."""
    __tablename__ = 'sync_sites'
    id = db.Column(db.Integer, primary_key=True)
    site_id = db.Column(db.String(50), unique=True, nullable=False)
    api_key_hash = db.Column(db.String(256), nullable=False)
    display_name = db.Column(db.String(100))
    is_active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    # Stamped by the hub itself on every authenticated call (see
    # sync_auth_required) so /sync/health can show which spokes are
    # actually checking in versus stale/offline, without the spoke needing
    # to report anything extra about its own health.
    last_push_at = db.Column(db.DateTime)
    last_pull_at = db.Column(db.DateTime)


class SyncPeerState(db.Model):
    """Local-instance-only bookkeeping for this server's own sync cursor —
    does not itself get synced. last_pull_at is set from the hub's returned
    server_time (never this machine's own clock), so clock skew between
    sites can't silently corrupt the pull watermark."""
    __tablename__ = 'sync_peer_state'
    id = db.Column(db.Integer, primary_key=True)
    peer_url = db.Column(db.String(255), unique=True, nullable=False)
    last_pull_at = db.Column(db.DateTime)
    last_push_at = db.Column(db.DateTime)
    last_error = db.Column(db.Text)
    updated_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))


class SyncConflict(db.Model):
    """A logged last-write-wins resolution where two sites edited the same
    row differently while both were offline. Both full payloads are kept —
    not just the diff — so a human reviewing later has everything needed to
    manually correct the losing side's change if it mattered (see the
    'last write wins, but log every conflict for review' policy)."""
    __tablename__ = 'sync_conflicts'
    id = db.Column(db.Integer, primary_key=True)
    table_name = db.Column(db.String(50), nullable=False)
    sync_uuid = db.Column(db.String(36), nullable=False)
    conflict_type = db.Column(db.String(30), nullable=False)  # 'lww' | 'duplicate_constraint' | 'fk_missing'
    winning_site_id = db.Column(db.String(50))
    losing_site_id = db.Column(db.String(50))
    winning_updated_at = db.Column(db.DateTime)
    losing_updated_at = db.Column(db.DateTime)
    winning_payload = db.Column(db.Text)  # JSON
    losing_payload = db.Column(db.Text)   # JSON
    detected_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    resolved = db.Column(db.Boolean, default=False)
    resolved_by = db.Column(db.Integer, db.ForeignKey('users.id'))
    resolved_at = db.Column(db.DateTime)
    resolution_notes = db.Column(db.Text)

    resolver = db.relationship('User', foreign_keys=[resolved_by])


class SpokeRelease(db.Model):
    """A published build of the spoke .exe (see spoke_build.spec), uploaded
    once on the hub and pulled by every spoke from there — this is what
    lets a spoke self-update instead of someone hand-carrying a USB drive
    to each site PC after every app change. Hub-only: this table only ever
    has rows on Central, since only an admin there uploads releases; a
    spoke's own copy of this table stays empty (see SpokeUpdateState for
    what a spoke tracks about its own update checks).

    Never part of the multi-site record sync (SYNC_TABLE_ORDER) — a
    release is a hub-authored artifact pulled over its own dedicated
    /api/spoke/* endpoints, not a row a spoke could ever legitimately own
    or edit itself."""
    __tablename__ = 'spoke_releases'
    id = db.Column(db.Integer, primary_key=True)
    version = db.Column(db.String(40), unique=True, nullable=False)
    filename = db.Column(db.String(255), nullable=False)
    sha256 = db.Column(db.String(64), nullable=False)
    file_size = db.Column(db.Integer, nullable=False)
    notes = db.Column(db.Text)
    # Exactly one release is "latest" at a time — the one every spoke's
    # next check-in will stage. Kept as an explicit flag rather than
    # "highest id" so a bad release can be rolled back by re-flagging an
    # older, known-good one without deleting the broken row's audit trail.
    is_latest = db.Column(db.Boolean, default=False, nullable=False)
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    created_by = db.Column(db.Integer, db.ForeignKey('users.id'))

    uploader = db.relationship('User', foreign_keys=[created_by])


class SpokeUpdateState(db.Model):
    """Local-instance-only bookkeeping for THIS spoke's self-update
    checks — mirrors SyncPeerState's role for data sync, but for the app
    binary itself. Never synced. Singleton: one row per spoke (id=1),
    same convention as the config a spoke keeps only about itself."""
    __tablename__ = 'spoke_update_state'
    id = db.Column(db.Integer, primary_key=True)
    last_checked_at = db.Column(db.DateTime)
    last_error = db.Column(db.Text)
    # Set once a newer version has been downloaded and unpacked into a
    # staging folder next to the install — the launcher script (not this
    # app process, which can't safely overwrite its own running files)
    # applies it on the next start. Left blank between checks/once applied.
    staged_version = db.Column(db.String(40))
    staged_dir = db.Column(db.String(500))
    updated_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))


class ImportBatch(db.Model):
    """One row per committed file import (ledger or franchise workbook). This
    is the structured audit trail for imports — separate from the generic
    AuditLog above — because it needs to carry the quarantined error rows
    and link to the records it created, so a bad import can be identified
    and reversed via ImportBatchRecord."""
    __tablename__ = 'import_batches'
    id = db.Column(db.Integer, primary_key=True)
    target_type = db.Column(db.String(30), nullable=False)   # 'ledger' | 'franchise_workbook'
    filename = db.Column(db.String(255), nullable=False)
    total_rows = db.Column(db.Integer, default=0)
    rows_imported = db.Column(db.Integer, default=0)
    rows_failed = db.Column(db.Integer, default=0)
    error_rows = db.Column(db.Text)   # JSON list of {**original_row_columns, 'System_Error': msg}
    status = db.Column(db.String(15), nullable=False, default='committed')  # 'committed' | 'reverted'
    created_by = db.Column(db.Integer, db.ForeignKey('users.id'))
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    reverted_at = db.Column(db.DateTime)
    reverted_by = db.Column(db.Integer, db.ForeignKey('users.id'))

    creator = db.relationship('User', foreign_keys=[created_by])
    reverter = db.relationship('User', foreign_keys=[reverted_by])
    records = db.relationship('ImportBatchRecord', backref='batch', lazy=True,
                               cascade='all, delete-orphan')

    @property
    def error_row_list(self):
        return json.loads(self.error_rows) if self.error_rows else []


class ImportBatchRecord(db.Model):
    """Links an ImportBatch to the actual row it created, so a revert can
    delete exactly those rows instead of guessing by timestamp."""
    __tablename__ = 'import_batch_records'
    id = db.Column(db.Integer, primary_key=True)
    batch_id = db.Column(db.Integer, db.ForeignKey('import_batches.id'), nullable=False)
    target_table = db.Column(db.String(30), nullable=False)   # 'daily_logs' | 'fuel_logs' | 'franchise_vehicles'
    record_id = db.Column(db.Integer, nullable=False)


# ─────────────────────────────────────────────────────────────
# Finance: loans, payables, receivables, capital, expenses, budget
# ─────────────────────────────────────────────────────────────
class Loan(db.Model):
    __tablename__ = 'loans'
    id = db.Column(db.Integer, primary_key=True)
    lender = db.Column(db.String(100), nullable=False)
    principal = db.Column(db.Float, nullable=False)
    interest_rate = db.Column(db.Float, default=0.0)
    start_date = db.Column(db.Date, nullable=False)
    term_months = db.Column(db.Integer)
    status = db.Column(db.String(20), default='active')
    notes = db.Column(db.Text)
    created_by = db.Column(db.Integer, db.ForeignKey('users.id'))
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    updated_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    sync_uuid = db.Column(db.String(36), unique=True, index=True)
    pending_push = db.Column(db.Boolean, default=False)
    last_modified_site = db.Column(db.String(50))
    deleted_at = db.Column(db.DateTime, nullable=True)
    last_synced_updated_at = db.Column(db.DateTime, nullable=True)
    server_touched_at = db.Column(db.DateTime, nullable=True)

    payments = db.relationship('LoanPayment', backref='loan', lazy=True,
                               cascade='all, delete-orphan')

    @property
    def total_repaid(self):
        return sum(p.amount for p in self.payments)

    @property
    def outstanding_balance(self):
        return self.principal - self.total_repaid


class LoanPayment(db.Model):
    __tablename__ = 'loan_payments'
    id = db.Column(db.Integer, primary_key=True)
    loan_id = db.Column(db.Integer, db.ForeignKey('loans.id'), nullable=False)
    payment_date = db.Column(db.Date, nullable=False, default=date.today)
    amount = db.Column(db.Float, nullable=False)
    notes = db.Column(db.Text)
    created_by = db.Column(db.Integer, db.ForeignKey('users.id'))
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    updated_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    sync_uuid = db.Column(db.String(36), unique=True, index=True)
    pending_push = db.Column(db.Boolean, default=False)
    last_modified_site = db.Column(db.String(50))
    deleted_at = db.Column(db.DateTime, nullable=True)
    last_synced_updated_at = db.Column(db.DateTime, nullable=True)
    server_touched_at = db.Column(db.DateTime, nullable=True)


class Payable(db.Model):
    """Accounts payable — amounts owed to suppliers OUTSIDE the fuel/maintenance
    logs (e.g. an insurance premium invoice), tracked on an accrual basis:
    the expense is recognized when the payable is created, not when it's paid."""
    __tablename__ = 'payables'
    id = db.Column(db.Integer, primary_key=True)
    supplier_name = db.Column(db.String(100), nullable=False)
    description = db.Column(db.Text)
    amount = db.Column(db.Float, nullable=False)
    invoice_date = db.Column(db.Date, nullable=False, default=date.today)
    due_date = db.Column(db.Date)
    status = db.Column(db.String(20), default='unpaid')
    paid_date = db.Column(db.Date)
    created_by = db.Column(db.Integer, db.ForeignKey('users.id'))
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    updated_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    sync_uuid = db.Column(db.String(36), unique=True, index=True)
    pending_push = db.Column(db.Boolean, default=False)
    last_modified_site = db.Column(db.String(50))
    deleted_at = db.Column(db.DateTime, nullable=True)
    last_synced_updated_at = db.Column(db.DateTime, nullable=True)
    server_touched_at = db.Column(db.DateTime, nullable=True)


class Receivable(db.Model):
    """Accounts receivable — revenue earned but not yet collected (e.g. an
    invoiced corporate/charter client), tracked on an accrual basis: revenue
    is recognized when the receivable is created, not when it's collected."""
    __tablename__ = 'receivables'
    id = db.Column(db.Integer, primary_key=True)
    client_name = db.Column(db.String(100), nullable=False)
    description = db.Column(db.Text)
    amount = db.Column(db.Float, nullable=False)
    invoice_date = db.Column(db.Date, nullable=False, default=date.today)
    due_date = db.Column(db.Date)
    status = db.Column(db.String(20), default='outstanding')
    collected_date = db.Column(db.Date)
    created_by = db.Column(db.Integer, db.ForeignKey('users.id'))
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    updated_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    sync_uuid = db.Column(db.String(36), unique=True, index=True)
    pending_push = db.Column(db.Boolean, default=False)
    last_modified_site = db.Column(db.String(50))
    deleted_at = db.Column(db.DateTime, nullable=True)
    last_synced_updated_at = db.Column(db.DateTime, nullable=True)
    server_touched_at = db.Column(db.DateTime, nullable=True)


class CommissionPayment(db.Model):
    """An actual cash payout of driver/conductor commission. Kept separate
    from the accrued commission figure (computed live from revenue, as the
    payroll report already does) so the two can be reconciled — accrued
    minus paid is the outstanding commission liability."""
    __tablename__ = 'commission_payments'
    __table_args__ = (
        db.Index('ix_commission_payments_driver_date', 'driver_id', 'payment_date'),
        db.Index('ix_commission_payments_date', 'payment_date'),
    )
    id = db.Column(db.Integer, primary_key=True)
    driver_id = db.Column(db.Integer, db.ForeignKey('drivers.id'), nullable=False)
    payment_date = db.Column(db.Date, nullable=False, default=date.today)
    amount = db.Column(db.Float, nullable=False)
    period_start = db.Column(db.Date)
    period_end = db.Column(db.Date)
    method = db.Column(db.String(30))
    notes = db.Column(db.Text)
    created_by = db.Column(db.Integer, db.ForeignKey('users.id'))
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    updated_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    sync_uuid = db.Column(db.String(36), unique=True, index=True)
    pending_push = db.Column(db.Boolean, default=False)
    last_modified_site = db.Column(db.String(50))
    deleted_at = db.Column(db.DateTime, nullable=True)
    last_synced_updated_at = db.Column(db.DateTime, nullable=True)
    server_touched_at = db.Column(db.DateTime, nullable=True)

    driver = db.relationship('Driver')


class PayrollDeduction(db.Model):
    """A withholding against a crew member's accrued commission for a given
    payslip period — loan repayments, advances, fines, damages, etc. Kept
    separate from CommissionPayment (an actual cash payout, which reduces
    what's still owed on top of this) and from DailyLog.garnish (a daily
    revenue shortfall netted off before commission is even calculated) —
    this instead reduces the commission already accrued, feeding into
    compute_payroll_earnings' net_pay/outstanding figures and printed on
    the per-driver payslip (see driver_payslip_pdf). reason is free text,
    same as DailyLog.reason_for_shortfall, rather than a managed category
    list — the deduction's date determines which payslip period it lands
    in, same as CommissionPayment.payment_date."""
    __tablename__ = 'payroll_deductions'
    __table_args__ = (
        db.Index('ix_payroll_deductions_driver_date', 'driver_id', 'deduction_date'),
        db.Index('ix_payroll_deductions_date', 'deduction_date'),
    )
    id = db.Column(db.Integer, primary_key=True)
    driver_id = db.Column(db.Integer, db.ForeignKey('drivers.id'), nullable=False)
    deduction_date = db.Column(db.Date, nullable=False, default=date.today)
    amount = db.Column(db.Float, nullable=False)
    reason = db.Column(db.Text, nullable=False)
    created_by = db.Column(db.Integer, db.ForeignKey('users.id'))
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    updated_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    sync_uuid = db.Column(db.String(36), unique=True, index=True)
    pending_push = db.Column(db.Boolean, default=False)
    last_modified_site = db.Column(db.String(50))
    deleted_at = db.Column(db.DateTime, nullable=True)
    last_synced_updated_at = db.Column(db.DateTime, nullable=True)
    server_touched_at = db.Column(db.DateTime, nullable=True)

    driver = db.relationship('Driver')


class CapitalContribution(db.Model):
    __tablename__ = 'capital_contributions'
    id = db.Column(db.Integer, primary_key=True)
    contributor = db.Column(db.String(100), nullable=False)
    amount = db.Column(db.Float, nullable=False)
    contribution_date = db.Column(db.Date, nullable=False, default=date.today)
    notes = db.Column(db.Text)
    created_by = db.Column(db.Integer, db.ForeignKey('users.id'))
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    updated_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    sync_uuid = db.Column(db.String(36), unique=True, index=True)
    pending_push = db.Column(db.Boolean, default=False)
    last_modified_site = db.Column(db.String(50))
    deleted_at = db.Column(db.DateTime, nullable=True)
    last_synced_updated_at = db.Column(db.DateTime, nullable=True)
    server_touched_at = db.Column(db.DateTime, nullable=True)


class OwnerDrawing(db.Model):
    __tablename__ = 'owner_drawings'
    id = db.Column(db.Integer, primary_key=True)
    amount = db.Column(db.Float, nullable=False)
    drawing_date = db.Column(db.Date, nullable=False, default=date.today)
    notes = db.Column(db.Text)
    created_by = db.Column(db.Integer, db.ForeignKey('users.id'))
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    updated_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    sync_uuid = db.Column(db.String(36), unique=True, index=True)
    pending_push = db.Column(db.Boolean, default=False)
    last_modified_site = db.Column(db.String(50))
    deleted_at = db.Column(db.DateTime, nullable=True)
    last_synced_updated_at = db.Column(db.DateTime, nullable=True)
    server_touched_at = db.Column(db.DateTime, nullable=True)


class ExpenseCategory(db.Model):
    """Two-level expense classification: a top-level heading (parent_id is
    None, e.g. "Maintenance") with optional sub-headings under it (e.g.
    "Engine Oil"). An Expense can be tagged to either a heading or a
    sub-heading — the hierarchy is for organization, not enforcement."""
    __tablename__ = 'expense_categories'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(60), nullable=False)
    parent_id = db.Column(db.Integer, db.ForeignKey('expense_categories.id'), nullable=True)
    # Pre-fills the Amount field on Add Expense when this category is picked
    # (see expense_form.html) — for a recurring, roughly-fixed charge like a
    # monthly garage fee, this saves re-typing the same figure once per
    # vehicle every month. Never enforced server-side: the typed amount on
    # the expense itself is still what gets saved, so a one-off exception
    # is just as easy to enter as the usual figure.
    default_amount = db.Column(db.Float, nullable=True)
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    updated_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    sync_uuid = db.Column(db.String(36), unique=True, index=True)
    pending_push = db.Column(db.Boolean, default=False)
    last_modified_site = db.Column(db.String(50))
    deleted_at = db.Column(db.DateTime, nullable=True)
    last_synced_updated_at = db.Column(db.DateTime, nullable=True)
    server_touched_at = db.Column(db.DateTime, nullable=True)

    parent = db.relationship('ExpenseCategory', remote_side=[id], backref='children')

    @property
    def display_name(self):
        return f'{self.parent.name} — {self.name}' if self.parent else self.name


class Expense(db.Model):
    __tablename__ = 'expenses'
    __table_args__ = (
        db.Index('ix_expenses_vehicle_date', 'vehicle_id', 'expense_date'),
        db.Index('ix_expenses_date', 'expense_date'),
    )
    id = db.Column(db.Integer, primary_key=True)
    category_id = db.Column(db.Integer, db.ForeignKey('expense_categories.id'), nullable=False)
    vehicle_id = db.Column(db.Integer, db.ForeignKey('vehicles.id'), nullable=True)
    # Attributes the expense to a driver — typically a road/trip expense
    # incurred while away from base. Optional: most expenses (insurance,
    # rent, etc.) aren't tied to any one driver. This is also what the
    # fleet reconciliation reports (report_fleet_reconciliation/
    # report_fleet_consolidated) sum as "expenses incurred while away".
    driver_id = db.Column(db.Integer, db.ForeignKey('drivers.id'), nullable=True)
    expense_date = db.Column(db.Date, nullable=False, default=date.today)
    description = db.Column(db.Text)
    amount = db.Column(db.Float, nullable=False)
    created_by = db.Column(db.Integer, db.ForeignKey('users.id'))
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    updated_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    sync_uuid = db.Column(db.String(36), unique=True, index=True)
    pending_push = db.Column(db.Boolean, default=False)
    last_modified_site = db.Column(db.String(50))
    deleted_at = db.Column(db.DateTime, nullable=True)
    last_synced_updated_at = db.Column(db.DateTime, nullable=True)
    server_touched_at = db.Column(db.DateTime, nullable=True)

    category = db.relationship('ExpenseCategory', backref='expenses')
    vehicle = db.relationship('Vehicle')
    driver = db.relationship('Driver')


class Budget(db.Model):
    __tablename__ = 'budgets'
    id = db.Column(db.Integer, primary_key=True)
    category = db.Column(db.String(60), nullable=False)
    month = db.Column(db.Date, nullable=False)
    amount = db.Column(db.Float, nullable=False)
    created_by = db.Column(db.Integer, db.ForeignKey('users.id'))
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    updated_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    sync_uuid = db.Column(db.String(36), unique=True, index=True)
    pending_push = db.Column(db.Boolean, default=False)
    last_modified_site = db.Column(db.String(50))
    deleted_at = db.Column(db.DateTime, nullable=True)
    last_synced_updated_at = db.Column(db.DateTime, nullable=True)
    server_touched_at = db.Column(db.DateTime, nullable=True)


class DriverDeposit(db.Model):
    """Cash banked for a given date, for reconciliation against the fleet's
    own net income (DailyLog.gross_revenue minus driver-attributed Expense
    rows — see report_fleet_reconciliation/report_fleet_consolidated).
    vehicle_id is optional and only used to attribute the deposit to a
    vehicle for the consolidated-by-vehicle view; a driver who drove more
    than one vehicle that day can leave it blank. driver_id is also optional:
    a row with driver_id (and vehicle_id) left blank is a fleet-wide total —
    one combined cash-deposited figure for all vehicles that date, for
    reconciliations that don't track deposits per driver — and is folded
    into the totals/variance on the reconciliation reports without being
    attributed to any single driver or vehicle row."""
    __tablename__ = 'driver_deposits'
    __table_args__ = (
        db.Index('ix_driver_deposits_date', 'deposit_date'),
    )
    id = db.Column(db.Integer, primary_key=True)
    driver_id = db.Column(db.Integer, db.ForeignKey('drivers.id'), nullable=True)
    vehicle_id = db.Column(db.Integer, db.ForeignKey('vehicles.id'), nullable=True)
    deposit_date = db.Column(db.Date, nullable=False, default=date.today)
    amount = db.Column(db.Float, nullable=False)
    notes = db.Column(db.Text)
    created_by = db.Column(db.Integer, db.ForeignKey('users.id'))
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    updated_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    sync_uuid = db.Column(db.String(36), unique=True, index=True)
    pending_push = db.Column(db.Boolean, default=False)
    last_modified_site = db.Column(db.String(50))
    deleted_at = db.Column(db.DateTime, nullable=True)
    last_synced_updated_at = db.Column(db.DateTime, nullable=True)
    server_touched_at = db.Column(db.DateTime, nullable=True)

    driver = db.relationship('Driver', backref='deposits')
    vehicle = db.relationship('Vehicle')


class FranchiseDailyIncome(db.Model):
    """One reconciliation record per calendar date per franchise vehicle for
    daily franchise fee collections — a date can have several entries, one
    per vehicle, since each vehicle's standalone fee is reconciled
    independently. vehicle_id may be null for a whole-franchise entry not
    attributable to one vehicle (e.g. a historical combined-total figure).
    Cash-in-hand, total expenditure, and variance are derived rather than
    stored, so they can never drift out of sync with the entered figures."""
    __tablename__ = 'franchise_daily_income'
    __table_args__ = (db.UniqueConstraint('entry_date', 'vehicle_id', name='uq_franchise_daily_income_date_vehicle'),)
    id = db.Column(db.Integer, primary_key=True)
    entry_date = db.Column(db.Date, nullable=False)
    vehicle_id = db.Column(db.Integer, db.ForeignKey('franchise_vehicles.id'), nullable=True)

    income = db.Column(db.Float, nullable=False, default=0)
    exp_traffic_fines = db.Column(db.Float, nullable=False, default=0)
    exp_facilitation_fees = db.Column(db.Float, nullable=False, default=0)
    exp_workshop = db.Column(db.Float, nullable=False, default=0)
    exp_wages = db.Column(db.Float, nullable=False, default=0)
    other_expenditure = db.Column(db.Float, nullable=False, default=0)

    deposited = db.Column(db.Float, nullable=False, default=0)
    description = db.Column(db.Text)
    created_by = db.Column(db.Integer, db.ForeignKey('users.id'))
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    updated_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    sync_uuid = db.Column(db.String(36), unique=True, index=True)
    pending_push = db.Column(db.Boolean, default=False)
    last_modified_site = db.Column(db.String(50))
    deleted_at = db.Column(db.DateTime, nullable=True)
    last_synced_updated_at = db.Column(db.DateTime, nullable=True)
    server_touched_at = db.Column(db.DateTime, nullable=True)

    @property
    def total_expenditure(self):
        return (self.exp_traffic_fines or 0) + (self.exp_facilitation_fees or 0) \
            + (self.exp_workshop or 0) + (self.exp_wages or 0) + (self.other_expenditure or 0)

    @property
    def cash_in_hand(self):
        return (self.income or 0) - self.total_expenditure

    @property
    def variance(self):
        return (self.deposited or 0) - self.cash_in_hand

    vehicle = db.relationship('FranchiseVehicle')
    creator = db.relationship('User', foreign_keys=[created_by])


class FranchiseWeeklyIncome(db.Model):
    """Same shape as FranchiseDailyIncome, but for the separate weekly
    franchise fee — kept as its own entity rather than extra columns on the
    same date row, since a vehicle's daily and weekly dues are independent
    obligations, not two figures on one entry. week_start holds the actual
    date the entry was recorded against (not normalized to that week's
    Monday) — reports that need to bucket entries by calendar week floor it
    to Monday themselves at read time instead. vehicle_id may be null for a
    whole-franchise entry, same as above.

    Unlike FranchiseDailyIncome, there is no (week_start, vehicle_id)
    uniqueness — a vehicle can log more than one entry for the same week
    (e.g. several partial payments through the week), so a repeat Add isn't
    blocked the way a repeat daily entry still is."""
    __tablename__ = 'franchise_weekly_income'
    __table_args__ = (
        db.Index('ix_franchise_weekly_income_vehicle_week', 'vehicle_id', 'week_start'),
        db.Index('ix_franchise_weekly_income_week', 'week_start'),
    )
    id = db.Column(db.Integer, primary_key=True)
    week_start = db.Column(db.Date, nullable=False)
    vehicle_id = db.Column(db.Integer, db.ForeignKey('franchise_vehicles.id'), nullable=True)

    income = db.Column(db.Float, nullable=False, default=0)
    exp_traffic_fines = db.Column(db.Float, nullable=False, default=0)
    exp_facilitation_fees = db.Column(db.Float, nullable=False, default=0)
    exp_workshop = db.Column(db.Float, nullable=False, default=0)
    exp_wages = db.Column(db.Float, nullable=False, default=0)
    other_expenditure = db.Column(db.Float, nullable=False, default=0)

    deposited = db.Column(db.Float, nullable=False, default=0)
    description = db.Column(db.Text)
    created_by = db.Column(db.Integer, db.ForeignKey('users.id'))
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    updated_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    sync_uuid = db.Column(db.String(36), unique=True, index=True)
    pending_push = db.Column(db.Boolean, default=False)
    last_modified_site = db.Column(db.String(50))
    deleted_at = db.Column(db.DateTime, nullable=True)
    last_synced_updated_at = db.Column(db.DateTime, nullable=True)
    server_touched_at = db.Column(db.DateTime, nullable=True)

    @property
    def total_expenditure(self):
        return (self.exp_traffic_fines or 0) + (self.exp_facilitation_fees or 0) \
            + (self.exp_workshop or 0) + (self.exp_wages or 0) + (self.other_expenditure or 0)

    @property
    def cash_in_hand(self):
        return (self.income or 0) - self.total_expenditure

    @property
    def variance(self):
        return (self.deposited or 0) - self.cash_in_hand

    vehicle = db.relationship('FranchiseVehicle')
    creator = db.relationship('User', foreign_keys=[created_by])


class FranchiseVehicle(db.Model):
    """A third-party vehicle paying to operate under the franchise — distinct
    from the company's own fleet (Vehicle above), which is operated directly
    rather than franchised out."""
    __tablename__ = 'franchise_vehicles'
    id = db.Column(db.Integer, primary_key=True)
    number_plate = db.Column(db.String(20), nullable=False, unique=True)
    franchisee_name = db.Column(db.String(100), nullable=False)
    status = db.Column(db.String(20), default='active')
    # The negotiated fee this vehicle owes per collection cycle — kept on the
    # vehicle rather than derived, since it's an agreed figure, not something
    # computed from collections. Separate daily/weekly columns (both nullable)
    # since the same vehicle can owe a daily due, a weekly fee, or both — see
    # FranchiseCollection.frequency.
    daily_fee = db.Column(db.Float, nullable=True)
    weekly_fee = db.Column(db.Float, nullable=True)
    # Running arrears balance — tracked manually rather than computed,
    # since "how much is owed" depends on a payment schedule this system
    # doesn't model (which days were actually due).
    amount_owed = db.Column(db.Float, nullable=False, default=0)
    notes = db.Column(db.Text)
    # True when this row was created via the quick "+ New Vehicle" inline
    # add on the Daily/Weekly Income pages rather than the full Franchise
    # Vehicles form — that path only collects plate + franchisee name, so
    # fees/status/notes are still unset and an admin should follow up.
    # Cleared by the "Mark Reviewed" action (franchise_vehicle_review).
    pending_review = db.Column(db.Boolean, nullable=False, default=False)
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    updated_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    sync_uuid = db.Column(db.String(36), unique=True, index=True)
    pending_push = db.Column(db.Boolean, default=False)
    last_modified_site = db.Column(db.String(50))
    deleted_at = db.Column(db.DateTime, nullable=True)
    last_synced_updated_at = db.Column(db.DateTime, nullable=True)
    server_touched_at = db.Column(db.DateTime, nullable=True)

    collections = db.relationship('FranchiseCollection', backref='vehicle', lazy=True,
                                  order_by='FranchiseCollection.entry_date.desc()')

    @property
    def total_income(self):
        """Sum of this vehicle's own Daily + Weekly Income entries — income
        is always recorded per vehicle, unlike expenditure, which is shared
        across the whole franchise (see FranchiseDailyIncome/
        FranchiseWeeklyIncome), so there's no per-vehicle expense/net figure
        to show alongside it."""
        daily = sum(e.income for e in FranchiseDailyIncome.query.filter_by(vehicle_id=self.id).all())
        weekly = sum(e.income for e in FranchiseWeeklyIncome.query.filter_by(vehicle_id=self.id).all())
        return daily + weekly


class FranchiseCollection(db.Model):
    """What one franchise vehicle paid on one date — the per-vehicle detail
    behind a day's franchise income total. Kept separate from the
    daily/weekly income entities (rather than forcing a sum-to-match)
    because in practice the two are reconciled by a person, not guaranteed
    to tie out line-for-line.

    frequency lives here rather than on FranchiseVehicle because the same
    vehicle can owe both a daily due and a separate weekly franchise fee —
    it's a property of the payment, not a fixed plan per vehicle.

    expense is that same vehicle's own cost for that day/week (fuel,
    fines, etc. it bears itself) — kept alongside amount so each
    franchisee's collection entry nets out to what it actually owes,
    without touching the company-wide franchise income expense figures."""
    __tablename__ = 'franchise_collections'
    id = db.Column(db.Integer, primary_key=True)
    vehicle_id = db.Column(db.Integer, db.ForeignKey('franchise_vehicles.id'), nullable=False)
    entry_date = db.Column(db.Date, nullable=False)
    frequency = db.Column(db.String(10), nullable=False, default='daily')  # 'daily' or 'weekly'
    amount = db.Column(db.Float, nullable=False)
    expense = db.Column(db.Float, nullable=False, default=0)
    notes = db.Column(db.Text)
    created_by = db.Column(db.Integer, db.ForeignKey('users.id'))
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    updated_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    sync_uuid = db.Column(db.String(36), unique=True, index=True)
    pending_push = db.Column(db.Boolean, default=False)
    last_modified_site = db.Column(db.String(50))
    deleted_at = db.Column(db.DateTime, nullable=True)
    last_synced_updated_at = db.Column(db.DateTime, nullable=True)
    server_touched_at = db.Column(db.DateTime, nullable=True)

    @property
    def net(self):
        return self.amount - (self.expense or 0)


class FranchiseExpenseCategory(db.Model):
    """Sub-heading under the Franchise module's single 'Operational Expenses'
    main heading (see FranchiseOperationalExpense) — e.g. Rent, Admin
    Salaries, Utilities. Flat list, no further nesting needed since there's
    only the one heading, unlike ExpenseCategory's two-level scheme."""
    __tablename__ = 'franchise_expense_categories'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(60), nullable=False, unique=True)
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    updated_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    sync_uuid = db.Column(db.String(36), unique=True, index=True)
    pending_push = db.Column(db.Boolean, default=False)
    last_modified_site = db.Column(db.String(50))
    deleted_at = db.Column(db.DateTime, nullable=True)
    last_synced_updated_at = db.Column(db.DateTime, nullable=True)
    server_touched_at = db.Column(db.DateTime, nullable=True)


class FranchiseOperationalExpense(db.Model):
    """Franchise-wide operating costs (rent, admin salaries, etc.) that
    aren't tied to one vehicle's cash reconciliation — unlike the exp_*
    columns on FranchiseDailyIncome/FranchiseWeeklyIncome, which are cash a
    specific vehicle's collection actually paid out that day/week. These
    roll up only into the Consolidated P&L (report_franchise_consolidated)
    as a separate 'Operational Expenses' section that reduces Net Profit —
    they deliberately don't touch the Daily/Weekly Income lists or the
    Cash Reconciliation (deposited vs. variance) figures, since no vehicle
    handled this cash."""
    __tablename__ = 'franchise_operational_expenses'
    id = db.Column(db.Integer, primary_key=True)
    expense_date = db.Column(db.Date, nullable=False, default=date.today)
    category_id = db.Column(db.Integer, db.ForeignKey('franchise_expense_categories.id'), nullable=False)
    amount = db.Column(db.Float, nullable=False)
    description = db.Column(db.Text)
    created_by = db.Column(db.Integer, db.ForeignKey('users.id'))
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    updated_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    sync_uuid = db.Column(db.String(36), unique=True, index=True)
    pending_push = db.Column(db.Boolean, default=False)
    last_modified_site = db.Column(db.String(50))
    deleted_at = db.Column(db.DateTime, nullable=True)
    last_synced_updated_at = db.Column(db.DateTime, nullable=True)
    server_touched_at = db.Column(db.DateTime, nullable=True)

    category = db.relationship('FranchiseExpenseCategory')


class FranchiseSuspenseResolution(db.Model):
    """Tracks manual clearance of a daily/weekly franchise reconciliation
    variance (see report_franchise_reconciliation / franchise_suspense_
    account) — the variance itself is never stored here, it's always
    recomputed live from FranchiseDailyIncome/FranchiseWeeklyIncome via
    _group_income_by_period, so this table can never drift out of sync
    with the source entries. A (source_type, period_date) pair with no row
    here is still open/unreconciled. Deliberately has no offline-sync
    fields, like AuditLog/ImportBatch above — this is a Central-side admin
    control record, not a synced business transaction."""
    __tablename__ = 'franchise_suspense_resolutions'
    __table_args__ = (db.UniqueConstraint('source_type', 'period_date',
                                          name='uq_franchise_suspense_source_period'),)
    id = db.Column(db.Integer, primary_key=True)
    source_type = db.Column(db.String(10), nullable=False)  # 'daily' | 'weekly'
    period_date = db.Column(db.Date, nullable=False)        # entry_date or week_start
    resolved_amount = db.Column(db.Float, nullable=False)   # variance snapshot at resolve time
    notes = db.Column(db.Text, nullable=False)
    resolved_by = db.Column(db.Integer, db.ForeignKey('users.id'))
    resolved_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))

    resolver = db.relationship('User')


class MaintenanceSchedule(db.Model):
    __tablename__ = 'maintenance_schedules'
    id = db.Column(db.Integer, primary_key=True)
    vehicle_id = db.Column(db.Integer, db.ForeignKey('vehicles.id'), nullable=False)
    description = db.Column(db.String(150), nullable=False)
    interval_days = db.Column(db.Integer)
    interval_km = db.Column(db.Float)
    last_done_date = db.Column(db.Date)
    last_done_odometer = db.Column(db.Float)
    next_due_date = db.Column(db.Date)
    next_due_odometer = db.Column(db.Float)
    status = db.Column(db.String(20), default='active')
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    updated_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    sync_uuid = db.Column(db.String(36), unique=True, index=True)
    pending_push = db.Column(db.Boolean, default=False)
    last_modified_site = db.Column(db.String(50))
    deleted_at = db.Column(db.DateTime, nullable=True)
    last_synced_updated_at = db.Column(db.DateTime, nullable=True)
    server_touched_at = db.Column(db.DateTime, nullable=True)

    vehicle = db.relationship('Vehicle')


# ─────────────────────────────────────────────────────────────
# Spares Store: parts inventory, purchases & marked-up sales
# ─────────────────────────────────────────────────────────────
class SparePart(db.Model):
    __tablename__ = 'spare_parts'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(120), nullable=False)
    part_number = db.Column(db.String(50))
    unit = db.Column(db.String(20), nullable=False, default='pc')
    # Running weighted-average unit cost, updated on every purchase — not
    # editable directly, it's a derived stock-valuation figure.
    cost_price = db.Column(db.Float, nullable=False, default=0.0)
    markup_percent = db.Column(db.Float, nullable=False, default=0.0)
    quantity_on_hand = db.Column(db.Float, nullable=False, default=0)
    reorder_level = db.Column(db.Integer, nullable=False, default=0)
    status = db.Column(db.String(20), default='active')
    notes = db.Column(db.Text)
    created_by = db.Column(db.Integer, db.ForeignKey('users.id'))
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    updated_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    sync_uuid = db.Column(db.String(36), unique=True, index=True)
    pending_push = db.Column(db.Boolean, default=False)
    last_modified_site = db.Column(db.String(50))
    deleted_at = db.Column(db.DateTime, nullable=True)
    last_synced_updated_at = db.Column(db.DateTime, nullable=True)
    server_touched_at = db.Column(db.DateTime, nullable=True)

    purchases = db.relationship('StorePurchase', backref='part', lazy=True,
                                cascade='all, delete-orphan')
    sales = db.relationship('StoreSale', backref='part', lazy=True,
                            cascade='all, delete-orphan')
    adjustments = db.relationship('StoreAdjustment', backref='part', lazy=True,
                                  cascade='all, delete-orphan')

    @property
    def selling_price(self):
        return round(self.cost_price * (1 + self.markup_percent / 100), 2)

    @property
    def stock_value(self):
        return self.cost_price * self.quantity_on_hand

    @property
    def low_stock(self):
        return self.quantity_on_hand <= self.reorder_level


class StorePurchase(db.Model):
    """A restock — adds quantity to the part and rolls the new cost into
    the part's weighted-average cost_price."""
    __tablename__ = 'store_purchases'
    __table_args__ = (
        db.Index('ix_store_purchases_date', 'purchase_date'),
    )
    id = db.Column(db.Integer, primary_key=True)
    part_id = db.Column(db.Integer, db.ForeignKey('spare_parts.id'), nullable=False)
    purchase_date = db.Column(db.Date, nullable=False, default=date.today)
    quantity = db.Column(db.Float, nullable=False)
    unit_cost = db.Column(db.Float, nullable=False)
    total_cost = db.Column(db.Float, nullable=False)
    supplier = db.Column(db.String(100))
    notes = db.Column(db.Text)
    created_by = db.Column(db.Integer, db.ForeignKey('users.id'))
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    updated_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    sync_uuid = db.Column(db.String(36), unique=True, index=True)
    pending_push = db.Column(db.Boolean, default=False)
    last_modified_site = db.Column(db.String(50))
    deleted_at = db.Column(db.DateTime, nullable=True)
    last_synced_updated_at = db.Column(db.DateTime, nullable=True)
    server_touched_at = db.Column(db.DateTime, nullable=True)

    creator = db.relationship('User', foreign_keys=[created_by])


class StoreSale(db.Model):
    """A sale at the part's marked-up selling price. Cost and price are
    snapshotted at sale time so historical profit doesn't shift when the
    part's cost or markup later changes."""
    __tablename__ = 'store_sales'
    __table_args__ = (
        db.Index('ix_store_sales_vehicle_date', 'vehicle_id', 'sale_date'),
        db.Index('ix_store_sales_date', 'sale_date'),
    )
    id = db.Column(db.Integer, primary_key=True)
    part_id = db.Column(db.Integer, db.ForeignKey('spare_parts.id'), nullable=False)
    # Set when this sale is to one of the fleet's own vehicles rather than an
    # outside customer — the sale amount then also counts as an expense on
    # that vehicle's income statement (see vehicle_income_totals). Mutually
    # exclusive with customer_name in practice, not enforced at the DB level.
    vehicle_id = db.Column(db.Integer, db.ForeignKey('vehicles.id'), nullable=True)
    sale_date = db.Column(db.Date, nullable=False, default=date.today)
    quantity = db.Column(db.Float, nullable=False)
    unit_cost = db.Column(db.Float, nullable=False)
    unit_price = db.Column(db.Float, nullable=False)
    total_amount = db.Column(db.Float, nullable=False)
    customer_name = db.Column(db.String(100))
    notes = db.Column(db.Text)
    created_by = db.Column(db.Integer, db.ForeignKey('users.id'))
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    updated_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    sync_uuid = db.Column(db.String(36), unique=True, index=True)
    pending_push = db.Column(db.Boolean, default=False)
    last_modified_site = db.Column(db.String(50))
    deleted_at = db.Column(db.DateTime, nullable=True)
    last_synced_updated_at = db.Column(db.DateTime, nullable=True)
    server_touched_at = db.Column(db.DateTime, nullable=True)

    creator = db.relationship('User', foreign_keys=[created_by])
    vehicle = db.relationship('Vehicle')

    @property
    def profit(self):
        return (self.unit_price - self.unit_cost) * self.quantity

    @property
    def customer_display(self):
        return self.vehicle.registration if self.vehicle else (self.customer_name or None)


ADJUSTMENT_REASONS = [
    ('physical_count', 'Physical Stock Count'),
    ('damage', 'Damaged / Written Off'),
    ('loss', 'Lost'),
    ('theft', 'Theft'),
    ('found', 'Found Surplus'),
    ('revaluation', 'Cost Revaluation'),
    ('other', 'Other'),
]
ADJUSTMENT_REASON_LABELS = dict(ADJUSTMENT_REASONS)


class StoreAdjustment(db.Model):
    """A manual correction to a part's quantity on hand and/or weighted-average
    unit cost — e.g. reconciling a physical stock count, writing off
    damage/loss/theft, recording found surplus, or revaluing stock. Unlike a
    purchase/sale this has no supplier/customer and no direct revenue/cost-of-
    sale impact; it just corrects the book figures to match reality, with the
    before/after snapshotted so the correction is auditable rather than
    silently overwriting history."""
    __tablename__ = 'store_adjustments'
    __table_args__ = (
        db.Index('ix_store_adjustments_date', 'adjustment_date'),
    )
    id = db.Column(db.Integer, primary_key=True)
    part_id = db.Column(db.Integer, db.ForeignKey('spare_parts.id'), nullable=False)
    adjustment_date = db.Column(db.Date, nullable=False, default=date.today)
    quantity_before = db.Column(db.Float, nullable=False)
    quantity_after = db.Column(db.Float, nullable=False)
    cost_price_before = db.Column(db.Float, nullable=False)
    cost_price_after = db.Column(db.Float, nullable=False)
    reason = db.Column(db.String(20), nullable=False, default='other')
    notes = db.Column(db.Text)
    created_by = db.Column(db.Integer, db.ForeignKey('users.id'))
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    updated_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    sync_uuid = db.Column(db.String(36), unique=True, index=True)
    pending_push = db.Column(db.Boolean, default=False)
    last_modified_site = db.Column(db.String(50))
    deleted_at = db.Column(db.DateTime, nullable=True)
    last_synced_updated_at = db.Column(db.DateTime, nullable=True)
    server_touched_at = db.Column(db.DateTime, nullable=True)

    creator = db.relationship('User', foreign_keys=[created_by])

    @property
    def quantity_delta(self):
        return self.quantity_after - self.quantity_before

    @property
    def reason_label(self):
        return ADJUSTMENT_REASON_LABELS.get(self.reason, self.reason)


# ─────────────────────────────────────────────────────────────
# Auth helpers
# ─────────────────────────────────────────────────────────────
@login_manager.user_loader
def load_user(user_id):
    return db.session.get(User, int(user_id))


PERMISSIONS = {
    'dashboard':    'Dashboard — overview stats, charts & recent activity',
    'crew_portal':  'Crew Portal — log daily income & view team performance leaderboard',
    'vehicles':     'Vehicles — view, add & edit vehicles',
    'drivers':      'Crew — view, add & edit drivers/conductors',
    'routes':       'Routes — view, add & edit routes',
    'daily_logs':   'Daily Transactions — view, record & edit vehicle transactions',
    'fuel_logs':    'Fuel Logs — view & record fuel entries',
    'maintenance':  'Maintenance — view & record maintenance logs',
    'reports':      'Finance & Reports — income statement, payroll, CSV exports',
    'compliance':   'Compliance — vehicle documents & expiry tracker',
    'finance':      'Finance Ledger — loans, payables, receivables, capital, expenses, budget',
    'store':        'Spares Store — parts inventory, purchases & marked-up sales',
    'franchise':    'Franchise Income — collection reconciliation entry & franchise P&L statements',
    'franchise_entry': 'Franchise Collections (Clerk) — enter income, register vehicles, and confirm vehicle payments; can see all clerks\' recorded income but not expenses, reports, or other franchise data',
}

PERMISSION_REDIRECTS = [
    ('dashboard',   'dashboard'),
    ('crew_portal', 'crew_leaderboard'),
    ('vehicles',    'vehicles'),
    ('drivers',     'drivers'),
    ('routes',      'routes_list'),
    ('daily_logs',  'driver_ledger'),
    ('fuel_logs',   'fuel_logs'),
    ('maintenance', 'maintenance_logs'),
    ('reports',     'report_income'),
    ('compliance',  'compliance'),
    ('finance',     'loans_list'),
    ('store',       'store_parts'),
    ('franchise',   'franchise_daily_income_list'),
    ('franchise_entry', 'franchise_my_collections'),
]


def first_permitted_url(user):
    if user.role == 'admin':
        return url_for('dashboard')
    for perm, endpoint in PERMISSION_REDIRECTS:
        if user.has_permission(perm):
            return url_for(endpoint)
    return url_for('no_access')


# Every main list/dashboard page worth having available offline even if
# this browser has never actually visited it — see api_precache_urls,
# which filters this down to what the current user can see and hands it
# to offline.js to proactively warm into the service worker's page cache
# in the background while online. Deliberately mirrors base.html's nav
# exactly (same (permission, endpoint) pairs, same set of pages) rather
# than a hand-picked subset, so a page added to the nav is automatically
# covered here too. Admin-only pages (User Management, Audit Log, Sync
# Sites, etc.) are left out on purpose — those are operator tooling, not
# the "whole business's data" a field user going offline actually needs.
PRECACHE_PAGES = [
    ('dashboard', 'dashboard'),
    ('crew_portal', 'driver_ledger'),
    ('crew_portal', 'crew_leaderboard'),
    ('vehicles', 'vehicles'),
    ('drivers', 'drivers'),
    ('drivers', 'driver_roster'),
    ('routes', 'routes_list'),
    ('fuel_logs', 'fuel_logs'),
    ('maintenance', 'maintenance_logs'),
    ('maintenance', 'maintenance_schedules'),
    ('reports', 'report_consolidated'),
    ('reports', 'report_income'),
    ('reports', 'report_cash_flow'),
    ('reports', 'report_payroll'),
    ('reports', 'report_shortfalls'),
    ('reports', 'report_financial_position'),
    ('reports', 'report_budget'),
    ('reports', 'report_fuel_efficiency'),
    ('reports', 'report_distance_travelled'),
    ('reports', 'report_route_profitability'),
    ('reports', 'report_vehicle_efficiency'),
    ('finance', 'loans_list'),
    ('finance', 'payables_list'),
    ('finance', 'receivables_list'),
    ('finance', 'capital_list'),
    ('finance', 'expenses_list'),
    ('franchise', 'franchise_daily_income_list'),
    ('franchise', 'franchise_weekly_income_list'),
    ('franchise', 'franchise_vehicles'),
    ('franchise', 'report_franchise_reconciliation'),
    ('franchise', 'report_franchise_dual_frequency'),
    ('franchise', 'report_franchise_weekly'),
    ('franchise', 'franchise_operational_expenses_list'),
    ('franchise', 'report_franchise_consolidated'),
    ('franchise_entry', 'franchise_my_collections'),
    ('franchise_entry', 'franchise_confirm_payments'),
    ('store', 'store_parts'),
    ('store', 'store_purchases'),
    ('store', 'store_sales'),
    ('store', 'store_movements'),
    ('store', 'store_adjustments'),
    ('store', 'store_trading_account'),
    ('compliance', 'compliance'),
]


def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not current_user.is_authenticated or current_user.role != 'admin':
            flash('Admin access required.', 'danger')
            return redirect(first_permitted_url(current_user))
        return f(*args, **kwargs)
    return decorated


def permission_required(perm):
    def decorator(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            if not current_user.has_permission(perm):
                flash('You do not have permission to access that section.', 'danger')
                return redirect(first_permitted_url(current_user))
            return f(*args, **kwargs)
        return decorated
    return decorator


def permission_required_any(*perms):
    def decorator(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            if not any(current_user.has_permission(p) for p in perms):
                flash('You do not have permission to access that section.', 'danger')
                return redirect(first_permitted_url(current_user))
            return f(*args, **kwargs)
        return decorated
    return decorator


def log_audit(action, table_name=None, record_id=None, description=None):
    entry = AuditLog(
        user_id=current_user.id if current_user.is_authenticated else None,
        action=action,
        table_name=table_name,
        record_id=record_id,
        description=description,
        ip_address=request.remote_addr,
    )
    db.session.add(entry)


def save_import_batch(target_type, filename, total_rows, imported, error_rows, created_records):
    """Record a committed file import as a structured ImportBatch (unlike
    log_audit's free-text trail, this carries the quarantined error rows and
    links to every record it created via ImportBatchRecord, so the import
    can be found later, its errors re-downloaded, and it can be reversed).
    Call right before db.session.commit() so the batch and the data it
    describes land in the same transaction."""
    batch = ImportBatch(
        target_type=target_type, filename=filename, total_rows=total_rows,
        rows_imported=imported, rows_failed=len(error_rows),
        error_rows=json.dumps(error_rows) if error_rows else None,
        created_by=current_user.id,
    )
    db.session.add(batch)
    db.session.flush()
    for table_name, record_id in created_records:
        db.session.add(ImportBatchRecord(batch_id=batch.id, target_table=table_name, record_id=record_id))
    return batch


def parse_date(s):
    try:
        return datetime.strptime(s, '%Y-%m-%d').date() if s else None
    except ValueError:
        raise ValueError(f'"{s}" is not a valid date (expected YYYY-MM-DD).')


def csv_export_response(filename, header, rows):
    """Build a CSV download response — shared by every /export route so each
    one only supplies its header row and data rows, not its own copy of the
    io.StringIO/csv.writer/Content-Disposition boilerplate."""
    out = io.StringIO()
    w = csv.writer(out)
    w.writerow(header)
    w.writerows(rows)
    out.seek(0)
    resp = make_response(out.getvalue())
    resp.headers['Content-Type'] = 'text/csv'
    resp.headers['Content-Disposition'] = f'attachment; filename={filename}'
    return resp


def _table_pdf_response(filename, title, subtitle, header, rows):
    """The PDF sibling of csv_export_response — same header+rows a caller
    already built for the CSV, turned into a PDF via the same _pdf_table
    look every report PDF uses (see _pdf_response/_pdf_table below), so a
    CSV export gains a PDF option without a separate query/computation to
    keep in sync. Landscape for wide tables so columns stay readable."""
    pagesize = landscape(A4) if len(header) > 6 else A4
    data = [header] + [[('' if c is None else str(c)) for c in row] for row in rows]
    elements = _pdf_section(title, subtitle, [_pdf_table(data, bold_last_row=False)])
    return _pdf_response(filename, elements, pagesize=pagesize)


def _detect_dayfirst(values):
    """Infer whether an uploaded file's slash/dash-separated dates are
    day-first (DD/MM/YYYY) or month-first (MM/DD/YYYY), by scanning for any
    value whose two leading numeric components can only be valid one way —
    e.g. 15/03/2026 must be day-first, since no month is 15. A file's dates
    all come from the same source/locale, so one determination covers every
    row in it; this is used instead of guessing per-row, which would silently
    swap day and month whenever both readings happen to be valid (03/04/2026
    could be 3 April or 4 March) and corrupt the date without ever raising an
    error. Returns True (dayfirst), False (monthfirst), or None if nothing in
    the sample disambiguates (every value was ISO, an Excel datetime, or
    genuinely ambiguous throughout) — callers should default to day-first."""
    for value in values:
        if isinstance(value, (datetime, date)):
            continue
        m = re.match(r'^(\d{1,2})[/-](\d{1,2})[/-]\d{4}$', str(value or '').strip())
        if not m:
            continue
        a, b = int(m.group(1)), int(m.group(2))
        if a > 12 and b <= 12:
            return True
        if b > 12 and a <= 12:
            return False
    return None


def parse_import_date(value, dayfirst=True):
    """Parse a date cell from an uploaded CSV/Excel row. Excel cells come
    through openpyxl as datetime objects already; CSV cells are plain
    strings, tried against the common formats a logbook might use.
    dayfirst picks which of the ambiguous DD/MM vs MM/DD slash/dash formats
    is tried first for a given value — callers importing a batch of rows
    should resolve this once per file via _detect_dayfirst (see there for
    why) rather than leaving every row to guess independently."""
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    if not s:
        raise ValueError('Date is required.')
    slash_dash = (['%d/%m/%Y', '%m/%d/%Y', '%d-%m-%Y', '%m-%d-%Y'] if dayfirst else
                  ['%m/%d/%Y', '%d/%m/%Y', '%m-%d-%Y', '%d-%m-%Y'])
    for fmt in ['%Y-%m-%d'] + slash_dash:
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    raise ValueError(f'"{s}" is not a recognized date (use YYYY-MM-DD).')


def parse_import_number(value, label):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip()
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        raise ValueError(f'{label} "{s}" is not a number.')


MAX_LEDGER_IMPORT_ROWS = 2000

# Canonical ledger import fields and the header synonyms an uploaded file might
# use for each — used to auto-map an arbitrary file's columns onto the shape
# the importer expects, instead of requiring an exact header match.
CANONICAL_LEDGER_FIELDS = [
    ('date', 'Date', ['date', 'log date', 'trip date', 'day']),
    ('driver', 'Driver', ['driver', 'driver name', 'crew', 'operator']),
    ('fare', 'Fare', ['fare', 'revenue', 'income', 'gross revenue', 'collection',
                      'collections', 'takings']),
    ('diesel_cost', 'Diesel (USD)', ['diesel cost', 'diesel usd', 'diesel amount',
                                     'petrol cost', 'petrol usd', 'petrol amount',
                                     'fuel cost', 'fuel usd', 'fuel amount',
                                     'diesel', 'petrol', 'fuel']),
    ('mileage', 'Mileage', ['mileage', 'odometer', 'odo', 'distance reading', 'km reading']),
]

# The raw-header key each canonical field must land on so the existing
# row-validation loop (which looks for these exact keys) keeps working unchanged.
CANONICAL_TO_ROW_KEY = {
    'date': 'date', 'driver': 'driver', 'fare': 'fare',
    'diesel_cost': 'diesel cost', 'mileage': 'mileage',
}

def _find_header_row(raw_rows, max_scan=10):
    """Locate the header row within the first few rows of an uploaded sheet,
    skipping any title/blank rows some logbooks put above the real columns
    (e.g. a "JULY 2026 DAILY TRANSACTIONS" banner row)."""
    date_synonyms = next(syns for key, _label, syns in CANONICAL_LEDGER_FIELDS if key == 'date')
    for idx, row in enumerate(raw_rows[:max_scan]):
        cells = [str(c).strip().lower() for c in row if c not in (None, '')]
        if len(cells) < 2:
            continue
        if any(c in date_synonyms or difflib.get_close_matches(c, date_synonyms, n=1, cutoff=0.75)
               for c in cells):
            return idx
    return 0


def _json_safe_cell(cell):
    # Excel date cells come through openpyxl as datetime/date objects,
    # which aren't JSON-serializable — the row data gets round-tripped
    # through a hidden form field, so normalize to the 'YYYY-MM-DD'
    # string parse_import_date already accepts.
    if isinstance(cell, datetime):
        return cell.date().isoformat()
    if isinstance(cell, date):
        return cell.isoformat()
    return cell


def _rows_from_raw(raw_rows):
    """Turn a sheet/CSV's raw rows into (headers, rows), locating the real
    header row (skipping any title/blank rows above it) and keying each row
    by its stripped, lowercased header."""
    header_idx = _find_header_row(raw_rows)
    headers = [str(h).strip().lower() if h is not None else '' for h in raw_rows[header_idx]]
    rows = [{h: _json_safe_cell(c) for h, c in zip(headers, r)}
            for r in raw_rows[header_idx + 1:] if any(c not in (None, '') for c in r)]
    return [h for h in headers if h], rows


def _select_sheet_for_vehicle(wb, vehicle):
    """Pick the right worksheet for a single-vehicle import. A single-sheet
    file is unambiguous, but a multi-sheet workbook (e.g. a fleet-wide master
    logbook) must NOT silently fall back to whichever sheet happened to be
    "active" when it was last saved — that would import a different vehicle's
    transactions under the one currently selected. Only proceed if a sheet's
    name matches the selected vehicle's registration; otherwise raise so the
    upload is rejected instead of misattributed."""
    if vehicle is None or len(wb.sheetnames) == 1:
        return wb.active
    target = _normalize_registration(vehicle.registration)
    for name in wb.sheetnames:
        if _normalize_registration(name) == target:
            return wb[name]
    raise ValueError(
        f'This workbook has {len(wb.sheetnames)} sheets and none is named "{vehicle.registration}" — '
        'to avoid importing the wrong vehicle\'s data, upload a file with just one sheet, or use '
        '"Import Fleet Workbook" below to import every vehicle\'s sheet at once.')


def read_uploaded_table(file, vehicle=None):
    """Parse an uploaded CSV/XLSX file into (headers, rows). Each row is a dict
    keyed by its stripped, lowercased header. Raises ValueError on an
    unsupported or unreadable file."""
    ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else ''
    try:
        if ext == 'csv':
            reader = csv.reader(io.StringIO(file.read().decode('utf-8-sig')))
            raw_rows = list(reader)
        elif ext in ('xlsx', 'xlsm'):
            wb = openpyxl.load_workbook(file, data_only=True)
            ws = _select_sheet_for_vehicle(wb, vehicle)
            raw_rows = [list(r) for r in ws.iter_rows(values_only=True)]
        else:
            raise ValueError('Unsupported file type — upload a .csv or .xlsx file.')
        if not raw_rows:
            raise ValueError('The file is empty.')
    except ValueError:
        raise
    except Exception:
        raise ValueError('Could not read that file. Make sure it is a valid CSV or Excel export.')

    headers, rows = _rows_from_raw(raw_rows)
    if len(rows) > MAX_LEDGER_IMPORT_ROWS:
        raise ValueError(f'That file has {len(rows)} rows — the importer previews at most '
                         f'{MAX_LEDGER_IMPORT_ROWS} at a time. Split it into smaller files.')
    return headers, rows


def read_uploaded_workbook_sheets(file):
    """Parse a multi-sheet XLSX/XLSM workbook into {sheet_name: (headers, rows)}
    for every sheet that has data — lets a fleet-wide workbook (one tab per
    vehicle, like a company's master logbook) be imported in a single pass.
    Raises ValueError on an unsupported or unreadable file."""
    ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else ''
    if ext not in ('xlsx', 'xlsm'):
        raise ValueError('Upload a .xlsx or .xlsm workbook — one sheet per vehicle.')
    try:
        wb = openpyxl.load_workbook(file, data_only=True)
    except Exception:
        raise ValueError('Could not read that file. Make sure it is a valid Excel workbook.')

    sheets = {}
    for ws in wb.worksheets:
        raw_rows = [list(r) for r in ws.iter_rows(values_only=True)]
        if not raw_rows:
            continue
        headers, rows = _rows_from_raw(raw_rows)
        if rows:
            sheets[ws.title] = (headers, rows[:MAX_LEDGER_IMPORT_ROWS])
    return sheets


def read_uploaded_workbook_raw_sheets(file):
    """Parse a multi-sheet XLSX/XLSM workbook into {sheet_name: raw_rows} —
    each sheet's untouched 2D grid (a list of row tuples), with no header
    detection or dict-keying applied. Unlike read_uploaded_workbook_sheets,
    which assumes a single flat table filling the whole sheet, this is for
    importers that need to locate one specific labeled block within a sheet
    that also holds other, unrelated tables side by side — see
    _find_franchise_reconciliation_block. Raises ValueError on an
    unsupported or unreadable file."""
    ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else ''
    if ext not in ('xlsx', 'xlsm'):
        raise ValueError('Upload a .xlsx or .xlsm workbook.')
    try:
        wb = openpyxl.load_workbook(file, data_only=True)
    except Exception:
        raise ValueError('Could not read that file. Make sure it is a valid Excel workbook.')

    sheets = {}
    for ws in wb.worksheets:
        raw_rows = [list(r) for r in ws.iter_rows(values_only=True)]
        if raw_rows:
            sheets[ws.title] = raw_rows
    return sheets


def _normalize_registration(name):
    return re.sub(r'\s+', ' ', str(name or '')).strip().upper()


def auto_map_columns(headers, fields=None):
    """Guess which uploaded header feeds each canonical field (ledger fields by
    default; pass a different fields list for another import shape), by exact
    synonym match first and fuzzy match second. Returns {field_key: header_or_None}."""
    fields = fields if fields is not None else CANONICAL_LEDGER_FIELDS
    mapping = {}
    claimed = set()
    for field_key, _label, synonyms in fields:
        match = next((h for h in headers if h in synonyms and h not in claimed), None)
        if not match:
            candidates = [h for h in headers if h not in claimed]
            # word-boundary substring match first, e.g. "daily fare" contains "fare"
            match = next((h for h in candidates
                          if any(re.search(rf'\b{re.escape(s)}\b', h) for s in synonyms)), None)
        if not match:
            # cutoff 0.72: high enough to reject coincidental look-alikes between
            # unrelated words (e.g. "registration" vs "description", "owner" vs
            # "other" both score ~0.6 by pure chance) while still catching genuine
            # misspellings of the synonym itself (e.g. "Vehcle" scores ~0.92).
            close = difflib.get_close_matches(synonyms[0], candidates, n=1, cutoff=0.72)
            if not close:
                for syn in synonyms[1:]:
                    close = difflib.get_close_matches(syn, candidates, n=1, cutoff=0.72)
                    if close:
                        break
            match = close[0] if close else None
        mapping[field_key] = match
        if match:
            claimed.add(match)
    return mapping


def apply_column_mapping(headers, raw_rows, mapping, row_key_map=None):
    """Rebuild each raw row (keyed by uploaded header) into the row shape an
    import loop expects (keyed by canonical field name), per `mapping`
    ({field_key: header_or_None}). row_key_map defaults to the ledger's."""
    row_key_map = row_key_map if row_key_map is not None else CANONICAL_TO_ROW_KEY
    mapped_rows = []
    for raw in raw_rows:
        row = {}
        for field_key, header in mapping.items():
            row_key = row_key_map[field_key]
            row[row_key] = raw.get(header) if header else None
        mapped_rows.append(row)
    return mapped_rows


def import_ledger_rows(file_rows, vehicle, auto_register_drivers=False):
    """Validate and persist already-mapped ledger rows (keyed by 'date',
    'driver', 'fare', 'diesel cost', 'mileage') as DailyLog/FuelLog entries
    for `vehicle`. Returns (imported_count, error_messages, created_driver_names);
    does not commit — the caller decides when to commit/rollback.

    If auto_register_drivers is True, a fare row naming a driver that isn't
    on file registers a new active Driver instead of erroring — used by the
    fleet-wide bulk import, where dozens of real driver names showing up for
    the first time is the expected case, not a data-entry mistake worth
    blocking on. The single-vehicle import leaves this off, since there a
    human already reviewed the file and can add the driver deliberately.

    created_records is a list of (target_table, record_id) tuples for every
    DailyLog/FuelLog row created — used to build an ImportBatchRecord trail
    so the import can be identified and reversed later. error_rows mirrors
    errors but keeps the original row data plus a 'System_Error' column, for
    a downloadable quarantine CSV."""
    driver_by_name = {d.name.strip().lower(): d for d in
                       Driver.query.filter_by(role='driver', status='active').all()}
    created_drivers = []
    created_records = []
    last_driver = None  # carries forward across blank-driver rows, see below
    dayfirst = _detect_dayfirst(r.get('date') for r in file_rows)
    if dayfirst is None:
        dayfirst = True

    imported = 0
    errors = []
    error_rows = []
    for i, raw_row in enumerate(file_rows, start=2):  # row 1 is the header
        row = {(k or '').strip().lower(): v for k, v in raw_row.items()}
        try:
            date_raw = row.get('date')
            if date_raw in (None, ''):
                continue
            log_date = parse_import_date(date_raw, dayfirst)

            driver_name = str(row.get('driver') or '').strip()
            driver = driver_by_name.get(driver_name.lower()) if driver_name else None

            cost_key = next((k for k in ('diesel cost', 'petrol cost', 'fuel cost') if k in row), 'diesel cost')
            fare = parse_import_number(row.get('fare'), 'Fare')
            diesel_cost = parse_import_number(row.get(cost_key), 'Diesel (USD)')
            mileage = parse_import_number(row.get('mileage'), 'Mileage')

            # Real-world logbooks often put a note ("GARAGE", "ARRESTED", "DRIVER
            # NOT FEELING WELL") in the driver column on no-fare days — only
            # require a recognized driver when there's actual revenue to attribute.
            if fare and not driver:
                if not driver_name and last_driver is not None:
                    # Some logbooks only write the driver's name once and leave
                    # it blank on the following days to mean "same as above".
                    driver = last_driver
                elif driver_name and auto_register_drivers:
                    driver = Driver(name=driver_name, role='driver', status='active')
                    db.session.add(driver)
                    db.session.flush()
                    driver_by_name[driver_name.lower()] = driver
                    created_drivers.append(driver_name)
                else:
                    raise ValueError(f'unknown driver "{driver_name}".' if driver_name else 'fare needs a driver name.')
            if fare is None and diesel_cost is None and mileage is None:
                continue

            if driver_name and driver:
                last_driver = driver

            if fare:
                conductor = resolve_conductor(driver, vehicle.id)
                log = DailyLog(
                    vehicle_id=vehicle.id, driver_id=driver.id,
                    conductor_id=conductor.id if conductor else None,
                    log_date=log_date, gross_revenue=fare, created_by=current_user.id,
                )
                db.session.add(log)
                db.session.flush()
                created_records.append(('daily_logs', log.id))
            if diesel_cost is not None:
                price = fuel_price_for(vehicle.fuel_type)
                fuel = FuelLog(
                    vehicle_id=vehicle.id, log_date=log_date,
                    liters=round(diesel_cost / price, 2) if price else 0,
                    cost_per_liter=price or 0.0,
                    total_cost=diesel_cost, odometer=mileage, created_by=current_user.id,
                )
                db.session.add(fuel)
                db.session.flush()
                created_records.append(('fuel_logs', fuel.id))
            elif mileage is not None:
                fuel = FuelLog(
                    vehicle_id=vehicle.id, log_date=log_date, liters=0,
                    odometer=mileage, created_by=current_user.id,
                )
                db.session.add(fuel)
                db.session.flush()
                created_records.append(('fuel_logs', fuel.id))
            imported += 1
        except ValueError as e:
            errors.append(f'Row {i}: {e}')
            error_rows.append({**row, 'System_Error': str(e)})

    return imported, errors, error_rows, created_drivers, created_records


# ─────────────────────────────────────────────────────────────
# Franchise Income Import — flat-row CSV/Excel import for the Daily/Weekly
# Income reconciliation pages (FranchiseDailyIncome/FranchiseWeeklyIncome),
# same two-step confirmed-mapping flow as the collections importer above.
# Each row is a whole-franchise (or, if a Vehicle column is mapped, a single
# vehicle's) reconciliation for one date — income, the four expense
# categories, other expenditure, and cash deposited.
# ─────────────────────────────────────────────────────────────
CANONICAL_FRANCHISE_INCOME_FIELDS = [
    ('date', 'Date', ['date', 'entry date', 'week start', 'week of', 'day']),
    ('vehicle', 'Vehicle / Number Plate', ['vehicle', 'number plate', 'plate']),
    ('income', 'Income', ['income', 'total income', 'revenue', 'collections', 'amount', 'amount paid']),
    ('exp_traffic_fines', 'Traffic Fines', ['traffic fines', 'fines']),
    ('exp_facilitation_fees', 'Facilitation Fees', ['facilitation fees', 'facilitation']),
    ('exp_workshop', 'Workshop', ['workshop', 'workshop all', 'repairs']),
    ('exp_wages', 'Wages', ['wages', 'salaries']),
    ('other_expenditure', 'Other Expenditure', ['other expenditure', 'other expenses', 'other', 'expense', 'expenses']),
    ('deposited', 'Deposited', ['deposited', 'deposit', 'banked', 'cash deposited']),
    ('description', 'Description', ['description', 'notes', 'remarks', 'comment', 'comments']),
]

FRANCHISE_INCOME_ROW_KEY_MAP = {key: key for key, _label, _syn in CANONICAL_FRANCHISE_INCOME_FIELDS}

# Used by the per-vehicle-tab import (Daily/Weekly Income pages) when a
# specific vehicle tab is selected — that vehicle is implied by the tab,
# mirroring the fleet ledger's single-vehicle import, so there's no
# Vehicle/Franchisee column to map. The Whole Franchise tab still gets the
# full field list (with Vehicle/Franchisee) below, since a flat file with
# many vehicles' rows in it has nowhere else to name each row's vehicle.
CANONICAL_FRANCHISE_INCOME_FIELDS_SCOPED = [
    f for f in CANONICAL_FRANCHISE_INCOME_FIELDS if f[0] not in ('vehicle', 'franchisee')]


def import_franchise_income_rows(file_rows, model_cls, date_field, week_normalize=False, forced_vehicle=None):
    """Validate and persist already-mapped franchise income/expense
    reconciliation rows (keyed by the CANONICAL_FRANCHISE_INCOME_FIELDS
    field names) into model_cls (FranchiseDailyIncome or
    FranchiseWeeklyIncome), one entry per date_field ('entry_date' or
    'week_start'). If week_normalize is True, each row's date is normalized
    to that week's Monday before being used as the key — mirrors
    franchise_weekly_income_add, since FranchiseWeeklyIncome holds one row
    per calendar week, not per day; a source file with several days in the
    same week must already be aggregated to one row per week, or later days
    will collide with the first on the unique (week_start, vehicle_id)
    constraint and be quarantined as duplicates rather than silently
    overwriting it.

    forced_vehicle, when passed (a FranchiseVehicle instance), scopes every
    row to that vehicle regardless of any Vehicle column — used by the
    per-vehicle-tab import and the fleet-workbook bulk import (sheet name IS
    the vehicle), mirroring import_ledger_rows' single-vehicle scoping. Left
    at its default of None ("no specific vehicle forced" — used for the
    Whole Franchise tab's import, where a flat file can carry many vehicles'
    rows), a row's own Vehicle column (if mapped) picks that row's vehicle,
    leaving the entry whole-franchise (vehicle_id=None) when blank —
    matching the manual entry form's default.

    Vehicles are never auto-registered here — a plate that doesn't match an
    existing FranchiseVehicle is rejected (row skipped with an error) rather
    than silently created. Franchise vehicles must be registered one at a
    time under Franchise Vehicles first; this keeps that registry
    deliberate, unlike the fleet ledger's bulk auto-registration.

    Returns (imported_count, error_messages, error_rows, created_vehicle_plates,
    created_records); does not commit — the caller decides when to commit/rollback.
    created_vehicle_plates is always empty — kept in the return shape for
    compatibility with the other importers' (row_key_map-based) return signature."""
    scoped = forced_vehicle is not None
    vehicle_by_plate = {} if scoped else {
        _normalize_registration(v.number_plate): v for v in FranchiseVehicle.query.all()}
    table_name = model_cls.__tablename__
    created_vehicles = []
    created_records = []
    dayfirst = _detect_dayfirst(r.get('date') for r in file_rows)
    if dayfirst is None:
        dayfirst = True

    imported = 0
    errors = []
    error_rows = []
    for i, raw_row in enumerate(file_rows, start=2):  # row 1 is the header
        row = {(k or '').strip().lower(): v for k, v in raw_row.items()}
        try:
            date_raw = row.get('date')
            if date_raw in (None, ''):
                continue
            entry_date = parse_import_date(date_raw, dayfirst)
            if week_normalize:
                entry_date = entry_date - timedelta(days=entry_date.weekday())

            if scoped:
                vehicle = forced_vehicle
            else:
                plate_raw = str(row.get('vehicle') or '').strip()
                if len(plate_raw) > 20:
                    raise ValueError(f'"{plate_raw[:40]}…" is too long to be a number plate — check the Vehicle '
                                     'column is mapped to the right column in your file, not a notes/description one.')
                vehicle = None
                if plate_raw:
                    plate = _normalize_registration(plate_raw)
                    vehicle = vehicle_by_plate.get(plate)
                    if not vehicle:
                        raise ValueError(f'unregistered vehicle "{plate_raw}" — register it under Franchise Vehicles '
                                         'first, or leave the Vehicle column blank for a whole-franchise entry.')

            existing = model_cls.query.execution_options(include_deleted=True).filter_by(
                **{date_field: entry_date}, vehicle_id=vehicle.id if vehicle else None).first()
            if existing and existing.deleted_at is None:
                label = vehicle.number_plate if vehicle else 'the franchise\'s shared expenditure'
                raise ValueError(f'an entry for {label} on {entry_date} already exists — skipped.')

            entry = existing or model_cls(vehicle_id=vehicle.id if vehicle else None)
            if existing:
                existing.deleted_at = None
            else:
                db.session.add(entry)
            setattr(entry, date_field, entry_date)
            entry.income = parse_import_number(row.get('income'), 'Income') or 0
            entry.exp_traffic_fines = parse_import_number(row.get('exp_traffic_fines'), 'Traffic Fines') or 0
            entry.exp_facilitation_fees = parse_import_number(row.get('exp_facilitation_fees'), 'Facilitation Fees') or 0
            entry.exp_workshop = parse_import_number(row.get('exp_workshop'), 'Workshop') or 0
            entry.exp_wages = parse_import_number(row.get('exp_wages'), 'Wages') or 0
            entry.other_expenditure = parse_import_number(row.get('other_expenditure'), 'Other Expenditure') or 0
            entry.deposited = parse_import_number(row.get('deposited'), 'Deposited') or 0
            entry.description = str(row.get('description') or '').strip()
            entry.created_by = current_user.id
            db.session.flush()
            created_records.append((table_name, entry.id))
            imported += 1
        except ValueError as e:
            errors.append(f'Row {i}: {e}')
            error_rows.append({**row, 'System_Error': str(e)})

    return imported, errors, error_rows, created_vehicles, created_records


# ─────────────────────────────────────────────────────────────
# Franchise Reconciliation Workbook Import — auto-located, auto-parsed
# alternative to the column-mapping importer above, for the franchise's own
# monthly Excel logbook shape: one sheet per month, each holding a
# "FRANCHISE COLLECTION RECONCILIATION SCHEDULE" block for whole-franchise
# daily income/expenses/deposits, side by side with an unrelated per-vehicle
# collections matrix (and sometimes a derived weekly-summary block) sharing
# the same rows. Since the shape is fixed and known — just shifted in
# position, because the matrix widens with the month's day count — this
# locates and parses it directly instead of asking the user to map columns.
# ─────────────────────────────────────────────────────────────
def _find_franchise_reconciliation_block(raw_rows):
    """Locate every 'FRANCHISE COLLECTION RECONCILIATION SCHEDULE'-shaped
    block within one sheet of a franchise monthly logbook and extract their
    rows, already shaped for import_franchise_income_rows.

    A real sheet can hold more than one such block — e.g. a "WEEKLY
    ANALYSIS" sheet in the wild turned out to have three, appended one below
    another as later weeks were added — and, seen across a real workbook,
    they don't always share one column layout (some have an OTHER
    expenditure column, some don't). So every block is found independently
    and parsed with its own column map, rather than detecting the first one
    and assuming the rest of the sheet matches it: reusing one block's
    layout for another's rows silently misreads them whenever the layouts
    differ (an OTHER column shifts CASH IN HAND/DEPOSITED over by one).

    Each block's own header is three rows deep — a group row (DATE / INCOME
    [$] / EXPENDITURE [$] / [OTHER] / CASH IN HAND [$] / DEPOSITED [$]), a
    DAILY/WEEKLY sub-row under INCOME and again under EXPENDITURE, and a
    category row (TRAFFIC FINES / FACILITATION FEES / WORKSHOP ALL / WAGES)
    repeated under each of those two EXPENDITURE sub-columns. That
    DAILY/WEEKLY split is two revenue/expense streams recorded side by side
    for the *same* date, not "today" vs "this week" — so each is summed
    into the single income/expense figure the app's Daily Income model
    actually stores; a caller never needs to know the split existed.

    A group row is found by content (a row containing cells matching 'date',
    'income', and 'expenditure') rather than assumed to sit at a fixed
    row/column: a real export like this puts its first block beside an
    unrelated per-vehicle collections matrix on a short (partial-month)
    sheet, but *below* that matrix — hundreds of rows down — on a full
    month's sheet, since the matrix grows one row per vehicle regardless of
    the month's length while a full month simply has more vehicle rows to
    list; its starting column shifts too, for the same reason. So the whole
    sheet is searched, not just the first few rows.

    Returns a list of row dicts (keys: date, income, exp_traffic_fines,
    exp_facilitation_fees, exp_workshop, exp_wages, other_expenditure,
    deposited) pooled from every block found, or None if the sheet has no
    such block at all (e.g. a sheet that's purely a derived pivot report,
    with nothing shaped like this to import) — distinct from returning an
    empty list, which means at least one block was found but none of its
    rows had a usable date."""
    def norm(v):
        return re.sub(r'\s+', ' ', str(v or '')).strip().lower()

    date_like_re = re.compile(r'^\d{1,4}[/-]\d{1,2}[/-]\d{1,4}$')

    def looks_like_date(v):
        # A block's row range runs to the next block's header (or end of
        # sheet) and can outrun where its own real data actually ends — e.g.
        # a short partial-month sheet's single block spans the rest of an
        # otherwise-unrelated sheet below it — so anything landing in the
        # date column has to look like a date, not just be non-blank, or
        # this would also scoop up day names, "TOTAL" rows, even vehicle
        # plates from an unrelated table sharing that column further down.
        if isinstance(v, (datetime, date)):
            return True
        return isinstance(v, str) and bool(date_like_re.match(v.strip()))

    group_rows = [r for r in range(len(raw_rows))
                  if (lambda cells: 'date' in cells and any('income' in c for c in cells)
                      and any('expenditure' in c for c in cells))([norm(c) for c in raw_rows[r]])]
    if not group_rows:
        return None

    def find_col(header, *needles):
        for i, c in enumerate(header):
            if any(n in norm(c) for n in needles):
                return i
        return None

    def cell_num(row, idx):
        if idx is None or idx >= len(row) or not isinstance(row[idx], (int, float)):
            return 0
        return row[idx]

    expense_categories = [
        ('exp_traffic_fines', ['traffic fines', 'traffic fine']),
        ('exp_facilitation_fees', ['facilitation fees', 'facilitation fee']),
        ('exp_workshop', ['workshop']),
        ('exp_wages', ['wages', 'wage']),
    ]

    rows = []
    for block_i, group_row in enumerate(group_rows):
        # A block's data ends where the next one's header begins (or end of
        # sheet, for the last one) — never spills into a differently-laid-out
        # block below it.
        block_end = group_rows[block_i + 1] if block_i + 1 < len(group_rows) else len(raw_rows)

        header = raw_rows[group_row]
        date_col = find_col(header, 'date')
        income_col = find_col(header, 'income')
        expenditure_col = find_col(header, 'expenditure')
        other_col = find_col(header, 'other')
        cash_col = find_col(header, 'cash in hand')
        deposited_col = find_col(header, 'deposited')
        if date_col is None or income_col is None or expenditure_col is None:
            continue
        end_col = (other_col if other_col is not None else cash_col if cash_col is not None
                   else deposited_col if deposited_col is not None else len(header))

        sub_row = raw_rows[group_row + 1] if group_row + 1 < block_end else []
        income_cols = [i for i, c in enumerate(sub_row)
                       if income_col <= i < expenditure_col and norm(c) in ('daily', 'weekly')]

        cat_row = raw_rows[group_row + 2] if group_row + 2 < block_end else []
        expense_cols = {
            key: [i for i, c in enumerate(cat_row) if expenditure_col <= i < end_col and any(n in norm(c) for n in needles)]
            for key, needles in expense_categories
        }

        for r in range(group_row + 3, block_end):
            row = raw_rows[r]
            date_val = row[date_col] if date_col < len(row) else None
            if not looks_like_date(date_val):
                continue  # skips blank rows and any unrelated table sharing these columns
            # Passed through as-is even when it's text (e.g. a manually-typed
            # "29/6/2026" cell rather than a real Excel date) rather than
            # parsed here — import_franchise_income_rows already parses
            # (and dayfirst-detects) whatever lands in 'date', the same way
            # it does for a CSV's string dates; a value that turns out not
            # to be parseable after all surfaces as a normal per-row import
            # error instead of silently vanishing.
            entry = {
                'date': date_val,
                'income': sum(cell_num(row, c) for c in income_cols),
                'other_expenditure': cell_num(row, other_col),
                'deposited': row[deposited_col] if deposited_col is not None and deposited_col < len(row) else None,
            }
            for key, cols in expense_cols.items():
                entry[key] = sum(cell_num(row, c) for c in cols)
            rows.append(entry)
    return rows


# ─────────────────────────────────────────────────────────────
# Franchise Vehicle Registration Import — bulk-register/update FranchiseVehicle
# rows by plate, same two-step confirmed-mapping flow as the imports above.
# This is the only import that can set a vehicle's daily/weekly fee, since the
# Collections/Income imports only carry transactional rows, not fee data —
# it's the franchise-side counterpart to the fleet ledger workbook's
# auto-register-by-sheet-name behavior, but as a first-class action.
# ─────────────────────────────────────────────────────────────
CANONICAL_FRANCHISE_VEHICLE_FIELDS = [
    ('vehicle', 'Vehicle / Number Plate', ['vehicle', 'number plate', 'plate', 'registration',
                                           'reg', 'vehicle reg', 'number']),
    ('franchisee', 'Franchisee Name', ['franchisee', 'franchisee name', 'owner', 'name']),
    ('daily_fee', 'Daily Fee', ['daily fee', 'daily rate', 'daily amount', 'daily']),
    ('weekly_fee', 'Weekly Fee', ['weekly fee', 'weekly rate', 'weekly amount', 'weekly']),
    ('status', 'Status', ['status']),
    ('notes', 'Notes', ['notes', 'note', 'remarks', 'comment', 'comments', 'details']),
]

FRANCHISE_VEHICLE_ROW_KEY_MAP = {key: key for key, _label, _syn in CANONICAL_FRANCHISE_VEHICLE_FIELDS}


def import_franchise_vehicle_rows(file_rows):
    """Validate and persist already-mapped franchise vehicle registration rows
    (keyed by the CANONICAL_FRANCHISE_VEHICLE_FIELDS field names) as
    FranchiseVehicle records.

    Upserts by normalized plate: a row naming a plate already on file (even
    soft-deleted) updates that vehicle in place — reviving it if needed —
    rather than erroring, so the same file can be re-imported later to
    update fees without duplicating vehicles. Only columns present in the
    file overwrite existing values; a blank/unmapped column leaves the
    existing value alone (a new vehicle gets 'active' status and no fees
    by default, matching the manual Add form's defaults).

    Returns (imported_count, error_messages, error_rows, created_records);
    does not commit — the caller decides when to commit/rollback."""
    vehicle_by_plate = {_normalize_registration(v.number_plate): v
                         for v in FranchiseVehicle.query.execution_options(include_deleted=True).all()}
    created_records = []

    imported = 0
    errors = []
    error_rows = []
    for i, raw_row in enumerate(file_rows, start=2):  # row 1 is the header
        row = {(k or '').strip().lower(): v for k, v in raw_row.items()}
        try:
            plate_raw = str(row.get('vehicle') or '').strip()
            if not plate_raw:
                continue
            if len(plate_raw) > 20:
                raise ValueError(f'"{plate_raw[:40]}…" is too long to be a number plate — check the Vehicle '
                                 'column is mapped to the right column in your file.')
            plate = _normalize_registration(plate_raw)

            franchisee_raw = str(row.get('franchisee') or '').strip()
            daily_fee = parse_import_number(row.get('daily_fee'), 'Daily Fee')
            weekly_fee = parse_import_number(row.get('weekly_fee'), 'Weekly Fee')
            status_raw = str(row.get('status') or '').strip().lower()
            if status_raw and status_raw not in ('active', 'inactive'):
                raise ValueError(f'status "{status_raw}" must be "active" or "inactive".')
            notes_raw = row.get('notes')

            vehicle = vehicle_by_plate.get(plate)
            is_new = vehicle is None
            if is_new:
                vehicle = FranchiseVehicle(number_plate=plate,
                                           franchisee_name=franchisee_raw or plate_raw.strip().title())
                db.session.add(vehicle)
            else:
                vehicle.deleted_at = None
                if franchisee_raw:
                    vehicle.franchisee_name = franchisee_raw
            if daily_fee is not None:
                vehicle.daily_fee = daily_fee
            if weekly_fee is not None:
                vehicle.weekly_fee = weekly_fee
            if status_raw:
                vehicle.status = status_raw
            elif is_new:
                vehicle.status = 'active'
            if notes_raw not in (None, ''):
                vehicle.notes = str(notes_raw).strip()
            touch_sync_fields(vehicle)
            db.session.flush()
            vehicle_by_plate[plate] = vehicle
            created_records.append(('franchise_vehicles', vehicle.id))
            imported += 1
        except ValueError as e:
            errors.append(f'Row {i}: {e}')
            error_rows.append({**row, 'System_Error': str(e)})

    return imported, errors, error_rows, created_records


# ─────────────────────────────────────────────────────────────
# Spares Store Stock Import — flat-row CSV/Excel import for restocking.
# Each row is one StorePurchase (a restock) for a named part on a given
# date, so — unlike the franchise collections/income importers, which
# reject a second row for the same date/vehicle — the same part can
# legitimately appear on many different dates (or even the same date
# twice, e.g. two separate deliveries) in a single file.
# ─────────────────────────────────────────────────────────────
CANONICAL_STOCK_FIELDS = [
    ('date', 'Date', ['date', 'purchase date', 'restock date', 'day']),
    ('part', 'Part Name / Part Number', ['part', 'part name', 'name', 'part number',
                                         'item', 'item name', 'description']),
    ('quantity', 'Quantity', ['quantity', 'qty', 'qty received', 'units', 'quantity received']),
    ('unit_cost', 'Unit Cost', ['unit cost', 'cost', 'price', 'unit price', 'cost price']),
    ('supplier', 'Supplier', ['supplier', 'vendor', 'source']),
    ('notes', 'Notes', ['notes', 'note', 'remarks', 'comment', 'comments']),
]

STOCK_ROW_KEY_MAP = {key: key for key, _label, _syn in CANONICAL_STOCK_FIELDS}


def import_stock_purchase_rows(file_rows, auto_create_parts=False):
    """Validate and persist already-mapped stock rows (keyed by 'date', 'part',
    'quantity', 'unit_cost', 'supplier', 'notes') as StorePurchase entries,
    rolling each into its part's quantity_on_hand and weighted-average
    cost_price exactly like the manual Record Purchase form does. Returns
    (imported_count, error_messages, error_rows, created_parts, created_records);
    does not commit — the caller decides when to commit/rollback.

    If auto_create_parts is True, a row naming a part that isn't on file
    (matched by part_number first, then by name) creates a new active
    SparePart instead of erroring — mirrors import_ledger_rows'
    auto_register_drivers behavior."""
    parts = SparePart.query.all()
    part_by_number = {p.part_number.strip().upper(): p for p in parts if p.part_number and p.part_number.strip()}
    part_by_name = {p.name.strip().lower(): p for p in parts}
    created_parts = []
    created_records = []
    dayfirst = _detect_dayfirst(r.get('date') for r in file_rows)
    if dayfirst is None:
        dayfirst = True

    imported = 0
    errors = []
    error_rows = []
    for i, raw_row in enumerate(file_rows, start=2):  # row 1 is the header
        row = {(k or '').strip().lower(): v for k, v in raw_row.items()}
        try:
            date_raw = row.get('date')
            if date_raw in (None, ''):
                continue
            purchase_date = parse_import_date(date_raw, dayfirst)

            part_raw = str(row.get('part') or '').strip()
            quantity_raw = parse_import_number(row.get('quantity'), 'Quantity')
            if not part_raw and quantity_raw is None:
                continue
            if not part_raw:
                raise ValueError('part name / part number is required.')
            if quantity_raw is None:
                raise ValueError('Quantity is required.')
            if quantity_raw <= 0:
                raise ValueError(f'Quantity "{quantity_raw}" must be greater than 0.')
            quantity = quantity_raw

            unit_cost = parse_import_number(row.get('unit_cost'), 'Unit Cost')
            if unit_cost is None:
                raise ValueError('Unit Cost is required.')
            if unit_cost < 0:
                raise ValueError('Unit Cost cannot be negative.')

            part = part_by_number.get(part_raw.upper()) or part_by_name.get(part_raw.lower())
            if not part:
                if auto_create_parts:
                    part = SparePart(name=part_raw, unit='pc', cost_price=0, quantity_on_hand=0,
                                     created_by=current_user.id)
                    db.session.add(part)
                    db.session.flush()
                    part_by_name[part_raw.lower()] = part
                    created_parts.append(part_raw)
                    created_records.append(('spare_parts', part.id))
                else:
                    raise ValueError(f'unknown part "{part_raw}" — add it under Spares Store first, '
                                     'or tick "Auto-create new parts".')

            purchase = StorePurchase(
                part_id=part.id, purchase_date=purchase_date, quantity=quantity, unit_cost=unit_cost,
                total_cost=quantity * unit_cost, supplier=str(row.get('supplier') or '').strip(),
                notes=str(row.get('notes') or '').strip(), created_by=current_user.id,
            )
            new_total_qty = part.quantity_on_hand + quantity
            part.cost_price = ((part.quantity_on_hand * part.cost_price) +
                               (quantity * unit_cost)) / new_total_qty
            part.quantity_on_hand = new_total_qty
            touch_sync_fields(part)

            db.session.add(purchase)
            db.session.flush()
            touch_sync_fields(purchase)
            created_records.append(('store_purchases', purchase.id))
            imported += 1
        except ValueError as e:
            errors.append(f'Row {i}: {e}')
            error_rows.append({**row, 'System_Error': str(e)})

    return imported, errors, error_rows, created_parts, created_records


def form_float(form, field, label=None, required=True, default=None, min_value=None):
    label = label or field.replace('_', ' ').capitalize()
    raw = (form.get(field) or '').strip()
    if not raw:
        if required:
            raise ValueError(f'{label} is required.')
        return default
    try:
        value = float(raw)
    except ValueError:
        raise ValueError(f'{label} must be a number.')
    if min_value is not None and value < min_value:
        raise ValueError(f'{label} cannot be less than {min_value}.')
    return value


def form_int(form, field, label=None, required=True, default=None, min_value=None):
    label = label or field.replace('_', ' ').capitalize()
    raw = (form.get(field) or '').strip()
    if not raw:
        if required:
            raise ValueError(f'{label} is required.')
        return default
    try:
        value = int(raw)
    except ValueError:
        raise ValueError(f'{label} must be a whole number.')
    if min_value is not None and value < min_value:
        raise ValueError(f'{label} cannot be less than {min_value}.')
    return value


def already_synced(client_id):
    """True if this client-generated offline submission id was already
    recorded — lets an offline-queue replay short-circuit instead of
    inserting a duplicate row."""
    return bool(client_id) and OfflineSyncLog.query.filter_by(client_id=client_id).first() is not None


def record_offline_sync(client_id, endpoint):
    if client_id:
        db.session.add(OfflineSyncLog(client_id=client_id, endpoint=endpoint, user_id=current_user.id))


def touch_sync_fields(obj):
    """Call on every create/edit of a multi-site-syncable model, right
    before commit. Mints a sync_uuid the first time (the cross-instance
    identity used instead of the local auto-increment id, since two offline
    sites can independently mint the same integer id), stamps updated_at
    (the LWW conflict-resolution timestamp — see /api/sync/push) and
    server_touched_at (a purely local, monotonic "I wrote this row just
    now" marker used only for /api/sync/pull's since= filter — see that
    field's docstring on why it has to be separate from updated_at), and
    marks the row pending_push so the local sync engine's outbox picks it
    up on its next cycle."""
    if not obj.sync_uuid:
        obj.sync_uuid = uuid.uuid4().hex
    now = datetime.now(timezone.utc)
    obj.updated_at = now
    obj.server_touched_at = now
    obj.pending_push = True
    obj.last_modified_site = app.config['SITE_ID']


def get_fuel_price_row():
    return FuelPrice.query.get(1)


def fuel_price_for(fuel_type):
    """Current per-liter price for 'diesel' or 'petrol', or None if the
    admin hasn't set one yet — callers treat that as "can't convert"."""
    row = get_fuel_price_row()
    if not row:
        return None
    return row.petrol_price if fuel_type == 'petrol' else row.diesel_price


def resolve_conductor(driver, vehicle_id):
    """Auto-attaches a conductor to a trip being logged: prefers the
    conductor paired directly to this driver (Driver.paired_driver_id),
    falling back to a conductor assigned to this vehicle
    (Driver.assigned_vehicle_id) — a conductor can be attached either way,
    and either attachment should flow through to DailyLog.conductor_id (and
    from there into payroll — see compute_payroll_earnings) without the
    person logging the trip having to pick the conductor by hand."""
    if driver and driver.paired_conductors:
        return driver.paired_conductors[0]
    if vehicle_id:
        return Driver.query.filter_by(assigned_vehicle_id=vehicle_id, role='conductor',
                                      status='active').first()
    return None


def check_unique(model, field_name, value, label=None, exclude_id=None):
    """Raise a friendly ValueError if another row already has this value,
    instead of letting the DB's UNIQUE constraint crash with a 500."""
    label = label or field_name.replace('_', ' ').capitalize()
    q = model.query.filter(getattr(model, field_name) == value)
    if exclude_id is not None:
        q = q.filter(model.id != exclude_id)
    if q.first():
        raise ValueError(f'{label} "{value}" is already in use.')


def handle_form_errors(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if request.method == 'POST':
            try:
                return f(*args, **kwargs)
            except KeyError as e:
                db.session.rollback()
                flash(f'Missing required field: {e}', 'danger')
                return redirect(request.url)
            except ValueError as e:
                db.session.rollback()
                flash(str(e), 'danger')
                return redirect(request.url)
        return f(*args, **kwargs)
    return decorated


def query_date_range(default_from=None, default_to=None):
    """Read date_from/date_to from the query string for report pages.
    Falls back to defaults on missing/invalid input and auto-swaps a
    reversed range, so any date (or combination) can be requested
    without raising — the caller always gets a usable (from, to) pair."""
    today = date.today()
    default_from = default_from or today.replace(day=1)
    default_to = default_to or today

    from_str = request.args.get('date_from', '').strip()
    to_str = request.args.get('date_to', '').strip()

    try:
        df = parse_date(from_str) if from_str else default_from
    except ValueError:
        flash(f'"{from_str}" is not a valid start date — showing {default_from} instead.', 'warning')
        df = default_from

    try:
        dt = parse_date(to_str) if to_str else default_to
    except ValueError:
        flash(f'"{to_str}" is not a valid end date — showing {default_to} instead.', 'warning')
        dt = default_to

    if df > dt:
        df, dt = dt, df
        flash('Start date was after end date — the range was swapped.', 'warning')

    return df, dt


def query_single_date(param='as_of', default=None):
    """Read a single date query param (e.g. 'as at' a point in time),
    falling back to a default on missing/invalid input instead of raising."""
    default = default or date.today()
    raw = request.args.get(param, '').strip()
    if not raw:
        return default
    try:
        return parse_date(raw)
    except ValueError:
        flash(f'"{raw}" is not a valid date — showing {default} instead.', 'warning')
        return default


def compute_commission_accrued(as_of):
    """Total driver/conductor commission earned (accrued) on all revenue
    up to as_of, using the same per-driver rate logic as the payroll report,
    but over the driver's entire history rather than one filter period."""
    dr_rate = app.config['COMMISSION_DRIVER_RATE']
    co_rate = app.config['COMMISSION_CONDUCTOR_RATE']
    total = 0.0
    for d in Driver.query.all():
        driven = db.session.query(func.sum(DailyLog.gross_revenue)).filter(
            DailyLog.driver_id == d.id, DailyLog.log_date <= as_of).scalar() or 0
        conducted = db.session.query(func.sum(DailyLog.gross_revenue)).filter(
            DailyLog.conductor_id == d.id, DailyLog.log_date <= as_of).scalar() or 0
        garnish_driven = db.session.query(func.sum(DailyLog.garnish)).filter(
            DailyLog.driver_id == d.id, DailyLog.log_date <= as_of).scalar() or 0
        garnish_conducted = db.session.query(func.sum(DailyLog.garnish)).filter(
            DailyLog.conductor_id == d.id, DailyLog.log_date <= as_of).scalar() or 0
        rate = d.commission_rate if d.commission_rate is not None else (
            dr_rate if d.role == 'driver' else co_rate)
        total += max(driven + conducted - garnish_driven - garnish_conducted, 0) * rate
    return total


def compute_financial_position(as_of):
    """Simplified statement of financial position as at a given date.

    This is built on real records (loans, payables, receivables, capital
    contributions, owner drawings, commission payments, expenses) rather
    than a single inferred "cash" figure, but it's still an approximation,
    not bookkeeping-grade — there's no real cash/bank reconciliation.
    Key modelling choices:
      - Vehicle purchases are a pure asset swap (cash down, fixed asset up)
        — they are NOT assumed to be capital-funded. If vehicles were
        bought before any Capital Contribution or Loan was recorded to
        explain the cash, Cash will legitimately show negative here. That's
        the system being honest about a genuine gap in what's recorded —
        fix it by adding an opening Capital Contribution or Loan entry.
      - Vehicles are depreciated straight-line over VEHICLE_USEFUL_LIFE_YEARS
        from when each was added to the fleet (no acquisition-date field
        exists separately from that).
      - Commission is accrued on all revenue earned (matching the payroll
        report's calculation), not just what's been paid out — the
        difference sits as a Commission Payable liability.
      - Payables/Receivables are accrual-based: recognized as an
        expense/revenue when created, not when settled.
      - Loan repayments are treated as pure principal reduction (interest
        is not separately expensed) — a deliberate simplification.
    """
    useful_life = app.config['VEHICLE_USEFUL_LIFE_YEARS']

    vehicle_rows = []
    total_vehicle_cost = 0.0
    total_accum_dep = 0.0
    for v in Vehicle.query.order_by(Vehicle.registration).all():
        acquired = v.created_at.date()
        if acquired > as_of:
            continue
        years_in_service = (as_of - acquired).days / 365.25
        accum_dep = min(v.acquisition_cost, v.acquisition_cost * years_in_service / useful_life) \
            if useful_life else 0.0
        nbv = v.acquisition_cost - accum_dep
        vehicle_rows.append({
            'vehicle': v,
            'cost': v.acquisition_cost,
            'accumulated_depreciation': accum_dep,
            'net_book_value': nbv,
            'years_in_service': years_in_service,
        })
        total_vehicle_cost += v.acquisition_cost
        total_accum_dep += accum_dep

    total_nbv = total_vehicle_cost - total_accum_dep

    total_revenue = db.session.query(func.sum(DailyLog.gross_revenue)).filter(
        DailyLog.log_date <= as_of).scalar() or 0
    total_maintenance = db.session.query(func.sum(MaintenanceLog.total_cost)).filter(
        MaintenanceLog.log_date <= as_of).scalar() or 0
    total_expenses = db.session.query(func.sum(Expense.amount)).filter(
        Expense.expense_date <= as_of).scalar() or 0

    commission_accrued = compute_commission_accrued(as_of)
    commission_paid = db.session.query(func.sum(CommissionPayment.amount)).filter(
        CommissionPayment.payment_date <= as_of).scalar() or 0
    commission_payable = commission_accrued - commission_paid

    loan_proceeds = db.session.query(func.sum(Loan.principal)).filter(
        Loan.start_date <= as_of).scalar() or 0
    loan_repayments = db.session.query(func.sum(LoanPayment.amount)).filter(
        LoanPayment.payment_date <= as_of).scalar() or 0
    loans_outstanding = loan_proceeds - loan_repayments

    payables_total = db.session.query(func.sum(Payable.amount)).filter(
        Payable.invoice_date <= as_of).scalar() or 0
    payables_paid = db.session.query(func.sum(Payable.amount)).filter(
        Payable.status == 'paid', Payable.paid_date.isnot(None), Payable.paid_date <= as_of).scalar() or 0
    payables_outstanding = payables_total - payables_paid

    receivables_total = db.session.query(func.sum(Receivable.amount)).filter(
        Receivable.invoice_date <= as_of).scalar() or 0
    receivables_collected = db.session.query(func.sum(Receivable.amount)).filter(
        Receivable.status == 'collected', Receivable.collected_date.isnot(None),
        Receivable.collected_date <= as_of).scalar() or 0
    receivables_outstanding = receivables_total - receivables_collected

    capital_contributions = db.session.query(func.sum(CapitalContribution.amount)).filter(
        CapitalContribution.contribution_date <= as_of).scalar() or 0
    owner_drawings = db.session.query(func.sum(OwnerDrawing.amount)).filter(
        OwnerDrawing.drawing_date <= as_of).scalar() or 0

    cash_and_equivalents = (
        total_revenue - total_maintenance - total_expenses
        - commission_paid - total_vehicle_cost
        + loan_proceeds - loan_repayments
        + capital_contributions - owner_drawings
        + receivables_collected - payables_paid
    )

    retained_earnings = (
        (total_revenue + receivables_total)
        - (total_maintenance + total_expenses + payables_total)
        - total_accum_dep - commission_accrued
    )
    owners_capital = capital_contributions - owner_drawings

    total_assets = total_nbv + cash_and_equivalents + receivables_outstanding
    total_liabilities = loans_outstanding + payables_outstanding + commission_payable
    total_equity = owners_capital + retained_earnings

    return {
        'as_of': as_of,
        'useful_life': useful_life,
        'vehicle_rows': vehicle_rows,
        'total_cost': total_vehicle_cost,
        'total_accum_dep': total_accum_dep,
        'total_nbv': total_nbv,
        'cash_and_equivalents': cash_and_equivalents,
        'receivables_outstanding': receivables_outstanding,
        'total_assets': total_assets,
        'loans_outstanding': loans_outstanding,
        'payables_outstanding': payables_outstanding,
        'commission_payable': commission_payable,
        'total_liabilities': total_liabilities,
        'owners_capital': owners_capital,
        'retained_earnings': retained_earnings,
        'total_equity': total_equity,
    }


# ─────────────────────────────────────────────────────────────
# Auth routes
# ─────────────────────────────────────────────────────────────
@app.route('/')
def index():
    return redirect(url_for('dashboard') if current_user.is_authenticated else url_for('login'))


@app.route('/sw.js')
def service_worker():
    """Served from the root (not /static/sw.js) so its default scope is the
    whole app, not just /static/ — required for it to intercept navigation
    requests across every page."""
    return send_from_directory('static', 'sw.js', mimetype='application/javascript')


@app.before_request
def require_first_run_setup():
    """A freshly launched spoke .exe (nothing pre-filled in .env, no
    admin account yet) goes straight to /setup instead of a bare login
    screen with no way in. Scoped to FROZEN only — Render's hub always
    comes up with ADMIN_PASSWORD pre-set, so this is purely a spoke
    first-run concern and never touches hub behavior."""
    if not FROZEN or request.endpoint in (None, 'setup', 'static', 'service_worker'):
        return
    if User.query.count() == 0:
        return redirect(url_for('setup'))


@app.route('/setup', methods=['GET', 'POST'])
def setup():
    """First-run wizard for a brand-new spoke .exe: pick a local admin
    password, optionally enroll with a hub right here instead of an
    admin visiting Sync Sites and copy-pasting an API key into .env (see
    api_sync_enroll). Locked out the moment any user exists — this is
    a first-boot-only endpoint, not a standing account-creation route."""
    if User.query.count() > 0:
        return redirect(url_for('login'))
    if request.method == 'POST':
        password = request.form.get('password') or ''
        confirm = request.form.get('confirm_password') or ''
        display_name = request.form.get('display_name', '').strip() or socket.gethostname()
        hub_url = request.form.get('hub_url', '').strip().rstrip('/')
        hub_username = request.form.get('hub_username', '').strip()
        hub_password = request.form.get('hub_password') or ''

        if len(password) < 8:
            flash('Password must be at least 8 characters.', 'danger')
            return render_template('auth/setup.html', hostname=display_name)
        if password != confirm:
            flash('Passwords do not match.', 'danger')
            return render_template('auth/setup.html', hostname=display_name)
        if hub_url and not (hub_username and hub_password):
            flash('Enter both the hub admin username and password to connect to a hub.', 'danger')
            return render_template('auth/setup.html', hostname=display_name)

        if hub_url:
            try:
                resp = requests.post(f'{hub_url}/api/sync/enroll', json={
                    'display_name': display_name,
                    'username': hub_username, 'password': hub_password,
                }, timeout=15)
            except requests.RequestException as e:
                flash(f"Could not reach that hub URL: {e}", 'danger')
                return render_template('auth/setup.html', hostname=display_name)
            if resp.status_code != 200:
                message = 'Enrollment failed.'
                try:
                    message = resp.json().get('error', message)
                except ValueError:
                    pass
                flash(message, 'danger')
                return render_template('auth/setup.html', hostname=display_name)
            payload = resp.json()
            app.config['SITE_ID'] = payload['site_id']
            app.config['SYNC_ENABLED'] = True
            app.config['SYNC_HUB_URL'] = hub_url
            app.config['SYNC_API_KEY'] = payload['api_key']
            persist_env_updates({
                'SITE_ID': payload['site_id'],
                'SYNC_ENABLED': 'true',
                'SYNC_HUB_URL': hub_url,
                'SYNC_API_KEY': payload['api_key'],
                'SYNC_INTERVAL_SECONDS': str(app.config['SYNC_INTERVAL_SECONDS']),
            })

        admin = User(username='admin', email='admin@transport.local', role='admin')
        admin.set_password(password)
        db.session.add(admin)
        db.session.commit()

        if app.config['SYNC_ENABLED']:
            start_sync_thread()

        login_user(admin)
        session.permanent = True
        flash('Setup complete — welcome to GRATZ.', 'success')
        return redirect(url_for('dashboard'))
    return render_template('auth/setup.html', hostname=socket.gethostname())


@app.route('/login', methods=['GET', 'POST'])
@limiter.limit('10 per minute')
def login():
    if current_user.is_authenticated:
        return redirect(first_permitted_url(current_user))
    if request.method == 'POST':
        username = request.form.get('username', '').strip().lower()
        password = request.form.get('password', '')
        user = User.query.filter_by(username=username).first()
        if user and user.check_password(password) and user.is_active:
            login_user(user)
            session.permanent = True
            log_audit('LOGIN', description=f'User {username} logged in')
            db.session.commit()
            return redirect(first_permitted_url(user))
        flash('Invalid username or password.', 'danger')
    return render_template('auth/login.html')


@app.route('/logout')
@login_required
def logout():
    log_audit('LOGOUT', description=f'User {current_user.username} logged out')
    db.session.commit()
    logout_user()
    flash('You have been signed out.', 'info')
    return redirect(url_for('login'))


@app.route('/account/change-password', methods=['GET', 'POST'])
@login_required
def change_password():
    if request.method == 'POST':
        current_pw = request.form.get('current_password', '')
        new_pw = request.form.get('new_password', '')
        confirm_pw = request.form.get('confirm_password', '')
        if not current_user.check_password(current_pw):
            flash('Current password is incorrect.', 'danger')
        elif len(new_pw) < 6:
            flash('New password must be at least 6 characters.', 'danger')
        elif new_pw != confirm_pw:
            flash('New passwords do not match.', 'danger')
        else:
            current_user.set_password(new_pw)
            log_audit('UPDATE', 'users', current_user.id, f'{current_user.username} changed their password')
            db.session.commit()
            flash('Password changed successfully.', 'success')
            return redirect(url_for('dashboard'))
    return render_template('auth/change_password.html')


# ─────────────────────────────────────────────────────────────
# Dashboard
# ─────────────────────────────────────────────────────────────
@app.route('/no-access')
@login_required
def no_access():
    return render_template('auth/no_access.html')


@app.route('/dashboard')
@login_required
@permission_required('dashboard')
def dashboard():
    today = date.today()
    month_start = today.replace(day=1)
    df, dt = query_date_range(default_from=month_start, default_to=today)
    date_from_str, date_to_str = df.strftime('%Y-%m-%d'), dt.strftime('%Y-%m-%d')

    today_revenue = db.session.query(func.sum(DailyLog.gross_revenue)).filter(
        DailyLog.log_date == today).scalar() or 0

    # Same fleet-wide total used by the Income Statement report: maintenance
    # + all Expense rows (vehicle-tagged and general overhead) + spares sold
    # to company vehicles — everything except fuel, which the driver pays for
    # out of the daily cash collected rather than it being a company expense.
    stmt = compute_income_statement(df, dt)
    month_revenue = stmt['gross_revenue']
    month_expenses = stmt['total_expenses']
    month_profit = stmt['net_profit']

    active_vehicles = Vehicle.query.filter_by(status='active').count()
    active_drivers = Driver.query.filter_by(status='active').count()
    active_routes = Route.query.count()

    expiry_threshold = today + timedelta(days=30)
    expiring_docs = VehicleDocument.query.filter(
        VehicleDocument.expiry_date.between(today, expiry_threshold)).count()
    expired_docs = VehicleDocument.query.filter(
        VehicleDocument.expiry_date < today).count()
    expiring_docs += Vehicle.query.filter(
        Vehicle.insurance_expiry.between(today, expiry_threshold)).count()
    expired_docs += Vehicle.query.filter(Vehicle.insurance_expiry < today).count()
    valid_docs = VehicleDocument.query.filter(VehicleDocument.expiry_date > expiry_threshold).count()
    valid_docs += Vehicle.query.filter(Vehicle.insurance_expiry > expiry_threshold).count()

    # Vehicles whose insurance is already expired or due within the same
    # 30-day window as the compliance counts above, expired-first so the
    # most urgent renewals are what the dashboard leads with.
    insurance_watch = Vehicle.query.filter(
        Vehicle.insurance_expiry.isnot(None),
        Vehicle.insurance_expiry <= expiry_threshold
    ).order_by(Vehicle.insurance_expiry).all()

    fuel_logs_mtd = FuelLog.query.filter(FuelLog.log_date.between(df, dt)).count()
    maintenance_logs_mtd = MaintenanceLog.query.filter(MaintenanceLog.log_date.between(df, dt)).count()

    unpaid_payables = Payable.query.filter(Payable.status != 'paid').count()
    outstanding_receivables = Receivable.query.filter(Receivable.status != 'collected').count()
    active_loans = Loan.query.filter(Loan.status == 'active').count()

    store_parts = SparePart.query.count()
    store_purchases_mtd = StorePurchase.query.filter(StorePurchase.purchase_date.between(df, dt)).count()
    store_sales_mtd = StoreSale.query.filter(StoreSale.sale_date.between(df, dt)).count()

    franchise_vehicles = FranchiseVehicle.query.count()
    franchise_daily_entries = FranchiseDailyIncome.query.filter(FranchiseDailyIncome.entry_date.between(df, dt)).count()
    franchise_weekly_entries = FranchiseWeeklyIncome.query.filter(FranchiseWeeklyIncome.week_start.between(df, dt)).count()
    franchise_deposited = db.session.query(func.sum(FranchiseDailyIncome.deposited)).filter(
        FranchiseDailyIncome.entry_date.between(df, dt)).scalar() or 0
    franchise_deposited += db.session.query(func.sum(FranchiseWeeklyIncome.deposited)).filter(
        FranchiseWeeklyIncome.week_start.between(df, dt)).scalar() or 0

    schedules_due = MaintenanceSchedule.query.filter(
        MaintenanceSchedule.next_due_date != None,
        MaintenanceSchedule.next_due_date <= expiry_threshold
    ).count()

    month_capital = db.session.query(func.sum(CapitalContribution.amount)).filter(
        CapitalContribution.contribution_date.between(df, dt)).scalar() or 0
    month_drawings = db.session.query(func.sum(OwnerDrawing.amount)).filter(
        OwnerDrawing.drawing_date.between(df, dt)).scalar() or 0
    month_operating_expenses = db.session.query(func.sum(Expense.amount)).filter(
        Expense.expense_date.between(df, dt)).scalar() or 0
    month_deposits = db.session.query(func.sum(DriverDeposit.amount)).filter(
        DriverDeposit.deposit_date.between(df, dt)).scalar() or 0

    recent_logs = DailyLog.query.filter(DailyLog.log_date.between(df, dt)).order_by(
        DailyLog.log_date.desc()).limit(6).all()

    rev_chart = []
    for i in range(6, -1, -1):
        d = dt - timedelta(days=i)
        rev = db.session.query(func.sum(DailyLog.gross_revenue)).filter(
            DailyLog.log_date == d).scalar() or 0
        rev_chart.append({'date': d.strftime('%d %b'), 'revenue': float(rev)})

    return render_template('dashboard.html',
        today_revenue=today_revenue, month_revenue=month_revenue,
        month_expenses=month_expenses, month_profit=month_profit,
        active_vehicles=active_vehicles, active_drivers=active_drivers,
        active_routes=active_routes,
        expiring_docs=expiring_docs, expired_docs=expired_docs, valid_docs=valid_docs,
        insurance_watch=insurance_watch,
        fuel_logs_mtd=fuel_logs_mtd, maintenance_logs_mtd=maintenance_logs_mtd,
        unpaid_payables=unpaid_payables, outstanding_receivables=outstanding_receivables,
        active_loans=active_loans,
        store_parts=store_parts, store_purchases_mtd=store_purchases_mtd, store_sales_mtd=store_sales_mtd,
        franchise_vehicles=franchise_vehicles, franchise_daily_entries=franchise_daily_entries,
        franchise_weekly_entries=franchise_weekly_entries, franchise_deposited=franchise_deposited,
        schedules_due=schedules_due,
        month_capital=month_capital, month_drawings=month_drawings,
        month_operating_expenses=month_operating_expenses, month_deposits=month_deposits,
        recent_logs=recent_logs, revenue_chart=json.dumps(rev_chart),
        date_from=date_from_str, date_to=date_to_str,
        today=today)


# ─────────────────────────────────────────────────────────────
# Vehicles
# ─────────────────────────────────────────────────────────────
@app.route('/vehicles')
@login_required
@permission_required('vehicles')
def vehicles():
    all_vehicles = Vehicle.query.order_by(Vehicle.registration).all()
    # One grouped query for every vehicle's lifetime revenue, instead of the
    # template calling v.total_revenue per row — that property (even now
    # that it's a SQL sum rather than a Python loop) would still be one
    # query per vehicle; this is one query total regardless of fleet size.
    revenue_by_vehicle = dict(
        db.session.query(DailyLog.vehicle_id, func.sum(DailyLog.gross_revenue))
        .group_by(DailyLog.vehicle_id).all())
    return render_template('vehicles/index.html', vehicles=all_vehicles,
                           revenue_by_vehicle=revenue_by_vehicle)


@app.route('/vehicles/add', methods=['GET', 'POST'])
@login_required
@permission_required('vehicles')
@handle_form_errors
def vehicle_add():
    if request.method == 'POST':
        registration = request.form['registration'].upper().strip()
        check_unique(Vehicle, 'registration', registration)
        fuel_type = request.form.get('fuel_type', 'diesel')
        if fuel_type not in ('diesel', 'petrol'):
            raise ValueError('Fuel type must be Diesel or Petrol.')
        v = Vehicle(
            registration=registration,
            make=request.form['make'].strip(),
            model=request.form['model'].strip(),
            year=form_int(request.form, 'year', min_value=1980),
            acquisition_cost=form_float(request.form, 'acquisition_cost', required=False, default=0, min_value=0),
            status=request.form.get('status', 'active'),
            fuel_type=fuel_type,
            daily_target=form_float(request.form, 'daily_target', required=False, min_value=0),
            insurance_provider=request.form.get('insurance_provider', '').strip() or None,
            insurance_policy_number=request.form.get('insurance_policy_number', '').strip() or None,
            insurance_type=request.form.get('insurance_type', '').strip() or None,
            insurance_expiry=parse_date(request.form.get('insurance_expiry')),
        )
        db.session.add(v)
        db.session.flush()
        log_audit('CREATE', 'vehicles', v.id, f'Added vehicle {v.registration}')
        touch_sync_fields(v)
        db.session.commit()
        flash(f'Vehicle {v.registration} registered successfully.', 'success')
        return redirect(url_for('vehicles'))
    return render_template('vehicles/form.html', vehicle=None, action='Register')


@app.route('/vehicles/<int:vid>')
@login_required
@permission_required('vehicles')
def vehicle_detail(vid):
    v = Vehicle.query.filter_by(id=vid).first_or_404()
    today = date.today()
    recent_logs = DailyLog.query.filter_by(vehicle_id=vid).order_by(DailyLog.log_date.desc()).limit(10).all()
    recent_fuel = FuelLog.query.filter_by(vehicle_id=vid).order_by(FuelLog.log_date.desc()).limit(5).all()
    recent_maint = MaintenanceLog.query.filter_by(vehicle_id=vid).order_by(MaintenanceLog.log_date.desc()).limit(5).all()
    return render_template('vehicles/detail.html', vehicle=v, today=today,
                           recent_logs=recent_logs, recent_fuel=recent_fuel,
                           recent_maint=recent_maint)


@app.route('/vehicles/<int:vid>/edit', methods=['GET', 'POST'])
@login_required
@permission_required('vehicles')
@handle_form_errors
def vehicle_edit(vid):
    v = Vehicle.query.filter_by(id=vid).first_or_404()
    if request.method == 'POST':
        registration = request.form['registration'].upper().strip()
        check_unique(Vehicle, 'registration', registration, exclude_id=v.id)
        v.registration = registration
        v.make = request.form['make'].strip()
        v.model = request.form['model'].strip()
        v.year = form_int(request.form, 'year', min_value=1980)
        v.acquisition_cost = form_float(request.form, 'acquisition_cost', required=False, default=0, min_value=0)
        v.status = request.form.get('status', 'active')
        fuel_type = request.form.get('fuel_type', 'diesel')
        if fuel_type not in ('diesel', 'petrol'):
            raise ValueError('Fuel type must be Diesel or Petrol.')
        v.fuel_type = fuel_type
        v.daily_target = form_float(request.form, 'daily_target', required=False, min_value=0)
        v.insurance_provider = request.form.get('insurance_provider', '').strip() or None
        v.insurance_policy_number = request.form.get('insurance_policy_number', '').strip() or None
        v.insurance_type = request.form.get('insurance_type', '').strip() or None
        v.insurance_expiry = parse_date(request.form.get('insurance_expiry'))
        log_audit('UPDATE', 'vehicles', v.id, f'Updated vehicle {v.registration}')
        touch_sync_fields(v)
        db.session.commit()
        flash(f'Vehicle {v.registration} updated.', 'success')
        return redirect(url_for('vehicle_detail', vid=vid))
    return render_template('vehicles/form.html', vehicle=v, action='Edit')


@app.route('/vehicles/<int:vid>/delete', methods=['POST'])
@login_required
@admin_required
def vehicle_delete(vid):
    v = Vehicle.query.filter_by(id=vid).first_or_404()
    reg = v.registration
    log_audit('DELETE', 'vehicles', vid, f'Deleted vehicle {reg}')
    now = datetime.now(timezone.utc)
    # cascade='all, delete-orphan' only fires on an actual ORM delete of
    # the parent, not on setting deleted_at — soft-delete documents
    # explicitly so they don't outlive their (now hidden) vehicle.
    for doc in v.documents:
        doc.deleted_at = now
        touch_sync_fields(doc)
    v.deleted_at = now
    touch_sync_fields(v)
    db.session.commit()
    flash(f'Vehicle {reg} removed.', 'warning')
    return redirect(url_for('vehicles'))


@app.route('/vehicles/<int:vehicle_id>/documents/add', methods=['GET', 'POST'])
@login_required
@permission_required('vehicles')
@handle_form_errors
def document_add(vehicle_id):
    vehicle = Vehicle.query.filter_by(id=vehicle_id).first_or_404()
    if request.method == 'POST':
        doc = VehicleDocument(
            vehicle_id=vehicle_id,
            doc_type=request.form['doc_type'],
            reference_number=request.form.get('reference_number', '').strip(),
            issue_date=parse_date(request.form.get('issue_date')),
            expiry_date=parse_date(request.form['expiry_date']),
            notes=request.form.get('notes', '').strip(),
        )
        db.session.add(doc)
        touch_sync_fields(doc)
        log_audit('CREATE', 'vehicle_documents', None,
                  f'Added {doc.doc_type} for {vehicle.registration}')
        db.session.commit()
        flash('Document added successfully.', 'success')
        return redirect(url_for('vehicle_detail', vid=vehicle_id))
    return render_template('vehicles/document_form.html', vehicle=vehicle)


@app.route('/documents/<int:did>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
@handle_form_errors
def document_edit(did):
    doc = VehicleDocument.query.filter_by(id=did).first_or_404()
    if request.method == 'POST':
        doc.doc_type = request.form['doc_type']
        doc.reference_number = request.form.get('reference_number', '').strip()
        doc.issue_date = parse_date(request.form.get('issue_date'))
        doc.expiry_date = parse_date(request.form['expiry_date'])
        doc.notes = request.form.get('notes', '').strip()
        log_audit('UPDATE', 'vehicle_documents', doc.id,
                  f'Updated {doc.doc_type} for {doc.vehicle.registration}')
        touch_sync_fields(doc)
        db.session.commit()
        flash('Document updated.', 'success')
        return redirect(url_for('vehicle_detail', vid=doc.vehicle_id))
    return render_template('vehicles/document_form.html', vehicle=doc.vehicle, doc=doc)


@app.route('/documents/<int:did>/delete', methods=['POST'])
@login_required
@permission_required('vehicles')
def document_delete(did):
    doc = VehicleDocument.query.filter_by(id=did).first_or_404()
    vid = doc.vehicle_id
    log_audit('DELETE', 'vehicle_documents', did, f'Deleted {doc.doc_type} document')
    doc.deleted_at = datetime.now(timezone.utc)
    touch_sync_fields(doc)
    db.session.commit()
    flash('Document deleted.', 'warning')
    return redirect(url_for('vehicle_detail', vid=vid))


# ─────────────────────────────────────────────────────────────
# Drivers
# ─────────────────────────────────────────────────────────────
@app.route('/drivers')
@login_required
@permission_required('drivers')
def drivers():
    all_drivers = Driver.query.order_by(Driver.name).all()
    return render_template('drivers/index.html', drivers=all_drivers)


@app.route('/drivers/add', methods=['GET', 'POST'])
@login_required
@permission_required('drivers')
@handle_form_errors
def driver_add():
    if request.method == 'POST':
        rate_input = form_float(request.form, 'commission_rate', required=False, min_value=0)
        role = request.form.get('role', 'driver')
        license_number = request.form.get('license_number', '').strip() or None
        if license_number:
            check_unique(Driver, 'license_number', license_number, label='License number')
        id_number = request.form.get('id_number', '').strip() or None
        if id_number:
            check_unique(Driver, 'id_number', id_number, label='ID number')
        paired_driver_id = form_int(request.form, 'paired_driver_id', required=False) if role == 'conductor' else None
        d = Driver(
            name=request.form['name'].strip(),
            license_number=license_number,
            id_number=id_number,
            phone=request.form.get('phone', '').strip(),
            role=role,
            commission_rate=rate_input / 100 if rate_input is not None else None,
            status=request.form.get('status', 'active'),
            paired_driver_id=paired_driver_id,
            assigned_vehicle_id=form_int(request.form, 'assigned_vehicle_id', required=False),
            next_of_kin_name=request.form.get('next_of_kin_name', '').strip(),
            next_of_kin_phone=request.form.get('next_of_kin_phone', '').strip(),
            next_of_kin_relationship=request.form.get('next_of_kin_relationship', '').strip(),
        )
        db.session.add(d)
        db.session.flush()
        log_audit('CREATE', 'drivers', d.id, f'Added driver {d.name}')
        touch_sync_fields(d)
        db.session.commit()
        flash(f'Driver {d.name} registered.', 'success')
        return redirect(url_for('drivers'))
    eligible_drivers = Driver.query.filter_by(role='driver', status='active').order_by(Driver.name).all()
    all_vehicles = Vehicle.query.filter_by(status='active').order_by(Vehicle.registration).all()
    return render_template('drivers/form.html', driver=None, action='Register',
                           eligible_drivers=eligible_drivers, vehicles=all_vehicles)


@app.route('/drivers/<int:did>/edit', methods=['GET', 'POST'])
@login_required
@permission_required('drivers')
@handle_form_errors
def driver_edit(did):
    d = Driver.query.filter_by(id=did).first_or_404()
    if request.method == 'POST':
        rate_input = form_float(request.form, 'commission_rate', required=False, min_value=0)
        role = request.form.get('role', 'driver')
        license_number = request.form.get('license_number', '').strip() or None
        if license_number:
            check_unique(Driver, 'license_number', license_number, label='License number', exclude_id=d.id)
        id_number = request.form.get('id_number', '').strip() or None
        if id_number:
            check_unique(Driver, 'id_number', id_number, label='ID number', exclude_id=d.id)
        paired_driver_id = form_int(request.form, 'paired_driver_id', required=False) if role == 'conductor' else None
        if paired_driver_id == d.id:
            raise ValueError('A conductor cannot be paired with themself.')
        d.name = request.form['name'].strip()
        d.license_number = license_number
        d.id_number = id_number
        d.phone = request.form.get('phone', '').strip()
        d.role = role
        d.commission_rate = rate_input / 100 if rate_input is not None else None
        d.status = request.form.get('status', 'active')
        d.paired_driver_id = paired_driver_id
        d.assigned_vehicle_id = form_int(request.form, 'assigned_vehicle_id', required=False)
        d.next_of_kin_name = request.form.get('next_of_kin_name', '').strip()
        d.next_of_kin_phone = request.form.get('next_of_kin_phone', '').strip()
        d.next_of_kin_relationship = request.form.get('next_of_kin_relationship', '').strip()
        log_audit('UPDATE', 'drivers', d.id, f'Updated driver {d.name}')
        touch_sync_fields(d)
        db.session.commit()
        flash(f'Driver {d.name} updated.', 'success')
        return redirect(url_for('drivers'))
    eligible_drivers = Driver.query.filter_by(role='driver', status='active').order_by(Driver.name).all()
    all_vehicles = Vehicle.query.filter_by(status='active').order_by(Vehicle.registration).all()
    return render_template('drivers/form.html', driver=d, action='Edit',
                           eligible_drivers=eligible_drivers, vehicles=all_vehicles)


@app.route('/drivers/<int:did>/delete', methods=['POST'])
@login_required
@admin_required
def driver_delete(did):
    d = Driver.query.filter_by(id=did).first_or_404()
    name = d.name
    log_audit('DELETE', 'drivers', did, f'Deleted driver {name}')
    d.deleted_at = datetime.now(timezone.utc)
    touch_sync_fields(d)
    db.session.commit()
    flash(f'Driver {name} removed.', 'warning')
    return redirect(url_for('drivers'))


@app.route('/drivers/roster')
@login_required
@permission_required('drivers')
def driver_roster():
    """Which vehicle each driver (and conductor) actually drove over a date
    range (a single day up to a full month or more) — read straight off
    each day's Daily Transaction entries rather than a separate
    planned-assignment table, since the vehicle is already captured
    alongside the fare/trips when the entry is logged."""
    df, dt = query_date_range()
    date_from_str, date_to_str = df.strftime('%Y-%m-%d'), dt.strftime('%Y-%m-%d')
    driver_id = request.args.get('driver_id', type=int)

    q = DailyLog.query.filter(DailyLog.log_date.between(df, dt), DailyLog.driver_id.isnot(None))
    if driver_id:
        q = q.filter(DailyLog.driver_id == driver_id)
    logs = q.join(Driver, DailyLog.driver_id == Driver.id).order_by(
        Driver.name, DailyLog.log_date, DailyLog.id).all()

    # How many distinct vehicles each driver used in the period, so a driver
    # who bounced between vehicles stands out at a glance in a multi-day view.
    vehicles_by_driver = {}
    for log in logs:
        vehicles_by_driver.setdefault(log.driver_id, set()).add(log.vehicle_id)

    all_drivers = Driver.query.filter_by(status='active').order_by(Driver.name).all()
    total_revenue = sum((log.gross_revenue or 0) for log in logs)
    return render_template('drivers/roster.html', logs=logs, drivers=all_drivers,
        vehicles_by_driver=vehicles_by_driver,
        date_from=date_from_str, date_to=date_to_str, driver_id=driver_id,
        total_revenue=total_revenue)


@app.route('/drivers/roster/export.xlsx')
@login_required
@permission_required('drivers')
def driver_roster_export_excel():
    df, dt = query_date_range()
    date_from_str, date_to_str = df.strftime('%Y-%m-%d'), dt.strftime('%Y-%m-%d')
    driver_id = request.args.get('driver_id', type=int)

    q = DailyLog.query.filter(DailyLog.log_date.between(df, dt), DailyLog.driver_id.isnot(None))
    if driver_id:
        q = q.filter(DailyLog.driver_id == driver_id)
    logs = q.join(Driver, DailyLog.driver_id == Driver.id).order_by(
        Driver.name, DailyLog.log_date, DailyLog.id).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Driver Roster'
    bold = Font(bold=True)
    money_fmt = '#,##0.00'

    ws.append(['Driver Roster'])
    ws['A1'].font = Font(bold=True, size=14)
    ws.append([f'Period: {date_from_str} to {date_to_str}'])
    if driver_id:
        driver = Driver.query.get(driver_id)
        ws.append([f'Driver: {driver.name if driver else "Unknown"}'])
    ws.append([f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")} by {current_user.username}'])
    ws.append([])

    headers = ['Driver', 'Role', 'Date', 'Vehicle', 'Conductor', 'Route', 'Trips', 'Revenue']
    ws.append(headers)
    for cell in ws[ws.max_row]:
        cell.font = bold

    total_revenue = 0
    for log in logs:
        ws.append([
            log.driver.name,
            log.driver.role.title(),
            log.log_date.strftime('%Y-%m-%d'),
            log.vehicle.registration if log.vehicle else '',
            log.conductor.name if log.conductor else '',
            log.route.name if log.route else '',
            log.trips_completed,
            log.gross_revenue or 0,
        ])
        row = ws.max_row
        ws[f'H{row}'].number_format = money_fmt
        total_revenue += float(log.gross_revenue or 0)

    ws.append([])
    ws.append(['', '', '', '', '', 'TOTAL', '', total_revenue])
    for cell in ws[ws.max_row]:
        cell.font = bold
    ws[f'H{ws.max_row}'].number_format = money_fmt

    widths = [28, 12, 14, 18, 18, 18, 10, 14]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    resp = make_response(out.getvalue())
    resp.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    resp.headers['Content-Disposition'] = f'attachment; filename=driver_roster_{date_from_str}_to_{date_to_str}.xlsx'
    return resp


@app.route('/drivers/roster/export.pdf')
@login_required
@permission_required('drivers')
def driver_roster_export_pdf():
    df, dt = query_date_range()
    date_from_str, date_to_str = df.strftime('%Y-%m-%d'), dt.strftime('%Y-%m-%d')
    driver_id = request.args.get('driver_id', type=int)

    q = DailyLog.query.filter(DailyLog.log_date.between(df, dt), DailyLog.driver_id.isnot(None))
    if driver_id:
        q = q.filter(DailyLog.driver_id == driver_id)
    logs = q.join(Driver, DailyLog.driver_id == Driver.id).order_by(
        Driver.name, DailyLog.log_date, DailyLog.id).all()

    styles = getSampleStyleSheet()
    elements = [
        Paragraph('Driver Roster', styles['Title']),
        Paragraph(f'Period: {date_from_str} to {date_to_str}', styles['Normal']),
    ]
    if driver_id:
        driver = Driver.query.get(driver_id)
        elements.append(Paragraph(f'Driver: {driver.name if driver else "Unknown"}', styles['Normal']))
    elements.append(Paragraph(f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")} by {current_user.username}', styles['Normal']))
    elements.append(Spacer(1, 10))

    headers = ['Driver', 'Role', 'Date', 'Vehicle', 'Conductor', 'Route', 'Trips', 'Revenue']
    data = [headers]
    total_revenue = 0
    for log in logs:
        row_revenue = float(log.gross_revenue or 0)
        data.append([
            log.driver.name,
            log.driver.role.title(),
            log.log_date.strftime('%Y-%m-%d'),
            log.vehicle.registration if log.vehicle else '',
            log.conductor.name if log.conductor else '',
            log.route.name if log.route else '',
            str(log.trips_completed),
            f'${row_revenue:,.2f}',
        ])
        total_revenue += row_revenue

    data.append(['', '', '', '', '', 'TOTAL', '', f'${total_revenue:,.2f}'])
    elements.append(_pdf_table(data))

    return _pdf_response(f'driver_roster_{date_from_str}_to_{date_to_str}.pdf', elements, pagesize=landscape(A4))


# ─────────────────────────────────────────────────────────────
# Routes
# ─────────────────────────────────────────────────────────────
@app.route('/routes')
@login_required
@permission_required('routes')
def routes_list():
    all_routes = Route.query.order_by(Route.name).all()
    return render_template('routes/index.html', routes=all_routes)


@app.route('/routes/add', methods=['GET', 'POST'])
@login_required
@permission_required('routes')
@handle_form_errors
def route_add():
    if request.method == 'POST':
        r = Route(
            name=request.form['name'].strip(),
            start_point=request.form['start_point'].strip(),
            end_point=request.form['end_point'].strip(),
            distance_km=form_float(request.form, 'distance_km', required=False, min_value=0),
            fare_rate=form_float(request.form, 'fare_rate', min_value=0),
            status=request.form.get('status', 'active'),
        )
        db.session.add(r)
        db.session.flush()
        log_audit('CREATE', 'routes', r.id, f'Added route {r.name}')
        touch_sync_fields(r)
        db.session.commit()
        flash(f'Route "{r.name}" added.', 'success')
        return redirect(url_for('routes_list'))
    return render_template('routes/form.html', route=None, action='Add')


@app.route('/routes/<int:rid>/edit', methods=['GET', 'POST'])
@login_required
@permission_required('routes')
@handle_form_errors
def route_edit(rid):
    r = Route.query.filter_by(id=rid).first_or_404()
    if request.method == 'POST':
        r.name = request.form['name'].strip()
        r.start_point = request.form['start_point'].strip()
        r.end_point = request.form['end_point'].strip()
        r.distance_km = form_float(request.form, 'distance_km', required=False, min_value=0)
        r.fare_rate = form_float(request.form, 'fare_rate', min_value=0)
        r.status = request.form.get('status', 'active')
        log_audit('UPDATE', 'routes', r.id, f'Updated route {r.name}')
        touch_sync_fields(r)
        db.session.commit()
        flash(f'Route "{r.name}" updated.', 'success')
        return redirect(url_for('routes_list'))
    return render_template('routes/form.html', route=r, action='Edit')


@app.route('/routes/<int:rid>/delete', methods=['POST'])
@login_required
@admin_required
def route_delete(rid):
    r = Route.query.filter_by(id=rid).first_or_404()
    name = r.name
    log_audit('DELETE', 'routes', rid, f'Deleted route {name}')
    r.deleted_at = datetime.now(timezone.utc)
    touch_sync_fields(r)
    db.session.commit()
    flash(f'Route "{name}" deleted.', 'warning')
    return redirect(url_for('routes_list'))


# ─────────────────────────────────────────────────────────────
# Daily Transactions — edit/delete a single vehicle/date entry from the
# Vehicle Ledger below. This replaces the old standalone Daily Logs
# CRUD pages, which duplicated the same DailyLog data behind a second,
# heavier form — everything now lives on one page: /logs/ledger.
# ─────────────────────────────────────────────────────────────
@app.route('/logs/ledger/<int:vehicle_id>/<log_date_str>/edit', methods=['GET', 'POST'])
@login_required
@permission_required_any('daily_logs', 'crew_portal')
@handle_form_errors
def ledger_entry_edit(vehicle_id, log_date_str):
    vehicle = Vehicle.query.filter_by(id=vehicle_id).first_or_404()
    try:
        log_date = parse_date(log_date_str)
    except ValueError:
        flash(f'"{log_date_str}" is not a valid date.', 'danger')
        return redirect(url_for('driver_ledger', vehicle_id=vehicle_id))

    period = request.values.get('period', 'month')
    date_from = request.values.get('date_from', '')
    date_to = request.values.get('date_to', '')
    daily_logs_for_date = DailyLog.query.filter_by(
        vehicle_id=vehicle_id, log_date=log_date).order_by(DailyLog.id).all()
    fuel_logs_for_date = FuelLog.query.filter_by(
        vehicle_id=vehicle_id, log_date=log_date).order_by(FuelLog.id).all()
    if len(daily_logs_for_date) > 1 or len(fuel_logs_for_date) > 1:
        flash('This day has more than one entry for this vehicle and can\'t be edited '
              'as a single row — delete it and re-enter instead.', 'warning')
        return redirect(url_for('driver_ledger', vehicle_id=vehicle_id, period=period, date_from=date_from, date_to=date_to))

    log = daily_logs_for_date[0] if daily_logs_for_date else None
    fuel = fuel_logs_for_date[0] if fuel_logs_for_date else None

    if request.method == 'POST':
        new_log_date = parse_date(request.form['log_date'])
        driver_id = form_int(request.form, 'driver_id', required=False)
        fare = form_float(request.form, 'fare', required=False, min_value=0)
        garnish = form_float(request.form, 'garnish', required=False, min_value=0)
        reason_for_shortfall = request.form.get('reason_for_shortfall', '').strip() or None
        diesel_cost = form_float(request.form, 'diesel_cost', required=False, min_value=0)
        mileage = form_float(request.form, 'mileage', required=False, min_value=0)

        if log is not None or fare is not None or garnish is not None:
            if fare is None:
                raise ValueError('Fare is required for this entry.')
        if garnish is not None and not driver_id:
            raise ValueError('Select a driver to record a garnish against.')

        if fare is None and diesel_cost is None and mileage is None and garnish is None:
            raise ValueError('Enter at least a fare, diesel cost, mileage reading, or garnish.')

        if new_log_date != log_date:
            conflict = (DailyLog.query.filter_by(vehicle_id=vehicle_id, log_date=new_log_date).first()
                        or FuelLog.query.filter_by(vehicle_id=vehicle_id, log_date=new_log_date).first())
            if conflict:
                raise ValueError(f'{vehicle.registration} already has an entry for '
                                  f'{new_log_date.strftime("%d %b %Y")} — edit that one instead.')

        if fare is not None:
            driver = Driver.query.filter_by(id=driver_id).first()
            conductor = resolve_conductor(driver, vehicle_id)
            if log is None:
                log = DailyLog(vehicle_id=vehicle_id, log_date=new_log_date, created_by=current_user.id)
                db.session.add(log)
            log.log_date = new_log_date
            log.driver_id = driver_id
            log.conductor_id = conductor.id if conductor else None
            log.gross_revenue = fare
            log.garnish = garnish or 0.0
            log.reason_for_shortfall = reason_for_shortfall
            log.updated_by = current_user.id
            log.updated_at = datetime.now(timezone.utc)
            touch_sync_fields(log)

        if diesel_cost is not None or mileage is not None:
            if fuel is None:
                fuel = FuelLog(vehicle_id=vehicle_id, log_date=new_log_date, liters=0, created_by=current_user.id)
                db.session.add(fuel)
            fuel.log_date = new_log_date
            price = fuel_price_for(vehicle.fuel_type)
            fuel.total_cost = diesel_cost or 0
            fuel.cost_per_liter = price or 0.0
            fuel.liters = round(fuel.total_cost / price, 2) if price and fuel.total_cost else 0
            fuel.odometer = mileage
            touch_sync_fields(fuel)
        elif fuel is not None:
            fuel.deleted_at = datetime.now(timezone.utc)
            touch_sync_fields(fuel)

        log_audit('UPDATE', 'daily_logs', log.id if log else None,
                  f'Edited ledger entry for {vehicle.registration} on {new_log_date}'
                  + (f' (moved from {log_date})' if new_log_date != log_date else ''))
        db.session.commit()
        flash('Entry updated.', 'success')
        return redirect(url_for('driver_ledger', vehicle_id=vehicle_id, period=period, date_from=date_from, date_to=date_to))

    all_drivers = Driver.query.filter_by(role='driver', status='active').order_by(Driver.name).all()
    return render_template('logs/ledger_entry_form.html', vehicle=vehicle, log=log, fuel=fuel,
                           log_date=log_date, drivers=all_drivers, period=period,
                           date_from=date_from, date_to=date_to)


@app.route('/logs/ledger/<int:vehicle_id>/<log_date_str>/delete', methods=['POST'])
@login_required
@admin_required
def ledger_entry_delete(vehicle_id, log_date_str):
    vehicle = Vehicle.query.filter_by(id=vehicle_id).first_or_404()
    try:
        log_date = parse_date(log_date_str)
    except ValueError:
        flash(f'"{log_date_str}" is not a valid date.', 'danger')
        return redirect(url_for('driver_ledger', vehicle_id=vehicle_id))

    period = request.form.get('period', 'month')
    date_from = request.form.get('date_from', '')
    date_to = request.form.get('date_to', '')
    # Soft-delete via a per-row loop, not a bulk .delete() — bulk deletes
    # issue a raw DELETE and bypass touch_sync_fields entirely, so the
    # tombstone would never propagate to other instances.
    now = datetime.now(timezone.utc)
    for log in DailyLog.query.filter_by(vehicle_id=vehicle_id, log_date=log_date).all():
        log.deleted_at = now
        touch_sync_fields(log)
    for fuel in FuelLog.query.filter_by(vehicle_id=vehicle_id, log_date=log_date).all():
        fuel.deleted_at = now
        touch_sync_fields(fuel)
    log_audit('DELETE', 'daily_logs', None, f'Deleted ledger entry for {vehicle.registration} on {log_date}')
    db.session.commit()
    flash('Entry deleted.', 'warning')
    return redirect(url_for('driver_ledger', vehicle_id=vehicle_id, period=period, date_from=date_from, date_to=date_to))


# ─────────────────────────────────────────────────────────────
# Vehicle Ledger — one running sheet per vehicle (date, driver, fare,
# diesel, mileage), matching how the fleet's paper/Excel logbooks are
# kept — one sheet per vehicle, driver rotating day to day. Posts to
# the same DailyLog/FuelLog tables the rest of the system uses.
# Replaces the old Crew Portal "Log Income" form. Filterable by
# day/week/month. Diesel is captured as a USD amount, not liters — crew
# report what they spent on fuel, not a metered liter reading. Liters are
# derived automatically from that amount using the vehicle's fuel-type
# price (see fuel_price_for) so the Fuel Efficiency report can still use
# these rows; if no price has been set yet, liters stays 0 and the row is
# skipped there rather than showing a false 0 L/100km.
# ─────────────────────────────────────────────────────────────
def vehicle_ledger_rows(vehicle_id, df=None, dt=None, daily_target=None):
    """Merge DailyLog (fare, any driver) and FuelLog (diesel cost/mileage)
    for this vehicle by date, with distance computed the same way the Fuel
    Efficiency report does — delta from the previous odometer reading.
    If df/dt are given, only rows in that range are returned, but the
    distance baseline still uses the last odometer reading before df so
    the first visible row isn't wrongly shown as having no distance.
    If daily_target is given, each day with a driver entry is flagged
    against it (see 'shortfall' on each row) — the same target used by
    the Revenue Shortfalls report, surfaced here at entry time too."""
    daily_q = DailyLog.query.filter_by(vehicle_id=vehicle_id)
    fuel_q = FuelLog.query.filter_by(vehicle_id=vehicle_id)
    if df:
        daily_q = daily_q.filter(DailyLog.log_date >= df)
        fuel_q = fuel_q.filter(FuelLog.log_date >= df)
    if dt:
        daily_q = daily_q.filter(DailyLog.log_date <= dt)
        fuel_q = fuel_q.filter(FuelLog.log_date <= dt)

    daily_by_date = {}
    for log in daily_q.order_by(DailyLog.log_date).all():
        daily_by_date.setdefault(log.log_date, []).append(log)

    fuel_by_date = {}
    for log in fuel_q.order_by(FuelLog.log_date).all():
        fuel_by_date.setdefault(log.log_date, []).append(log)

    prev_odometer = None
    if df:
        baseline = FuelLog.query.filter(
            FuelLog.vehicle_id == vehicle_id, FuelLog.log_date < df,
            FuelLog.odometer.isnot(None)).order_by(FuelLog.log_date.desc()).first()
        prev_odometer = baseline.odometer if baseline else None

    all_dates = sorted(set(daily_by_date) | set(fuel_by_date))
    rows = []
    total_fare = 0.0
    total_diesel_cost = 0.0
    total_garnish = 0.0
    for d in all_dates:
        daily_logs = daily_by_date.get(d, [])
        fare = sum(l.gross_revenue for l in daily_logs)
        garnish = sum(l.garnish for l in daily_logs)
        reason_for_shortfall = '; '.join(n for n in (l.reason_for_shortfall for l in daily_logs) if n) or None
        driver_names = ', '.join(sorted({l.driver.name for l in daily_logs if l.driver})) or None
        fuel_logs = fuel_by_date.get(d, [])
        diesel_cost = sum(f.total_cost for f in fuel_logs)
        odometer = max((f.odometer for f in fuel_logs if f.odometer is not None), default=None)

        distance = None
        if odometer is not None and prev_odometer is not None:
            distance = odometer - prev_odometer
        if odometer is not None:
            prev_odometer = odometer

        total_fare += fare
        total_diesel_cost += diesel_cost
        total_garnish += garnish
        # Only flag a day that actually has a driver/fare entry — a
        # fuel-only row shouldn't read as a missed revenue target.
        shortfall = None
        if daily_target and daily_logs and fare < daily_target:
            shortfall = daily_target - fare
        rows.append({
            'date': d, 'driver_names': driver_names, 'fare': fare,
            'garnish': garnish, 'reason_for_shortfall': reason_for_shortfall,
            'shortfall': shortfall,
            'diesel_cost': diesel_cost,
            'odometer': odometer, 'distance': distance,
            # More than one entry on the same day (e.g. a driver change
            # mid-shift) can't be represented/edited as a single row —
            # the UI falls back to "delete all, re-enter" for these.
            'multiple': len(daily_logs) > 1 or len(fuel_logs) > 1,
        })
    return rows, total_fare, total_diesel_cost, total_garnish


def resolve_ledger_period(period, today):
    if period == 'today':
        df = dt = today
    elif period == 'week':
        df, dt = today - timedelta(days=today.weekday()), today
    elif period == 'all':
        df = dt = None
    elif period == 'custom':
        from_str = request.args.get('date_from', '').strip()
        to_str = request.args.get('date_to', '').strip()
        try:
            df = parse_date(from_str) if from_str else today.replace(day=1)
        except ValueError:
            flash(f'"{from_str}" is not a valid start date — showing {today.replace(day=1)} instead.', 'warning')
            df = today.replace(day=1)
        try:
            dt = parse_date(to_str) if to_str else today
        except ValueError:
            flash(f'"{to_str}" is not a valid end date — showing {today} instead.', 'warning')
            dt = today
        if df > dt:
            df, dt = dt, df
    else:
        period = 'month'
        df, dt = today.replace(day=1), today
    return period, df, dt


@app.route('/logs/ledger')
@login_required
@permission_required_any('daily_logs', 'crew_portal')
def driver_ledger():
    today = date.today()

    period, df, dt = resolve_ledger_period(request.args.get('period', 'month'), today)
    date_from_str = df.strftime('%Y-%m-%d') if df else ''
    date_to_str = dt.strftime('%Y-%m-%d') if dt else ''

    all_vehicles = Vehicle.query.filter_by(status='active').order_by(Vehicle.registration).all()
    vehicle_id = request.args.get('vehicle_id', '')
    vehicle = Vehicle.query.filter_by(id=vehicle_id).first() if vehicle_id else (all_vehicles[0] if all_vehicles else None)

    rows, total_fare, total_diesel_cost, total_garnish = [], 0.0, 0.0, 0.0
    latest_odometer = None
    if vehicle:
        rows, total_fare, total_diesel_cost, total_garnish = vehicle_ledger_rows(
            vehicle.id, df, dt, daily_target=vehicle.daily_target)
        latest_fuel = FuelLog.query.filter(
            FuelLog.vehicle_id == vehicle.id, FuelLog.odometer.isnot(None)
        ).order_by(FuelLog.log_date.desc(), FuelLog.id.desc()).first()
        latest_odometer = latest_fuel.odometer if latest_fuel else None

    all_drivers = Driver.query.filter_by(role='driver', status='active').order_by(Driver.name).all()
    return render_template('logs/ledger.html', vehicles=all_vehicles, vehicle=vehicle,
        drivers=all_drivers,
        rows=rows, total_fare=total_fare, total_diesel_cost=total_diesel_cost,
        total_garnish=total_garnish,
        period=period, date_from=date_from_str, date_to=date_to_str,
        today=today.strftime('%Y-%m-%d'), latest_odometer=latest_odometer)


@app.route('/logs/ledger/add', methods=['POST'])
@login_required
@permission_required_any('daily_logs', 'crew_portal')
def driver_ledger_add():
    # This is a POST-only route with no GET counterpart at the same URL, so
    # (unlike the other forms in the app) errors can't redirect back to
    # request.url — that would GET this same POST-only URL and 405. Errors
    # are handled locally here and always redirect to the GET ledger page.
    period = request.form.get('period', 'month')
    date_from = request.form.get('date_from', '')
    date_to = request.form.get('date_to', '')
    vehicle_id = form_int(request.form, 'vehicle_id', required=False)
    client_id = request.form.get('_client_id')
    if already_synced(client_id):
        flash('Already recorded.', 'info')
        return redirect(url_for('driver_ledger', vehicle_id=vehicle_id, period=period, date_from=date_from, date_to=date_to))
    try:
        vehicle = Vehicle.query.filter_by(id=vehicle_id).first() if vehicle_id else None
        if not vehicle:
            raise ValueError('Select a vehicle.')

        log_date = parse_date(request.form['log_date'])
        driver_id = form_int(request.form, 'driver_id', required=False)
        fare = form_float(request.form, 'fare', required=False, min_value=0)
        garnish = form_float(request.form, 'garnish', required=False, min_value=0)
        reason_for_shortfall = request.form.get('reason_for_shortfall', '').strip() or None
        diesel_cost = form_float(request.form, 'diesel_cost', required=False, min_value=0)
        mileage = form_float(request.form, 'mileage', required=False, min_value=0)

        if garnish is not None and not driver_id:
            raise ValueError('Select a driver to record a garnish against.')
        if driver_id and fare is None and garnish is None:
            raise ValueError('Fare or a garnish is required when a driver is selected.')

        if fare is None and diesel_cost is None and mileage is None and garnish is None:
            raise ValueError('Enter at least a fare, diesel cost, mileage reading, or garnish.')

        if fare is not None or garnish is not None:
            # Reject an exact repeat of an already-recorded entry for this
            # vehicle, date and driver — same fare and garnish as an
            # existing (non-deleted) row is virtually certainly the same
            # transaction re-entered (a double-tapped submit, or someone
            # re-keying a fare they already logged), not a second genuine
            # shift. A different driver, fare, or garnish for the same
            # vehicle/date still goes through — multiple genuine entries
            # per day (different drivers/shifts) are expected and already
            # relied on elsewhere (see the day-total shortfall check below).
            dup = DailyLog.query.filter_by(vehicle_id=vehicle_id, log_date=log_date, driver_id=driver_id,
                                            gross_revenue=fare or 0.0, garnish=garnish or 0.0).first()
            if dup:
                raise ValueError(f'A ledger entry for {vehicle.registration} on {log_date} with the same '
                                  f'driver, fare and garnish already exists — this looks like a duplicate submission.')

            driver = Driver.query.filter_by(id=driver_id).first()
            conductor = resolve_conductor(driver, vehicle_id)
            daily = DailyLog(
                vehicle_id=vehicle_id, driver_id=driver_id, conductor_id=conductor.id if conductor else None,
                log_date=log_date, gross_revenue=fare or 0.0,
                garnish=garnish or 0.0, reason_for_shortfall=reason_for_shortfall,
                created_by=current_user.id,
            )
            db.session.add(daily)
            touch_sync_fields(daily)
            log_audit('CREATE', 'daily_logs', None,
                       f'Ledger entry for {vehicle.registration} on {log_date}: fare {fare or 0.0}' +
                       (f', garnish {garnish} ({reason_for_shortfall})' if garnish else ''))

        if diesel_cost is not None:
            # Same principle as the fare/garnish check above — an identical
            # diesel cost already logged for this vehicle/date is a
            # duplicate submission, not a second genuine fill-up (a real
            # second fill-up on the same day would almost never cost the
            # exact same amount).
            dup_fuel = FuelLog.query.filter_by(vehicle_id=vehicle_id, log_date=log_date, total_cost=diesel_cost).first()
            if dup_fuel:
                raise ValueError(f'A fuel entry for {vehicle.registration} on {log_date} with the same diesel '
                                  f'cost already exists — this looks like a duplicate submission.')

            price = fuel_price_for(vehicle.fuel_type)
            fuel = FuelLog(
                vehicle_id=vehicle_id, log_date=log_date,
                liters=round(diesel_cost / price, 2) if price else 0,
                cost_per_liter=price or 0.0,
                total_cost=diesel_cost, odometer=mileage, created_by=current_user.id,
            )
            db.session.add(fuel)
            touch_sync_fields(fuel)
            log_audit('CREATE', 'fuel_logs', None, f'Ledger entry for {vehicle.registration} on {log_date}: diesel ${diesel_cost}')
        elif mileage is not None:
            dup_fuel = FuelLog.query.filter_by(vehicle_id=vehicle_id, log_date=log_date, liters=0, odometer=mileage).first()
            if dup_fuel:
                raise ValueError(f'An odometer reading for {vehicle.registration} on {log_date} with the same '
                                  f'value already exists — this looks like a duplicate submission.')

            fuel = FuelLog(
                vehicle_id=vehicle_id, log_date=log_date, liters=0, odometer=mileage, created_by=current_user.id,
            )
            db.session.add(fuel)
            touch_sync_fields(fuel)

        record_offline_sync(client_id, 'driver_ledger_add')
        db.session.commit()
        flash('Ledger entry recorded.', 'success')

        # Auto-flag a shortfall right at entry time, not just later on the
        # Revenue Shortfalls report — checks the day's total fare (this entry
        # plus any others already logged for the same vehicle/date).
        if fare is not None and vehicle.daily_target:
            day_total = db.session.query(func.sum(DailyLog.gross_revenue)).filter(
                DailyLog.vehicle_id == vehicle_id, DailyLog.log_date == log_date).scalar() or 0
            if day_total < vehicle.daily_target:
                gap = vehicle.daily_target - day_total
                flash(f'Flagged: {vehicle.registration} is ${gap:,.2f} short of its ${vehicle.daily_target:,.2f} '
                      f'daily target on {log_date} — see Revenue Shortfalls to garnish.', 'warning')
    except KeyError as e:
        db.session.rollback()
        flash(f'Missing required field: {e}', 'danger')
    except ValueError as e:
        db.session.rollback()
        flash(str(e), 'danger')

    return redirect(url_for('driver_ledger', vehicle_id=vehicle_id, period=period, date_from=date_from, date_to=date_to))


@app.route('/logs/ledger/export')
@login_required
@permission_required_any('daily_logs', 'crew_portal')
def driver_ledger_export():
    vehicle_id = request.args.get('vehicle_id', '')
    vehicle = Vehicle.query.filter_by(id=vehicle_id).first() if vehicle_id else None
    if not vehicle:
        flash('Select a vehicle to export.', 'danger')
        return redirect(url_for('driver_ledger'))

    period, df, dt = resolve_ledger_period(request.args.get('period', 'month'), date.today())
    ledger_rows, _, _, _ = vehicle_ledger_rows(vehicle.id, df, dt)

    fuel_label = vehicle.fuel_type.capitalize()
    header = ['Date', 'Driver', 'Fare', f'{fuel_label} (USD)', 'Mileage', 'Distance',
              'Garnish', 'Reason for Shortfall']
    rows = [[
        row['date'], row['driver_names'] or '',
        f"{row['fare']:.2f}" if row['fare'] else '',
        f"{row['diesel_cost']:.2f}" if row['diesel_cost'] else '',
        row['odometer'] if row['odometer'] is not None else '',
        row['distance'] if row['distance'] is not None else '',
        f"{row['garnish']:.2f}" if row['garnish'] else '',
        row['reason_for_shortfall'] or '',
    ] for row in ledger_rows]

    safe_reg = vehicle.registration.replace(' ', '_')
    if request.args.get('format') == 'pdf':
        return _table_pdf_response(f'{safe_reg}_ledger_{period}_{date.today()}.pdf',
            f'Driver Ledger — {vehicle.registration}', f'Period: {df} to {dt}', header, rows)
    return csv_export_response(f'{safe_reg}_ledger_{period}_{date.today()}.csv', header, rows)


# ─────────────────────────────────────────────────────────────
# Driver Deposits — cash a driver banks for a date, reconciled against the
# fleet's own net income (DailyLog.gross_revenue minus driver-attributed
# Expense rows) on the Fleet Reconciliation reports below.
# ─────────────────────────────────────────────────────────────
@app.route('/logs/deposits')
@login_required
@permission_required('daily_logs')
def driver_deposits_list():
    page = request.args.get('page', 1, type=int)
    driver_id = request.args.get('driver_id', type=int)
    df, dt = query_date_range()

    query = DriverDeposit.query.filter(DriverDeposit.deposit_date.between(df, dt))
    if driver_id:
        query = query.filter_by(driver_id=driver_id)
    deposits = query.order_by(DriverDeposit.deposit_date.desc(), DriverDeposit.id.desc()).paginate(page=page, per_page=20)
    total = query.with_entities(func.sum(DriverDeposit.amount)).scalar() or 0

    all_drivers = Driver.query.filter_by(role='driver', status='active').order_by(Driver.name).all()
    all_vehicles = Vehicle.query.filter_by(status='active').order_by(Vehicle.registration).all()
    return render_template('fleet/deposits.html', deposits=deposits, drivers=all_drivers, vehicles=all_vehicles,
                           driver_id=driver_id, date_from=df.strftime('%Y-%m-%d'), date_to=dt.strftime('%Y-%m-%d'),
                           total=total, today=date.today().strftime('%Y-%m-%d'))


@app.route('/logs/deposits/add', methods=['POST'])
@login_required
@permission_required('daily_logs')
def driver_deposit_add():
    # POST-only, no GET counterpart at this URL (see driver_ledger_add) —
    # errors are handled locally here and always redirect to the GET list page.
    client_id = request.form.get('_client_id')
    if already_synced(client_id):
        flash('Already recorded.', 'info')
        return redirect(url_for('driver_deposits_list'))
    try:
        driver_id = form_int(request.form, 'driver_id', required=False)
        driver = None
        vehicle_id = None
        if driver_id:
            driver = Driver.query.filter_by(id=driver_id).first()
            if not driver:
                raise ValueError('Select a valid driver.')
            # Falls back to the driver's normally assigned vehicle when the form
            # leaves it blank — keeps the vehicle-consolidated report accurate
            # without forcing every entry to pick a vehicle explicitly.
            vehicle_id = form_int(request.form, 'vehicle_id', required=False) or driver.assigned_vehicle_id
        # driver_id left blank means a fleet-wide total — one combined cash
        # figure for all vehicles that date, not attributed to a driver.

        deposit = DriverDeposit(
            driver_id=driver.id if driver else None,
            vehicle_id=vehicle_id,
            deposit_date=parse_date(request.form['deposit_date']),
            amount=form_float(request.form, 'amount', min_value=0),
            notes=request.form.get('notes', '').strip(),
            created_by=current_user.id,
        )
        db.session.add(deposit)
        db.session.flush()
        depositor = driver.name if driver else 'all vehicles (fleet total)'
        log_audit('CREATE', 'driver_deposits', deposit.id,
                  f'Deposit of {deposit.amount} for {depositor} on {deposit.deposit_date}')
        record_offline_sync(client_id, 'driver_deposit_add')
        touch_sync_fields(deposit)
        db.session.commit()
        flash('Deposit recorded.', 'success')
    except KeyError as e:
        db.session.rollback()
        flash(f'Missing required field: {e}', 'danger')
    except ValueError as e:
        db.session.rollback()
        flash(str(e), 'danger')

    return redirect(url_for('driver_deposits_list'))


@app.route('/logs/deposits/<int:did>/delete', methods=['POST'])
@login_required
@admin_required
def driver_deposit_delete(did):
    deposit = DriverDeposit.query.filter_by(id=did).first_or_404()
    depositor = deposit.driver.name if deposit.driver else 'all vehicles (fleet total)'
    log_audit('DELETE', 'driver_deposits', did, f'Deleted deposit of {deposit.amount} for {depositor}')
    deposit.deleted_at = datetime.now(timezone.utc)
    touch_sync_fields(deposit)
    db.session.commit()
    flash('Deposit deleted.', 'warning')
    return redirect(url_for('driver_deposits_list'))


def _resolve_ledger_import_vehicle():
    period = request.form.get('period', 'month')
    vehicle_id = form_int(request.form, 'vehicle_id', required=False)
    vehicle = Vehicle.query.filter_by(id=vehicle_id).first() if vehicle_id else None
    return vehicle, vehicle_id, period


@app.route('/logs/ledger/import/preview', methods=['POST'])
@login_required
@permission_required_any('daily_logs', 'crew_portal')
def driver_ledger_import_preview():
    vehicle, vehicle_id, period = _resolve_ledger_import_vehicle()
    if not vehicle:
        flash('Select a vehicle before importing.', 'danger')
        return redirect(url_for('driver_ledger', period=period))

    file = request.files.get('file')
    if file and file.filename:
        filename = file.filename
        try:
            headers, raw_rows = read_uploaded_table(file, vehicle=vehicle)
        except ValueError as e:
            flash(str(e), 'danger')
            return redirect(url_for('driver_ledger', vehicle_id=vehicle_id, period=period))
        mapping = auto_map_columns(headers)
    else:
        # Re-preview after the user adjusted the mapping — the file itself
        # isn't resubmitted, the previously parsed rows travel via raw_data.
        try:
            filename = request.form.get('filename', 'uploaded file')
            payload = json.loads(request.form.get('raw_data') or '{}')
            headers, raw_rows = payload.get('headers') or [], payload.get('rows') or []
            if not headers or not raw_rows:
                raise ValueError('Choose a CSV or Excel file to import.')
            mapping = {field_key: (request.form.get(f'map_{field_key}') or None)
                       for field_key, _label, _syn in CANONICAL_LEDGER_FIELDS}
        except (ValueError, json.JSONDecodeError, TypeError):
            flash('Choose a CSV or Excel file to import — the previous preview session expired.', 'danger')
            return redirect(url_for('driver_ledger', vehicle_id=vehicle_id, period=period))

    if not raw_rows:
        flash('That file has no data rows to import — it only has a header row. '
              'Add rows with a Date and Fare/Diesel (USD)/Mileage, then re-import.', 'warning')
        return redirect(url_for('driver_ledger', vehicle_id=vehicle_id, period=period))

    preview_rows = apply_column_mapping(headers, raw_rows[:10], mapping)
    return render_template('logs/ledger_import_preview.html',
                           vehicle=vehicle, period=period, filename=filename,
                           headers=headers, mapping=mapping, fields=CANONICAL_LEDGER_FIELDS,
                           preview_rows=preview_rows, row_count=len(raw_rows),
                           raw_data=json.dumps({'headers': headers, 'rows': raw_rows}))


@app.route('/logs/ledger/import/confirm', methods=['POST'])
@login_required
@permission_required_any('daily_logs', 'crew_portal')
def driver_ledger_import_confirm():
    vehicle, vehicle_id, period = _resolve_ledger_import_vehicle()
    if not vehicle:
        flash('Select a vehicle before importing.', 'danger')
        return redirect(url_for('driver_ledger', period=period))

    filename = request.form.get('filename', 'uploaded file')
    try:
        payload = json.loads(request.form.get('raw_data') or '{}')
        headers, raw_rows = payload.get('headers') or [], payload.get('rows') or []
        if not raw_rows:
            raise ValueError('empty')
    except (ValueError, json.JSONDecodeError, TypeError):
        flash('That preview session expired — please choose the file again.', 'danger')
        return redirect(url_for('driver_ledger', vehicle_id=vehicle_id, period=period))

    mapping = {field_key: (request.form.get(f'map_{field_key}') or None)
               for field_key, _label, _syn in CANONICAL_LEDGER_FIELDS}
    file_rows = apply_column_mapping(headers, raw_rows, mapping)
    imported, errors, error_rows, created_drivers, created_records = import_ledger_rows(
        file_rows, vehicle, auto_register_drivers=True)

    if imported or error_rows:
        # Commit even when imported == 0: a batch made only of failed rows
        # still needs to persist so its quarantine CSV can be downloaded.
        save_import_batch('ledger', filename, len(raw_rows), imported, error_rows, created_records)
        if imported:
            log_audit('CREATE', 'daily_logs', None,
                       f'Imported {imported} ledger row(s) for {vehicle.registration} from {filename}')
        for driver in Driver.query.filter(Driver.name.in_(created_drivers)).all():
            log_audit('CREATE', 'drivers', driver.id,
                       f'Auto-registered driver "{driver.name}" from ledger import ({filename}) — '
                       f'not on file, added because a fare row named them.')
        db.session.commit()
    else:
        db.session.rollback()

    if imported:
        flash(f'Imported {imported} row(s) for {vehicle.registration}.', 'success')
    if created_drivers:
        flash(f'Auto-registered new driver(s): {", ".join(created_drivers)}.', 'success')
    if errors:
        shown = errors[:10]
        more = f' (+{len(errors) - 10} more)' if len(errors) > 10 else ''
        flash('Skipped rows — ' + '; '.join(shown) + more, 'warning')
    if not imported and not errors:
        flash('No rows found to import.', 'warning')

    return redirect(url_for('driver_ledger', vehicle_id=vehicle_id, period=period))


@app.route('/logs/ledger/import/bulk', methods=['POST'])
@login_required
@permission_required_any('daily_logs', 'crew_portal')
def driver_ledger_import_bulk():
    """Import a fleet-wide workbook with one sheet per vehicle (sheet name
    matched against each vehicle's registration) in a single pass. Unlike the
    single-vehicle import there's no manual column-mapping step — with dozens
    of sheets that isn't practical, so each sheet is auto-mapped the same way
    and the results page reports what happened per vehicle."""
    period = request.form.get('period', 'month')
    file = request.files.get('file')
    if not file or not file.filename:
        flash('Choose an Excel workbook to import.', 'danger')
        return redirect(url_for('driver_ledger', period=period))

    try:
        sheets = read_uploaded_workbook_sheets(file)
    except ValueError as e:
        flash(str(e), 'danger')
        return redirect(url_for('driver_ledger', period=period))

    if not sheets:
        flash('That workbook has no sheets with data rows to import.', 'warning')
        return redirect(url_for('driver_ledger', period=period))

    auto_register = request.form.get('auto_register') == '1'
    vehicles_by_reg = {_normalize_registration(v.registration): v for v in Vehicle.query.all()}

    results = []
    for sheet_name, (headers, rows) in sheets.items():
        savepoint = db.session.begin_nested()
        try:
            mapping = auto_map_columns(headers)
            vehicle = vehicles_by_reg.get(_normalize_registration(sheet_name))
            vehicle_created = False

            # Distinguish an actual per-vehicle ledger tab from a fleet-wide
            # summary sheet like "DAILY TOTAL INCOME" before ever auto-registering
            # a "vehicle" for it. A summary sheet also has a Date column and its
            # "INCOME" header even fuzzy-matches the Fare synonym list, so Fare
            # alone isn't a safe signal — a real per-vehicle logbook always
            # names a Driver or logs Mileage; a financial rollup never does.
            looks_like_ledger = bool(mapping.get('date')) and bool(
                mapping.get('driver') or mapping.get('mileage'))

            if not vehicle and looks_like_ledger and auto_register:
                registration = re.sub(r'\s+', ' ', sheet_name).strip().upper()
                vehicle = Vehicle(registration=registration, make='Unknown', model='Unknown',
                                   year=date.today().year, status='active', fuel_type='diesel')
                db.session.add(vehicle)
                db.session.flush()
                vehicles_by_reg[_normalize_registration(registration)] = vehicle
                vehicle_created = True
                log_audit('CREATE', 'vehicles', vehicle.id,
                          f'Auto-registered vehicle {registration} from fleet workbook '
                          f'import (sheet "{sheet_name}") — make/model/year are placeholders.')

            if not vehicle:
                reason = ('No vehicle matches this sheet name.' if looks_like_ledger else
                          "Doesn't look like a per-vehicle ledger sheet.")
                results.append({'sheet': sheet_name, 'vehicle': None, 'mapping': {},
                                 'imported': 0, 'errors': [], 'created_drivers': [],
                                 'vehicle_created': False, 'skip_reason': reason})
                savepoint.commit()
                continue

            if not mapping.get('date'):
                results.append({'sheet': sheet_name, 'vehicle': vehicle, 'mapping': mapping,
                                 'imported': 0, 'errors': [], 'created_drivers': [],
                                 'vehicle_created': vehicle_created,
                                 'skip_reason': 'No Date column detected.'})
                savepoint.commit()
                continue

            mapped_rows = apply_column_mapping(headers, rows, mapping)
            imported, errors, _error_rows, created_drivers, _created_records = import_ledger_rows(
                mapped_rows, vehicle, auto_register_drivers=auto_register)
            if imported:
                log_audit('CREATE', 'daily_logs', None,
                           f'Imported {imported} ledger row(s) for {vehicle.registration} '
                           f'from {file.filename} (sheet "{sheet_name}")')
            for driver in Driver.query.filter(Driver.name.in_(created_drivers)).all():
                log_audit('CREATE', 'drivers', driver.id,
                           f'Auto-registered driver "{driver.name}" from fleet workbook import '
                           f'({file.filename}, sheet "{sheet_name}") — not on file, added because '
                           f'a fare row named them.')
            results.append({'sheet': sheet_name, 'vehicle': vehicle, 'mapping': mapping,
                             'imported': imported, 'errors': errors,
                             'created_drivers': created_drivers,
                             'vehicle_created': vehicle_created, 'skip_reason': None})
            savepoint.commit()
        except Exception as e:
            savepoint.rollback()
            results.append({'sheet': sheet_name, 'vehicle': None, 'mapping': {},
                             'imported': 0, 'errors': [], 'created_drivers': [],
                             'vehicle_created': False, 'skip_reason': f'Unexpected error: {e}'})

    total_imported = sum(r['imported'] for r in results)
    total_registered = (sum(1 for r in results if r['vehicle_created']) +
                        sum(len(r['created_drivers']) for r in results))
    if total_imported or total_registered:
        db.session.commit()
    else:
        db.session.rollback()

    return render_template('logs/ledger_bulk_import_result.html',
                           filename=file.filename, results=results,
                           total_imported=total_imported, period=period,
                           fields=CANONICAL_LEDGER_FIELDS)


# ─────────────────────────────────────────────────────────────
# Fuel Prices — admin-set current price per liter, used to auto-derive
# cost/liters on every fuel log (see fuel_log_add/edit and driver_ledger_add
# below) so fuel efficiency can be tracked without extra data entry.
# ─────────────────────────────────────────────────────────────
@app.route('/settings/fuel-prices', methods=['GET', 'POST'])
@login_required
@admin_required
@handle_form_errors
def fuel_prices():
    price = get_fuel_price_row()
    if not price:
        price = FuelPrice(id=1)
        db.session.add(price)
        db.session.flush()
    if request.method == 'POST':
        price.diesel_price = form_float(request.form, 'diesel_price', label='Diesel price',
                                        required=False, min_value=0)
        price.petrol_price = form_float(request.form, 'petrol_price', label='Petrol price',
                                        required=False, min_value=0)
        price.updated_by = current_user.id
        price.updated_at = datetime.now(timezone.utc)
        log_audit('UPDATE', 'fuel_prices', price.id,
                  f'Set fuel prices: diesel {price.diesel_price}, petrol {price.petrol_price}')
        db.session.commit()
        flash('Fuel prices updated.', 'success')
        return redirect(url_for('fuel_prices'))
    return render_template('admin/fuel_prices.html', price=price)


# ─────────────────────────────────────────────────────────────
# Fuel Logs
# ─────────────────────────────────────────────────────────────
@app.route('/logs/fuel')
@login_required
@permission_required('fuel_logs')
def fuel_logs():
    page = request.args.get('page', 1, type=int)
    vehicle_id = request.args.get('vehicle_id', '')
    q = FuelLog.query
    if vehicle_id:
        q = q.filter(FuelLog.vehicle_id == vehicle_id)
    logs = q.order_by(FuelLog.log_date.desc()).paginate(page=page, per_page=20)
    all_vehicles = Vehicle.query.order_by(Vehicle.registration).all()
    return render_template('logs/fuel/index.html', logs=logs, vehicles=all_vehicles,
                           vehicle_id=vehicle_id)


@app.route('/logs/fuel/export')
@login_required
@permission_required('fuel_logs')
def fuel_logs_export():
    vehicle_id = request.args.get('vehicle_id', '')
    q = FuelLog.query
    if vehicle_id:
        q = q.filter(FuelLog.vehicle_id == vehicle_id)
    logs = q.order_by(FuelLog.log_date.desc()).all()
    rows = [[l.log_date, l.vehicle.registration, l.liters,
             f'{l.cost_per_liter:.2f}' if l.cost_per_liter is not None else '',
             f'{l.total_cost:.2f}' if l.total_cost is not None else '',
             l.odometer if l.odometer is not None else '', l.supplier or '', l.notes or '']
            for l in logs]
    header = ['Date', 'Vehicle', 'Liters', 'Cost/Liter', 'Total Cost', 'Odometer', 'Supplier', 'Notes']
    if request.args.get('format') == 'pdf':
        return _table_pdf_response(f'fuel_logs_{date.today()}.pdf', 'Fuel Logs',
            f'Generated {date.today()}', header, rows)
    return csv_export_response(f'fuel_logs_{date.today()}.csv', header, rows)


@app.route('/logs/fuel/add', methods=['GET', 'POST'])
@login_required
@permission_required('fuel_logs')
@handle_form_errors
def fuel_log_add():
    if request.method == 'POST':
        client_id = request.form.get('_client_id')
        if already_synced(client_id):
            flash('Already recorded.', 'info')
            return redirect(url_for('fuel_logs'))
        liters = form_float(request.form, 'liters', min_value=0)
        vehicle_id = form_int(request.form, 'vehicle_id')
        vehicle = Vehicle.query.filter_by(id=vehicle_id).first()
        if not vehicle:
            raise ValueError('Select a vehicle.')
        price = fuel_price_for(vehicle.fuel_type) or 0.0
        log = FuelLog(
            vehicle_id=vehicle_id,
            log_date=parse_date(request.form['log_date']),
            liters=liters,
            cost_per_liter=price,
            total_cost=liters * price,
            odometer=form_float(request.form, 'odometer', required=False, min_value=0),
            supplier=request.form.get('supplier', '').strip(),
            notes=request.form.get('notes', '').strip(),
            created_by=current_user.id,
        )
        db.session.add(log)
        db.session.flush()
        log_audit('CREATE', 'fuel_logs', log.id,
                  f'Fuel log for {log.vehicle.registration}: {liters}L')
        record_offline_sync(client_id, 'fuel_log_add')
        touch_sync_fields(log)
        db.session.commit()
        flash('Fuel log recorded.', 'success')
        return redirect(url_for('fuel_logs'))

    all_vehicles = Vehicle.query.filter_by(status='active').order_by(Vehicle.registration).all()
    return render_template('logs/fuel/form.html', vehicles=all_vehicles,
                           today=date.today().strftime('%Y-%m-%d'))


@app.route('/logs/fuel/<int:lid>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
@handle_form_errors
def fuel_log_edit(lid):
    log = FuelLog.query.filter_by(id=lid).first_or_404()
    if request.method == 'POST':
        log.vehicle_id = form_int(request.form, 'vehicle_id')
        vehicle = Vehicle.query.filter_by(id=log.vehicle_id).first()
        if not vehicle:
            raise ValueError('Select a vehicle.')
        log.log_date = parse_date(request.form['log_date'])
        log.liters = form_float(request.form, 'liters', min_value=0)
        price = fuel_price_for(vehicle.fuel_type) or 0.0
        log.cost_per_liter = price
        log.total_cost = log.liters * price
        log.odometer = form_float(request.form, 'odometer', required=False, min_value=0)
        log.supplier = request.form.get('supplier', '').strip()
        log.notes = request.form.get('notes', '').strip()
        log_audit('UPDATE', 'fuel_logs', log.id, f'Updated fuel log for {log.vehicle.registration}')
        touch_sync_fields(log)
        db.session.commit()
        flash('Fuel log updated.', 'success')
        return redirect(url_for('fuel_logs'))
    # Not filtered to active-only, unlike the Add form — an edit must still
    # be able to show/keep the log's own vehicle even if it's since been
    # marked inactive.
    all_vehicles = Vehicle.query.order_by(Vehicle.registration).all()
    return render_template('logs/fuel/form.html', vehicles=all_vehicles, log=log,
                           today=log.log_date.strftime('%Y-%m-%d'))


@app.route('/logs/fuel/<int:lid>/delete', methods=['POST'])
@login_required
@admin_required
def fuel_log_delete(lid):
    log = FuelLog.query.filter_by(id=lid).first_or_404()
    log_audit('DELETE', 'fuel_logs', lid, f'Deleted fuel log {lid}')
    log.deleted_at = datetime.now(timezone.utc)
    touch_sync_fields(log)
    db.session.commit()
    flash('Fuel log deleted.', 'warning')
    return redirect(url_for('fuel_logs'))


# ─────────────────────────────────────────────────────────────
# Maintenance Logs
# ─────────────────────────────────────────────────────────────
@app.route('/logs/maintenance')
@login_required
@permission_required('maintenance')
def maintenance_logs():
    page = request.args.get('page', 1, type=int)
    vehicle_id = request.args.get('vehicle_id', '')
    q = MaintenanceLog.query
    if vehicle_id:
        q = q.filter(MaintenanceLog.vehicle_id == vehicle_id)
    logs = q.order_by(MaintenanceLog.log_date.desc()).paginate(page=page, per_page=20)
    all_vehicles = Vehicle.query.order_by(Vehicle.registration).all()
    return render_template('logs/maintenance/index.html', logs=logs, vehicles=all_vehicles,
                           vehicle_id=vehicle_id)


@app.route('/logs/maintenance/export')
@login_required
@permission_required('maintenance')
def maintenance_logs_export():
    vehicle_id = request.args.get('vehicle_id', '')
    q = MaintenanceLog.query
    if vehicle_id:
        q = q.filter(MaintenanceLog.vehicle_id == vehicle_id)
    logs = q.order_by(MaintenanceLog.log_date.desc()).all()
    rows = [[l.log_date, l.vehicle.registration, l.description, f'{l.parts_cost:.2f}',
             f'{l.labor_cost:.2f}', f'{l.total_cost:.2f}', l.mechanic or '', l.notes or '']
            for l in logs]
    header = ['Date', 'Vehicle', 'Description', 'Parts Cost', 'Labor Cost', 'Total Cost', 'Mechanic', 'Notes']
    if request.args.get('format') == 'pdf':
        return _table_pdf_response(f'maintenance_logs_{date.today()}.pdf', 'Maintenance Logs',
            f'Generated {date.today()}', header, rows)
    return csv_export_response(f'maintenance_logs_{date.today()}.csv', header, rows)


@app.route('/logs/maintenance/add', methods=['GET', 'POST'])
@login_required
@permission_required('maintenance')
@handle_form_errors
def maintenance_log_add():
    if request.method == 'POST':
        client_id = request.form.get('_client_id')
        if already_synced(client_id):
            flash('Already recorded.', 'info')
            return redirect(url_for('maintenance_logs'))
        parts = form_float(request.form, 'parts_cost', required=False, default=0, min_value=0)
        labor = form_float(request.form, 'labor_cost', required=False, default=0, min_value=0)
        log = MaintenanceLog(
            vehicle_id=form_int(request.form, 'vehicle_id'),
            log_date=parse_date(request.form['log_date']),
            description=request.form['description'].strip(),
            parts_cost=parts,
            labor_cost=labor,
            total_cost=parts + labor,
            mechanic=request.form.get('mechanic', '').strip(),
            notes=request.form.get('notes', '').strip(),
            created_by=current_user.id,
        )
        db.session.add(log)
        db.session.flush()
        log_audit('CREATE', 'maintenance_logs', log.id,
                  f'Maintenance for {log.vehicle.registration}')
        record_offline_sync(client_id, 'maintenance_log_add')
        touch_sync_fields(log)
        db.session.commit()
        flash('Maintenance log recorded.', 'success')
        return redirect(url_for('maintenance_logs'))

    all_vehicles = Vehicle.query.filter_by(status='active').order_by(Vehicle.registration).all()
    return render_template('logs/maintenance/form.html', vehicles=all_vehicles,
                           today=date.today().strftime('%Y-%m-%d'))


@app.route('/logs/maintenance/<int:lid>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
@handle_form_errors
def maintenance_log_edit(lid):
    log = MaintenanceLog.query.filter_by(id=lid).first_or_404()
    if request.method == 'POST':
        parts = form_float(request.form, 'parts_cost', required=False, default=0, min_value=0)
        labor = form_float(request.form, 'labor_cost', required=False, default=0, min_value=0)
        log.vehicle_id = form_int(request.form, 'vehicle_id')
        log.log_date = parse_date(request.form['log_date'])
        log.description = request.form['description'].strip()
        log.parts_cost = parts
        log.labor_cost = labor
        log.total_cost = parts + labor
        log.mechanic = request.form.get('mechanic', '').strip()
        log.notes = request.form.get('notes', '').strip()
        log_audit('UPDATE', 'maintenance_logs', log.id, f'Updated maintenance log for {log.vehicle.registration}')
        touch_sync_fields(log)
        db.session.commit()
        flash('Maintenance log updated.', 'success')
        return redirect(url_for('maintenance_logs'))
    all_vehicles = Vehicle.query.order_by(Vehicle.registration).all()
    return render_template('logs/maintenance/form.html', vehicles=all_vehicles, log=log,
                           today=log.log_date.strftime('%Y-%m-%d'))


@app.route('/logs/maintenance/<int:lid>/delete', methods=['POST'])
@login_required
@admin_required
def maintenance_log_delete(lid):
    log = MaintenanceLog.query.filter_by(id=lid).first_or_404()
    log_audit('DELETE', 'maintenance_logs', lid, f'Deleted maintenance log {lid}')
    log.deleted_at = datetime.now(timezone.utc)
    touch_sync_fields(log)
    db.session.commit()
    flash('Maintenance log deleted.', 'warning')
    return redirect(url_for('maintenance_logs'))


def latest_odometer(vehicle_id):
    latest_fuel = FuelLog.query.filter_by(vehicle_id=vehicle_id).filter(
        FuelLog.odometer.isnot(None)).order_by(FuelLog.log_date.desc()).first()
    return latest_fuel.odometer if latest_fuel else None


def schedule_status(sched):
    today = date.today()
    odometer = latest_odometer(sched.vehicle_id)
    due_by_date = sched.next_due_date is not None and sched.next_due_date <= today
    due_by_km = (sched.next_due_odometer is not None and odometer is not None
                 and odometer >= sched.next_due_odometer)
    if due_by_date or due_by_km:
        return 'overdue'
    soon_by_date = sched.next_due_date is not None and (sched.next_due_date - today).days <= 14
    soon_by_km = (sched.next_due_odometer is not None and odometer is not None
                  and sched.next_due_odometer - odometer <= 500)
    if soon_by_date or soon_by_km:
        return 'due_soon'
    return 'ok'


# ─────────────────────────────────────────────────────────────
# Preventive Maintenance Scheduling
# ─────────────────────────────────────────────────────────────
@app.route('/maintenance/schedules')
@login_required
@permission_required('maintenance')
def maintenance_schedules():
    schedules = MaintenanceSchedule.query.filter_by(status='active').join(Vehicle).order_by(
        Vehicle.registration).all()
    rows = [{'schedule': s, 'status': schedule_status(s), 'odometer': latest_odometer(s.vehicle_id)}
            for s in schedules]
    rows.sort(key=lambda r: {'overdue': 0, 'due_soon': 1, 'ok': 2}[r['status']])
    all_vehicles = Vehicle.query.filter_by(status='active').order_by(Vehicle.registration).all()
    return render_template('maintenance/schedules.html', rows=rows, vehicles=all_vehicles,
                           today=date.today().strftime('%Y-%m-%d'))


@app.route('/maintenance/schedules/add', methods=['GET', 'POST'])
@login_required
@permission_required('maintenance')
@handle_form_errors
def maintenance_schedule_add():
    if request.method == 'POST':
        interval_days = form_int(request.form, 'interval_days', required=False, min_value=1)
        interval_km = form_float(request.form, 'interval_km', required=False, min_value=1)
        last_done_date = parse_date(request.form.get('last_done_date')) or date.today()
        last_done_odometer = form_float(request.form, 'last_done_odometer', required=False, min_value=0)

        sched = MaintenanceSchedule(
            vehicle_id=form_int(request.form, 'vehicle_id'),
            description=request.form['description'].strip(),
            interval_days=interval_days,
            interval_km=interval_km,
            last_done_date=last_done_date,
            last_done_odometer=last_done_odometer,
            next_due_date=(last_done_date + timedelta(days=interval_days)) if interval_days else None,
            next_due_odometer=(last_done_odometer + interval_km)
                if (interval_km and last_done_odometer is not None) else None,
        )
        db.session.add(sched)
        db.session.flush()
        log_audit('CREATE', 'maintenance_schedules', sched.id, f'Added maintenance schedule: {sched.description}')
        touch_sync_fields(sched)
        db.session.commit()
        flash('Maintenance schedule added.', 'success')
        return redirect(url_for('maintenance_schedules'))
    all_vehicles = Vehicle.query.filter_by(status='active').order_by(Vehicle.registration).all()
    return render_template('maintenance/schedule_form.html', vehicles=all_vehicles,
                           today=date.today().strftime('%Y-%m-%d'))


@app.route('/maintenance/schedules/<int:sid>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
@handle_form_errors
def maintenance_schedule_edit(sid):
    sched = MaintenanceSchedule.query.filter_by(id=sid).first_or_404()
    if request.method == 'POST':
        interval_days = form_int(request.form, 'interval_days', required=False, min_value=1)
        interval_km = form_float(request.form, 'interval_km', required=False, min_value=1)
        last_done_date = parse_date(request.form.get('last_done_date')) or date.today()
        last_done_odometer = form_float(request.form, 'last_done_odometer', required=False, min_value=0)

        sched.vehicle_id = form_int(request.form, 'vehicle_id')
        sched.description = request.form['description'].strip()
        sched.interval_days = interval_days
        sched.interval_km = interval_km
        sched.last_done_date = last_done_date
        sched.last_done_odometer = last_done_odometer
        sched.next_due_date = (last_done_date + timedelta(days=interval_days)) if interval_days else None
        sched.next_due_odometer = (last_done_odometer + interval_km) \
            if (interval_km and last_done_odometer is not None) else None
        log_audit('UPDATE', 'maintenance_schedules', sched.id, f'Updated maintenance schedule: {sched.description}')
        touch_sync_fields(sched)
        db.session.commit()
        flash('Maintenance schedule updated.', 'success')
        return redirect(url_for('maintenance_schedules'))
    all_vehicles = Vehicle.query.order_by(Vehicle.registration).all()
    return render_template('maintenance/schedule_form.html', vehicles=all_vehicles, sched=sched,
                           today=date.today().strftime('%Y-%m-%d'))


@app.route('/maintenance/schedules/<int:sid>/done', methods=['POST'])
@login_required
@permission_required('maintenance')
@handle_form_errors
def maintenance_schedule_done(sid):
    sched = MaintenanceSchedule.query.filter_by(id=sid).first_or_404()
    done_date = parse_date(request.form.get('done_date')) or date.today()
    done_odometer = form_float(request.form, 'done_odometer', required=False, min_value=0)

    sched.last_done_date = done_date
    sched.last_done_odometer = done_odometer if done_odometer is not None else sched.last_done_odometer
    sched.next_due_date = (done_date + timedelta(days=sched.interval_days)) if sched.interval_days else None
    sched.next_due_odometer = (sched.last_done_odometer + sched.interval_km) \
        if (sched.interval_km and sched.last_done_odometer is not None) else None

    log_audit('UPDATE', 'maintenance_schedules', sid, f'Marked "{sched.description}" as done')
    touch_sync_fields(sched)
    db.session.commit()
    flash('Marked as done — next due date recalculated.', 'success')
    return redirect(url_for('maintenance_schedules'))


@app.route('/maintenance/schedules/<int:sid>/delete', methods=['POST'])
@login_required
@admin_required
def maintenance_schedule_delete(sid):
    sched = MaintenanceSchedule.query.filter_by(id=sid).first_or_404()
    log_audit('DELETE', 'maintenance_schedules', sid, f'Deleted maintenance schedule: {sched.description}')
    sched.deleted_at = datetime.now(timezone.utc)
    touch_sync_fields(sched)
    db.session.commit()
    flash('Maintenance schedule deleted.', 'warning')
    return redirect(url_for('maintenance_schedules'))


# ─────────────────────────────────────────────────────────────
# Spares Store: parts inventory, purchases & marked-up sales
# ─────────────────────────────────────────────────────────────
@app.route('/store/parts')
@login_required
@permission_required('store')
def store_parts():
    parts = SparePart.query.order_by(SparePart.name).all()
    total_stock_value = sum(p.stock_value for p in parts)
    low_stock_count = sum(1 for p in parts if p.status == 'active' and p.low_stock)
    month_start = date.today().replace(day=1)
    month_profit = sum(s.profit for s in
                       StoreSale.query.filter(StoreSale.sale_date >= month_start).all())
    return render_template('store/parts.html', parts=parts,
                           total_stock_value=total_stock_value,
                           low_stock_count=low_stock_count, month_profit=month_profit)


@app.route('/store/parts/add', methods=['GET', 'POST'])
@login_required
@permission_required('store')
@handle_form_errors
def store_part_add():
    if request.method == 'POST':
        part = SparePart(
            name=request.form['name'].strip(),
            part_number=request.form.get('part_number', '').strip(),
            unit=request.form.get('unit', 'pc').strip() or 'pc',
            markup_percent=form_float(request.form, 'markup_percent', required=False,
                                      default=0, min_value=0),
            reorder_level=form_int(request.form, 'reorder_level', required=False,
                                   default=0, min_value=0),
            notes=request.form.get('notes', '').strip(),
            created_by=current_user.id,
        )
        db.session.add(part)
        db.session.flush()
        log_audit('CREATE', 'spare_parts', part.id, f'Added spare part: {part.name}')
        touch_sync_fields(part)
        db.session.commit()
        flash('Spare part added. Record a purchase to bring in stock.', 'success')
        return redirect(url_for('store_parts'))
    return render_template('store/part_form.html', part=None)


@app.route('/store/parts/<int:pid>/edit', methods=['GET', 'POST'])
@login_required
@permission_required('store')
@handle_form_errors
def store_part_edit(pid):
    part = SparePart.query.filter_by(id=pid).first_or_404()
    if request.method == 'POST':
        part.name = request.form['name'].strip()
        part.part_number = request.form.get('part_number', '').strip()
        part.unit = request.form.get('unit', 'pc').strip() or 'pc'
        part.markup_percent = form_float(request.form, 'markup_percent', required=False,
                                         default=0, min_value=0)
        part.reorder_level = form_int(request.form, 'reorder_level', required=False,
                                      default=0, min_value=0)
        part.status = request.form.get('status', 'active')
        part.notes = request.form.get('notes', '').strip()
        log_audit('UPDATE', 'spare_parts', part.id, f'Updated spare part: {part.name}')
        touch_sync_fields(part)
        db.session.commit()
        flash('Spare part updated.', 'success')
        return redirect(url_for('store_parts'))
    return render_template('store/part_form.html', part=part)


@app.route('/store/parts/<int:pid>/delete', methods=['POST'])
@login_required
@admin_required
def store_part_delete(pid):
    part = SparePart.query.filter_by(id=pid).first_or_404()
    log_audit('DELETE', 'spare_parts', pid, f'Deleted spare part: {part.name}')
    now = datetime.now(timezone.utc)
    # cascade='all, delete-orphan' only fires on an actual ORM delete of
    # the parent, not on setting deleted_at — soft-delete purchase/sale
    # history explicitly so it doesn't outlive its (now hidden) part.
    for purchase in part.purchases:
        purchase.deleted_at = now
        touch_sync_fields(purchase)
    for sale in part.sales:
        sale.deleted_at = now
        touch_sync_fields(sale)
    for adjustment in part.adjustments:
        adjustment.deleted_at = now
        touch_sync_fields(adjustment)
    part.deleted_at = now
    touch_sync_fields(part)
    db.session.commit()
    flash('Spare part deleted.', 'warning')
    return redirect(url_for('store_parts'))


@app.route('/store/purchases')
@login_required
@permission_required('store')
def store_purchases():
    page = request.args.get('page', 1, type=int)
    part_id = request.args.get('part_id', '')
    q = StorePurchase.query
    if part_id:
        q = q.filter(StorePurchase.part_id == part_id)
    purchases = q.order_by(StorePurchase.purchase_date.desc()).paginate(page=page, per_page=20)
    all_parts = SparePart.query.order_by(SparePart.name).all()
    return render_template('store/purchases.html', purchases=purchases, parts=all_parts,
                           part_id=part_id)


@app.route('/store/purchases/export')
@login_required
@permission_required('store')
def store_purchases_export():
    part_id = request.args.get('part_id', '')
    q = StorePurchase.query
    if part_id:
        q = q.filter(StorePurchase.part_id == part_id)
    purchases = q.order_by(StorePurchase.purchase_date.desc()).all()
    rows = [[p.purchase_date, p.part.name, qty_filter(p.quantity), f'{p.unit_cost:.2f}',
             f'{p.total_cost:.2f}', p.supplier or '', p.notes or ''] for p in purchases]
    header = ['Date', 'Part', 'Quantity', 'Unit Cost', 'Total Cost', 'Supplier', 'Notes']
    if request.args.get('format') == 'pdf':
        return _table_pdf_response(f'store_purchases_{date.today()}.pdf', 'Spares Store Purchases',
            f'Generated {date.today()}', header, rows)
    return csv_export_response(f'store_purchases_{date.today()}.csv', header, rows)


@app.route('/store/purchases/add', methods=['GET', 'POST'])
@login_required
@permission_required('store')
@handle_form_errors
def store_purchase_add():
    if request.method == 'POST':
        client_id = request.form.get('_client_id')
        if already_synced(client_id):
            flash('Already recorded.', 'info')
            return redirect(url_for('store_purchases'))
        part = SparePart.query.filter_by(id=form_int(request.form, 'part_id')).first_or_404()
        quantity = form_float(request.form, 'quantity', min_value=0.01)
        unit_cost = form_float(request.form, 'unit_cost', min_value=0)

        purchase = StorePurchase(
            part_id=part.id,
            purchase_date=parse_date(request.form['purchase_date']),
            quantity=quantity,
            unit_cost=unit_cost,
            total_cost=quantity * unit_cost,
            supplier=request.form.get('supplier', '').strip(),
            notes=request.form.get('notes', '').strip(),
            created_by=current_user.id,
        )
        new_total_qty = part.quantity_on_hand + quantity
        part.cost_price = ((part.quantity_on_hand * part.cost_price) +
                           (quantity * unit_cost)) / new_total_qty
        part.quantity_on_hand = new_total_qty

        db.session.add(purchase)
        db.session.flush()
        log_audit('CREATE', 'store_purchases', purchase.id,
                  f'Purchased {quantity} x {part.name} @ {unit_cost}')
        record_offline_sync(client_id, 'store_purchase_add')
        touch_sync_fields(purchase)
        touch_sync_fields(part)
        db.session.commit()
        flash('Purchase recorded and stock updated.', 'success')
        return redirect(url_for('store_purchases'))

    all_parts = SparePart.query.filter_by(status='active').order_by(SparePart.name).all()
    return render_template('store/purchase_form.html', parts=all_parts,
                           today=date.today().strftime('%Y-%m-%d'))


@app.route('/store/purchases/<int:pid>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
@handle_form_errors
def store_purchase_edit(pid):
    purchase = StorePurchase.query.filter_by(id=pid).first_or_404()
    if request.method == 'POST':
        part = SparePart.query.filter_by(id=form_int(request.form, 'part_id')).first_or_404()
        quantity = form_float(request.form, 'quantity', min_value=0.01)
        unit_cost = form_float(request.form, 'unit_cost', min_value=0)

        # Reverse this purchase's old quantity off whichever part it was
        # against, then apply the new quantity to the (possibly different)
        # part — same "adjust quantity_on_hand only" approach as the delete
        # route; average cost_price is not recomputed for the same reason
        # (see store_purchase_delete).
        old_part = purchase.part
        if part.id == old_part.id:
            part.quantity_on_hand = max(0, part.quantity_on_hand - purchase.quantity + quantity)
            touch_sync_fields(part)
        else:
            old_part.quantity_on_hand = max(0, old_part.quantity_on_hand - purchase.quantity)
            part.quantity_on_hand = part.quantity_on_hand + quantity
            touch_sync_fields(old_part)
            touch_sync_fields(part)

        purchase.part_id = part.id
        purchase.purchase_date = parse_date(request.form['purchase_date'])
        purchase.quantity = quantity
        purchase.unit_cost = unit_cost
        purchase.total_cost = quantity * unit_cost
        purchase.supplier = request.form.get('supplier', '').strip()
        purchase.notes = request.form.get('notes', '').strip()
        log_audit('UPDATE', 'store_purchases', purchase.id,
                  f'Updated purchase to {quantity} x {part.name} @ {unit_cost}')
        touch_sync_fields(purchase)
        db.session.commit()
        flash('Purchase updated and stock adjusted accordingly. Note: this does not '
             'recompute historical average cost.', 'success')
        return redirect(url_for('store_purchases'))
    all_parts = SparePart.query.order_by(SparePart.name).all()
    return render_template('store/purchase_form.html', parts=all_parts, purchase=purchase,
                           today=purchase.purchase_date.strftime('%Y-%m-%d'))


@app.route('/store/purchases/<int:pid>/delete', methods=['POST'])
@login_required
@admin_required
def store_purchase_delete(pid):
    purchase = StorePurchase.query.filter_by(id=pid).first_or_404()
    part = purchase.part
    part.quantity_on_hand = max(0, part.quantity_on_hand - purchase.quantity)
    log_audit('DELETE', 'store_purchases', pid,
              f'Deleted purchase of {purchase.quantity} x {part.name}')
    purchase.deleted_at = datetime.now(timezone.utc)
    touch_sync_fields(purchase)
    touch_sync_fields(part)
    db.session.commit()
    flash('Purchase deleted and stock reduced accordingly. Note: this does not '
         'recompute historical average cost.', 'warning')
    return redirect(url_for('store_purchases'))


@app.route('/store/purchases/import/preview', methods=['POST'])
@login_required
@permission_required('store')
def store_purchases_import_preview():
    """Same two-step confirmed-mapping flow as the Franchise Collections
    importer: parse the file, auto-map its columns onto the canonical stock
    fields, and show a confirmation/adjustment step before writing anything.
    Also re-entered (without a fresh file) when the user adjusts the mapping
    and clicks Re-preview — the parsed rows travel via raw_data."""
    file = request.files.get('file')
    if file and file.filename:
        filename = file.filename
        try:
            headers, raw_rows = read_uploaded_table(file)
        except ValueError as e:
            flash(str(e), 'danger')
            return redirect(url_for('store_purchases'))
        mapping = auto_map_columns(headers, fields=CANONICAL_STOCK_FIELDS)
    else:
        try:
            filename = request.form.get('filename', 'uploaded file')
            payload = json.loads(request.form.get('raw_data') or '{}')
            headers, raw_rows = payload.get('headers') or [], payload.get('rows') or []
            if not headers or not raw_rows:
                raise ValueError('Choose a CSV or Excel file to import.')
            mapping = {field_key: (request.form.get(f'map_{field_key}') or None)
                       for field_key, _label, _syn in CANONICAL_STOCK_FIELDS}
        except (ValueError, json.JSONDecodeError, TypeError):
            flash('Choose a CSV or Excel file to import — the previous preview session expired.', 'danger')
            return redirect(url_for('store_purchases'))

    if not raw_rows:
        flash('That file has no data rows to import — it only has a header row. '
              'Add rows with a Date, Part and Quantity, then re-import.', 'warning')
        return redirect(url_for('store_purchases'))

    preview_rows = apply_column_mapping(headers, raw_rows[:10], mapping, row_key_map=STOCK_ROW_KEY_MAP)
    return render_template('store/purchases_import_preview.html',
                           filename=filename, headers=headers, mapping=mapping,
                           fields=CANONICAL_STOCK_FIELDS, preview_rows=preview_rows,
                           row_count=len(raw_rows),
                           raw_data=json.dumps({'headers': headers, 'rows': raw_rows}))


@app.route('/store/purchases/import/confirm', methods=['POST'])
@login_required
@permission_required('store')
def store_purchases_import_confirm():
    filename = request.form.get('filename', 'uploaded file')
    try:
        payload = json.loads(request.form.get('raw_data') or '{}')
        headers, raw_rows = payload.get('headers') or [], payload.get('rows') or []
        if not raw_rows:
            raise ValueError('empty')
    except (ValueError, json.JSONDecodeError, TypeError):
        flash('That preview session expired — please choose the file again.', 'danger')
        return redirect(url_for('store_purchases'))

    mapping = {field_key: (request.form.get(f'map_{field_key}') or None)
               for field_key, _label, _syn in CANONICAL_STOCK_FIELDS}
    file_rows = apply_column_mapping(headers, raw_rows, mapping, row_key_map=STOCK_ROW_KEY_MAP)
    auto_create = request.form.get('auto_create_parts') == '1'
    imported, errors, error_rows, created_parts, created_records = import_stock_purchase_rows(
        file_rows, auto_create_parts=auto_create)

    if imported or error_rows:
        # Commit even when imported == 0: a batch made only of failed rows
        # still needs to persist so its quarantine CSV can be downloaded.
        save_import_batch('store_purchases', filename, len(raw_rows), imported, error_rows, created_records)
        if imported:
            log_audit('CREATE', 'store_purchases', None,
                      f'Imported {imported} stock purchase row(s) from {filename}')
        created_part_ids = [rid for table, rid in created_records if table == 'spare_parts']
        for part_name, part_id in zip(created_parts, created_part_ids):
            log_audit('CREATE', 'spare_parts', part_id,
                      f'Auto-created spare part "{part_name}" from stock import ({filename}) — '
                      'not on file, added because a purchase row named it.')
        db.session.commit()
    else:
        db.session.rollback()

    if imported:
        flash(f'Imported {imported} stock purchase row(s).', 'success')
    if created_parts:
        flash(f'Auto-created new part(s): {", ".join(created_parts)}.', 'success')
    if errors:
        shown = errors[:10]
        more = f' (+{len(errors) - 10} more)' if len(errors) > 10 else ''
        flash('Skipped rows — ' + '; '.join(shown) + more, 'warning')
    if not imported and not errors:
        flash('No rows found to import.', 'warning')

    return redirect(url_for('store_purchases'))


@app.route('/store/sales')
@login_required
@permission_required('store')
def store_sales():
    page = request.args.get('page', 1, type=int)
    part_id = request.args.get('part_id', '')
    q = StoreSale.query
    if part_id:
        q = q.filter(StoreSale.part_id == part_id)
    sales = q.order_by(StoreSale.sale_date.desc()).paginate(page=page, per_page=20)
    all_parts = SparePart.query.order_by(SparePart.name).all()
    return render_template('store/sales.html', sales=sales, parts=all_parts, part_id=part_id)


@app.route('/store/sales/export')
@login_required
@permission_required('store')
def store_sales_export():
    part_id = request.args.get('part_id', '')
    q = StoreSale.query
    if part_id:
        q = q.filter(StoreSale.part_id == part_id)
    sales = q.order_by(StoreSale.sale_date.desc()).all()
    rows = [[s.sale_date, s.part.name, qty_filter(s.quantity), f'{s.unit_cost:.2f}', f'{s.unit_price:.2f}',
             f'{s.total_amount:.2f}', s.vehicle.registration if s.vehicle else '',
             s.customer_name or '', s.notes or ''] for s in sales]
    header = ['Date', 'Part', 'Quantity', 'Unit Cost', 'Unit Price', 'Total Amount', 'Sold To Vehicle', 'Customer', 'Notes']
    if request.args.get('format') == 'pdf':
        return _table_pdf_response(f'store_sales_{date.today()}.pdf', 'Spares Store Sales',
            f'Generated {date.today()}', header, rows)
    return csv_export_response(f'store_sales_{date.today()}.csv', header, rows)


@app.route('/store/sales/add', methods=['GET', 'POST'])
@login_required
@permission_required('store')
@handle_form_errors
def store_sale_add():
    if request.method == 'POST':
        client_id = request.form.get('_client_id')
        if already_synced(client_id):
            flash('Already recorded.', 'info')
            return redirect(url_for('store_sales'))
        part = SparePart.query.filter_by(id=form_int(request.form, 'part_id')).first_or_404()
        quantity = form_float(request.form, 'quantity', min_value=0.01)
        if quantity > part.quantity_on_hand:
            raise ValueError(f'Only {part.quantity_on_hand} {part.unit}(s) of {part.name} in stock.')
        # Total Price lets the total be entered directly (e.g. a negotiated
        # round figure) instead of via a per-unit price — when given, it's
        # the authoritative amount and unit_price is only backed out from it
        # for the per-unit records; otherwise unit_price x quantity governs.
        total_price = form_float(request.form, 'total_price', required=False, min_value=0)
        if total_price is not None:
            unit_price = total_price / quantity
            total_amount = total_price
        else:
            unit_price = form_float(request.form, 'unit_price', required=False,
                                    default=part.selling_price, min_value=0)
            total_amount = quantity * unit_price
        vehicle_id = form_int(request.form, 'vehicle_id', required=False)

        sale = StoreSale(
            part_id=part.id,
            vehicle_id=vehicle_id,
            sale_date=parse_date(request.form['sale_date']),
            quantity=quantity,
            unit_cost=part.cost_price,
            unit_price=unit_price,
            total_amount=total_amount,
            customer_name=request.form.get('customer_name', '').strip() if not vehicle_id else None,
            notes=request.form.get('notes', '').strip(),
            created_by=current_user.id,
        )
        part.quantity_on_hand -= quantity

        db.session.add(sale)
        db.session.flush()
        log_audit('CREATE', 'store_sales', sale.id,
                  f'Sold {quantity} x {part.name} @ {unit_price}' +
                  (f' to vehicle {sale.vehicle.registration} (booked as an expense on that vehicle)'
                   if sale.vehicle else ''))
        record_offline_sync(client_id, 'store_sale_add')
        touch_sync_fields(sale)
        touch_sync_fields(part)
        db.session.commit()
        flash('Sale recorded and stock updated.', 'success')
        return redirect(url_for('store_sales'))

    all_parts = SparePart.query.filter(SparePart.status == 'active',
                                       SparePart.quantity_on_hand > 0).order_by(SparePart.name).all()
    all_vehicles = Vehicle.query.filter_by(status='active').order_by(Vehicle.registration).all()
    return render_template('store/sale_form.html', parts=all_parts, vehicles=all_vehicles,
                           today=date.today().strftime('%Y-%m-%d'))


@app.route('/store/sales/<int:sid>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
@handle_form_errors
def store_sale_edit(sid):
    sale = StoreSale.query.filter_by(id=sid).first_or_404()
    if request.method == 'POST':
        part = SparePart.query.filter_by(id=form_int(request.form, 'part_id')).first_or_404()
        quantity = form_float(request.form, 'quantity', min_value=0.01)

        # Put the old sale's quantity back onto its old part before checking
        # stock for the new quantity/part — same reversal approach as
        # store_purchase_edit.
        old_part = sale.part
        available = part.quantity_on_hand + (sale.quantity if part.id == old_part.id else 0)
        if quantity > available:
            raise ValueError(f'Only {available} {part.unit}(s) of {part.name} in stock.')
        total_price = form_float(request.form, 'total_price', required=False, min_value=0)
        if total_price is not None:
            unit_price = total_price / quantity
            total_amount = total_price
        else:
            unit_price = form_float(request.form, 'unit_price', required=False,
                                    default=part.selling_price, min_value=0)
            total_amount = quantity * unit_price
        vehicle_id = form_int(request.form, 'vehicle_id', required=False)

        if part.id == old_part.id:
            part.quantity_on_hand = part.quantity_on_hand + sale.quantity - quantity
            touch_sync_fields(part)
        else:
            old_part.quantity_on_hand += sale.quantity
            part.quantity_on_hand -= quantity
            touch_sync_fields(old_part)
            touch_sync_fields(part)

        sale.part_id = part.id
        sale.vehicle_id = vehicle_id
        sale.sale_date = parse_date(request.form['sale_date'])
        sale.quantity = quantity
        sale.unit_cost = part.cost_price
        sale.unit_price = unit_price
        sale.total_amount = total_amount
        sale.customer_name = request.form.get('customer_name', '').strip() if not vehicle_id else None
        sale.notes = request.form.get('notes', '').strip()
        log_audit('UPDATE', 'store_sales', sale.id, f'Updated sale to {quantity} x {part.name} @ {unit_price}')
        touch_sync_fields(sale)
        db.session.commit()
        flash('Sale updated and stock adjusted accordingly.', 'success')
        return redirect(url_for('store_sales'))

    all_parts = SparePart.query.order_by(SparePart.name).all()
    all_vehicles = Vehicle.query.order_by(Vehicle.registration).all()
    return render_template('store/sale_form.html', parts=all_parts, vehicles=all_vehicles, sale=sale,
                           today=sale.sale_date.strftime('%Y-%m-%d'))


@app.route('/store/sales/<int:sid>/delete', methods=['POST'])
@login_required
@admin_required
def store_sale_delete(sid):
    sale = StoreSale.query.filter_by(id=sid).first_or_404()
    part = sale.part
    part.quantity_on_hand += sale.quantity
    log_audit('DELETE', 'store_sales', sid, f'Deleted sale of {sale.quantity} x {part.name}')
    sale.deleted_at = datetime.now(timezone.utc)
    touch_sync_fields(sale)
    touch_sync_fields(part)
    db.session.commit()
    flash('Sale deleted and stock restored.', 'warning')
    return redirect(url_for('store_sales'))


@app.route('/store/adjustments')
@login_required
@permission_required('store')
def store_adjustments():
    page = request.args.get('page', 1, type=int)
    part_id = request.args.get('part_id', '')
    q = StoreAdjustment.query
    if part_id:
        q = q.filter(StoreAdjustment.part_id == part_id)
    adjustments = q.order_by(StoreAdjustment.adjustment_date.desc(),
                             StoreAdjustment.id.desc()).paginate(page=page, per_page=20)
    all_parts = SparePart.query.order_by(SparePart.name).all()
    return render_template('store/adjustments.html', adjustments=adjustments, parts=all_parts,
                           part_id=part_id, reasons=ADJUSTMENT_REASONS)


@app.route('/store/adjustments/export')
@login_required
@permission_required('store')
def store_adjustments_export():
    part_id = request.args.get('part_id', '')
    q = StoreAdjustment.query
    if part_id:
        q = q.filter(StoreAdjustment.part_id == part_id)
    adjustments = q.order_by(StoreAdjustment.adjustment_date.desc(), StoreAdjustment.id.desc()).all()
    rows = [[a.adjustment_date, a.part.name, qty_filter(a.quantity_before), qty_filter(a.quantity_after),
             qty_filter(a.quantity_delta), f'{a.cost_price_before:.2f}', f'{a.cost_price_after:.2f}',
             a.reason_label, a.notes or ''] for a in adjustments]
    header = ['Date', 'Part', 'Qty Before', 'Qty After', 'Qty Change', 'Cost Before', 'Cost After',
              'Reason', 'Notes']
    if request.args.get('format') == 'pdf':
        return _table_pdf_response(f'store_adjustments_{date.today()}.pdf', 'Spares Store Adjustments',
            f'Generated {date.today()}', header, rows)
    return csv_export_response(f'store_adjustments_{date.today()}.csv', header, rows)


@app.route('/store/adjustments/add', methods=['GET', 'POST'])
@login_required
@admin_required
@handle_form_errors
def store_adjustment_add():
    """Corrects a part's quantity on hand and/or weighted-average unit cost
    to match a physical count or a write-off/write-up — admin-only since,
    unlike a purchase or sale, it overrides the book figures directly rather
    than following the normal buy/sell trail."""
    if request.method == 'POST':
        client_id = request.form.get('_client_id')
        if already_synced(client_id):
            flash('Already recorded.', 'info')
            return redirect(url_for('store_adjustments'))
        part = SparePart.query.filter_by(id=form_int(request.form, 'part_id')).first_or_404()
        new_quantity = form_float(request.form, 'quantity', min_value=0)
        new_cost_price = form_float(request.form, 'cost_price', min_value=0)

        adjustment = StoreAdjustment(
            part_id=part.id,
            adjustment_date=parse_date(request.form['adjustment_date']),
            quantity_before=part.quantity_on_hand,
            quantity_after=new_quantity,
            cost_price_before=part.cost_price,
            cost_price_after=new_cost_price,
            reason=request.form.get('reason', 'other'),
            notes=request.form.get('notes', '').strip(),
            created_by=current_user.id,
        )
        part.quantity_on_hand = new_quantity
        part.cost_price = new_cost_price

        db.session.add(adjustment)
        db.session.flush()
        log_audit('CREATE', 'store_adjustments', adjustment.id,
                  f'Adjusted {part.name}: qty {adjustment.quantity_before} -> {new_quantity}, '
                  f'cost {adjustment.cost_price_before} -> {new_cost_price} ({adjustment.reason_label})')
        record_offline_sync(client_id, 'store_adjustment_add')
        touch_sync_fields(adjustment)
        touch_sync_fields(part)
        db.session.commit()
        flash('Stock corrected.', 'success')
        return redirect(url_for('store_adjustments'))

    all_parts = SparePart.query.filter_by(status='active').order_by(SparePart.name).all()
    part_id = request.args.get('part_id', type=int)
    return render_template('store/adjustment_form.html', parts=all_parts, reasons=ADJUSTMENT_REASONS,
                           today=date.today().strftime('%Y-%m-%d'), preselect_part_id=part_id)


@app.route('/store/adjustments/<int:aid>/delete', methods=['POST'])
@login_required
@admin_required
def store_adjustment_delete(aid):
    adjustment = StoreAdjustment.query.filter_by(id=aid).first_or_404()
    part = adjustment.part
    # Reverses this adjustment's own delta rather than resetting to
    # quantity_before/cost_price_before outright — same "undo just this
    # record's effect" approach as store_purchase_delete, so it stays
    # correct even if other transactions happened on the part since.
    part.quantity_on_hand = max(0, part.quantity_on_hand - adjustment.quantity_delta)
    log_audit('DELETE', 'store_adjustments', aid,
              f'Deleted adjustment of {part.name} ({adjustment.reason_label})')
    adjustment.deleted_at = datetime.now(timezone.utc)
    touch_sync_fields(adjustment)
    touch_sync_fields(part)
    db.session.commit()
    flash('Adjustment deleted and stock reverted accordingly.', 'warning')
    return redirect(url_for('store_adjustments'))


@app.route('/store/trading-account')
@login_required
@permission_required('store')
def store_trading_account():
    """Standalone Trading & Profit or Loss Account for the Spares Store,
    run as its own cost centre. Revenue is ALL sales — including internal
    sales to company vehicles at the same marked-up price a walk-in
    customer pays — because from the store's side that's a real sale.
    That same amount also lands as a maintenance expense on the buying
    vehicle's income statement (see vehicle_income_totals): the store
    recognizes revenue, the vehicle recognizes a cost — two sides of one
    internal transfer, not a double-count within either statement.
    Cost of sales is summed from each sale's snapshotted unit_cost, which
    is more accurate than an opening/closing-stock estimate here since
    cost_price is a live weighted average, not a point-in-time figure."""
    df, dt = query_date_range()
    date_from_str, date_to_str = df.strftime('%Y-%m-%d'), dt.strftime('%Y-%m-%d')

    sales = StoreSale.query.filter(StoreSale.sale_date.between(df, dt)).all()

    sales_revenue = sum(s.total_amount for s in sales)
    cost_of_sales = sum(s.unit_cost * s.quantity for s in sales)
    gross_profit = sales_revenue - cost_of_sales
    gross_margin = (gross_profit / sales_revenue * 100) if sales_revenue else 0

    external = [s for s in sales if not s.vehicle_id]
    internal = [s for s in sales if s.vehicle_id]
    external_sales = sum(s.total_amount for s in external)
    external_cogs = sum(s.unit_cost * s.quantity for s in external)
    internal_sales = sum(s.total_amount for s in internal)
    internal_cogs = sum(s.unit_cost * s.quantity for s in internal)

    purchases_total = db.session.query(func.sum(StorePurchase.total_cost)).filter(
        StorePurchase.purchase_date.between(df, dt)).scalar() or 0
    closing_stock_value = sum(p.stock_value for p in SparePart.query.all())

    by_part = {}
    for s in sales:
        row = by_part.setdefault(s.part_id,
            {'part': s.part, 'quantity': 0, 'sales': 0.0, 'cost_of_sales': 0.0})
        row['quantity'] += s.quantity
        row['sales'] += s.total_amount
        row['cost_of_sales'] += s.unit_cost * s.quantity
    part_breakdown = list(by_part.values())
    for row in part_breakdown:
        row['gross_profit'] = row['sales'] - row['cost_of_sales']
        row['margin'] = (row['gross_profit'] / row['sales'] * 100) if row['sales'] else 0
    part_breakdown.sort(key=lambda r: r['gross_profit'], reverse=True)

    return render_template('store/trading_account.html',
        sales_revenue=sales_revenue, cost_of_sales=cost_of_sales,
        gross_profit=gross_profit, gross_margin=gross_margin,
        external_sales=external_sales, external_cogs=external_cogs,
        external_profit=external_sales - external_cogs,
        internal_sales=internal_sales, internal_cogs=internal_cogs,
        internal_profit=internal_sales - internal_cogs,
        purchases_total=purchases_total, closing_stock_value=closing_stock_value,
        part_breakdown=part_breakdown,
        date_from=date_from_str, date_to=date_to_str)


@app.route('/store/trading-account/export')
@login_required
@permission_required('store')
def store_trading_account_export():
    df, dt = query_date_range()
    sales = StoreSale.query.filter(StoreSale.sale_date.between(df, dt)).all()
    by_part = {}
    for s in sales:
        row = by_part.setdefault(s.part_id, {'part': s.part, 'quantity': 0, 'sales': 0.0, 'cost_of_sales': 0.0})
        row['quantity'] += s.quantity
        row['sales'] += s.total_amount
        row['cost_of_sales'] += s.unit_cost * s.quantity
    rows = []
    for row in sorted(by_part.values(), key=lambda r: r['sales'] - r['cost_of_sales'], reverse=True):
        gross_profit = row['sales'] - row['cost_of_sales']
        margin = (gross_profit / row['sales'] * 100) if row['sales'] else 0
        rows.append([row['part'].name, row['quantity'], f"{row['sales']:.2f}",
                     f"{row['cost_of_sales']:.2f}", f"{gross_profit:.2f}", f"{margin:.1f}"])
    header = ['Part', 'Quantity Sold', 'Sales', 'Cost of Sales', 'Gross Profit', 'Margin %']
    if request.args.get('format') == 'pdf':
        return _table_pdf_response(f'store_trading_account_{df}_to_{dt}.pdf', 'Spares Store Trading Account',
            f'Period: {df} to {dt}', header, rows)
    return csv_export_response(f'store_trading_account_{df}_to_{dt}.csv', header, rows)


def _stock_movement_summary(df, dt, part_id=None):
    """Per-part opening/in/out/adjustments/closing quantities for the
    period, run off every purchase/sale/adjustment ever recorded (not just
    ones in range) so the opening balance as of df is exact rather than
    estimated. Adjustments are tracked as a signed net delta rather than
    split into in/out — a correction isn't a restock or a sale."""
    parts = SparePart.query.order_by(SparePart.name).all()
    if part_id:
        parts = [p for p in parts if str(p.id) == str(part_id)]

    purchases = StorePurchase.query.filter(StorePurchase.purchase_date <= dt).all()
    sales = StoreSale.query.filter(StoreSale.sale_date <= dt).all()
    adjustments = StoreAdjustment.query.filter(StoreAdjustment.adjustment_date <= dt).all()

    rows = []
    for part in parts:
        p_purchases = [p for p in purchases if p.part_id == part.id]
        p_sales = [s for s in sales if s.part_id == part.id]
        p_adjustments = [a for a in adjustments if a.part_id == part.id]
        opening = (sum(p.quantity for p in p_purchases if p.purchase_date < df) -
                   sum(s.quantity for s in p_sales if s.sale_date < df) +
                   sum(a.quantity_delta for a in p_adjustments if a.adjustment_date < df))
        stock_in = sum(p.quantity for p in p_purchases if p.purchase_date >= df)
        stock_out = sum(s.quantity for s in p_sales if s.sale_date >= df)
        net_adjustments = sum(a.quantity_delta for a in p_adjustments if a.adjustment_date >= df)
        closing = opening + stock_in - stock_out + net_adjustments
        rows.append({'part': part, 'opening': opening, 'stock_in': stock_in,
                     'stock_out': stock_out, 'net_adjustments': net_adjustments, 'closing': closing})
    return rows


def _stock_movement_ledger(part, df, dt):
    """Chronological IN/OUT/ADJ ledger for one part with a running balance,
    seeded by the exact opening balance as of df."""
    purchases = StorePurchase.query.filter(StorePurchase.part_id == part.id,
                                            StorePurchase.purchase_date < df).all()
    sales = StoreSale.query.filter(StoreSale.part_id == part.id,
                                    StoreSale.sale_date < df).all()
    adjustments = StoreAdjustment.query.filter(StoreAdjustment.part_id == part.id,
                                               StoreAdjustment.adjustment_date < df).all()
    balance = (sum(p.quantity for p in purchases) - sum(s.quantity for s in sales) +
               sum(a.quantity_delta for a in adjustments))

    entries = []
    for p in StorePurchase.query.filter(StorePurchase.part_id == part.id,
                                         StorePurchase.purchase_date.between(df, dt)).all():
        entries.append({'date': p.purchase_date, 'type': 'IN', 'quantity': p.quantity,
                        'reference': p.supplier or '—', 'notes': p.notes, '_seq': p.id})
    for s in StoreSale.query.filter(StoreSale.part_id == part.id,
                                     StoreSale.sale_date.between(df, dt)).all():
        entries.append({'date': s.sale_date, 'type': 'OUT', 'quantity': s.quantity,
                        'reference': s.vehicle.registration if s.vehicle else (s.customer_name or '—'),
                        'notes': s.notes, '_seq': s.id})
    for a in StoreAdjustment.query.filter(StoreAdjustment.part_id == part.id,
                                          StoreAdjustment.adjustment_date.between(df, dt)).all():
        entries.append({'date': a.adjustment_date, 'type': 'ADJ', 'quantity': a.quantity_delta,
                        'reference': a.reason_label, 'notes': a.notes, '_seq': a.id})
    entries.sort(key=lambda e: (e['date'], e['_seq']))

    for e in entries:
        balance += e['quantity'] if e['type'] in ('IN', 'ADJ') else -e['quantity']
        e['balance'] = balance
    return entries


@app.route('/store/movements')
@login_required
@permission_required('store')
def store_movements():
    """Stock movement report: opening/in/out/closing per part for the
    period, with a chronological running-balance ledger when a single
    part is selected — the spares-store equivalent of a stock card."""
    df, dt = query_date_range()
    date_from_str, date_to_str = df.strftime('%Y-%m-%d'), dt.strftime('%Y-%m-%d')
    part_id = request.args.get('part_id', '')

    summary = _stock_movement_summary(df, dt, part_id)
    ledger = None
    ledger_part = None
    if part_id:
        ledger_part = SparePart.query.filter_by(id=part_id).first_or_404()
        ledger = _stock_movement_ledger(ledger_part, df, dt)

    all_parts = SparePart.query.order_by(SparePart.name).all()
    return render_template('store/movements.html', summary=summary, ledger=ledger,
                           ledger_part=ledger_part, parts=all_parts, part_id=part_id,
                           date_from=date_from_str, date_to=date_to_str)


@app.route('/store/movements/export')
@login_required
@permission_required('store')
def store_movements_export():
    df, dt = query_date_range()
    part_id = request.args.get('part_id', '')

    if part_id:
        part = SparePart.query.filter_by(id=part_id).first_or_404()
        ledger = _stock_movement_ledger(part, df, dt)
        rows = [[e['date'], e['type'], qty_filter(e['quantity']), e['reference'] or '',
                 qty_filter(e['balance'])] for e in ledger]
        header = ['Date', 'Type', 'Quantity', 'Reference', 'Balance']
        title = f'Stock Movement — {part.name}'
    else:
        summary = _stock_movement_summary(df, dt)
        rows = [[row['part'].name, qty_filter(row['opening']), qty_filter(row['stock_in']),
                 qty_filter(row['stock_out']), qty_filter(row['net_adjustments']),
                 qty_filter(row['closing'])] for row in summary]
        header = ['Part', 'Opening', 'Stock In', 'Stock Out', 'Adjustments', 'Closing']
        title = 'Stock Movement Summary'

    if request.args.get('format') == 'pdf':
        return _table_pdf_response(f'store_movements_{df}_to_{dt}.pdf', title,
            f'Period: {df} to {dt}', header, rows)
    return csv_export_response(f'store_movements_{df}_to_{dt}.csv', header, rows)


# ─────────────────────────────────────────────────────────────
# Reports
# ─────────────────────────────────────────────────────────────
# ─────────────────────────────────────────────────────────────
# Crew Portal
# ─────────────────────────────────────────────────────────────
# Income logging moved to the Driver Ledger (see driver_ledger / driver_ledger_add
# above) — it replaced this page so fare, diesel and mileage are captured together.


@app.route('/crew/leaderboard')
@login_required
@permission_required('crew_portal')
def crew_leaderboard():
    today = date.today()
    df, dt = query_date_range()
    date_from_str, date_to_str = df.strftime('%Y-%m-%d'), dt.strftime('%Y-%m-%d')

    dr_rate = app.config['COMMISSION_DRIVER_RATE']
    co_rate = app.config['COMMISSION_CONDUCTOR_RATE']

    rows = []
    for d in Driver.query.filter_by(status='active').order_by(Driver.name).all():
        driven = db.session.query(
            func.sum(DailyLog.gross_revenue),
            func.sum(DailyLog.trips_completed),
            func.count(DailyLog.id),
        ).filter(DailyLog.driver_id == d.id,
                 DailyLog.log_date.between(df, dt)).first()
        conducted = db.session.query(
            func.sum(DailyLog.gross_revenue),
            func.sum(DailyLog.trips_completed),
            func.count(DailyLog.id),
        ).filter(DailyLog.conductor_id == d.id,
                 DailyLog.log_date.between(df, dt)).first()

        total_rev   = (driven[0] or 0) + (conducted[0] or 0)
        total_trips = (driven[1] or 0) + (conducted[1] or 0)
        days_worked = (driven[2] or 0) + (conducted[2] or 0)
        if days_worked == 0:
            continue
        rate = d.commission_rate if d.commission_rate is not None else (
            dr_rate if d.role == 'driver' else co_rate)
        rows.append({
            'driver': d,
            'total_revenue': total_rev,
            'total_trips': total_trips,
            'days_worked': days_worked,
            'avg_per_day': total_rev / days_worked if days_worked else 0,
            'commission': total_rev * rate,
            'rate_pct': rate * 100,
        })

    rows.sort(key=lambda r: r['total_revenue'], reverse=True)
    for i, r in enumerate(rows):
        r['rank'] = i + 1

    my_driver = current_user.linked_driver
    my_row = next((r for r in rows if my_driver and r['driver'].id == my_driver.id), None)

    return render_template('crew/leaderboard.html',
                           rows=rows, my_row=my_row,
                           date_from=date_from_str, date_to=date_to_str, today=today)


def vehicle_income_totals(df, dt, vehicle_id=None):
    """Revenue/maintenance/expense/spares totals for one vehicle (or the
    whole fleet if vehicle_id is None) over [df, dt]. Only expenses (and
    store sales) explicitly tagged to a vehicle count toward that vehicle's
    statement — general overhead (untagged expenses) only appears in the
    consolidated total. Spares sold from the in-house store to a company
    vehicle are booked as an expense on that vehicle here, at the same
    marked-up price the store charges any other customer — see StoreSale."""
    rev_q = db.session.query(func.sum(DailyLog.gross_revenue)).filter(DailyLog.log_date.between(df, dt))
    maint_q = db.session.query(func.sum(MaintenanceLog.total_cost)).filter(MaintenanceLog.log_date.between(df, dt))
    exp_q = db.session.query(func.sum(Expense.amount)).filter(Expense.expense_date.between(df, dt))
    spares_q = db.session.query(func.sum(StoreSale.total_amount)).filter(StoreSale.sale_date.between(df, dt))

    if vehicle_id:
        rev_q = rev_q.filter(DailyLog.vehicle_id == vehicle_id)
        maint_q = maint_q.filter(MaintenanceLog.vehicle_id == vehicle_id)
        exp_q = exp_q.filter(Expense.vehicle_id == vehicle_id)
        spares_q = spares_q.filter(StoreSale.vehicle_id == vehicle_id)
    else:
        exp_q = exp_q.filter(Expense.vehicle_id.is_(None))
        spares_q = spares_q.filter(StoreSale.vehicle_id.isnot(None))

    revenue = rev_q.scalar() or 0
    maintenance = maint_q.scalar() or 0
    expenses = exp_q.scalar() or 0
    spares = spares_q.scalar() or 0
    return revenue, maintenance, expenses, spares


def expense_breakdown_by_category(df, dt, vehicle_id=None):
    """Group Expense amounts by heading/sub-heading for the period and scope,
    for the itemized breakdown on the income statement. Consolidated scope
    (vehicle_id=None) includes both vehicle-tagged and general expenses,
    matching the consolidated statement's total; per-vehicle scope only
    includes that vehicle's tagged expenses."""
    q = db.session.query(Expense.category_id, func.sum(Expense.amount)).filter(
        Expense.expense_date.between(df, dt))
    if vehicle_id:
        q = q.filter(Expense.vehicle_id == vehicle_id)
    totals_by_cat = dict(q.group_by(Expense.category_id).all())

    rows = []
    for h in ExpenseCategory.query.filter_by(parent_id=None).order_by(ExpenseCategory.name).all():
        h_direct = totals_by_cat.get(h.id, 0)
        children = []
        for c in sorted(h.children, key=lambda x: x.name):
            amt = totals_by_cat.get(c.id, 0)
            if amt:
                children.append({'name': c.name, 'amount': amt})
        heading_total = h_direct + sum(c['amount'] for c in children)
        if heading_total:
            rows.append({'name': h.name, 'direct': h_direct, 'children': children, 'total': heading_total})
    rows.sort(key=lambda r: r['total'], reverse=True)
    return rows


# The 'Wages' ExpenseCategory keeps its DB name (imports/other reports key
# off it), but the income statement displays it under the fuller label —
# applied consistently wherever the statement's category names surface
# (report page, PDF report pack, WhatsApp summary) via this shared map.
INCOME_STATEMENT_LABELS = {'Wages': 'Wages and Salaries'}


def statement_expense_line_items(df, dt, vehicle_id=None):
    """The actual source rows behind each Statement Summary category —
    MaintenanceLog entries and vehicle-tagged StoreSale spares under
    Maintenance, Expense rows grouped by their top-level heading (or
    'Other' for a custom heading) everywhere else. Powers the income
    statement's click-to-drill-down UI so a category total isn't a dead end."""
    standard_names = ('Maintenance', 'Wages', 'Traffic Fines', 'Insurance', 'Admin')
    items = {name: [] for name in standard_names}
    items['Other'] = []

    maint_q = MaintenanceLog.query.filter(MaintenanceLog.log_date.between(df, dt))
    if vehicle_id:
        maint_q = maint_q.filter(MaintenanceLog.vehicle_id == vehicle_id)
    for m in maint_q.all():
        items['Maintenance'].append({
            'date': m.log_date, 'source': 'Maintenance Log',
            'description': m.description or '—', 'vehicle': m.vehicle.registration,
            'amount': m.total_cost,
        })

    spares_q = StoreSale.query.filter(StoreSale.sale_date.between(df, dt))
    if vehicle_id:
        spares_q = spares_q.filter(StoreSale.vehicle_id == vehicle_id)
    else:
        spares_q = spares_q.filter(StoreSale.vehicle_id.isnot(None))
    for s in spares_q.all():
        items['Maintenance'].append({
            'date': s.sale_date, 'source': 'Store Sale',
            'description': s.part.name, 'vehicle': s.vehicle.registration if s.vehicle else '—',
            'amount': s.total_amount,
        })

    heading_name_by_cat_id = {}
    for h in ExpenseCategory.query.filter_by(parent_id=None).all():
        heading_name_by_cat_id[h.id] = h.name
        for c in h.children:
            heading_name_by_cat_id[c.id] = h.name

    exp_q = Expense.query.filter(Expense.expense_date.between(df, dt))
    if vehicle_id:
        exp_q = exp_q.filter(Expense.vehicle_id == vehicle_id)
    for e in exp_q.all():
        heading = heading_name_by_cat_id.get(e.category_id)
        bucket = heading if heading in items else 'Other'
        items[bucket].append({
            'date': e.expense_date, 'source': e.category.display_name,
            'description': e.description or '—',
            'vehicle': e.vehicle.registration if e.vehicle else '—',
            'amount': e.amount,
        })

    for rows in items.values():
        rows.sort(key=lambda r: r['date'], reverse=True)
    return {INCOME_STATEMENT_LABELS.get(name, name): rows for name, rows in items.items()}


def statement_revenue_line_items(df, dt, vehicle_id=None):
    """The DailyLog rows behind the Gross Revenue line — same purpose as
    statement_expense_line_items but for the one revenue line, so it can
    drill down to its source the same way instead of being a dead end."""
    q = DailyLog.query.filter(DailyLog.log_date.between(df, dt))
    if vehicle_id:
        q = q.filter(DailyLog.vehicle_id == vehicle_id)
    items = []
    for log in q.order_by(DailyLog.log_date.desc()).all():
        if not log.gross_revenue:
            continue
        crew = ' / '.join(n.name for n in (log.driver, log.conductor) if n) or '—'
        items.append({
            'date': log.log_date, 'source': 'Daily Transaction',
            'vehicle': log.vehicle.registration, 'description': crew,
            'amount': log.gross_revenue,
        })
    return items


def compute_income_statement(df, dt, vehicle_id=None):
    """Fleet-wide (vehicle_id=None) or per-vehicle income statement for
    [df, dt] — shared by the Income Statement report page and the Full
    Report Pack PDF so the two can't drift apart on the numbers."""
    if vehicle_id:
        # Per-vehicle statement: only costs directly attributable to this vehicle.
        gross_revenue, maintenance_cost, vehicle_expenses, spares_cost = vehicle_income_totals(df, dt, vehicle_id)
        general_expenses = 0
    else:
        # Consolidated statement: fleet-wide maintenance plus ALL expenses
        # (both vehicle-tagged and general overhead), plus spares sold from
        # the store to any company vehicle.
        gross_revenue, maintenance_cost, general_expenses, spares_cost = vehicle_income_totals(df, dt, None)
        vehicle_expenses = db.session.query(func.sum(Expense.amount)).filter(
            Expense.expense_date.between(df, dt), Expense.vehicle_id.isnot(None)).scalar() or 0

    # Wages and Salaries is driven by the actual crew commission payroll —
    # same math as the Payroll/Wages & Salaries reports (see
    # compute_payroll_earnings) — rather than manually-booked 'Wages'
    # Expense rows, which in practice are rarely entered (see report_wages).
    # Any Wages-category Expense rows that do exist still count too, added
    # to this below once the category breakdown is built.
    _, payroll_commissions, _, _, _, _ = compute_payroll_earnings(df, dt, vehicle_id or None)

    total_expenses = maintenance_cost + vehicle_expenses + general_expenses + spares_cost + payroll_commissions
    net_profit = gross_revenue - total_expenses
    profit_margin = (net_profit / gross_revenue * 100) if gross_revenue else 0

    # Same four totals as vehicle_income_totals(df, dt, v.id), but grouped
    # by vehicle in one query each instead of four queries PER vehicle —
    # this used to be 4×(fleet size) round-trips on every call (including
    # the dashboard, which doesn't even use vehicle_breakdown and was
    # paying that cost for nothing on its every single load).
    rev_by_vehicle = dict(db.session.query(DailyLog.vehicle_id, func.sum(DailyLog.gross_revenue))
                          .filter(DailyLog.log_date.between(df, dt)).group_by(DailyLog.vehicle_id).all())
    maint_by_vehicle = dict(db.session.query(MaintenanceLog.vehicle_id, func.sum(MaintenanceLog.total_cost))
                            .filter(MaintenanceLog.log_date.between(df, dt)).group_by(MaintenanceLog.vehicle_id).all())
    exp_by_vehicle = dict(db.session.query(Expense.vehicle_id, func.sum(Expense.amount))
                          .filter(Expense.expense_date.between(df, dt), Expense.vehicle_id.isnot(None))
                          .group_by(Expense.vehicle_id).all())
    spares_by_vehicle = dict(db.session.query(StoreSale.vehicle_id, func.sum(StoreSale.total_amount))
                             .filter(StoreSale.sale_date.between(df, dt), StoreSale.vehicle_id.isnot(None))
                             .group_by(StoreSale.vehicle_id).all())

    vehicle_breakdown = []
    for v in Vehicle.query.order_by(Vehicle.registration).all():
        v_rev = rev_by_vehicle.get(v.id) or 0
        v_maint = maint_by_vehicle.get(v.id) or 0
        v_exp = exp_by_vehicle.get(v.id) or 0
        v_spares = spares_by_vehicle.get(v.id) or 0
        if v_rev == 0 and v_maint == 0 and v_exp == 0 and v_spares == 0:
            continue
        v_total_cost = v_maint + v_exp + v_spares
        v_net = v_rev - v_total_cost
        vehicle_breakdown.append({
            # Spares bought from the in-house store for this vehicle count as
            # a maintenance cost, alongside logged maintenance-job spend.
            'vehicle': v, 'revenue': v_rev, 'maintenance': v_maint + v_spares,
            'expenses': v_exp, 'net_profit': v_net,
            'margin': (v_net / v_rev * 100) if v_rev else 0,
        })
    vehicle_breakdown.sort(key=lambda r: r['net_profit'], reverse=True)
    expense_breakdown = expense_breakdown_by_category(df, dt, vehicle_id or None)

    # Statement Summary classifies expenses under five fixed headings
    # (matching create_default_expense_categories) rather than the
    # tagged/untagged split — Maintenance folds in logged maintenance-job
    # costs, spares sold from the store to a company vehicle, and any
    # Expense rows booked to the Maintenance category, all together.
    # Anything booked under a custom heading a user has added falls into
    # "Other" so it still counts toward Total Operating Expenses.
    statement_category_names = ('Maintenance', 'Wages', 'Traffic Fines', 'Insurance', 'Admin')
    category_totals = {name: 0.0 for name in statement_category_names}
    other_expenses = 0.0
    for row in expense_breakdown:
        if row['name'] in category_totals:
            category_totals[row['name']] = row['total']
        else:
            other_expenses += row['total']
    category_totals['Maintenance'] += maintenance_cost + spares_cost
    category_totals['Wages'] += payroll_commissions

    statement_expenses = [(INCOME_STATEMENT_LABELS.get(name, name), category_totals[name])
                           for name in statement_category_names]
    if other_expenses:
        statement_expenses.append(('Other', other_expenses))

    # Relabel for display only, after the name-matching above (which relies
    # on the raw ExpenseCategory.name) is done.
    expense_breakdown = [
        {**row, 'name': INCOME_STATEMENT_LABELS.get(row['name'], row['name'])}
        for row in expense_breakdown
    ]

    return dict(
        gross_revenue=gross_revenue,
        maintenance_cost=maintenance_cost, vehicle_expenses=vehicle_expenses,
        general_expenses=general_expenses, total_expenses=total_expenses,
        statement_expenses=statement_expenses,
        net_profit=net_profit, profit_margin=profit_margin,
        vehicle_breakdown=vehicle_breakdown, expense_breakdown=expense_breakdown,
    )


@app.route('/reports/income')
@login_required
@permission_required('reports')
def report_income():
    vehicle_id = request.args.get('vehicle_id', '')
    df, dt = query_date_range()
    date_from_str, date_to_str = df.strftime('%Y-%m-%d'), dt.strftime('%Y-%m-%d')

    stmt = compute_income_statement(df, dt, vehicle_id)
    statement_expense_items = statement_expense_line_items(df, dt, vehicle_id or None)
    revenue_items = statement_revenue_line_items(df, dt, vehicle_id or None)

    all_vehicles = Vehicle.query.order_by(Vehicle.registration).all()
    return render_template('reports/income.html',
        statement_expense_items=statement_expense_items,
        revenue_items=revenue_items,
        vehicles=all_vehicles,
        date_from=date_from_str, date_to=date_to_str, vehicle_id=vehicle_id,
        **stmt)


def compute_payroll_earnings(df, dt, vehicle_id=None):
    """Crew commission breakdown for [df, dt] — shared by the payroll report
    page, its Excel/PDF exports, and the vehicle-scoped Wages & Salaries
    report (see wages_expense_for_vehicle) so none of them can drift apart
    on the commission math. vehicle_id narrows every DailyLog-based figure
    to that one vehicle's rows; CommissionPayment/PayrollDeduction (paid,
    deductions) stay driver-level regardless, since a payment isn't tied to
    which vehicle earned it."""
    dr_rate = app.config['COMMISSION_DRIVER_RATE']
    co_rate = app.config['COMMISSION_CONDUCTOR_RATE']

    earnings = []
    for d in Driver.query.filter_by(status='active').order_by(Driver.name).all():
        driven_q = db.session.query(func.sum(DailyLog.gross_revenue),
                                    func.count(DailyLog.id)).filter(
            DailyLog.driver_id == d.id,
            DailyLog.log_date.between(df, dt))
        conducted_q = db.session.query(func.sum(DailyLog.gross_revenue),
                                       func.count(DailyLog.id)).filter(
            DailyLog.conductor_id == d.id,
            DailyLog.log_date.between(df, dt))
        garnish_driven_q = db.session.query(func.sum(DailyLog.garnish)).filter(
            DailyLog.driver_id == d.id,
            DailyLog.log_date.between(df, dt))
        garnish_conducted_q = db.session.query(func.sum(DailyLog.garnish)).filter(
            DailyLog.conductor_id == d.id,
            DailyLog.log_date.between(df, dt))
        if vehicle_id:
            driven_q = driven_q.filter(DailyLog.vehicle_id == vehicle_id)
            conducted_q = conducted_q.filter(DailyLog.vehicle_id == vehicle_id)
            garnish_driven_q = garnish_driven_q.filter(DailyLog.vehicle_id == vehicle_id)
            garnish_conducted_q = garnish_conducted_q.filter(DailyLog.vehicle_id == vehicle_id)

        driven = driven_q.first()
        conducted = conducted_q.first()
        garnish_driven = garnish_driven_q.scalar() or 0
        garnish_conducted = garnish_conducted_q.scalar() or 0

        rev = (driven[0] or 0) + (conducted[0] or 0)
        days = (driven[1] or 0) + (conducted[1] or 0)
        garnish = garnish_driven + garnish_conducted
        if days == 0 and not garnish:
            continue
        rate = d.commission_rate if d.commission_rate is not None else (
            dr_rate if d.role == 'driver' else co_rate)
        # Garnish is netted off revenue before commission is calculated, not
        # deducted from the commission afterwards — the commission percentage
        # applies to what the crew member actually brought in after garnish.
        commission = max(rev - garnish, 0) * rate
        payment_rows = CommissionPayment.query.filter(
            CommissionPayment.driver_id == d.id,
            CommissionPayment.payment_date.between(df, dt)).order_by(CommissionPayment.payment_date).all()
        paid = sum(p.amount for p in payment_rows)
        deduction_rows = PayrollDeduction.query.filter(
            PayrollDeduction.driver_id == d.id,
            PayrollDeduction.deduction_date.between(df, dt)).order_by(PayrollDeduction.deduction_date).all()
        deductions = sum(x.amount for x in deduction_rows)
        net_pay = commission - deductions
        earnings.append({
            'driver': d,
            'total_revenue': rev,
            'days_worked': days,
            'rate_pct': rate * 100,
            'commission': commission,
            'garnish': garnish,
            'deductions': deductions,
            'deduction_rows': deduction_rows,
            'net_pay': net_pay,
            'paid': paid,
            'payments': payment_rows,
            'outstanding': net_pay - paid,
            'conductors': [],
        })

    total_commissions = sum(e['commission'] for e in earnings)
    total_garnish = sum(e['garnish'] for e in earnings)
    total_deductions = sum(e['deductions'] for e in earnings)
    total_paid = sum(e['paid'] for e in earnings)
    total_outstanding = sum(e['outstanding'] for e in earnings)

    # Nest each conductor's row under their paired driver so payroll reads
    # crew-by-crew (driver + the conductor who rode with them) instead of one
    # flat alphabetical list mixing roles. A conductor only nests if their
    # paired driver also has an earnings row this period — otherwise (no
    # pairing set, or the paired driver didn't earn anything) it stays as
    # its own top-level row, same as before. If a driver has no conductor
    # to nest, a placeholder row still prints under them labeled "Conductor",
    # with commission projected off the driver's own revenue at the standard
    # conductor rate — there's no CommissionPayment target without a real
    # person to pay it to yet, but the conductor's cut is still owed on that
    # revenue, so it's folded into the totals below (just not payable via
    # the per-row "Pay" action, which needs a real driver_id).
    grouped, nested_ids = [], set()
    placeholder_total = 0.0
    for e in earnings:
        if e['driver'].role == 'conductor':
            continue
        for ce in earnings:
            if ce['driver'].role == 'conductor' and ce['driver'].paired_driver_id == e['driver'].id:
                e['conductors'].append(ce)
                nested_ids.add(ce['driver'].id)
        if not e['conductors']:
            placeholder_commission = max(e['total_revenue'] - e['garnish'], 0) * co_rate
            e['conductors'].append({
                'driver': None, 'is_placeholder': True,
                'total_revenue': e['total_revenue'], 'days_worked': e['days_worked'],
                'rate_pct': co_rate * 100,
                'commission': placeholder_commission, 'garnish': e['garnish'],
                'deductions': 0, 'deduction_rows': [], 'net_pay': placeholder_commission, 'paid': 0, 'payments': [],
                'outstanding': placeholder_commission,
            })
            placeholder_total += placeholder_commission
        grouped.append(e)
    grouped.extend(e for e in earnings if e['driver'].role == 'conductor' and e['driver'].id not in nested_ids)
    earnings = grouped

    # total_commissions/total_outstanding above only summed named people
    # (drivers, plus any conductor who has their own Driver record) — a
    # driver with no named conductor still owes a conductor's cut on that
    # revenue (the placeholder row above), so it must count toward the
    # period totals too, not just commission tied to a named person.
    total_commissions += placeholder_total
    total_outstanding += placeholder_total

    return earnings, total_commissions, total_garnish, total_deductions, total_paid, total_outstanding


@app.route('/reports/payroll')
@login_required
@permission_required('reports')
def report_payroll():
    df, dt = query_date_range()
    date_from_str, date_to_str = df.strftime('%Y-%m-%d'), dt.strftime('%Y-%m-%d')
    earnings, total_commissions, total_garnish, total_deductions, total_paid, total_outstanding = compute_payroll_earnings(df, dt)

    return render_template('reports/payroll.html',
        earnings=earnings, total_commissions=total_commissions,
        total_garnish=total_garnish, total_deductions=total_deductions,
        total_paid=total_paid, total_outstanding=total_outstanding,
        date_from=date_from_str, date_to=date_to_str)


@app.route('/reports/payroll/export.xlsx')
@login_required
@permission_required('reports')
def export_payroll_excel():
    df, dt = query_date_range()
    date_from_str, date_to_str = df.strftime('%Y-%m-%d'), dt.strftime('%Y-%m-%d')
    earnings, total_commissions, total_garnish, total_deductions, total_paid, total_outstanding = compute_payroll_earnings(df, dt)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Payroll'
    bold = Font(bold=True)
    money_fmt = '#,##0.00'

    ws.append(['Crew Payroll / Commissions'])
    ws['A1'].font = Font(bold=True, size=14)
    ws.append([f'Period: {date_from_str} to {date_to_str}'])
    ws.append([f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")} by {current_user.username}'])
    ws.append([])

    headers = ['Crew Member', 'Role', 'Days Worked', 'Revenue Generated', 'Garnish',
               'Rate %', 'Accrued', 'Deductions', 'Net Pay', 'Paid', 'Outstanding']
    ws.append(headers)
    for cell in ws[ws.max_row]:
        cell.font = bold

    def write_row(name, role, e):
        ws.append([name, role, e['days_worked'], e['total_revenue'], e['garnish'],
                   round(e['rate_pct'], 1), e['commission'], e['deductions'], e['net_pay'],
                   e['paid'], e['outstanding']])
        r = ws.max_row
        for col in ('D', 'E', 'G', 'H', 'I', 'J', 'K'):
            ws[f'{col}{r}'].number_format = money_fmt

    for e in earnings:
        write_row(e['driver'].name, e['driver'].role.title(), e)
        for ce in e['conductors']:
            name = ce['driver'].name if ce['driver'] else 'Conductor (placeholder)'
            write_row(f'  {name}', 'Conductor', ce)

    ws.append([])
    ws.append(['TOTAL', '', '', '', total_garnish, '', total_commissions, total_deductions,
               total_commissions - total_deductions, total_paid, total_outstanding])
    r = ws.max_row
    for cell in ws[r]:
        cell.font = bold
    for col in ('E', 'G', 'H', 'I', 'J', 'K'):
        ws[f'{col}{r}'].number_format = money_fmt

    for i, width in enumerate([28, 12, 12, 17, 12, 8, 14, 14, 14, 14, 14], start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    resp = make_response(out.getvalue())
    resp.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    resp.headers['Content-Disposition'] = f'attachment; filename=payroll_{date_from_str}_to_{date_to_str}.xlsx'
    return resp


@app.route('/reports/payroll/export.pdf')
@login_required
@permission_required('reports')
def export_payroll_pdf():
    df, dt = query_date_range()
    date_from_str, date_to_str = df.strftime('%Y-%m-%d'), dt.strftime('%Y-%m-%d')
    earnings, total_commissions, total_garnish, total_deductions, total_paid, total_outstanding = compute_payroll_earnings(df, dt)

    styles = getSampleStyleSheet()
    elements = [
        Paragraph('Crew Payroll / Commissions', styles['Title']),
        Paragraph(f'Period: {date_from_str} to {date_to_str}', styles['Normal']),
        Paragraph(f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")} by {current_user.username}', styles['Normal']),
        Spacer(1, 10),
    ]

    headers = ['Crew Member', 'Role', 'Days', 'Revenue', 'Garnish', 'Rate', 'Accrued', 'Deductions', 'Net Pay', 'Paid', 'Outstanding']
    data = [headers]

    def data_row(name, role, e):
        data.append([name, role, str(e['days_worked']), f"${e['total_revenue']:,.2f}",
                     f"${e['garnish']:,.2f}", f"{e['rate_pct']:.1f}%", f"${e['commission']:,.2f}",
                     f"${e['deductions']:,.2f}", f"${e['net_pay']:,.2f}",
                     f"${e['paid']:,.2f}", f"${e['outstanding']:,.2f}"])

    for e in earnings:
        data_row(e['driver'].name, e['driver'].role.title(), e)
        for ce in e['conductors']:
            name = ce['driver'].name if ce['driver'] else 'Conductor (placeholder)'
            data_row(f'  {name}', 'Conductor', ce)

    data.append(['TOTAL', '', '', '', f'${total_garnish:,.2f}', '', f'${total_commissions:,.2f}',
                 f'${total_deductions:,.2f}', f'${total_commissions - total_deductions:,.2f}',
                 f'${total_paid:,.2f}', f'${total_outstanding:,.2f}'])
    elements.append(_pdf_table(data))

    return _pdf_response(f'payroll_{date_from_str}_to_{date_to_str}.pdf', elements, pagesize=landscape(A4))


@app.route('/reports/payroll/export-paid.pdf')
@login_required
@permission_required('reports')
def export_payroll_paid_pdf():
    """Crew who actually received a commission payment this period (paid >
    0 — see compute_payroll_earnings' CommissionPayment sum), with their
    deductions and amount paid — as opposed to export_payroll_pdf, which
    lists everyone with accrued commission whether or not they've been
    paid yet."""
    df, dt = query_date_range()
    date_from_str, date_to_str = df.strftime('%Y-%m-%d'), dt.strftime('%Y-%m-%d')
    earnings, *_ = compute_payroll_earnings(df, dt)
    paid_rows = [(name, role, e) for name, role, e in _flatten_wages_earnings(earnings) if e['paid'] > 0]

    if not paid_rows:
        flash(f'No paid crew for {date_from_str} to {date_to_str}.', 'warning')
        return redirect(url_for('report_payroll', date_from=date_from_str, date_to=date_to_str))

    styles = _pdf_styles()
    elements = [
        Paragraph('Payroll — Paid Crew', styles['Title']),
        Paragraph(f'Period: {date_from_str} to {date_to_str}', styles['Normal']),
        Paragraph(f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")} by {current_user.username}', styles['Normal']),
        Spacer(1, 10),
    ]

    data = [['Crew Member', 'Role', 'Deductions', 'Net Pay', 'Paid', 'Outstanding']]
    total_deductions = total_net_pay = total_paid = total_outstanding = 0.0
    for name, role, e in paid_rows:
        data.append([name, role, f"${e['deductions']:,.2f}", f"${e['net_pay']:,.2f}",
                     f"${e['paid']:,.2f}", f"${e['outstanding']:,.2f}"])
        total_deductions += e['deductions']
        total_net_pay += e['net_pay']
        total_paid += e['paid']
        total_outstanding += e['outstanding']
    data.append(['TOTAL', '', f'${total_deductions:,.2f}', f'${total_net_pay:,.2f}',
                 f'${total_paid:,.2f}', f'${total_outstanding:,.2f}'])
    elements.append(_pdf_table(data))

    return _pdf_response(f'payroll_paid_{date_from_str}_to_{date_to_str}.pdf', elements)


@app.route('/finance/payroll/pay-placeholder-conductor', methods=['POST'])
@login_required
@permission_required('finance')
def payroll_pay_placeholder_conductor():
    """Mark a Payroll placeholder conductor paid in one step. compute_payroll_earnings
    prints a placeholder for a driver with no named conductor on file —
    there's no Driver row to attach a CommissionPayment to until one
    exists, so that row normally has no Pay button. This auto-names the
    conductor "<Driver's first name>'s Conductor" (same convention as the
    placeholder label itself — see _flatten_wages_earnings/payslip naming)
    rather than asking for a name up front, since at pay time the actual
    person often isn't known/settled yet; the auto-created Driver record
    can always be renamed later from Drivers once it is. Reuses an
    existing one of that name under the same driver if this isn't its
    first payment, and records the payment against them in the same
    submission, so "mark paid" works the same one-step way it does for
    a named conductor."""
    date_from = request.form.get('date_from', '')
    date_to = request.form.get('date_to', '')
    back = lambda: redirect(url_for('report_payroll', date_from=date_from, date_to=date_to))
    try:
        paired_driver_id = form_int(request.form, 'paired_driver_id')
        amount = form_float(request.form, 'amount', min_value=0)
        payment_date = parse_date(request.form['payment_date'])
    except KeyError as e:
        flash(f'Missing required field: {e}', 'danger')
        return back()
    except ValueError as e:
        flash(str(e), 'danger')
        return back()

    driver = Driver.query.filter_by(id=paired_driver_id).first()
    if not driver:
        flash('Driver not found.', 'danger')
        return back()
    name = f"{driver.name.split(' ')[0]}'s Conductor" if driver.name else 'Conductor'

    conductor = Driver.query.filter(
        Driver.paired_driver_id == paired_driver_id, Driver.role == 'conductor',
        func.lower(Driver.name) == name.lower()).first()
    if not conductor:
        conductor = Driver(name=name, role='conductor', status='active', paired_driver_id=paired_driver_id)
        db.session.add(conductor)
        db.session.flush()
        log_audit('CREATE', 'drivers', conductor.id,
                  f'Added conductor {conductor.name} (paired to driver #{paired_driver_id}) from Payroll')
        touch_sync_fields(conductor)

    # The placeholder's commission is projected off the driver's own
    # DailyLog rows (see compute_payroll_earnings) — none of which name a
    # conductor yet, which is exactly why this was a placeholder. Attribute
    # this period's still-unassigned rows to the conductor now being paid,
    # so compute_payroll_earnings picks them up as *their* conducted revenue
    # from here on (matching the figure just paid) instead of the row
    # staying invisible next load — without this, the conductor has 0 days
    # worked, gets skipped by compute_payroll_earnings' "no activity"
    # filter entirely, and both the payment and any Cancel/Pay controls for
    # it would vanish from the report despite the payment existing in the DB.
    if date_from and date_to:
        unassigned_logs = DailyLog.query.filter(
            DailyLog.driver_id == paired_driver_id,
            DailyLog.conductor_id.is_(None),
            DailyLog.deleted_at.is_(None),
            DailyLog.log_date.between(parse_date(date_from), parse_date(date_to)),
        ).all()
        for log in unassigned_logs:
            log.conductor_id = conductor.id
            touch_sync_fields(log)

    payment = CommissionPayment(
        driver_id=conductor.id, payment_date=payment_date, amount=amount,
        period_start=parse_date(date_from) if date_from else None,
        period_end=parse_date(date_to) if date_to else None,
        created_by=current_user.id,
    )
    db.session.add(payment)
    db.session.flush()
    log_audit('CREATE', 'commission_payments', payment.id,
              f'Commission payment of {payment.amount} to conductor #{conductor.id} ({conductor.name})')
    touch_sync_fields(payment)
    db.session.commit()
    flash(f'{conductor.name} marked paid {amount:,.2f}.', 'success')
    return back()


@app.route('/reports/payroll/paid-sheet.pdf')
@login_required
@permission_required('reports')
def payroll_paid_sheet_pdf():
    """Payslip-style breakdown for crew who have actually been paid this
    period (filtered to paid > 0) — gross salary (accrued commission) down
    through deductions to net salary, so it can double as the payroll
    record kept on file (as opposed to the Pay Sheet, which is the blank
    sheet crew sign as cash is first handed out)."""
    df, dt = query_date_range()
    date_from_str, date_to_str = df.strftime('%Y-%m-%d'), dt.strftime('%Y-%m-%d')
    earnings, *_ = compute_payroll_earnings(df, dt)
    crew_rows = [(name, role, row) for name, role, row in _flatten_wages_earnings(earnings) if row['paid'] > 0]

    if not crew_rows:
        flash(f'No paid crew for {date_from_str} to {date_to_str}.', 'warning')
        return redirect(url_for('report_payroll', date_from=date_from_str, date_to=date_to_str))

    styles = _pdf_styles()
    elements = [
        Paragraph('Payroll Paid Sheet', styles['Title']),
        Paragraph(f'Period: {date_from_str} to {date_to_str}', styles['Normal']),
        Paragraph(f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")} by {current_user.username}', styles['Normal']),
        Spacer(1, 10),
    ]

    def reason_text(row):
        return '; '.join(f"{d.reason or 'Deduction'} (${d.amount:,.2f})" for d in row['deduction_rows']) or '—'

    data = [['#', 'Crew Member', 'Role', 'Gross Salary', 'Deductions', 'Reason', 'Net Salary']]
    total_gross = total_deductions = total_net_pay = 0.0
    for i, (name, role, row) in enumerate(crew_rows, start=1):
        data.append([str(i), name, role, f"${row['commission']:,.2f}", f"${row['deductions']:,.2f}",
                     reason_text(row), f"${row['net_pay']:,.2f}"])
        total_gross += row['commission']
        total_deductions += row['deductions']
        total_net_pay += row['net_pay']
    data.append(['', '', 'TOTAL', f"${total_gross:,.2f}", f"${total_deductions:,.2f}", '', f"${total_net_pay:,.2f}"])

    table = _pdf_table(data, bold_last_row=True, col_widths=[18, 115, 48, 62, 55, 120, 62])
    table.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 0.75, colors.black),
    ]))
    elements.append(table)

    return _pdf_response(f'payroll_paid_sheet_{date_from_str}_to_{date_to_str}.pdf', elements)


@app.route('/reports/payroll/paid-sheet.csv')
@login_required
@permission_required('reports')
def payroll_paid_sheet_csv():
    """CSV twin of payroll_paid_sheet_pdf — same paid > 0 filter and
    columns, for whoever wants the paid roster in a spreadsheet rather
    than a printable PDF."""
    df, dt = query_date_range()
    date_from_str, date_to_str = df.strftime('%Y-%m-%d'), dt.strftime('%Y-%m-%d')
    earnings, *_ = compute_payroll_earnings(df, dt)
    crew_rows = [(name, role, row) for name, role, row in _flatten_wages_earnings(earnings) if row['paid'] > 0]

    if not crew_rows:
        flash(f'No paid crew for {date_from_str} to {date_to_str}.', 'warning')
        return redirect(url_for('report_payroll', date_from=date_from_str, date_to=date_to_str))

    def reason_text(row):
        return '; '.join(f"{d.reason or 'Deduction'} (${d.amount:,.2f})" for d in row['deduction_rows'])

    header = ['#', 'Crew Member', 'Role', 'Gross Salary', 'Deductions', 'Deduction Reason', 'Net Salary']
    out_rows = [[i, name, role, f'{row["commission"]:.2f}', f'{row["deductions"]:.2f}', reason_text(row), f'{row["net_pay"]:.2f}']
                for i, (name, role, row) in enumerate(crew_rows, start=1)]
    out_rows.append(['', '', 'TOTAL', f'{sum(row["commission"] for _, _, row in crew_rows):.2f}',
                      f'{sum(row["deductions"] for _, _, row in crew_rows):.2f}', '',
                      f'{sum(row["net_pay"] for _, _, row in crew_rows):.2f}'])
    return csv_export_response(f'payroll_paid_sheet_{date_from_str}_to_{date_to_str}.csv', header, out_rows)


def _flatten_wages_earnings(earnings):
    """Flatten compute_payroll_earnings' driver-with-nested-conductors
    structure into one ordered list of (name, role, row) for a simple
    per-crew-member table — the Wages & Salaries report doesn't need the
    Payroll page's Pay/Deduct actions, just the accrued figures."""
    flat = []
    for e in earnings:
        flat.append((e['driver'].name, e['driver'].role.title(), e))
        for ce in e['conductors']:
            name = ce['driver'].name if ce['driver'] else f"{e['driver'].name.split(' ')[0]} Conductor"
            flat.append((name, 'Conductor', ce))
    return flat


@app.route('/reports/wages')
@login_required
@permission_required('reports')
def report_wages():
    """Wages & salaries expense for one vehicle (or the whole fleet) over a
    period — deliberately the *same* driver/conductor commission math as
    the Payroll report (see compute_payroll_earnings), just scoped to a
    vehicle's own DailyLog rows, so this always reconciles with what
    Payroll shows rather than depending on wage amounts being separately
    booked as Expense rows (which in practice are rarely entered)."""
    vehicle_id = request.args.get('vehicle_id', '')
    df, dt = query_date_range()
    date_from_str, date_to_str = df.strftime('%Y-%m-%d'), dt.strftime('%Y-%m-%d')

    earnings, total_commissions, total_garnish, _, _, _ = compute_payroll_earnings(df, dt, vehicle_id or None)
    crew_rows = _flatten_wages_earnings(earnings)

    all_vehicles = Vehicle.query.order_by(Vehicle.registration).all()
    return render_template('reports/wages.html',
        crew_rows=crew_rows, total_wages=total_commissions, total_garnish=total_garnish,
        vehicles=all_vehicles, vehicle_id=vehicle_id,
        date_from=date_from_str, date_to=date_to_str)


@app.route('/reports/wages/export')
@login_required
@permission_required('reports')
def export_wages():
    vehicle_id = request.args.get('vehicle_id', '')
    df, dt = query_date_range()
    date_from_str, date_to_str = df.strftime('%Y-%m-%d'), dt.strftime('%Y-%m-%d')

    earnings, total_commissions, total_garnish, _, _, _ = compute_payroll_earnings(df, dt, vehicle_id or None)
    crew_rows = _flatten_wages_earnings(earnings)

    vehicle_label = 'All Vehicles'
    if vehicle_id:
        v = Vehicle.query.filter_by(id=vehicle_id).first()
        vehicle_label = f'{v.registration} — {v.make} {v.model}' if v else f'Vehicle #{vehicle_id}'

    header = ['Crew Member', 'Role', 'Days Worked', 'Revenue Generated (USD)', 'Garnish (USD)',
               'Rate %', 'Wages / Commission (USD)']
    rows = [[name, role, r['days_worked'], f"{r['total_revenue']:.2f}", f"{r['garnish']:.2f}",
             f"{r['rate_pct']:.1f}", f"{r['commission']:.2f}"] for name, role, r in crew_rows]
    rows.append(['', '', '', '', f'{total_garnish:.2f}', 'TOTAL WAGES & SALARIES', f'{total_commissions:.2f}'])

    scope_suffix = f'_vehicle{vehicle_id}' if vehicle_id else '_all_vehicles'
    if request.args.get('format') == 'pdf':
        return _table_pdf_response(f'wages{scope_suffix}_{date_from_str}_to_{date_to_str}.pdf',
            'Wages & Salaries Expense', f'Scope: {vehicle_label} — Period: {date_from_str} to {date_to_str}',
            header, rows)
    return csv_export_response(f'wages{scope_suffix}_{date_from_str}_to_{date_to_str}.csv', header, rows)


def compute_consolidated_overview(df, dt):
    """Company-wide P&L for [df, dt], combining the three standalone income
    statements (Fleet, Franchise, Spares Store) into one set of segment
    totals, pulled from the same helpers each of those pages already uses
    so this can never drift from what they show individually.

    Naively summing revenue/expenses across segments is still correct here
    even though a store sale to a company vehicle shows up twice — once as
    Store revenue, once as a Fleet expense (see vehicle_income_totals) — at
    the same marked-up price. On combination the markup cancels out (Store
    revenue includes it, Fleet expense includes it), leaving only the
    store's real supplier cost as the company-wide expense, which is the
    economically correct combined figure."""
    gross_revenue, maintenance_cost, general_expenses, spares_cost = vehicle_income_totals(df, dt, None)
    vehicle_expenses = db.session.query(func.sum(Expense.amount)).filter(
        Expense.expense_date.between(df, dt), Expense.vehicle_id.isnot(None)).scalar() or 0
    # Crew commission payroll, same as the Fleet income statement — see
    # compute_income_statement.
    _, payroll_commissions, _, _, _, _ = compute_payroll_earnings(df, dt)
    fleet_expenses = maintenance_cost + vehicle_expenses + general_expenses + spares_cost + payroll_commissions

    daily_entries = FranchiseDailyIncome.query.filter(FranchiseDailyIncome.entry_date.between(df, dt)).all()
    weekly_entries = FranchiseWeeklyIncome.query.filter(FranchiseWeeklyIncome.week_start.between(df, dt)).all()
    daily_totals = _income_entry_totals(daily_entries)
    weekly_totals = _income_entry_totals(weekly_entries)
    franchise_income = daily_totals['income'] + weekly_totals['income']
    franchise_expenses = daily_totals['total_expenditure'] + weekly_totals['total_expenditure']

    sales = StoreSale.query.filter(StoreSale.sale_date.between(df, dt)).all()
    store_revenue = sum(s.total_amount for s in sales)
    store_cost = sum(s.unit_cost * s.quantity for s in sales)

    segments = [
        {'name': 'Fleet Operations', 'revenue': gross_revenue, 'expenses': fleet_expenses,
         'net_profit': gross_revenue - fleet_expenses,
         'count': DailyLog.query.filter(DailyLog.log_date.between(df, dt)).count()},
        {'name': 'Franchise', 'revenue': franchise_income, 'expenses': franchise_expenses,
         'net_profit': franchise_income - franchise_expenses,
         'count': len(daily_entries) + len(weekly_entries)},
        {'name': 'Spares Store', 'revenue': store_revenue, 'expenses': store_cost,
         'net_profit': store_revenue - store_cost, 'count': len(sales)},
    ]
    totals = {
        'revenue': sum(s['revenue'] for s in segments),
        'expenses': sum(s['expenses'] for s in segments),
        'net_profit': sum(s['net_profit'] for s in segments),
    }
    return segments, totals


@app.route('/reports/consolidated')
@login_required
@permission_required('reports')
def report_consolidated():
    df, dt = query_date_range()
    date_from_str, date_to_str = df.strftime('%Y-%m-%d'), dt.strftime('%Y-%m-%d')
    segments, totals = compute_consolidated_overview(df, dt)
    return render_template('reports/consolidated.html',
        segments=segments, totals=totals,
        date_from=date_from_str, date_to=date_to_str)


@app.route('/reports/consolidated/export.xlsx')
@login_required
@permission_required('reports')
def export_consolidated_excel():
    df, dt = query_date_range()
    date_from_str, date_to_str = df.strftime('%Y-%m-%d'), dt.strftime('%Y-%m-%d')
    segments, totals = compute_consolidated_overview(df, dt)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Consolidated'
    bold = Font(bold=True)
    money_fmt = '#,##0.00'

    ws.append(['Company-Wide Consolidated Statement'])
    ws['A1'].font = Font(bold=True, size=14)
    ws.append([f'Period: {date_from_str} to {date_to_str}'])
    ws.append([f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")} by {current_user.username}'])
    ws.append([])

    headers = ['Segment', 'Entries', 'Revenue', 'Expenses', 'Net Profit']
    ws.append(headers)
    for cell in ws[ws.max_row]:
        cell.font = bold

    for s in segments:
        ws.append([s['name'], s['count'], s['revenue'], s['expenses'], s['net_profit']])
        r = ws.max_row
        for col in ('C', 'D', 'E'):
            ws[f'{col}{r}'].number_format = money_fmt

    ws.append([])
    ws.append(['TOTAL', '', totals['revenue'], totals['expenses'], totals['net_profit']])
    r = ws.max_row
    for cell in ws[r]:
        cell.font = bold
    for col in ('C', 'D', 'E'):
        ws[f'{col}{r}'].number_format = money_fmt

    for i, width in enumerate([22, 10, 16, 16, 16], start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    resp = make_response(out.getvalue())
    resp.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    resp.headers['Content-Disposition'] = f'attachment; filename=consolidated_{date_from_str}_to_{date_to_str}.xlsx'
    return resp


@app.route('/reports/consolidated/export.pdf')
@login_required
@permission_required('reports')
def export_consolidated_pdf():
    df, dt = query_date_range()
    date_from_str, date_to_str = df.strftime('%Y-%m-%d'), dt.strftime('%Y-%m-%d')
    segments, totals = compute_consolidated_overview(df, dt)

    styles = getSampleStyleSheet()
    elements = [
        Paragraph('Company-Wide Consolidated Statement', styles['Title']),
        Paragraph(f'Period: {date_from_str} to {date_to_str}', styles['Normal']),
        Paragraph(f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")} by {current_user.username}', styles['Normal']),
        Spacer(1, 10),
    ]

    headers = ['Segment', 'Entries', 'Revenue', 'Expenses', 'Net Profit']
    data = [headers]
    for s in segments:
        data.append([s['name'], str(s['count']), f"${s['revenue']:,.2f}",
                     f"${s['expenses']:,.2f}", f"${s['net_profit']:,.2f}"])
    data.append(['TOTAL', '', f"${totals['revenue']:,.2f}",
                 f"${totals['expenses']:,.2f}", f"${totals['net_profit']:,.2f}"])
    elements.append(_pdf_table(data))

    return _pdf_response(f'consolidated_{date_from_str}_to_{date_to_str}.pdf', elements)


# ─────────────────────────────────────────────────────────────
# Full Report Pack — every standalone report bundled into one PDF, each
# as its own self-contained, page-broken section (a "one document with
# everything, each report standing alone" handover pack). All
# period-based sections share one [df, dt] from query_date_range(); the
# point-in-time ones (Financial Position, Distance Travelled) are taken
# "as at" dt, and Compliance is always "as at today" since it's a
# live status check, not a historical figure.
# ─────────────────────────────────────────────────────────────
_PDF_LOGO_PATH = os.path.join(app.root_path, 'static', 'img', 'logo-horizontal-dark.png')


def _draw_pdf_letterhead(canvas, doc):
    """onFirstPage/onLaterPages callback (see _pdf_response) — draws the
    GRATZ logo at the top of every page of every generated PDF, plus a
    page-number footer. Centralized here rather than as a flowable so
    every current and future PDF export picks it up automatically just by
    going through _pdf_response, with no per-report code."""
    canvas.saveState()
    logo_h = 12 * mm
    canvas.drawImage(_PDF_LOGO_PATH, doc.leftMargin, doc.pagesize[1] - doc.topMargin + 6 * mm,
                     height=logo_h, width=logo_h * (543 / 331),
                     preserveAspectRatio=True, mask='auto')
    canvas.setFont('Helvetica', 7.5)
    canvas.setFillColor(colors.HexColor('#64748b'))
    canvas.drawRightString(doc.pagesize[0] - doc.rightMargin, 10 * mm, f'Page {doc.page}')
    canvas.restoreState()


def _pdf_response(filename, elements, pagesize=A4, margins=14 * mm):
    """One response builder for every PDF export — BytesIO buffer,
    SimpleDocTemplate with the shared letterhead wired in, and the
    Content-Type/Content-Disposition headers. topMargin is padded beyond
    the caller's margin to leave room for the logo _draw_pdf_letterhead
    draws on every page."""
    out = io.BytesIO()
    doc = SimpleDocTemplate(out, pagesize=pagesize, leftMargin=margins, rightMargin=margins,
                            topMargin=margins + 14 * mm, bottomMargin=margins)
    doc.build(elements, onFirstPage=_draw_pdf_letterhead, onLaterPages=_draw_pdf_letterhead)
    out.seek(0)
    resp = make_response(out.getvalue())
    resp.headers['Content-Type'] = 'application/pdf'
    resp.headers['Content-Disposition'] = f'attachment; filename={filename}'
    return resp


def _pdf_styles():
    return getSampleStyleSheet()


def _pdf_table(data, bold_last_row=True, col_widths=None):
    """One consistent look for every header+rows report table — mirrors
    export_payroll_pdf/export_consolidated_pdf's inline style so every
    section in the Full Report Pack (and those two standalone PDFs)
    matches. First row is the header; if bold_last_row, the final row is
    treated as a bold, shaded TOTAL row."""
    table = Table(data, repeatRows=1, colWidths=col_widths)
    style = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#5f1015')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8.5),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
        ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]
    if bold_last_row:
        style.append(('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'))
        style.append(('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#f1f5f9')))
        style.append(('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f8fafc')]))
    else:
        style.append(('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]))
    table.setStyle(TableStyle(style))
    return table


def _pdf_statement_table(rows, bold_indices=()):
    """Two-column label/value table for statement-style sections (Income
    Statement, Cash Flow, Financial Position, Trading Account, Franchise
    P&L) — no header row; rows in bold_indices (subtotal/total lines) are
    bolded and shaded."""
    table = Table(rows, colWidths=[320, 140])
    style = [
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('LINEBELOW', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]
    for i in bold_indices:
        style.append(('FONTNAME', (0, i), (-1, i), 'Helvetica-Bold'))
        style.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor('#f1f5f9')))
    table.setStyle(TableStyle(style))
    return table


def _pdf_section(title, subtitle, flowables, note=None):
    """One report's worth of flowables — heading, subtitle, optional note,
    then whatever the caller built — ending in a PageBreak so each report
    in the pack starts on its own page ("standing alone")."""
    styles = _pdf_styles()
    elements = [Paragraph(title, styles['Heading1'])]
    if subtitle:
        elements.append(Paragraph(subtitle, styles['Normal']))
    elements.append(Spacer(1, 8))
    if note:
        elements.append(Paragraph(note, styles['Italic']))
        elements.append(Spacer(1, 6))
    elements.extend(flowables if isinstance(flowables, list) else [flowables])
    elements.append(PageBreak())
    return elements


def _full_pack_cover(df, dt, included):
    styles = _pdf_styles()
    elements = [
        Paragraph('Full Report Pack', styles['Title']),
        Paragraph(f'Period: {df} to {dt}', styles['Normal']),
        Paragraph(f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")} by {current_user.username}', styles['Normal']),
        Spacer(1, 16),
        Paragraph('Reports included in this document:', styles['Heading3']),
        Spacer(1, 4),
    ]
    for name in included:
        elements.append(Paragraph(f'&bull; {name}', styles['Normal']))
    elements.append(PageBreak())
    return elements


def _consolidated_overview_pdf(df, dt):
    segments, totals = compute_consolidated_overview(df, dt)
    headers = ['Segment', 'Entries', 'Revenue', 'Expenses', 'Net Profit']
    data = [headers] + [[s['name'], str(s['count']), f"${s['revenue']:,.2f}",
                         f"${s['expenses']:,.2f}", f"${s['net_profit']:,.2f}"] for s in segments]
    data.append(['TOTAL', '', f"${totals['revenue']:,.2f}", f"${totals['expenses']:,.2f}",
                f"${totals['net_profit']:,.2f}"])
    return _pdf_section('Consolidated Overview', f'Period: {df} to {dt} — Fleet, Franchise and Spares Store combined',
                        [_pdf_table(data)])


def _income_statement_pdf(df, dt):
    s = compute_income_statement(df, dt)
    rows = [['Gross Revenue', f"${s['gross_revenue']:,.2f}"]]
    for name, amt in s['statement_expenses']:
        rows.append([f'  {name}', f"${amt:,.2f}"])
    rows.append(['Total Operating Expenses', f"${s['total_expenses']:,.2f}"])
    rows.append(['NET PROFIT', f"${s['net_profit']:,.2f}"])
    rows.append(['Profit Margin', f"{s['profit_margin']:.1f}%"])
    bold_idx = {len(rows) - 3, len(rows) - 2}
    flowables = [_pdf_statement_table(rows, bold_indices=bold_idx)]

    if s['vehicle_breakdown']:
        styles = _pdf_styles()
        flowables += [Spacer(1, 14), Paragraph('Per-Vehicle Breakdown', styles['Heading3']), Spacer(1, 6)]
        headers = ['Vehicle', 'Revenue', 'Maintenance', 'Expenses', 'Net Profit', 'Margin']
        vdata = [headers] + [[
            v['vehicle'].registration, f"${v['revenue']:,.2f}", f"${v['maintenance']:,.2f}",
            f"${v['expenses']:,.2f}", f"${v['net_profit']:,.2f}", f"{v['margin']:.1f}%",
        ] for v in s['vehicle_breakdown']]
        flowables.append(_pdf_table(vdata, bold_last_row=False))
    return _pdf_section('Income Statement', f'Period: {df} to {dt} — fleet-wide', flowables)


def _cash_flow_pdf(df, dt):
    cf = compute_cash_flow(df, dt)
    rows = [
        ['Cash from Operations', ''],
        ['  Revenue collected', f"${cf['operating_in']:,.2f}"],
        ['  Receivables collected', f"${cf['receivables_in']:,.2f}"],
        ['  Maintenance paid', f"-${cf['maint_out']:,.2f}"],
        ['  Expenses paid', f"-${cf['expenses_out']:,.2f}"],
        ['  Commissions paid', f"-${cf['commission_out']:,.2f}"],
        ['  Payables paid', f"-${cf['payables_out']:,.2f}"],
        ['Net Cash from Operations', f"${cf['net_operating']:,.2f}"],
        ['Cash from Investing', ''],
        ['  Vehicles purchased', f"-${cf['investing_out']:,.2f}"],
        ['Net Cash from Investing', f"${cf['net_investing']:,.2f}"],
        ['Cash from Financing', ''],
        ['  Loan proceeds', f"${cf['loan_proceeds_in']:,.2f}"],
        ['  Loan repayments', f"-${cf['loan_repay_out']:,.2f}"],
        ['  Capital contributed', f"${cf['capital_in']:,.2f}"],
        ['  Owner drawings', f"-${cf['drawings_out']:,.2f}"],
        ['Net Cash from Financing', f"${cf['net_financing']:,.2f}"],
        ['Net Change in Cash', f"${cf['net_change']:,.2f}"],
        ['Opening Cash', f"${cf['opening_cash']:,.2f}"],
        ['Closing Cash', f"${cf['closing_cash']:,.2f}"],
    ]
    bold_labels = {'Net Cash from Operations', 'Net Cash from Investing', 'Net Cash from Financing',
                   'Net Change in Cash', 'Closing Cash'}
    bold_idx = {i for i, r in enumerate(rows) if r[0] in bold_labels}
    return _pdf_section('Cash Flow Statement', f'Period: {df} to {dt}',
                        [_pdf_statement_table(rows, bold_indices=bold_idx)])


def _financial_position_pdf(dt):
    fp = compute_financial_position(dt)
    rows = [
        ['Non-Current Assets', ''],
        ['  Vehicles at cost', f"${fp['total_cost']:,.2f}"],
        ['  Accumulated depreciation', f"-${fp['total_accum_dep']:,.2f}"],
        ['  Net book value', f"${fp['total_nbv']:,.2f}"],
        ['Current Assets', ''],
        ['  Cash and equivalents', f"${fp['cash_and_equivalents']:,.2f}"],
        ['  Receivables outstanding', f"${fp['receivables_outstanding']:,.2f}"],
        ['TOTAL ASSETS', f"${fp['total_assets']:,.2f}"],
        ['Liabilities', ''],
        ['  Loans outstanding', f"${fp['loans_outstanding']:,.2f}"],
        ['  Payables outstanding', f"${fp['payables_outstanding']:,.2f}"],
        ['  Commission payable', f"${fp['commission_payable']:,.2f}"],
        ['TOTAL LIABILITIES', f"${fp['total_liabilities']:,.2f}"],
        ['Equity', ''],
        ["  Owner's capital", f"${fp['owners_capital']:,.2f}"],
        ['  Retained earnings', f"${fp['retained_earnings']:,.2f}"],
        ['TOTAL EQUITY', f"${fp['total_equity']:,.2f}"],
    ]
    bold_labels = {'TOTAL ASSETS', 'TOTAL LIABILITIES', 'TOTAL EQUITY'}
    bold_idx = {i for i, r in enumerate(rows) if r[0] in bold_labels}
    return _pdf_section('Statement of Financial Position', f'As at {dt}',
                        [_pdf_statement_table(rows, bold_indices=bold_idx)])


def _budget_pdf(dt):
    month_start = dt.replace(day=1)
    month_end = (month_start.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)
    actuals = {
        'Revenue': db.session.query(func.sum(DailyLog.gross_revenue)).filter(
            DailyLog.log_date.between(month_start, month_end)).scalar() or 0,
        'Maintenance': db.session.query(func.sum(MaintenanceLog.total_cost)).filter(
            MaintenanceLog.log_date.between(month_start, month_end)).scalar() or 0,
    }
    for cat in ExpenseCategory.query.all():
        actuals[f'Expense: {cat.display_name}'] = db.session.query(func.sum(Expense.amount)).filter(
            Expense.category_id == cat.id, Expense.expense_date.between(month_start, month_end)).scalar() or 0
    budgets = {b.category: b.amount for b in Budget.query.filter_by(month=month_start).all()}
    all_categories = sorted(set(list(actuals.keys()) + list(budgets.keys())))
    if not all_categories:
        return _pdf_section('Budget vs Actual', f'Month: {month_start.strftime("%B %Y")}', [],
                            note='No budgets or actuals recorded for this month.')
    headers = ['Category', 'Budget', 'Actual', 'Variance']
    data = [headers] + [[cat, f"${budgets.get(cat, 0):,.2f}", f"${actuals.get(cat, 0):,.2f}",
                         f"${actuals.get(cat, 0) - budgets.get(cat, 0):,.2f}"] for cat in all_categories]
    return _pdf_section('Budget vs Actual', f'Month: {month_start.strftime("%B %Y")}',
                        [_pdf_table(data, bold_last_row=False)])


def _shortfalls_pdf(df, dt):
    rows = []
    targeted_vehicles = Vehicle.query.filter(
        Vehicle.daily_target.isnot(None), Vehicle.daily_target > 0).order_by(Vehicle.registration).all()
    for v in targeted_vehicles:
        logs = DailyLog.query.filter(DailyLog.vehicle_id == v.id, DailyLog.log_date.between(df, dt)).all()
        by_date = {}
        for log in logs:
            by_date.setdefault(log.log_date, []).append(log)
        for d, day_logs in by_date.items():
            fare = sum(l.gross_revenue for l in day_logs)
            if fare >= v.daily_target:
                continue
            garnish = sum(l.garnish for l in day_logs)
            shortfall = v.daily_target - fare
            rows.append({'vehicle': v, 'date': d, 'target': v.daily_target, 'fare': fare,
                        'shortfall': shortfall, 'garnish': garnish, 'remaining': shortfall - garnish})
    rows.sort(key=lambda r: r['date'], reverse=True)
    if not rows:
        return _pdf_section('Revenue Shortfalls', f'Period: {df} to {dt}', [], note='No shortfalls in this period.')
    headers = ['Date', 'Vehicle', 'Target', 'Fare', 'Shortfall', 'Garnish', 'Remaining']
    data = [headers] + [[
        r['date'], r['vehicle'].registration, f"${r['target']:,.2f}", f"${r['fare']:,.2f}",
        f"${r['shortfall']:,.2f}", f"${r['garnish']:,.2f}", f"${r['remaining']:,.2f}",
    ] for r in rows]
    data.append(['TOTAL', '', '', '', f"${sum(r['shortfall'] for r in rows):,.2f}",
                f"${sum(r['garnish'] for r in rows):,.2f}", f"${sum(max(r['remaining'], 0) for r in rows):,.2f}"])
    return _pdf_section('Revenue Shortfalls', f'Period: {df} to {dt}', [_pdf_table(data)])


def _payroll_pdf(df, dt):
    earnings, total_commissions, total_garnish, total_deductions, total_paid, total_outstanding = compute_payroll_earnings(df, dt)
    headers = ['Crew Member', 'Role', 'Days', 'Revenue', 'Garnish', 'Rate', 'Accrued', 'Deductions', 'Net Pay', 'Paid', 'Outstanding']
    data = [headers]

    def add_row(name, role, e):
        data.append([name, role, str(e['days_worked']), f"${e['total_revenue']:,.2f}", f"${e['garnish']:,.2f}",
                    f"{e['rate_pct']:.1f}%", f"${e['commission']:,.2f}", f"${e['deductions']:,.2f}",
                    f"${e['net_pay']:,.2f}", f"${e['paid']:,.2f}", f"${e['outstanding']:,.2f}"])

    for e in earnings:
        add_row(e['driver'].name, e['driver'].role.title(), e)
        for ce in e['conductors']:
            name = ce['driver'].name if ce['driver'] else 'Conductor (placeholder)'
            add_row(f'  {name}', 'Conductor', ce)
    data.append(['TOTAL', '', '', '', f"${total_garnish:,.2f}", '', f"${total_commissions:,.2f}",
                f"${total_deductions:,.2f}", f"${total_commissions - total_deductions:,.2f}",
                f"${total_paid:,.2f}", f"${total_outstanding:,.2f}"])
    return _pdf_section('Crew Payroll / Commissions', f'Period: {df} to {dt}', [_pdf_table(data)])


def _vehicle_performance_pdf(df, dt):
    rows = []
    for v in Vehicle.query.filter_by(status='active').order_by(Vehicle.registration).all():
        agg = db.session.query(func.sum(DailyLog.gross_revenue), func.sum(DailyLog.trips_completed),
                               func.count(DailyLog.id)).filter(
            DailyLog.vehicle_id == v.id, DailyLog.log_date.between(df, dt)).first()
        revenue, trips, days = (agg[0] or 0), (agg[1] or 0), (agg[2] or 0)
        if days == 0:
            continue
        avg_per_day = revenue / days if days else 0
        target = v.daily_target or 0
        achievement = (avg_per_day / target * 100) if target else None
        rows.append({'vehicle': v, 'days': days, 'trips': trips, 'revenue': revenue,
                    'avg_per_day': avg_per_day, 'target': target, 'achievement': achievement})
    rows.sort(key=lambda r: r['revenue'], reverse=True)
    if not rows:
        return _pdf_section('Vehicle Performance', f'Period: {df} to {dt}', [], note='No activity logged in this period.')
    headers = ['Vehicle', 'Days Logged', 'Trips', 'Revenue', 'Avg/Day', 'Target', 'Achievement']
    data = [headers] + [[
        r['vehicle'].registration, str(r['days']), str(r['trips']), f"${r['revenue']:,.2f}",
        f"${r['avg_per_day']:,.2f}", f"${r['target']:,.2f}" if r['target'] else '—',
        f"{r['achievement']:.0f}%" if r['achievement'] is not None else '—',
    ] for r in rows]
    return _pdf_section('Vehicle Performance', f'Period: {df} to {dt}', [_pdf_table(data, bold_last_row=False)])


def _fuel_efficiency_pdf(df, dt):
    rows = []
    for v in Vehicle.query.order_by(Vehicle.registration).all():
        logs = FuelLog.query.filter(FuelLog.vehicle_id == v.id, FuelLog.log_date.between(df, dt),
                                    FuelLog.odometer.isnot(None)).order_by(FuelLog.odometer).all()
        if len(logs) < 2:
            continue
        total_distance = total_liters = 0
        for prev, curr in zip(logs, logs[1:]):
            distance = curr.odometer - prev.odometer
            if distance <= 0 or not curr.liters:
                continue
            total_distance += distance
            total_liters += curr.liters
        if not total_distance:
            continue
        rows.append({'vehicle': v, 'distance': total_distance, 'liters': total_liters,
                    'avg': (total_liters / total_distance) * 100})
    rows.sort(key=lambda r: r['avg'])
    if not rows:
        return _pdf_section('Fuel Efficiency', f'Period: {df} to {dt}', [],
                            note='Not enough fuel/odometer data in this period.')
    headers = ['Vehicle', 'Distance (km)', 'Fuel (L)', 'Avg L/100km']
    data = [headers] + [[r['vehicle'].registration, f"{r['distance']:,.0f}", f"{r['liters']:,.1f}",
                         f"{r['avg']:.1f}"] for r in rows]
    fleet_distance = sum(r['distance'] for r in rows)
    fleet_liters = sum(r['liters'] for r in rows)
    fleet_avg = (fleet_liters / fleet_distance * 100) if fleet_distance else 0
    data.append(['FLEET AVERAGE', f"{fleet_distance:,.0f}", f"{fleet_liters:,.1f}", f"{fleet_avg:.1f}"])
    return _pdf_section('Fuel Efficiency', f'Period: {df} to {dt}', [_pdf_table(data)])


def _distance_travelled_pdf(dt):
    rows = []
    for v in Vehicle.query.order_by(Vehicle.registration).all():
        odometer = db.session.query(func.max(FuelLog.odometer)).filter(
            FuelLog.vehicle_id == v.id, FuelLog.log_date == dt, FuelLog.odometer.isnot(None)).scalar()
        distance = None
        if odometer is not None:
            prev = FuelLog.query.filter(FuelLog.vehicle_id == v.id, FuelLog.log_date < dt,
                                        FuelLog.odometer.isnot(None)).order_by(FuelLog.log_date.desc()).first()
            if prev:
                distance = odometer - prev.odometer
        if distance is not None:
            rows.append({'vehicle': v, 'odometer': odometer, 'distance': distance})
    if not rows:
        return _pdf_section('Distance Travelled', f'As at {dt}', [], note='No odometer readings for this date.')
    headers = ['Vehicle', 'Odometer', 'Distance Since Last Reading (km)']
    data = [headers] + [[r['vehicle'].registration, f"{r['odometer']:,.0f}", f"{r['distance']:,.0f}"] for r in rows]
    data.append(['FLEET TOTAL', '', f"{sum(r['distance'] for r in rows):,.0f}"])
    return _pdf_section('Distance Travelled', f'As at {dt}', [_pdf_table(data)])


def _route_profitability_pdf(df, dt):
    total_revenue = db.session.query(func.sum(DailyLog.gross_revenue)).filter(
        DailyLog.log_date.between(df, dt)).scalar() or 0
    total_costs = db.session.query(func.sum(MaintenanceLog.total_cost)).filter(
        MaintenanceLog.log_date.between(df, dt)).scalar() or 0

    route_data = db.session.query(
        Route.id, Route.name, func.sum(DailyLog.gross_revenue).label('revenue'),
        func.count(DailyLog.id).label('log_days'),
    ).join(DailyLog, Route.id == DailyLog.route_id).filter(
        DailyLog.log_date.between(df, dt)).group_by(Route.id).all()

    rows = []
    for r in route_data:
        revenue = r.revenue or 0
        allocated_cost = (revenue / total_revenue * total_costs) if total_revenue else 0
        rows.append({'name': r.name, 'revenue': revenue, 'log_days': r.log_days,
                    'allocated_cost': allocated_cost, 'net_profit': revenue - allocated_cost})
    unrouted = db.session.query(
        func.sum(DailyLog.gross_revenue).label('revenue'), func.count(DailyLog.id).label('log_days'),
    ).filter(DailyLog.route_id.is_(None), DailyLog.log_date.between(df, dt)).first()
    if unrouted and unrouted.log_days:
        revenue = unrouted.revenue or 0
        allocated_cost = (revenue / total_revenue * total_costs) if total_revenue else 0
        rows.append({'name': '(No Route)', 'revenue': revenue, 'log_days': unrouted.log_days,
                    'allocated_cost': allocated_cost, 'net_profit': revenue - allocated_cost})
    rows.sort(key=lambda x: x['net_profit'], reverse=True)
    if not rows:
        return _pdf_section('Route Profitability', f'Period: {df} to {dt}', [], note='No route activity in this period.')
    headers = ['Route', 'Log Days', 'Revenue', 'Allocated Cost', 'Net Profit']
    data = [headers] + [[r['name'], str(r['log_days']), f"${r['revenue']:,.2f}",
                         f"${r['allocated_cost']:,.2f}", f"${r['net_profit']:,.2f}"] for r in rows]
    data.append(['TOTAL', '', f"${total_revenue:,.2f}", f"${total_costs:,.2f}", f"${total_revenue - total_costs:,.2f}"])
    return _pdf_section('Route Profitability', f'Period: {df} to {dt}', [_pdf_table(data)])


def _vehicle_efficiency_pdf(df, dt):
    rows = _compute_vehicle_efficiency_rows(df, dt)
    if not rows:
        return _pdf_section('Vehicle Efficiency & Profitability', f'Period: {df} to {dt}', [],
                            note='No revenue, cost or fuel/odometer data in this period.')
    headers = ['Vehicle', 'Revenue', 'Maintenance', 'Other Exp.', 'Total Cost', 'Net Profit', 'Km/Liter']
    data = [headers] + [[
        r['vehicle'].registration, f"${r['revenue']:,.2f}",
        f"${r['maintenance_cost']:,.2f}", f"${r['other_expenses']:,.2f}", f"${r['total_cost']:,.2f}",
        f"${r['net_profit']:,.2f}", f"{r['km_per_liter']:.1f}" if r['distance'] else '—',
    ] for r in rows]
    fleet_revenue = sum(r['revenue'] for r in rows)
    fleet_maint = sum(r['maintenance_cost'] for r in rows)
    fleet_other = sum(r['other_expenses'] for r in rows)
    fleet_cost = sum(r['total_cost'] for r in rows)
    data.append(['TOTAL', f"${fleet_revenue:,.2f}", f"${fleet_maint:,.2f}",
                 f"${fleet_other:,.2f}", f"${fleet_cost:,.2f}", f"${fleet_revenue - fleet_cost:,.2f}", ''])
    return _pdf_section('Vehicle Efficiency & Profitability', f'Period: {df} to {dt}', [_pdf_table(data)])


def _daily_transactions_pdf(df, dt):
    logs = DailyLog.query.filter(DailyLog.log_date.between(df, dt)).order_by(DailyLog.log_date.desc()).all()
    if not logs:
        return _pdf_section('Daily Transactions', f'Period: {df} to {dt}', [], note='No daily transactions in this period.')
    headers = ['Date', 'Vehicle', 'Driver', 'Route', 'Trips', 'Revenue', 'Garnish']
    data = [headers] + [[
        l.log_date, l.vehicle.registration, l.driver.name if l.driver else '—',
        l.route.name if l.route else '—', str(l.trips_completed), f"${l.gross_revenue:,.2f}",
        f"${l.garnish:,.2f}" if l.garnish else '—',
    ] for l in logs]
    data.append(['TOTAL', '', '', '', str(len(logs)), f"${sum(l.gross_revenue for l in logs):,.2f}",
                f"${sum(l.garnish or 0 for l in logs):,.2f}"])
    return _pdf_section('Daily Transactions', f'Period: {df} to {dt} — {len(logs)} entries', [_pdf_table(data)])


def _franchise_reconciliation_pdf(df, dt):
    daily_entries = FranchiseDailyIncome.query.filter(FranchiseDailyIncome.entry_date.between(df, dt)).all()
    weekly_entries = FranchiseWeeklyIncome.query.filter(FranchiseWeeklyIncome.week_start.between(df, dt)).all()
    daily_rows = _group_income_by_period(daily_entries, 'entry_date')
    weekly_rows = _group_income_by_period(weekly_entries, 'week_start')
    daily_totals = _income_entry_totals(daily_entries)
    weekly_totals = _income_entry_totals(weekly_entries)

    def rows_table(rows, totals, period_label):
        headers = [period_label, 'Vehicles', 'Income', 'Expenditure', 'Net Income', 'Deposited', 'Variance']
        data = [headers] + [[
            r['period'], str(r['vehicle_count']), f"${r['income']:,.2f}", f"${r['total_expenditure']:,.2f}",
            f"${r['cash_in_hand']:,.2f}", f"${r['deposited']:,.2f}", f"${r['variance']:,.2f}",
        ] for r in rows]
        data.append(['TOTAL', '', f"${totals['income']:,.2f}", f"${totals['total_expenditure']:,.2f}",
                    f"${totals['cash_in_hand']:,.2f}", f"${totals['deposited']:,.2f}", f"${totals['variance']:,.2f}"])
        return _pdf_table(data)

    styles = _pdf_styles()
    flowables = [Paragraph('Daily Franchise Income', styles['Heading3']), Spacer(1, 4)]
    flowables.append(rows_table(daily_rows, daily_totals, 'Date') if daily_rows
                     else Paragraph('No daily entries in this period.', styles['Normal']))
    flowables += [Spacer(1, 14), Paragraph('Weekly Franchise Income', styles['Heading3']), Spacer(1, 4)]
    flowables.append(rows_table(weekly_rows, weekly_totals, 'Week Of') if weekly_rows
                     else Paragraph('No weekly entries in this period.', styles['Normal']))
    return _pdf_section('Franchise Reconciliation Schedule', f'Period: {df} to {dt}', flowables)


def _franchise_weekly_analysis_pdf(df, dt):
    daily_entries = FranchiseDailyIncome.query.filter(
        FranchiseDailyIncome.entry_date.between(df, dt)).order_by(FranchiseDailyIncome.entry_date.asc()).all()
    weekly_entries = FranchiseWeeklyIncome.query.filter(
        FranchiseWeeklyIncome.week_start.between(df, dt)).order_by(FranchiseWeeklyIncome.week_start.asc()).all()
    weekly_by_week = {}
    for e in weekly_entries:
        week_start = e.week_start - timedelta(days=e.week_start.weekday())
        weekly_by_week.setdefault(week_start, []).append(e)
    daily_by_week = {}
    for e in daily_entries:
        week_start = e.entry_date - timedelta(days=e.entry_date.weekday())
        daily_by_week.setdefault(week_start, []).append(e)
    week_starts = sorted(set(daily_by_week.keys()) | set(weekly_by_week.keys()))
    if not week_starts:
        return _pdf_section('Franchise Weekly Analysis', f'Period: {df} to {dt}', [],
                            note='No franchise income in this period.')
    rows = []
    for start in week_starts:
        dtot = _income_entry_totals(daily_by_week.get(start, []))
        wtot = _income_entry_totals(weekly_by_week.get(start, []))
        total_income = dtot['income'] + wtot['income']
        total_exp = dtot['total_expenditure'] + wtot['total_expenditure']
        rows.append({'week_start': start, 'total_income': total_income, 'total_expenditure': total_exp,
                    'net_profit': total_income - total_exp})
    headers = ['Week Of', 'Total Income', 'Total Expenditure', 'Net Profit']
    data = [headers] + [[r['week_start'], f"${r['total_income']:,.2f}", f"${r['total_expenditure']:,.2f}",
                         f"${r['net_profit']:,.2f}"] for r in rows]
    data.append(['TOTAL', f"${sum(r['total_income'] for r in rows):,.2f}",
                f"${sum(r['total_expenditure'] for r in rows):,.2f}", f"${sum(r['net_profit'] for r in rows):,.2f}"])
    return _pdf_section('Franchise Weekly Analysis', f'Period: {df} to {dt}', [_pdf_table(data)])


def _franchise_consolidated_pdf(df, dt):
    daily_entries = FranchiseDailyIncome.query.filter(FranchiseDailyIncome.entry_date.between(df, dt)).all()
    weekly_entries = FranchiseWeeklyIncome.query.filter(FranchiseWeeklyIncome.week_start.between(df, dt)).all()
    daily_totals = _income_entry_totals(daily_entries)
    weekly_totals = _income_entry_totals(weekly_entries)
    total_income = daily_totals['income'] + weekly_totals['income']
    total_expenditure = daily_totals['total_expenditure'] + weekly_totals['total_expenditure']
    rows = [
        ['Daily Franchise Income', f"${daily_totals['income']:,.2f}"],
        ['Weekly Franchise Income', f"${weekly_totals['income']:,.2f}"],
        ['TOTAL INCOME', f"${total_income:,.2f}"],
        ['Total Expenditure', f"${total_expenditure:,.2f}"],
        ['NET PROFIT', f"${total_income - total_expenditure:,.2f}"],
    ]
    return _pdf_section('Franchise Consolidated P&L', f'Period: {df} to {dt}',
                        [_pdf_statement_table(rows, bold_indices={2, 4})])


def _store_trading_account_pdf(df, dt):
    sales = StoreSale.query.filter(StoreSale.sale_date.between(df, dt)).all()
    sales_revenue = sum(s.total_amount for s in sales)
    cost_of_sales = sum(s.unit_cost * s.quantity for s in sales)
    gross_profit = sales_revenue - cost_of_sales
    purchases_total = db.session.query(func.sum(StorePurchase.total_cost)).filter(
        StorePurchase.purchase_date.between(df, dt)).scalar() or 0
    closing_stock_value = sum(p.stock_value for p in SparePart.query.all())
    rows = [
        ['Sales Revenue', f"${sales_revenue:,.2f}"],
        ['Cost of Sales', f"-${cost_of_sales:,.2f}"],
        ['GROSS PROFIT', f"${gross_profit:,.2f}"],
        ['Purchases This Period', f"${purchases_total:,.2f}"],
        ['Closing Stock Value', f"${closing_stock_value:,.2f}"],
    ]
    return _pdf_section('Spares Store Trading Account', f'Period: {df} to {dt}',
                        [_pdf_statement_table(rows, bold_indices={2})])


def _compliance_pdf():
    today = date.today()
    threshold = today + timedelta(days=30)
    expired = VehicleDocument.query.filter(VehicleDocument.expiry_date < today).order_by(
        VehicleDocument.expiry_date).all()
    expiring = VehicleDocument.query.filter(VehicleDocument.expiry_date.between(today, threshold)).order_by(
        VehicleDocument.expiry_date).all()
    for v in Vehicle.query.filter(Vehicle.insurance_expiry.isnot(None)).all():
        doc_type = f'Insurance ({v.insurance_type})' if v.insurance_type else 'Insurance'
        entry = {'vehicle': v, 'doc_type': doc_type, 'expiry_date': v.insurance_expiry}
        if v.insurance_status == 'expired':
            expired.append(entry)
        elif v.insurance_status == 'warning':
            expiring.append(entry)
    if not expired and not expiring:
        return _pdf_section('Compliance', f'As at {today}', [], note='No expired or soon-to-expire documents.')

    def doc_row(d):
        if isinstance(d, dict):
            return [d['vehicle'].registration, d['doc_type'], str(d['expiry_date'])]
        return [d.vehicle.registration, d.doc_type.title(), str(d.expiry_date)]

    styles = _pdf_styles()
    flowables = []
    if expired:
        flowables += [Paragraph(f'Expired ({len(expired)})', styles['Heading3']), Spacer(1, 4),
                     _pdf_table([['Vehicle', 'Document', 'Expired On']] + [doc_row(d) for d in expired],
                                bold_last_row=False), Spacer(1, 14)]
    if expiring:
        flowables += [Paragraph(f'Expiring Within 30 Days ({len(expiring)})', styles['Heading3']), Spacer(1, 4),
                     _pdf_table([['Vehicle', 'Document', 'Expires On']] + [doc_row(d) for d in expiring],
                                bold_last_row=False)]
    return _pdf_section('Compliance', f'As at {today}', flowables)


@app.route('/reports/full-pack/export.pdf')
@login_required
@permission_required('reports')
def export_full_report_pack():
    """Every standalone report the current user has access to, bundled
    into one PDF — each report is its own page-broken section, so the
    document is a handover/filing pack rather than one merged table."""
    df, dt = query_date_range()
    date_from_str, date_to_str = df.strftime('%Y-%m-%d'), dt.strftime('%Y-%m-%d')

    sections = [
        ('Consolidated Overview', True, lambda: _consolidated_overview_pdf(df, dt)),
        ('Income Statement', True, lambda: _income_statement_pdf(df, dt)),
        ('Cash Flow Statement', True, lambda: _cash_flow_pdf(df, dt)),
        ('Statement of Financial Position', True, lambda: _financial_position_pdf(dt)),
        ('Budget vs Actual', True, lambda: _budget_pdf(dt)),
        ('Revenue Shortfalls', True, lambda: _shortfalls_pdf(df, dt)),
        ('Crew Payroll / Commissions', True, lambda: _payroll_pdf(df, dt)),
        ('Vehicle Performance', True, lambda: _vehicle_performance_pdf(df, dt)),
        ('Fuel Efficiency', True, lambda: _fuel_efficiency_pdf(df, dt)),
        ('Distance Travelled', True, lambda: _distance_travelled_pdf(dt)),
        ('Route Profitability', True, lambda: _route_profitability_pdf(df, dt)),
        ('Vehicle Efficiency & Profitability', True, lambda: _vehicle_efficiency_pdf(df, dt)),
        ('Daily Transactions', True, lambda: _daily_transactions_pdf(df, dt)),
        ('Franchise Reconciliation Schedule', current_user.has_permission('franchise'),
         lambda: _franchise_reconciliation_pdf(df, dt)),
        ('Franchise Weekly Analysis', current_user.has_permission('franchise'),
         lambda: _franchise_weekly_analysis_pdf(df, dt)),
        ('Franchise Consolidated P&L', current_user.has_permission('franchise'),
         lambda: _franchise_consolidated_pdf(df, dt)),
        ('Spares Store Trading Account', current_user.has_permission('store'),
         lambda: _store_trading_account_pdf(df, dt)),
        ('Compliance', current_user.has_permission('compliance'), lambda: _compliance_pdf()),
    ]
    included = [(name, build) for name, allowed, build in sections if allowed]
    if not included:
        flash('No reports available for your permissions.', 'danger')
        return redirect(url_for('report_consolidated'))

    elements = _full_pack_cover(df, dt, [name for name, _ in included])
    for _name, build in included:
        elements.extend(build())
    if elements and isinstance(elements[-1], PageBreak):
        elements.pop()

    return _pdf_response(f'full_report_pack_{date_from_str}_to_{date_to_str}.pdf', elements)


def _compute_shortfall_rows(df, dt):
    """Flags every vehicle/day where actual fare fell below that vehicle's
    admin-set daily_target — vehicles with no target set are skipped
    entirely. Each flagged day shows how much garnish (if any) has already
    been applied against the shortfall, so the admin can see at a glance
    what's still unresolved and act on it inline. Shared by the report page
    and its CSV export so both stay in sync with one computation."""
    rows = []
    targeted_vehicles = Vehicle.query.filter(
        Vehicle.daily_target.isnot(None), Vehicle.daily_target > 0
    ).order_by(Vehicle.registration).all()
    for v in targeted_vehicles:
        logs = DailyLog.query.filter(
            DailyLog.vehicle_id == v.id, DailyLog.log_date.between(df, dt)
        ).all()
        by_date = {}
        for log in logs:
            by_date.setdefault(log.log_date, []).append(log)
        for d, day_logs in by_date.items():
            fare = sum(l.gross_revenue for l in day_logs)
            if fare >= v.daily_target:
                continue
            garnish = sum(l.garnish for l in day_logs)
            reasons = '; '.join(n for n in (l.reason_for_shortfall for l in day_logs) if n) or None
            drivers = sorted({l.driver for l in day_logs if l.driver}, key=lambda dr: dr.name)
            shortfall = v.daily_target - fare
            rows.append({
                'vehicle': v, 'date': d, 'drivers': drivers,
                'target': v.daily_target, 'fare': fare, 'shortfall': shortfall,
                'garnish': garnish, 'remaining': shortfall - garnish,
                'reason_for_shortfall': reasons,
            })
    rows.sort(key=lambda r: r['date'], reverse=True)
    return rows


@app.route('/reports/shortfalls')
@login_required
@permission_required('reports')
def report_shortfalls():
    df, dt = query_date_range()
    date_from_str, date_to_str = df.strftime('%Y-%m-%d'), dt.strftime('%Y-%m-%d')

    rows = _compute_shortfall_rows(df, dt)
    total_shortfall = sum(r['shortfall'] for r in rows)
    total_garnish = sum(r['garnish'] for r in rows)
    total_remaining = sum(max(r['remaining'], 0) for r in rows)
    pending_count = sum(1 for r in rows if r['remaining'] > 0)

    return render_template('reports/shortfalls.html', rows=rows,
        total_shortfall=total_shortfall, total_garnish=total_garnish,
        total_remaining=total_remaining, pending_count=pending_count,
        date_from=date_from_str, date_to=date_to_str)


@app.route('/reports/shortfalls/export')
@login_required
@permission_required('reports')
def report_shortfalls_export():
    df, dt = query_date_range()
    rows = _compute_shortfall_rows(df, dt)
    out_rows = [[r['date'], r['vehicle'].registration, ', '.join(d.name for d in r['drivers']),
                 f"{r['target']:.2f}", f"{r['fare']:.2f}", f"{r['shortfall']:.2f}",
                 f"{r['garnish']:.2f}", f"{max(r['remaining'], 0):.2f}", r['reason_for_shortfall'] or '']
                for r in rows]
    header = ['Date', 'Vehicle', 'Driver(s)', 'Target', 'Fare', 'Shortfall', 'Garnish Applied', 'Remaining', 'Reason']
    if request.args.get('format') == 'pdf':
        return _table_pdf_response(f'revenue_shortfalls_{df}_to_{dt}.pdf', 'Revenue Shortfalls',
            f'Period: {df} to {dt}', header, out_rows)
    return csv_export_response(f'revenue_shortfalls_{df}_to_{dt}.csv', header, out_rows)


@app.route('/finance/commission-payments')
@login_required
@permission_required('finance')
def commission_payments_list():
    """Flat list of individual CommissionPayment records with edit/delete —
    the Payroll report only shows accrued/paid/outstanding totals per
    driver, with no per-payment row to attach those actions to."""
    page = request.args.get('page', 1, type=int)
    payments = CommissionPayment.query.join(Driver).order_by(
        CommissionPayment.payment_date.desc()).paginate(page=page, per_page=30)
    return render_template('reports/commission_payments.html', payments=payments)


@app.route('/finance/commission-payments/<int:pid>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
@handle_form_errors
def commission_payment_edit(pid):
    payment = CommissionPayment.query.filter_by(id=pid).first_or_404()
    if request.method == 'POST':
        payment.driver_id = form_int(request.form, 'driver_id')
        payment.payment_date = parse_date(request.form['payment_date'])
        payment.amount = form_float(request.form, 'amount', min_value=0)
        payment.period_start = parse_date(request.form.get('period_start'))
        payment.period_end = parse_date(request.form.get('period_end'))
        payment.method = request.form.get('method', '').strip()
        payment.notes = request.form.get('notes', '').strip()
        log_audit('UPDATE', 'commission_payments', payment.id,
                  f'Updated commission payment to driver #{payment.driver_id}')
        touch_sync_fields(payment)
        db.session.commit()
        flash('Commission payment updated.', 'success')
        return redirect(url_for('commission_payments_list'))
    all_drivers = Driver.query.order_by(Driver.name).all()
    return render_template('reports/commission_payment_form.html', payment=payment, drivers=all_drivers)


@app.route('/finance/commission-payments/add', methods=['POST'])
@login_required
@permission_required('finance')
@handle_form_errors
def commission_payment_add():
    payment = CommissionPayment(
        driver_id=form_int(request.form, 'driver_id'),
        payment_date=parse_date(request.form['payment_date']),
        amount=form_float(request.form, 'amount', min_value=0),
        period_start=parse_date(request.form.get('period_start')),
        period_end=parse_date(request.form.get('period_end')),
        method=request.form.get('method', '').strip(),
        notes=request.form.get('notes', '').strip(),
        created_by=current_user.id,
    )
    db.session.add(payment)
    db.session.flush()
    log_audit('CREATE', 'commission_payments', payment.id,
              f'Commission payment of {payment.amount} to driver #{payment.driver_id}')
    touch_sync_fields(payment)
    db.session.commit()
    flash('Commission payment recorded.', 'success')
    return redirect(request.referrer or url_for('report_payroll'))


@app.route('/finance/commission-payments/<int:pid>/delete', methods=['POST'])
@login_required
@admin_required
def commission_payment_delete(pid):
    payment = CommissionPayment.query.filter_by(id=pid).first_or_404()
    log_audit('DELETE', 'commission_payments', pid, f'Deleted commission payment of {payment.amount}')
    payment.deleted_at = datetime.now(timezone.utc)
    touch_sync_fields(payment)
    db.session.commit()
    flash('Commission payment deleted.', 'warning')
    return redirect(request.referrer or url_for('report_payroll'))


# ─────────────────────────────────────────────────────────────
# Payroll Deductions — withholdings against a crew member's accrued
# commission (see PayrollDeduction, compute_payroll_earnings), and the
# per-driver Payslip PDF that itemizes them. CRUD mirrors CommissionPayment
# above exactly (same permission split, same inline-add-from-Payroll-report
# pattern) since the two are companion concepts on the same report.
# ─────────────────────────────────────────────────────────────
@app.route('/finance/payroll-deductions')
@login_required
@permission_required('finance')
def payroll_deductions_list():
    """Flat list of individual PayrollDeduction records with edit/delete —
    the Payroll report only shows the total deducted per driver, with no
    per-deduction row to attach those actions to."""
    page = request.args.get('page', 1, type=int)
    deductions = PayrollDeduction.query.join(Driver).order_by(
        PayrollDeduction.deduction_date.desc()).paginate(page=page, per_page=30)
    return render_template('reports/payroll_deductions.html', deductions=deductions)


@app.route('/finance/payroll-deductions/<int:did>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
@handle_form_errors
def payroll_deduction_edit(did):
    deduction = PayrollDeduction.query.filter_by(id=did).first_or_404()
    if request.method == 'POST':
        deduction.driver_id = form_int(request.form, 'driver_id')
        deduction.deduction_date = parse_date(request.form['deduction_date'])
        deduction.amount = form_float(request.form, 'amount', min_value=0)
        deduction.reason = request.form.get('reason', '').strip()
        if not deduction.reason:
            raise ValueError('Reason is required.')
        log_audit('UPDATE', 'payroll_deductions', deduction.id,
                  f'Updated payroll deduction for driver #{deduction.driver_id}')
        touch_sync_fields(deduction)
        db.session.commit()
        flash('Deduction updated.', 'success')
        return redirect(url_for('payroll_deductions_list'))
    all_drivers = Driver.query.order_by(Driver.name).all()
    return render_template('reports/payroll_deduction_form.html', deduction=deduction, drivers=all_drivers)


@app.route('/finance/payroll-deductions/add', methods=['POST'])
@login_required
@permission_required('finance')
@handle_form_errors
def payroll_deduction_add():
    """Records one or more deductions in a single submission — each
    amount/reason pair (payroll.html's Deduct form lets you add several
    before saving) becomes its own PayrollDeduction row, since each
    deducted amount stands on its own with its own reason rather than
    being lumped into one figure. Falls back to the singular amount/reason
    fields for any caller that only ever posts one (e.g. a future
    integration), so the multi-row list isn't the only supported shape."""
    driver_id = form_int(request.form, 'driver_id')
    deduction_date = parse_date(request.form['deduction_date'])
    amounts = request.form.getlist('amount[]') or [request.form.get('amount', '')]
    reasons = request.form.getlist('reason[]') or [request.form.get('reason', '')]

    created = []
    for amount_raw, reason in zip(amounts, reasons):
        amount_raw = (amount_raw or '').strip()
        reason = (reason or '').strip()
        if not amount_raw and not reason:
            continue
        if not amount_raw:
            raise ValueError('Each deduction needs an amount.')
        try:
            amount = float(amount_raw)
        except ValueError:
            raise ValueError('Amount must be a number.')
        if amount < 0:
            raise ValueError('Amount cannot be negative.')
        if not reason:
            raise ValueError('Each deduction needs its own reason.')
        deduction = PayrollDeduction(driver_id=driver_id, deduction_date=deduction_date,
                                      amount=amount, reason=reason, created_by=current_user.id)
        db.session.add(deduction)
        created.append(deduction)

    if not created:
        raise ValueError('Enter at least one deduction (amount + reason).')

    db.session.flush()
    for deduction in created:
        log_audit('CREATE', 'payroll_deductions', deduction.id,
                  f'Payroll deduction of {deduction.amount} for driver #{driver_id}: {deduction.reason}')
        touch_sync_fields(deduction)
    db.session.commit()
    flash(f'{len(created)} deductions recorded.' if len(created) > 1 else 'Deduction recorded.', 'success')
    return redirect(request.referrer or url_for('report_payroll'))


@app.route('/finance/payroll-deductions/<int:did>/delete', methods=['POST'])
@login_required
@admin_required
def payroll_deduction_delete(did):
    deduction = PayrollDeduction.query.filter_by(id=did).first_or_404()
    log_audit('DELETE', 'payroll_deductions', did, f'Deleted payroll deduction of {deduction.amount}')
    deduction.deleted_at = datetime.now(timezone.utc)
    touch_sync_fields(deduction)
    db.session.commit()
    flash('Deduction deleted.', 'warning')
    return redirect(request.referrer or url_for('report_payroll'))


@app.route('/reports/payroll/pay-sheet.pdf')
@login_required
@permission_required('reports')
def payroll_pay_sheet_pdf():
    """One-page roster of every driver/conductor with their net pay for the
    period and a blank Signature/Date cell each — the physical sheet crew
    sign as they're handed their cash, as opposed to the individual
    Payslip PDFs (see _payslip_elements) which are each person's own
    itemized copy. Net Pay here (not gross Accrued) is what's actually
    being handed over and signed for."""
    df, dt = query_date_range()
    date_from_str, date_to_str = df.strftime('%Y-%m-%d'), dt.strftime('%Y-%m-%d')
    earnings, *_ = compute_payroll_earnings(df, dt)
    crew_rows = _flatten_wages_earnings(earnings)

    if not crew_rows:
        flash(f'No payroll activity for {date_from_str} to {date_to_str}.', 'warning')
        return redirect(url_for('report_payroll', date_from=date_from_str, date_to=date_to_str))

    styles = _pdf_styles()
    elements = [
        Paragraph('Payroll Pay Sheet', styles['Title']),
        Paragraph(f'Period: {date_from_str} to {date_to_str}', styles['Normal']),
        Paragraph(f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")} by {current_user.username}', styles['Normal']),
        Spacer(1, 10),
    ]

    data = [['#', 'Crew Member', 'Role', 'Net Pay (USD)', 'Signature', 'Date']]
    total_net_pay = 0.0
    for i, (name, role, row) in enumerate(crew_rows, start=1):
        data.append([str(i), name, role, f"${row['net_pay']:,.2f}", '', ''])
        total_net_pay += row['net_pay']
    data.append(['', '', 'TOTAL', f"${total_net_pay:,.2f}", '', ''])

    table = _pdf_table(data, bold_last_row=True, col_widths=[24, 140, 60, 70, 140, 70])
    # Bold black grid (over _pdf_table's default light-gray one) so the row/
    # column lines are actually visible on a printed copy, plus extra
    # vertical padding so the blank Signature/Date cells are tall enough to
    # sign in, not just wide.
    table.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 0.75, colors.black),
        ('TOPPADDING', (0, 1), (-1, -1), 12),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 12),
    ]))
    elements.append(table)

    return _pdf_response(f'payroll_pay_sheet_{date_from_str}_to_{date_to_str}.pdf', elements)


@app.route('/reports/payroll/payslip/<int:driver_id>.pdf')
@login_required
@permission_required('reports')
def driver_payslip_pdf(driver_id):
    """A single crew member's printable payslip for the selected period —
    gross commission, itemized deductions with their reasons, net pay, and
    what's already been paid vs. still owed. Reuses compute_payroll_earnings
    so the figures can never drift from what the Payroll report itself
    shows for this driver."""
    driver = Driver.query.filter_by(id=driver_id).first_or_404()
    df, dt = query_date_range()
    date_from_str, date_to_str = df.strftime('%Y-%m-%d'), dt.strftime('%Y-%m-%d')

    earnings, *_ = compute_payroll_earnings(df, dt)
    all_rows = []
    for row in earnings:
        all_rows.append(row)
        all_rows.extend(row['conductors'])
    e = next((r for r in all_rows if r.get('driver') and r['driver'].id == driver_id), None)
    if e is None:
        flash(f'{driver.name} has no payroll activity for {date_from_str} to {date_to_str}.', 'warning')
        return redirect(url_for('report_payroll', date_from=date_from_str, date_to=date_to_str))

    deduction_rows, payment_rows = _crew_deduction_payment_rows(driver_id, df, dt)

    return _payslip_pdf_response(driver.name, driver.role.title(), e, deduction_rows, payment_rows,
                                  date_from_str, date_to_str)


@app.route('/reports/payroll/payslip/<int:driver_id>/conductor.pdf')
@login_required
@permission_required('reports')
def conductor_payslip_pdf(driver_id):
    """Payslip for a driver's placeholder conductor row — printed when no
    named conductor is on file for them (see compute_payroll_earnings'
    is_placeholder rows), so there's still something to hand the person who
    actually rode that day. Printed under a generic "<driver's first name>
    Conductor" label since there's no real Driver record to name it after —
    and for the same reason, no deductions/payments can exist against them
    yet (both need a real driver_id), so those sections always print empty."""
    driver = Driver.query.filter_by(id=driver_id).first_or_404()
    df, dt = query_date_range()
    date_from_str, date_to_str = df.strftime('%Y-%m-%d'), dt.strftime('%Y-%m-%d')

    earnings, *_ = compute_payroll_earnings(df, dt)
    row = next((r for r in earnings if r['driver'] and r['driver'].id == driver_id), None)
    ce = next((c for c in row['conductors'] if c.get('is_placeholder')), None) if row else None
    if ce is None:
        flash(f'No placeholder conductor slip to print for {driver.name} in {date_from_str} to {date_to_str}.', 'warning')
        return redirect(url_for('report_payroll', date_from=date_from_str, date_to=date_to_str))

    first_name = driver.name.split()[0] if driver.name else 'Conductor'
    display_name = f'{first_name} Conductor'
    return _payslip_pdf_response(display_name, 'Conductor', ce, [], [], date_from_str, date_to_str)


def _crew_deduction_payment_rows(driver_id, df, dt):
    deduction_rows = PayrollDeduction.query.filter(
        PayrollDeduction.driver_id == driver_id,
        PayrollDeduction.deduction_date.between(df, dt)).order_by(PayrollDeduction.deduction_date).all()
    payment_rows = CommissionPayment.query.filter(
        CommissionPayment.driver_id == driver_id,
        CommissionPayment.payment_date.between(df, dt)).order_by(CommissionPayment.payment_date).all()
    return deduction_rows, payment_rows


@app.route('/reports/payroll/payslips/print-all.pdf')
@login_required
@permission_required('reports')
def payroll_payslips_print_all():
    """Every driver and conductor's payslip for the selected period, printed
    as one combined PDF — same per-crew-member layout as the individual
    Payslip downloads (including the signature line), one per page, so the
    whole batch can be run off and handed out in one go instead of printing
    each crew member's slip separately."""
    df, dt = query_date_range()
    date_from_str, date_to_str = df.strftime('%Y-%m-%d'), dt.strftime('%Y-%m-%d')
    earnings, *_ = compute_payroll_earnings(df, dt)

    if not earnings:
        flash(f'No payroll activity for {date_from_str} to {date_to_str}.', 'warning')
        return redirect(url_for('report_payroll', date_from=date_from_str, date_to=date_to_str))

    elements = []
    for e in earnings:
        deduction_rows, payment_rows = _crew_deduction_payment_rows(e['driver'].id, df, dt)
        elements += _payslip_elements(e['driver'].name, e['driver'].role.title(), e,
                                       deduction_rows, payment_rows, date_from_str, date_to_str)
        elements.append(PageBreak())

        for ce in e['conductors']:
            if ce.get('is_placeholder'):
                first_name = e['driver'].name.split()[0] if e['driver'].name else 'Conductor'
                elements += _payslip_elements(f'{first_name} Conductor', 'Conductor', ce, [], [],
                                               date_from_str, date_to_str)
            else:
                c_deduction_rows, c_payment_rows = _crew_deduction_payment_rows(ce['driver'].id, df, dt)
                elements += _payslip_elements(ce['driver'].name, 'Conductor', ce,
                                               c_deduction_rows, c_payment_rows, date_from_str, date_to_str)
            elements.append(PageBreak())

    if elements and isinstance(elements[-1], PageBreak):
        elements.pop()

    return _pdf_response(f'payslips_all_{date_from_str}_to_{date_to_str}.pdf', elements, margins=18 * mm)


def _payslip_pdf_response(display_name, role_label, e, deduction_rows, payment_rows, date_from_str, date_to_str):
    """Single-payslip download — driver_payslip_pdf/conductor_payslip_pdf's
    response wrapper around the shared element builder below."""
    elements = _payslip_elements(display_name, role_label, e, deduction_rows, payment_rows,
                                  date_from_str, date_to_str)
    safe_name = display_name.replace(' ', '_')
    return _pdf_response(f'payslip_{safe_name}_{date_from_str}_to_{date_to_str}.pdf', elements, margins=18 * mm)


def _payslip_elements(display_name, role_label, e, deduction_rows, payment_rows, date_from_str, date_to_str):
    """One crew member's payslip as a list of flowables — shared by the
    single-payslip download (_payslip_pdf_response) and the print-all-at-once
    combined PDF (payroll_payslips_print_all), which concatenates one of
    these per crew member with a PageBreak in between. Ends with a
    signature line so the printed copy can be signed by the driver/
    conductor and kept on file as proof of receipt."""
    styles = _pdf_styles()
    elements = [
        Paragraph('Payslip', styles['Title']),
        Spacer(1, 4),
        Paragraph(f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")} by {current_user.username}', styles['Normal']),
        Spacer(1, 10),
    ]

    elements.append(_pdf_statement_table([
        ['Crew Member', display_name],
        ['Role', role_label],
        ['Pay Period', f'{date_from_str} to {date_to_str}'],
        ['Days Worked', str(e['days_worked'])],
    ]))
    elements.append(Spacer(1, 14))

    elements.append(Paragraph('Earnings', styles['Heading3']))
    elements.append(_pdf_statement_table([
        ['Revenue Generated', f"${e['total_revenue']:,.2f}"],
        ['Garnish (netted off before commission)', f"${e['garnish']:,.2f}"],
        ['Commission Rate', f"{e['rate_pct']:.1f}%"],
        ['Gross Commission Accrued', f"${e['commission']:,.2f}"],
    ], bold_indices=(3,)))
    elements.append(Spacer(1, 14))

    elements.append(Paragraph('Deductions', styles['Heading3']))
    if deduction_rows:
        ded_data = [['Date', 'Reason', 'Amount']]
        for dd in deduction_rows:
            ded_data.append([dd.deduction_date.strftime('%Y-%m-%d'), dd.reason, f"${dd.amount:,.2f}"])
        ded_data.append(['', 'TOTAL DEDUCTIONS', f"${e['deductions']:,.2f}"])
        elements.append(_pdf_table(ded_data, col_widths=[75, 305, 90]))
    else:
        elements.append(Paragraph('No deductions this period.', styles['Normal']))
    elements.append(Spacer(1, 14))

    elements.append(Paragraph('Summary', styles['Heading3']))
    elements.append(_pdf_statement_table([
        ['Gross Commission Accrued', f"${e['commission']:,.2f}"],
        ['Total Deductions', f"${e['deductions']:,.2f}"],
        ['Net Pay', f"${e['net_pay']:,.2f}"],
        ['Already Paid This Period', f"${e['paid']:,.2f}"],
        ['Balance Due', f"${e['outstanding']:,.2f}"],
    ], bold_indices=(2, 4)))

    if payment_rows:
        elements.append(Spacer(1, 14))
        elements.append(Paragraph('Payments Made This Period', styles['Heading3']))
        pay_data = [['Date', 'Amount', 'Method', 'Notes']]
        for p in payment_rows:
            pay_data.append([p.payment_date.strftime('%Y-%m-%d'), f"${p.amount:,.2f}", p.method or '—', p.notes or '—'])
        elements.append(_pdf_table(pay_data, bold_last_row=False, col_widths=[75, 75, 90, 230]))

    # Signature line — this printed copy is signed by the driver/conductor
    # acknowledging receipt of the net pay above, then kept on file.
    elements.append(Spacer(1, 30))
    elements.append(Paragraph(
        'I acknowledge receipt of the net pay shown above for this period.', styles['Normal']))
    elements.append(Spacer(1, 26))
    sig_table = Table([['', '', ''], [f'{role_label} Signature', '', 'Date']], colWidths=[260, 40, 140])
    sig_table.setStyle(TableStyle([
        ('LINEABOVE', (0, 0), (0, 0), 0.75, colors.black),
        ('LINEABOVE', (2, 0), (2, 0), 0.75, colors.black),
        ('TOPPADDING', (0, 1), (-1, 1), 3),
        ('FONTSIZE', (0, 1), (-1, 1), 8.5),
        ('TEXTCOLOR', (0, 1), (-1, 1), colors.HexColor('#64748b')),
    ]))
    elements.append(sig_table)

    return elements


def compute_cash_flow(df, dt):
    """Cash Flow statement for [df, dt] — shared by the report page and
    the Full Report Pack PDF."""
    def in_range(col):
        return col.between(df, dt)

    operating_in = db.session.query(func.sum(DailyLog.gross_revenue)).filter(
        in_range(DailyLog.log_date)).scalar() or 0
    receivables_in = db.session.query(func.sum(Receivable.amount)).filter(
        Receivable.status == 'collected', in_range(Receivable.collected_date)).scalar() or 0

    maint_out = db.session.query(func.sum(MaintenanceLog.total_cost)).filter(
        in_range(MaintenanceLog.log_date)).scalar() or 0
    expenses_out = db.session.query(func.sum(Expense.amount)).filter(
        in_range(Expense.expense_date)).scalar() or 0
    commission_out = db.session.query(func.sum(CommissionPayment.amount)).filter(
        in_range(CommissionPayment.payment_date)).scalar() or 0
    payables_out = db.session.query(func.sum(Payable.amount)).filter(
        Payable.status == 'paid', in_range(Payable.paid_date)).scalar() or 0

    net_operating = operating_in + receivables_in - maint_out - expenses_out - commission_out - payables_out

    vehicles_bought = [v for v in Vehicle.query.all() if df <= v.created_at.date() <= dt]
    investing_out = sum(v.acquisition_cost for v in vehicles_bought)
    net_investing = -investing_out

    loan_proceeds_in = db.session.query(func.sum(Loan.principal)).filter(
        in_range(Loan.start_date)).scalar() or 0
    loan_repay_out = db.session.query(func.sum(LoanPayment.amount)).filter(
        in_range(LoanPayment.payment_date)).scalar() or 0
    capital_in = db.session.query(func.sum(CapitalContribution.amount)).filter(
        in_range(CapitalContribution.contribution_date)).scalar() or 0
    drawings_out = db.session.query(func.sum(OwnerDrawing.amount)).filter(
        in_range(OwnerDrawing.drawing_date)).scalar() or 0
    net_financing = loan_proceeds_in - loan_repay_out + capital_in - drawings_out

    net_change = net_operating + net_investing + net_financing

    opening_cash = compute_financial_position(df - timedelta(days=1))['cash_and_equivalents']
    closing_cash = compute_financial_position(dt)['cash_and_equivalents']

    return dict(
        operating_in=operating_in, receivables_in=receivables_in,
        maint_out=maint_out, expenses_out=expenses_out,
        commission_out=commission_out, payables_out=payables_out, net_operating=net_operating,
        investing_out=investing_out, net_investing=net_investing, vehicles_bought=vehicles_bought,
        loan_proceeds_in=loan_proceeds_in, loan_repay_out=loan_repay_out,
        capital_in=capital_in, drawings_out=drawings_out, net_financing=net_financing,
        net_change=net_change, opening_cash=opening_cash, closing_cash=closing_cash,
    )


@app.route('/reports/cash-flow')
@login_required
@permission_required('reports')
def report_cash_flow():
    df, dt = query_date_range()
    date_from_str, date_to_str = df.strftime('%Y-%m-%d'), dt.strftime('%Y-%m-%d')
    cf = compute_cash_flow(df, dt)
    return render_template('reports/cash_flow.html',
        date_from=date_from_str, date_to=date_to_str, **cf)


@app.route('/reports/cash-flow/export')
@login_required
@permission_required('reports')
def report_cash_flow_export():
    df, dt = query_date_range()
    if request.args.get('format') == 'pdf':
        return _pdf_response(f'cash_flow_{df}_to_{dt}.pdf', _cash_flow_pdf(df, dt))
    cf = compute_cash_flow(df, dt)
    rows = [
        ['Operating — Fare/Trip Revenue', f"{cf['operating_in']:.2f}"],
        ['Operating — Receivables Collected', f"{cf['receivables_in']:.2f}"],
        ['Operating — Maintenance', f"-{cf['maint_out']:.2f}"],
        ['Operating — Expenses', f"-{cf['expenses_out']:.2f}"],
        ['Operating — Commission Paid', f"-{cf['commission_out']:.2f}"],
        ['Operating — Payables Paid', f"-{cf['payables_out']:.2f}"],
        ['Net Cash from Operating Activities', f"{cf['net_operating']:.2f}"],
        ['Investing — Vehicles Acquired', f"-{cf['investing_out']:.2f}"],
        ['Net Cash from Investing Activities', f"{cf['net_investing']:.2f}"],
        ['Financing — Loan Proceeds', f"{cf['loan_proceeds_in']:.2f}"],
        ['Financing — Loan Repayments', f"-{cf['loan_repay_out']:.2f}"],
        ['Financing — Capital Contributed', f"{cf['capital_in']:.2f}"],
        ['Financing — Owner Drawings', f"-{cf['drawings_out']:.2f}"],
        ['Net Cash from Financing Activities', f"{cf['net_financing']:.2f}"],
        ['Net Change in Cash', f"{cf['net_change']:.2f}"],
        ['Opening Cash', f"{cf['opening_cash']:.2f}"],
        ['Closing Cash', f"{cf['closing_cash']:.2f}"],
    ]
    return csv_export_response(f'cash_flow_{df}_to_{dt}.csv', ['Line Item', 'Amount'], rows)


def _resolve_budget_month(month_str):
    today = date.today()
    month_str = month_str or today.strftime('%Y-%m')
    try:
        month_start = datetime.strptime(month_str, '%Y-%m').date().replace(day=1)
    except ValueError:
        flash(f'"{month_str}" is not a valid month — showing {today.strftime("%Y-%m")} instead.', 'warning')
        month_start = today.replace(day=1)
        month_str = month_start.strftime('%Y-%m')
    return month_start, month_str


def _compute_budget_rows(month_start):
    """Budget vs. Actual for one calendar month — shared by the report page
    and its CSV export. Expense category labels are prefixed to avoid
    colliding with the fixed Revenue/Maintenance keys below — an admin
    could otherwise name a heading "Maintenance" (as in the worked example)
    and silently shadow the MaintenanceLog-derived figure."""
    month_end = (month_start.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)
    actuals = {
        'Revenue': db.session.query(func.sum(DailyLog.gross_revenue)).filter(
            DailyLog.log_date.between(month_start, month_end)).scalar() or 0,
        'Maintenance': db.session.query(func.sum(MaintenanceLog.total_cost)).filter(
            MaintenanceLog.log_date.between(month_start, month_end)).scalar() or 0,
    }
    for cat in ExpenseCategory.query.all():
        actuals[f'Expense: {cat.display_name}'] = db.session.query(func.sum(Expense.amount)).filter(
            Expense.category_id == cat.id,
            Expense.expense_date.between(month_start, month_end)).scalar() or 0

    budgets = {b.category: b.amount for b in Budget.query.filter_by(month=month_start).all()}
    all_categories = sorted(set(list(actuals.keys()) + list(budgets.keys())))
    rows = []
    for cat in all_categories:
        budget_amt = budgets.get(cat, 0)
        actual_amt = actuals.get(cat, 0)
        rows.append({'category': cat, 'budget': budget_amt, 'actual': actual_amt,
                     'variance': actual_amt - budget_amt})
    return rows


@app.route('/reports/budget')
@login_required
@permission_required('reports')
def report_budget():
    month_start, month_str = _resolve_budget_month(request.args.get('month', ''))
    rows = _compute_budget_rows(month_start)
    categories_available = ['Revenue', 'Maintenance'] + \
        [f'Expense: {c.display_name}' for c in ExpenseCategory.query.all()]
    return render_template('reports/budget.html', rows=rows, month=month_str,
        month_label=month_start.strftime('%B %Y'), categories=categories_available)


@app.route('/reports/budget/export')
@login_required
@permission_required('reports')
def report_budget_export():
    month_start, month_str = _resolve_budget_month(request.args.get('month', ''))
    rows = _compute_budget_rows(month_start)
    out_rows = [[r['category'], f"{r['budget']:.2f}", f"{r['actual']:.2f}", f"{r['variance']:.2f}"] for r in rows]
    header = ['Category', 'Budget', 'Actual', 'Variance']
    if request.args.get('format') == 'pdf':
        return _table_pdf_response(f'budget_vs_actual_{month_str}.pdf', 'Budget vs Actual',
            f'Month: {month_str}', header, out_rows)
    return csv_export_response(f'budget_vs_actual_{month_str}.csv', header, out_rows)


@app.route('/reports/budget/set', methods=['POST'])
@login_required
@permission_required('finance')
@handle_form_errors
def budget_set():
    month_str = request.form.get('month', '')
    month_start = datetime.strptime(month_str, '%Y-%m').date().replace(day=1)
    category = request.form['category'].strip()
    amount = form_float(request.form, 'amount', min_value=0)

    existing = Budget.query.filter_by(category=category, month=month_start).first()
    if existing:
        existing.amount = amount
        touch_sync_fields(existing)
    else:
        b = Budget(category=category, month=month_start, amount=amount, created_by=current_user.id)
        db.session.add(b)
        db.session.flush()
        touch_sync_fields(b)
    log_audit('UPDATE', 'budgets', None, f'Set budget for {category} in {month_str}: {amount}')
    db.session.commit()
    flash(f'Budget for {category} set.', 'success')
    return redirect(url_for('report_budget', month=month_str))


def _compute_fuel_efficiency_rows(df, dt):
    rows = []
    for v in Vehicle.query.order_by(Vehicle.registration).all():
        # Chronological order, not odometer order: fill-ups must pair with
        # whichever one actually came right before them in time. Sorting by
        # odometer instead used to mean one mis-keyed reading (a misread
        # gauge, a transposed digit — routine in a real fleet) would pair
        # with whatever OTHER reading happened to be numerically nearby,
        # not the true previous fill-up, producing nonsense segments (seen
        # on real data: one vehicle came out at 109 L/100km, another at
        # 7,400 L/100km — both physically impossible).
        logs = FuelLog.query.filter(
            FuelLog.vehicle_id == v.id, FuelLog.log_date.between(df, dt),
            FuelLog.odometer.isnot(None)).order_by(FuelLog.log_date, FuelLog.id).all()
        if len(logs) < 2:
            continue
        segments = []
        for prev, curr in zip(logs, logs[1:]):
            distance = curr.odometer - prev.odometer
            if distance <= 0:
                continue
            # Skip fill-ups with no liters recorded (e.g. a diesel entry logged
            # only as a cash amount, or a mileage-only reading) — we genuinely
            # don't know how much fuel covered this distance, so it can't be
            # counted rather than being shown as an implausible 0 L/100km.
            if not curr.liters:
                continue
            # Fuel burned to cover this segment is the fill-up that ends it.
            l_per_100km = (curr.liters / distance) * 100
            segments.append({'from_date': prev.log_date, 'to_date': curr.log_date,
                             'distance': distance, 'liters': curr.liters,
                             'l_per_100km': l_per_100km})
        if not segments:
            continue
        total_distance = sum(s['distance'] for s in segments)
        total_liters = sum(s['liters'] for s in segments)
        # Aggregate consumption over the whole period (distance-weighted), which
        # is more accurate than averaging each segment's ratio equally.
        overall_l_per_100km = (total_liters / total_distance) * 100 if total_distance else 0
        km_per_liter = (total_distance / total_liters) if total_liters else 0
        rows.append({'vehicle': v, 'segments': segments,
                     'avg_l_per_100km': overall_l_per_100km,
                     'total_distance': total_distance, 'total_liters': total_liters,
                     'km_per_liter': km_per_liter})
    return rows


@app.route('/reports/fuel-efficiency')
@login_required
@permission_required('reports')
def report_fuel_efficiency():
    df, dt = query_date_range()
    date_from_str, date_to_str = df.strftime('%Y-%m-%d'), dt.strftime('%Y-%m-%d')
    rows = _compute_fuel_efficiency_rows(df, dt)

    # Fleet-wide figures aggregated across all measured distance/fuel.
    fleet_distance = sum(r['total_distance'] for r in rows)
    fleet_liters = sum(r['total_liters'] for r in rows)
    fleet_avg = (fleet_liters / fleet_distance) * 100 if fleet_distance else 0
    return render_template('reports/fuel_efficiency.html', rows=rows, fleet_avg=fleet_avg,
        fleet_distance=fleet_distance, fleet_liters=fleet_liters,
        date_from=date_from_str, date_to=date_to_str)


@app.route('/reports/fuel-efficiency/export')
@login_required
@permission_required('reports')
def report_fuel_efficiency_export():
    df, dt = query_date_range()
    rows = _compute_fuel_efficiency_rows(df, dt)
    out_rows = [[r['vehicle'].registration, f"{r['total_distance']:.1f}", f"{r['total_liters']:.2f}",
                 f"{r['avg_l_per_100km']:.2f}", f"{r['km_per_liter']:.2f}"] for r in rows]
    header = ['Vehicle', 'Total Distance (km)', 'Total Liters', 'Avg L/100km', 'Km/Liter']
    if request.args.get('format') == 'pdf':
        return _table_pdf_response(f'fuel_efficiency_{df}_to_{dt}.pdf', 'Fuel Efficiency',
            f'Period: {df} to {dt}', header, out_rows)
    return csv_export_response(f'fuel_efficiency_{df}_to_{dt}.csv', header, out_rows)


@app.route('/reports/distance-travelled')
@login_required
@permission_required('reports')
def report_distance_travelled():
    df, dt = query_date_range()
    date_from_str, date_to_str = df.strftime('%Y-%m-%d'), dt.strftime('%Y-%m-%d')

    rows = []
    for v in Vehicle.query.order_by(Vehicle.registration).all():
        # Odometer reading logged for any date in the requested range
        odometer = db.session.query(func.max(FuelLog.odometer)).filter(
            FuelLog.vehicle_id == v.id, FuelLog.log_date.between(df, dt),
            FuelLog.odometer.isnot(None)).scalar()

        distance = prev_odometer = prev_date = None
        if odometer is not None:
            prev = FuelLog.query.filter(
                FuelLog.vehicle_id == v.id, FuelLog.log_date < df,
                FuelLog.odometer.isnot(None)).order_by(FuelLog.log_date.desc()).first()
            if prev:
                prev_odometer, prev_date = prev.odometer, prev.log_date
                distance = odometer - prev_odometer

        rows.append({'vehicle': v, 'odometer': odometer,
                     'prev_odometer': prev_odometer, 'prev_date': prev_date,
                     'distance': distance})

    fleet_distance = sum(r['distance'] for r in rows if r['distance'] is not None)
    vehicles_reporting = sum(1 for r in rows if r['distance'] is not None)
    # Fleet-wide revenue for the requested range
    fleet_revenue = db.session.query(func.sum(DailyLog.gross_revenue)).filter(
        DailyLog.log_date.between(df, dt)).scalar() or 0.0
    return render_template('reports/distance_travelled.html', rows=rows,
        date_from=date_from_str, date_to=date_to_str, fleet_distance=fleet_distance,
        vehicles_reporting=vehicles_reporting, fleet_size=len(rows), fleet_revenue=fleet_revenue)


@app.route('/reports/route-profitability')
@login_required
@permission_required('reports')
def report_route_profitability():
    df, dt = query_date_range()
    date_from_str, date_to_str = df.strftime('%Y-%m-%d'), dt.strftime('%Y-%m-%d')
    rows, total_revenue, total_costs = _compute_route_profitability(df, dt)
    return render_template('reports/route_profitability.html', rows=rows,
        total_revenue=total_revenue, total_costs=total_costs,
        date_from=date_from_str, date_to=date_to_str)


@app.route('/reports/route-profitability/export')
@login_required
@permission_required('reports')
def report_route_profitability_export():
    df, dt = query_date_range()
    rows, _, _ = _compute_route_profitability(df, dt)
    out_rows = [[r['route'].name if r['route'] else 'Unrouted', f"{r['revenue']:.2f}", r['trips'],
                 r['log_days'], f"{r['allocated_cost']:.2f}", f"{r['net_profit']:.2f}"] for r in rows]
    header = ['Route', 'Revenue', 'Trips', 'Log Days', 'Allocated Cost', 'Net Profit']
    if request.args.get('format') == 'pdf':
        return _table_pdf_response(f'route_profitability_{df}_to_{dt}.pdf', 'Route Profitability',
            f'Period: {df} to {dt}', header, out_rows)
    return csv_export_response(f'route_profitability_{df}_to_{dt}.csv', header, out_rows)


def _compute_route_profitability(df, dt):
    total_revenue = db.session.query(func.sum(DailyLog.gross_revenue)).filter(
        DailyLog.log_date.between(df, dt)).scalar() or 0
    total_maintenance = db.session.query(func.sum(MaintenanceLog.total_cost)).filter(
        MaintenanceLog.log_date.between(df, dt)).scalar() or 0
    total_costs = total_maintenance

    route_data = db.session.query(
        Route.id, Route.name, Route.start_point, Route.end_point,
        func.sum(DailyLog.gross_revenue).label('revenue'),
        func.sum(DailyLog.trips_completed).label('trips'),
        func.count(DailyLog.id).label('log_days'),
    ).join(DailyLog, Route.id == DailyLog.route_id).filter(
        DailyLog.log_date.between(df, dt)
    ).group_by(Route.id).all()

    rows = []
    for r in route_data:
        revenue = r.revenue or 0
        allocated_cost = (revenue / total_revenue * total_costs) if total_revenue else 0
        rows.append({
            'route': r, 'revenue': revenue, 'trips': r.trips or 0, 'log_days': r.log_days,
            'allocated_cost': allocated_cost, 'net_profit': revenue - allocated_cost,
        })

    # Entries with no route (e.g. from the Vehicle Ledger, which doesn't ask
    # for one) would otherwise be silently dropped by the INNER JOIN above,
    # leaving the per-route rows short of the fleet-wide total shown.
    unrouted = db.session.query(
        func.sum(DailyLog.gross_revenue).label('revenue'),
        func.sum(DailyLog.trips_completed).label('trips'),
        func.count(DailyLog.id).label('log_days'),
    ).filter(DailyLog.route_id.is_(None), DailyLog.log_date.between(df, dt)).first()
    if unrouted and unrouted.log_days:
        revenue = unrouted.revenue or 0
        allocated_cost = (revenue / total_revenue * total_costs) if total_revenue else 0
        rows.append({
            'route': None, 'revenue': revenue, 'trips': unrouted.trips or 0, 'log_days': unrouted.log_days,
            'allocated_cost': allocated_cost, 'net_profit': revenue - allocated_cost,
        })

    rows.sort(key=lambda x: x['net_profit'], reverse=True)
    return rows, total_revenue, total_costs


def _compute_vehicle_efficiency_rows(df, dt):
    """Per-vehicle profitability: revenue and tracked costs (maintenance,
    other expenses) set against fuel efficiency and distance covered.
    Fuel isn't counted as a cost here — like on the Income Statement, it's
    paid by the driver out of the daily cash collected rather than being a
    company expense; km/L and L/100km still show fuel efficiency itself.
    Every active vehicle gets a row, even an all-zero one — this reads as a
    fleet roster (which vehicles logged nothing this period is itself
    useful to see), not just a list of whichever vehicles happened to have
    activity."""
    efficiency_by_vehicle = {r['vehicle'].id: r for r in _compute_fuel_efficiency_rows(df, dt)}

    revenue_by_vehicle = dict(
        db.session.query(DailyLog.vehicle_id, func.sum(DailyLog.gross_revenue))
        .filter(DailyLog.log_date.between(df, dt)).group_by(DailyLog.vehicle_id).all())
    maintenance_by_vehicle = dict(
        db.session.query(MaintenanceLog.vehicle_id, func.sum(MaintenanceLog.total_cost))
        .filter(MaintenanceLog.log_date.between(df, dt)).group_by(MaintenanceLog.vehicle_id).all())
    expenses_by_vehicle = dict(
        db.session.query(Expense.vehicle_id, func.sum(Expense.amount))
        .filter(Expense.expense_date.between(df, dt), Expense.vehicle_id.isnot(None))
        .group_by(Expense.vehicle_id).all())

    rows = []
    for v in Vehicle.query.filter_by(status='active').order_by(Vehicle.registration).all():
        revenue = revenue_by_vehicle.get(v.id) or 0.0
        maintenance_cost = maintenance_by_vehicle.get(v.id) or 0.0
        other_expenses = expenses_by_vehicle.get(v.id) or 0.0
        total_cost = maintenance_cost + other_expenses
        net_profit = revenue - total_cost

        eff = efficiency_by_vehicle.get(v.id)
        distance = eff['total_distance'] if eff else 0
        liters = eff['total_liters'] if eff else 0

        rows.append({
            'vehicle': v, 'revenue': revenue,
            'maintenance_cost': maintenance_cost, 'other_expenses': other_expenses,
            'total_cost': total_cost, 'net_profit': net_profit,
            'distance': distance, 'liters': liters,
            'km_per_liter': eff['km_per_liter'] if eff else 0,
            'l_per_100km': eff['avg_l_per_100km'] if eff else 0,
            'cost_per_km': (total_cost / distance) if distance else None,
            'revenue_per_km': (revenue / distance) if distance else None,
            'profit_margin': (net_profit / revenue * 100) if revenue else None,
        })
    rows.sort(key=lambda r: r['net_profit'], reverse=True)
    return rows


@app.route('/reports/vehicle-efficiency')
@login_required
@permission_required('reports')
def report_vehicle_efficiency():
    df, dt = query_date_range()
    date_from_str, date_to_str = df.strftime('%Y-%m-%d'), dt.strftime('%Y-%m-%d')
    rows = _compute_vehicle_efficiency_rows(df, dt)
    fleet_revenue = sum(r['revenue'] for r in rows)
    fleet_cost = sum(r['total_cost'] for r in rows)
    fleet_distance = sum(r['distance'] for r in rows)
    return render_template('reports/vehicle_efficiency.html', rows=rows,
        fleet_revenue=fleet_revenue, fleet_cost=fleet_cost,
        fleet_profit=fleet_revenue - fleet_cost, fleet_distance=fleet_distance,
        date_from=date_from_str, date_to=date_to_str)


@app.route('/reports/vehicle-efficiency/export')
@login_required
@permission_required('reports')
def report_vehicle_efficiency_export():
    df, dt = query_date_range()
    rows = _compute_vehicle_efficiency_rows(df, dt)
    out_rows = [[
        r['vehicle'].registration, f"{r['revenue']:.2f}",
        f"{r['maintenance_cost']:.2f}", f"{r['other_expenses']:.2f}", f"{r['total_cost']:.2f}",
        f"{r['net_profit']:.2f}", f"{r['distance']:.1f}", f"{r['liters']:.2f}",
        f"{r['km_per_liter']:.2f}",
        f"{r['cost_per_km']:.2f}" if r['cost_per_km'] is not None else '',
        f"{r['revenue_per_km']:.2f}" if r['revenue_per_km'] is not None else '',
    ] for r in rows]
    header = ['Vehicle', 'Revenue', 'Maintenance Cost', 'Other Expenses', 'Total Cost',
              'Net Profit', 'Distance (km)', 'Liters', 'Km/Liter', 'Cost/km', 'Revenue/km']
    if request.args.get('format') == 'pdf':
        return _table_pdf_response(f'vehicle_efficiency_{df}_to_{dt}.pdf', 'Vehicle Efficiency & Profitability',
            f'Period: {df} to {dt}', header, out_rows)
    return csv_export_response(f'vehicle_efficiency_{df}_to_{dt}.csv', header, out_rows)


@app.route('/reports/financial-position')
@login_required
@permission_required('reports')
def report_financial_position():
    as_of = query_single_date('as_of')
    fp = compute_financial_position(as_of)
    return render_template('reports/financial_position.html',
        as_of_str=as_of.strftime('%Y-%m-%d'), **fp)


# ─────────────────────────────────────────────────────────────
# CSV Exports
# ─────────────────────────────────────────────────────────────
@app.route('/reports/export/daily-logs')
@login_required
@permission_required_any('daily_logs', 'crew_portal')
def export_daily_logs():
    df = request.args.get('date_from', '')
    dt = request.args.get('date_to', '')
    q = DailyLog.query.order_by(DailyLog.log_date.desc())
    try:
        if df:
            q = q.filter(DailyLog.log_date >= parse_date(df))
        if dt:
            q = q.filter(DailyLog.log_date <= parse_date(dt))
    except ValueError as e:
        flash(str(e), 'danger')
        return redirect(url_for('driver_ledger'))

    header = ['Date', 'Vehicle', 'Driver', 'Conductor', 'Route',
              'Trips', 'Gross Revenue (USD)', 'Garnish', 'Reason for Shortfall',
              'Entered By', 'Notes']
    rows = [[log.log_date, log.vehicle.registration, log.driver.name if log.driver else '',
             log.conductor.name if log.conductor else '',
             log.route.name if log.route else '', log.trips_completed,
             f'{log.gross_revenue:.2f}',
             f'{log.garnish:.2f}' if log.garnish else '',
             log.reason_for_shortfall or '',
             log.creator.username if log.creator else '',
             log.notes or ''] for log in q.all()]

    subtitle = f'{df or "earliest"} to {dt or "latest"}'
    if request.args.get('format') == 'pdf':
        return _table_pdf_response(f'daily_transactions_{date.today()}.pdf',
            'Daily Transactions', subtitle, header, rows)
    return csv_export_response(f'daily_transactions_{date.today()}.csv', header, rows)


@app.route('/reports/export/income')
@login_required
@permission_required('reports')
def export_income():
    vehicle_id = request.args.get('vehicle_id', '')
    df, dt = query_date_range()
    df_str, dt_str = df.strftime('%Y-%m-%d'), dt.strftime('%Y-%m-%d')

    daily_q = DailyLog.query.filter(DailyLog.log_date.between(df, dt))
    fuel_q = FuelLog.query.filter(FuelLog.log_date.between(df, dt))
    maint_q = MaintenanceLog.query.filter(MaintenanceLog.log_date.between(df, dt))
    exp_q = Expense.query.filter(Expense.expense_date.between(df, dt))
    spares_q = StoreSale.query.filter(StoreSale.sale_date.between(df, dt))

    vehicle_label = 'Consolidated (fleet-wide)'
    if vehicle_id:
        v = Vehicle.query.filter_by(id=vehicle_id).first()
        vehicle_label = f'{v.registration} — {v.make} {v.model}' if v else f'Vehicle #{vehicle_id}'
        daily_q = daily_q.filter(DailyLog.vehicle_id == vehicle_id)
        fuel_q = fuel_q.filter(FuelLog.vehicle_id == vehicle_id)
        maint_q = maint_q.filter(MaintenanceLog.vehicle_id == vehicle_id)
        exp_q = exp_q.filter(Expense.vehicle_id == vehicle_id)
        spares_q = spares_q.filter(StoreSale.vehicle_id == vehicle_id)
    else:
        spares_q = spares_q.filter(StoreSale.vehicle_id.isnot(None))

    daily = daily_q.order_by(DailyLog.log_date).all()
    fuel = fuel_q.order_by(FuelLog.log_date).all()
    maintenance = maint_q.order_by(MaintenanceLog.log_date).all()
    expenses = exp_q.order_by(Expense.expense_date).all()
    spares = spares_q.order_by(StoreSale.sale_date).all()

    total_rev = sum(l.gross_revenue for l in daily)
    total_fuel_liters = sum(f.liters for f in fuel)
    total_maint = sum(m.total_cost for m in maintenance)
    total_exp = sum(e.amount for e in expenses)
    total_spares = sum(s.total_amount for s in spares)

    if request.args.get('format') == 'pdf':
        styles = _pdf_styles()
        flowables = []

        def section(title, headers, data_rows, total_row):
            flowables.append(Paragraph(title, styles['Heading3']))
            flowables.append(Spacer(1, 4))
            flowables.append(_pdf_table([headers] + data_rows + [total_row]))
            flowables.append(Spacer(1, 14))

        section('Revenue', ['Date', 'Vehicle', 'Route', 'Trips', 'Gross Revenue (USD)'],
                [[l.log_date, l.vehicle.registration, l.route.name if l.route else '',
                  str(l.trips_completed), f'{l.gross_revenue:.2f}'] for l in daily],
                ['', '', '', 'TOTAL REVENUE', f'{total_rev:.2f}'])
        section('Fuel Consumption (not a cost — tracked in liters only)', ['Date', 'Vehicle', 'Liters', 'Supplier'],
                [[f.log_date, f.vehicle.registration, f.liters, f.supplier or ''] for f in fuel],
                ['', '', 'TOTAL LITERS', f'{total_fuel_liters:.1f}'])
        section('Maintenance Expenses', ['Date', 'Vehicle', 'Description', 'Parts (USD)', 'Labor (USD)', 'Total (USD)'],
                [[m.log_date, m.vehicle.registration, m.description,
                  f'{m.parts_cost:.2f}', f'{m.labor_cost:.2f}', f'{m.total_cost:.2f}'] for m in maintenance],
                ['', '', '', '', 'TOTAL MAINTENANCE', f'{total_maint:.2f}'])
        section('Other Expenses', ['Date', 'Category', 'Vehicle', 'Description', 'Amount (USD)'],
                [[e.expense_date, e.category.display_name, e.vehicle.registration if e.vehicle else '(general)',
                  e.description or '', f'{e.amount:.2f}'] for e in expenses],
                ['', '', '', 'TOTAL OTHER EXPENSES', f'{total_exp:.2f}'])
        section('Spares Sold to Company Vehicles (booked as a maintenance expense on that vehicle)',
                ['Date', 'Part', 'Vehicle', 'Qty', 'Unit Price (USD)', 'Total (USD)'],
                [[s.sale_date, s.part.name, s.vehicle.registration if s.vehicle else '',
                  s.quantity, f'{s.unit_price:.2f}', f'{s.total_amount:.2f}'] for s in spares],
                ['', '', '', '', 'TOTAL SPARES', f'{total_spares:.2f}'])

        flowables.append(_pdf_statement_table(
            [['NET PROFIT', f'${total_rev - total_maint - total_exp - total_spares:,.2f}']], bold_indices=(0,)))
        elements = _pdf_section('Transport Fleet Income Statement (ZIMRA Compliant)',
                                f'Scope: {vehicle_label} — Period: {df_str} to {dt_str}', flowables)
        scope_suffix = f'_vehicle{vehicle_id}' if vehicle_id else '_consolidated'
        return _pdf_response(f'income{scope_suffix}_{df_str}_to_{dt_str}.pdf', elements, pagesize=landscape(A4))

    out = io.StringIO()
    w = csv.writer(out)
    w.writerow(['TRANSPORT FLEET INCOME STATEMENT (ZIMRA COMPLIANT)'])
    w.writerow([f'Scope: {vehicle_label}'])
    w.writerow([f'Period: {df_str} to {dt_str}'])
    w.writerow([f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")} by {current_user.username}'])
    w.writerow([])

    w.writerow(['REVENUE'])
    w.writerow(['Date', 'Vehicle', 'Route', 'Trips', 'Gross Revenue (USD)'])
    total_rev = 0
    for l in daily:
        w.writerow([l.log_date, l.vehicle.registration, l.route.name if l.route else '',
                    l.trips_completed, f'{l.gross_revenue:.2f}'])
        total_rev += l.gross_revenue
    w.writerow(['', '', '', 'TOTAL REVENUE', f'{total_rev:.2f}'])
    w.writerow([])

    w.writerow(['FUEL CONSUMPTION (not a cost — tracked in liters only)'])
    w.writerow(['Date', 'Vehicle', 'Liters', 'Supplier'])
    total_fuel_liters = 0
    for f in fuel:
        w.writerow([f.log_date, f.vehicle.registration, f.liters, f.supplier or ''])
        total_fuel_liters += f.liters
    w.writerow(['', '', 'TOTAL LITERS', f'{total_fuel_liters:.1f}'])
    w.writerow([])

    w.writerow(['MAINTENANCE EXPENSES'])
    w.writerow(['Date', 'Vehicle', 'Description', 'Parts (USD)', 'Labor (USD)', 'Total (USD)'])
    total_maint = 0
    for m in maintenance:
        w.writerow([m.log_date, m.vehicle.registration, m.description,
                    f'{m.parts_cost:.2f}', f'{m.labor_cost:.2f}', f'{m.total_cost:.2f}'])
        total_maint += m.total_cost
    w.writerow(['', '', '', '', 'TOTAL MAINTENANCE', f'{total_maint:.2f}'])
    w.writerow([])

    w.writerow(['OTHER EXPENSES'])
    w.writerow(['Date', 'Category', 'Vehicle', 'Description', 'Amount (USD)'])
    total_exp = 0
    for e in expenses:
        w.writerow([e.expense_date, e.category.display_name, e.vehicle.registration if e.vehicle else '(general)',
                    e.description or '', f'{e.amount:.2f}'])
        total_exp += e.amount
    w.writerow(['', '', '', 'TOTAL OTHER EXPENSES', f'{total_exp:.2f}'])
    w.writerow([])

    w.writerow(['SPARES SOLD TO COMPANY VEHICLES (booked as a maintenance expense on that vehicle)'])
    w.writerow(['Date', 'Part', 'Vehicle', 'Qty', 'Unit Price (USD)', 'Total (USD)'])
    total_spares = 0
    for s in spares:
        w.writerow([s.sale_date, s.part.name, s.vehicle.registration if s.vehicle else '',
                    s.quantity, f'{s.unit_price:.2f}', f'{s.total_amount:.2f}'])
        total_spares += s.total_amount
    w.writerow(['', '', '', '', 'TOTAL SPARES', f'{total_spares:.2f}'])
    w.writerow([])

    w.writerow(['NET PROFIT', f'{total_rev - total_maint - total_exp - total_spares:.2f}'])

    out.seek(0)
    resp = make_response(out.getvalue())
    resp.headers['Content-Type'] = 'text/csv'
    scope_suffix = f'_vehicle{vehicle_id}' if vehicle_id else '_consolidated'
    resp.headers['Content-Disposition'] = f'attachment; filename=income{scope_suffix}_{df_str}_to_{dt_str}.csv'
    return resp


@app.route('/reports/export/financial-position')
@login_required
@permission_required('reports')
def export_financial_position():
    as_of = query_single_date('as_of')
    fp = compute_financial_position(as_of)

    if request.args.get('format') == 'pdf':
        return _pdf_response(f'financial_position_{as_of}.pdf', _financial_position_pdf(as_of))

    out = io.StringIO()
    w = csv.writer(out)
    w.writerow(['STATEMENT OF FINANCIAL POSITION (SIMPLIFIED — SEE NOTES)'])
    w.writerow([f'As at: {as_of}'])
    w.writerow([f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")} by {current_user.username}'])
    w.writerow([])

    w.writerow(['ASSETS'])
    w.writerow(['Non-Current Assets — Vehicles'])
    w.writerow(['Vehicle', 'Cost (USD)', 'Accumulated Depreciation (USD)', 'Net Book Value (USD)'])
    for row in fp['vehicle_rows']:
        w.writerow([row['vehicle'].registration, f"{row['cost']:.2f}",
                    f"{row['accumulated_depreciation']:.2f}", f"{row['net_book_value']:.2f}"])
    w.writerow(['', '', 'TOTAL VEHICLES (NBV)', f"{fp['total_nbv']:.2f}"])
    w.writerow([])
    w.writerow(['Current Assets'])
    w.writerow(['Cash & Cash Equivalents', f"{fp['cash_and_equivalents']:.2f}"])
    w.writerow(['Receivables Outstanding', f"{fp['receivables_outstanding']:.2f}"])
    w.writerow([])
    w.writerow(['TOTAL ASSETS', f"{fp['total_assets']:.2f}"])
    w.writerow([])

    w.writerow(['LIABILITIES'])
    w.writerow(['Loans Outstanding', f"{fp['loans_outstanding']:.2f}"])
    w.writerow(['Payables Outstanding', f"{fp['payables_outstanding']:.2f}"])
    w.writerow(['Commission Payable (accrued, unpaid)', f"{fp['commission_payable']:.2f}"])
    w.writerow(['TOTAL LIABILITIES', f"{fp['total_liabilities']:.2f}"])
    w.writerow([])

    w.writerow(['EQUITY'])
    w.writerow(["Owner's Capital", f"{fp['owners_capital']:.2f}"])
    w.writerow(['Retained Earnings', f"{fp['retained_earnings']:.2f}"])
    w.writerow(['TOTAL EQUITY', f"{fp['total_equity']:.2f}"])
    w.writerow([])
    w.writerow(['TOTAL LIABILITIES + EQUITY', f"{fp['total_liabilities'] + fp['total_equity']:.2f}"])
    w.writerow([])
    w.writerow(['NOTE: Simplified statement. Vehicle purchases are a pure asset swap, not assumed '
                 'capital-funded — negative cash means a vehicle was bought before a Capital '
                 'Contribution or Loan was recorded to explain the funding. Vehicles are '
                 f"straight-line depreciated over {fp['useful_life']} years from the date each was "
                 'added to the fleet. Commission is accrued on all revenue earned, not just what has '
                 'been paid. Loan repayments are treated as pure principal reduction.'])

    out.seek(0)
    resp = make_response(out.getvalue())
    resp.headers['Content-Type'] = 'text/csv'
    resp.headers['Content-Disposition'] = f'attachment; filename=financial_position_{as_of}.csv'
    return resp


@app.route('/reports/export/distance-travelled')
@login_required
@permission_required('reports')
def export_distance_travelled():
    df, dt = query_date_range()

    # Compute fleet revenue for the range
    fleet_revenue = db.session.query(func.sum(DailyLog.gross_revenue)).filter(
        DailyLog.log_date.between(df, dt)).scalar() or 0.0

    total_distance = 0.0
    rows_out = []
    for v in Vehicle.query.order_by(Vehicle.registration).all():
        odometer = db.session.query(func.max(FuelLog.odometer)).filter(
            FuelLog.vehicle_id == v.id, FuelLog.log_date.between(df, dt),
            FuelLog.odometer.isnot(None)).scalar()
        prev_odometer = prev_date = distance = None
        if odometer is not None:
            prev = FuelLog.query.filter(
                FuelLog.vehicle_id == v.id, FuelLog.log_date < df,
                FuelLog.odometer.isnot(None)).order_by(FuelLog.log_date.desc()).first()
            if prev:
                prev_odometer, prev_date = prev.odometer, prev.log_date
                distance = odometer - prev_odometer
                total_distance += distance or 0.0
        rows_out.append((v.registration, prev_date or '', prev_odometer or '',
                         odometer if odometer is not None else '',
                         f'{distance:.0f}' if distance is not None else ''))

    if request.args.get('format') == 'pdf':
        styles = _pdf_styles()
        flowables = [_pdf_statement_table([
            ['Fleet Total Distance (km)', f'{total_distance:,.0f}'],
            ['Fleet Total Revenue (USD)', f'${fleet_revenue:,.2f}'],
        ]), Spacer(1, 14)]
        table_header = ['Vehicle', 'Previous Reading Date', 'Previous Odometer (km)',
                        'Odometer in Range (km)', 'Distance Travelled (km)']
        flowables.append(_pdf_table([table_header] + [list(r) for r in rows_out], bold_last_row=False))
        elements = _pdf_section('Distance Travelled', f'Period: {df} to {dt}', flowables)
        return _pdf_response(f'distance_travelled_{df}_to_{dt}.pdf', elements)

    out = io.StringIO()
    w = csv.writer(out)
    w.writerow([f'DISTANCE TRAVELLED — {df} to {dt}'])
    w.writerow([])
    w.writerow(['Fleet Total Distance (km)', f'{total_distance:.0f}'])
    w.writerow(['Fleet Total Revenue (USD)', f'{fleet_revenue:.2f}'])
    w.writerow([])
    w.writerow(['Vehicle', 'Previous Reading Date', 'Previous Odometer (km)',
                'Odometer in Range (km)', 'Distance Travelled (km)'])
    for out_row in rows_out:
        w.writerow(out_row)

    out.seek(0)
    resp = make_response(out.getvalue())
    resp.headers['Content-Type'] = 'text/csv'
    resp.headers['Content-Disposition'] = f'attachment; filename=distance_travelled_{df}_to_{dt}.csv'
    return resp


# ─────────────────────────────────────────────────────────────
# Finance Ledger: Loans
# ─────────────────────────────────────────────────────────────
@app.route('/finance/loans')
@login_required
@permission_required('finance')
def loans_list():
    all_loans = Loan.query.order_by(Loan.start_date.desc()).all()
    return render_template('finance/loans.html', loans=all_loans, today=date.today().strftime('%Y-%m-%d'))


@app.route('/finance/loans/export')
@login_required
@permission_required('finance')
def loans_export():
    all_loans = Loan.query.order_by(Loan.start_date.desc()).all()
    rows = [[l.lender, f'{l.principal:.2f}', l.interest_rate, l.start_date, l.term_months or '',
             l.status, f'{sum(p.amount for p in l.payments):.2f}', l.notes or ''] for l in all_loans]
    header = ['Lender', 'Principal', 'Interest Rate %', 'Start Date', 'Term (months)', 'Status', 'Repaid to Date', 'Notes']
    if request.args.get('format') == 'pdf':
        return _table_pdf_response(f'loans_{date.today()}.pdf', 'Loans', f'Generated {date.today()}', header, rows)
    return csv_export_response(f'loans_{date.today()}.csv', header, rows)


@app.route('/finance/loans/add', methods=['GET', 'POST'])
@login_required
@permission_required('finance')
@handle_form_errors
def loan_add():
    if request.method == 'POST':
        loan = Loan(
            lender=request.form['lender'].strip(),
            principal=form_float(request.form, 'principal', min_value=0),
            interest_rate=form_float(request.form, 'interest_rate', required=False, default=0, min_value=0),
            start_date=parse_date(request.form['start_date']),
            term_months=form_int(request.form, 'term_months', required=False),
            status='active',
            notes=request.form.get('notes', '').strip(),
            created_by=current_user.id,
        )
        db.session.add(loan)
        db.session.flush()
        log_audit('CREATE', 'loans', loan.id, f'Added loan from {loan.lender} for {loan.principal}')
        touch_sync_fields(loan)
        db.session.commit()
        flash('Loan recorded.', 'success')
        return redirect(url_for('loans_list'))
    return render_template('finance/loan_form.html', today=date.today().strftime('%Y-%m-%d'))


@app.route('/finance/loans/<int:lid>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
@handle_form_errors
def loan_edit(lid):
    loan = Loan.query.filter_by(id=lid).first_or_404()
    if request.method == 'POST':
        loan.lender = request.form['lender'].strip()
        loan.principal = form_float(request.form, 'principal', min_value=0)
        loan.interest_rate = form_float(request.form, 'interest_rate', required=False, default=0, min_value=0)
        loan.start_date = parse_date(request.form['start_date'])
        loan.term_months = form_int(request.form, 'term_months', required=False)
        loan.notes = request.form.get('notes', '').strip()
        log_audit('UPDATE', 'loans', loan.id, f'Updated loan from {loan.lender}')
        touch_sync_fields(loan)
        db.session.commit()
        flash('Loan updated.', 'success')
        return redirect(url_for('loans_list'))
    return render_template('finance/loan_form.html', loan=loan, today=loan.start_date.strftime('%Y-%m-%d'))


@app.route('/finance/loans/<int:lid>/delete', methods=['POST'])
@login_required
@admin_required
def loan_delete(lid):
    loan = Loan.query.filter_by(id=lid).first_or_404()
    log_audit('DELETE', 'loans', lid, f'Deleted loan from {loan.lender}')
    now = datetime.now(timezone.utc)
    # The model's cascade='all, delete-orphan' only fires on an actual ORM
    # delete of the parent, not on setting deleted_at — soft-delete
    # payments explicitly so they don't outlive their (now hidden) loan.
    for payment in loan.payments:
        payment.deleted_at = now
        touch_sync_fields(payment)
    loan.deleted_at = now
    touch_sync_fields(loan)
    db.session.commit()
    flash('Loan deleted.', 'warning')
    return redirect(url_for('loans_list'))


@app.route('/finance/loans/<int:lid>/payment', methods=['POST'])
@login_required
@permission_required('finance')
@handle_form_errors
def loan_payment_add(lid):
    loan = Loan.query.filter_by(id=lid).first_or_404()
    payment = LoanPayment(
        loan_id=lid,
        payment_date=parse_date(request.form['payment_date']),
        amount=form_float(request.form, 'amount', min_value=0),
        notes=request.form.get('notes', '').strip(),
        created_by=current_user.id,
    )
    db.session.add(payment)
    db.session.flush()
    log_audit('CREATE', 'loan_payments', None, f'Repayment of {payment.amount} on loan from {loan.lender}')
    touch_sync_fields(payment)
    db.session.commit()
    flash('Loan repayment recorded.', 'success')
    return redirect(url_for('loans_list'))


@app.route('/finance/loan-payments/<int:pid>/edit', methods=['POST'])
@login_required
@admin_required
@handle_form_errors
def loan_payment_edit(pid):
    payment = LoanPayment.query.filter_by(id=pid).first_or_404()
    payment.payment_date = parse_date(request.form['payment_date'])
    payment.amount = form_float(request.form, 'amount', min_value=0)
    payment.notes = request.form.get('notes', '').strip()
    log_audit('UPDATE', 'loan_payments', payment.id, f'Updated repayment on loan from {payment.loan.lender}')
    touch_sync_fields(payment)
    db.session.commit()
    flash('Loan repayment updated.', 'success')
    return redirect(url_for('loans_list'))


@app.route('/finance/loan-payments/<int:pid>/delete', methods=['POST'])
@login_required
@admin_required
def loan_payment_delete(pid):
    payment = LoanPayment.query.filter_by(id=pid).first_or_404()
    log_audit('DELETE', 'loan_payments', pid, f'Deleted loan repayment of {payment.amount}')
    payment.deleted_at = datetime.now(timezone.utc)
    touch_sync_fields(payment)
    db.session.commit()
    flash('Loan repayment deleted.', 'warning')
    return redirect(url_for('loans_list'))


# ─────────────────────────────────────────────────────────────
# Finance Ledger: Payables (Accounts Payable)
# ─────────────────────────────────────────────────────────────
@app.route('/finance/payables')
@login_required
@permission_required('finance')
def payables_list():
    all_payables = Payable.query.order_by(Payable.invoice_date.desc()).all()
    return render_template('finance/payables.html', payables=all_payables)


@app.route('/finance/payables/export')
@login_required
@permission_required('finance')
def payables_export():
    all_payables = Payable.query.order_by(Payable.invoice_date.desc()).all()
    rows = [[p.supplier_name, p.description or '', f'{p.amount:.2f}', p.invoice_date,
             p.due_date or '', p.status, p.paid_date or ''] for p in all_payables]
    header = ['Supplier', 'Description', 'Amount', 'Invoice Date', 'Due Date', 'Status', 'Paid Date']
    if request.args.get('format') == 'pdf':
        return _table_pdf_response(f'payables_{date.today()}.pdf', 'Payables', f'Generated {date.today()}', header, rows)
    return csv_export_response(f'payables_{date.today()}.csv', header, rows)


@app.route('/finance/payables/add', methods=['GET', 'POST'])
@login_required
@permission_required('finance')
@handle_form_errors
def payable_add():
    if request.method == 'POST':
        p = Payable(
            supplier_name=request.form['supplier_name'].strip(),
            description=request.form.get('description', '').strip(),
            amount=form_float(request.form, 'amount', min_value=0),
            invoice_date=parse_date(request.form['invoice_date']),
            due_date=parse_date(request.form.get('due_date')),
            created_by=current_user.id,
        )
        db.session.add(p)
        db.session.flush()
        log_audit('CREATE', 'payables', p.id, f'Payable to {p.supplier_name}: {p.amount}')
        touch_sync_fields(p)
        db.session.commit()
        flash('Payable recorded.', 'success')
        return redirect(url_for('payables_list'))
    return render_template('finance/payable_form.html', today=date.today().strftime('%Y-%m-%d'))


@app.route('/finance/payables/<int:pid>/mark-paid', methods=['POST'])
@login_required
@permission_required('finance')
def payable_mark_paid(pid):
    p = Payable.query.filter_by(id=pid).first_or_404()
    p.status = 'paid'
    p.paid_date = date.today()
    log_audit('UPDATE', 'payables', pid, f'Marked payable to {p.supplier_name} as paid')
    touch_sync_fields(p)
    db.session.commit()
    flash('Payable marked as paid.', 'success')
    return redirect(url_for('payables_list'))


@app.route('/finance/payables/<int:pid>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
@handle_form_errors
def payable_edit(pid):
    p = Payable.query.filter_by(id=pid).first_or_404()
    if request.method == 'POST':
        p.supplier_name = request.form['supplier_name'].strip()
        p.description = request.form.get('description', '').strip()
        p.amount = form_float(request.form, 'amount', min_value=0)
        p.invoice_date = parse_date(request.form['invoice_date'])
        p.due_date = parse_date(request.form.get('due_date'))
        log_audit('UPDATE', 'payables', p.id, f'Updated payable to {p.supplier_name}')
        touch_sync_fields(p)
        db.session.commit()
        flash('Payable updated.', 'success')
        return redirect(url_for('payables_list'))
    return render_template('finance/payable_form.html', payable=p, today=p.invoice_date.strftime('%Y-%m-%d'))


@app.route('/finance/payables/<int:pid>/delete', methods=['POST'])
@login_required
@admin_required
def payable_delete(pid):
    p = Payable.query.filter_by(id=pid).first_or_404()
    log_audit('DELETE', 'payables', pid, f'Deleted payable to {p.supplier_name}')
    p.deleted_at = datetime.now(timezone.utc)
    touch_sync_fields(p)
    db.session.commit()
    flash('Payable deleted.', 'warning')
    return redirect(url_for('payables_list'))


# ─────────────────────────────────────────────────────────────
# Finance Ledger: Receivables (Accounts Receivable)
# ─────────────────────────────────────────────────────────────
@app.route('/finance/receivables')
@login_required
@permission_required('finance')
def receivables_list():
    all_receivables = Receivable.query.order_by(Receivable.invoice_date.desc()).all()
    return render_template('finance/receivables.html', receivables=all_receivables)


@app.route('/finance/receivables/export')
@login_required
@permission_required('finance')
def receivables_export():
    all_receivables = Receivable.query.order_by(Receivable.invoice_date.desc()).all()
    rows = [[r.client_name, r.description or '', f'{r.amount:.2f}', r.invoice_date,
             r.due_date or '', r.status, r.collected_date or ''] for r in all_receivables]
    header = ['Client', 'Description', 'Amount', 'Invoice Date', 'Due Date', 'Status', 'Collected Date']
    if request.args.get('format') == 'pdf':
        return _table_pdf_response(f'receivables_{date.today()}.pdf', 'Receivables', f'Generated {date.today()}', header, rows)
    return csv_export_response(f'receivables_{date.today()}.csv', header, rows)


@app.route('/finance/receivables/add', methods=['GET', 'POST'])
@login_required
@permission_required('finance')
@handle_form_errors
def receivable_add():
    if request.method == 'POST':
        r = Receivable(
            client_name=request.form['client_name'].strip(),
            description=request.form.get('description', '').strip(),
            amount=form_float(request.form, 'amount', min_value=0),
            invoice_date=parse_date(request.form['invoice_date']),
            due_date=parse_date(request.form.get('due_date')),
            created_by=current_user.id,
        )
        db.session.add(r)
        db.session.flush()
        log_audit('CREATE', 'receivables', r.id, f'Receivable from {r.client_name}: {r.amount}')
        touch_sync_fields(r)
        db.session.commit()
        flash('Receivable recorded.', 'success')
        return redirect(url_for('receivables_list'))
    return render_template('finance/receivable_form.html', today=date.today().strftime('%Y-%m-%d'))


@app.route('/finance/receivables/<int:rid>/mark-collected', methods=['POST'])
@login_required
@permission_required('finance')
def receivable_mark_collected(rid):
    r = Receivable.query.filter_by(id=rid).first_or_404()
    r.status = 'collected'
    r.collected_date = date.today()
    log_audit('UPDATE', 'receivables', rid, f'Marked receivable from {r.client_name} as collected')
    touch_sync_fields(r)
    db.session.commit()
    flash('Receivable marked as collected.', 'success')
    return redirect(url_for('receivables_list'))


@app.route('/finance/receivables/<int:rid>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
@handle_form_errors
def receivable_edit(rid):
    r = Receivable.query.filter_by(id=rid).first_or_404()
    if request.method == 'POST':
        r.client_name = request.form['client_name'].strip()
        r.description = request.form.get('description', '').strip()
        r.amount = form_float(request.form, 'amount', min_value=0)
        r.invoice_date = parse_date(request.form['invoice_date'])
        r.due_date = parse_date(request.form.get('due_date'))
        log_audit('UPDATE', 'receivables', r.id, f'Updated receivable from {r.client_name}')
        touch_sync_fields(r)
        db.session.commit()
        flash('Receivable updated.', 'success')
        return redirect(url_for('receivables_list'))
    return render_template('finance/receivable_form.html', receivable=r, today=r.invoice_date.strftime('%Y-%m-%d'))


@app.route('/finance/receivables/<int:rid>/delete', methods=['POST'])
@login_required
@admin_required
def receivable_delete(rid):
    r = Receivable.query.filter_by(id=rid).first_or_404()
    log_audit('DELETE', 'receivables', rid, f'Deleted receivable from {r.client_name}')
    r.deleted_at = datetime.now(timezone.utc)
    touch_sync_fields(r)
    db.session.commit()
    flash('Receivable deleted.', 'warning')
    return redirect(url_for('receivables_list'))


# ─────────────────────────────────────────────────────────────
# Finance Ledger: Owner's Capital & Drawings
# ─────────────────────────────────────────────────────────────
@app.route('/finance/capital')
@login_required
@permission_required('finance')
def capital_list():
    contributions = CapitalContribution.query.order_by(CapitalContribution.contribution_date.desc()).all()
    drawings = OwnerDrawing.query.order_by(OwnerDrawing.drawing_date.desc()).all()
    return render_template('finance/capital.html', contributions=contributions, drawings=drawings)


@app.route('/finance/capital/export')
@login_required
@permission_required('finance')
def capital_export():
    contributions = CapitalContribution.query.order_by(CapitalContribution.contribution_date.desc()).all()
    drawings = OwnerDrawing.query.order_by(OwnerDrawing.drawing_date.desc()).all()
    rows = [['Contribution', c.contribution_date, c.contributor, f'{c.amount:.2f}', c.notes or ''] for c in contributions]
    rows += [['Drawing', d.drawing_date, '', f'{d.amount:.2f}', d.notes or ''] for d in drawings]
    rows.sort(key=lambda r: r[1], reverse=True)
    header = ['Type', 'Date', 'Contributor', 'Amount', 'Notes']
    if request.args.get('format') == 'pdf':
        return _table_pdf_response(f'capital_drawings_{date.today()}.pdf', 'Capital / Drawings',
            f'Generated {date.today()}', header, rows)
    return csv_export_response(f'capital_drawings_{date.today()}.csv', header, rows)


@app.route('/finance/capital/contributions/add', methods=['GET', 'POST'])
@login_required
@permission_required('finance')
@handle_form_errors
def capital_contribution_add():
    if request.method == 'POST':
        c = CapitalContribution(
            contributor=request.form['contributor'].strip(),
            amount=form_float(request.form, 'amount', min_value=0),
            contribution_date=parse_date(request.form['contribution_date']),
            notes=request.form.get('notes', '').strip(),
            created_by=current_user.id,
        )
        db.session.add(c)
        db.session.flush()
        log_audit('CREATE', 'capital_contributions', c.id, f'Capital contribution from {c.contributor}: {c.amount}')
        touch_sync_fields(c)
        db.session.commit()
        flash('Capital contribution recorded.', 'success')
        return redirect(url_for('capital_list'))
    return render_template('finance/capital_contribution_form.html', today=date.today().strftime('%Y-%m-%d'))


@app.route('/finance/capital/contributions/<int:cid>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
@handle_form_errors
def capital_contribution_edit(cid):
    c = CapitalContribution.query.filter_by(id=cid).first_or_404()
    if request.method == 'POST':
        c.contributor = request.form['contributor'].strip()
        c.amount = form_float(request.form, 'amount', min_value=0)
        c.contribution_date = parse_date(request.form['contribution_date'])
        c.notes = request.form.get('notes', '').strip()
        log_audit('UPDATE', 'capital_contributions', c.id, f'Updated capital contribution from {c.contributor}')
        touch_sync_fields(c)
        db.session.commit()
        flash('Capital contribution updated.', 'success')
        return redirect(url_for('capital_list'))
    return render_template('finance/capital_contribution_form.html', contribution=c,
                           today=c.contribution_date.strftime('%Y-%m-%d'))


@app.route('/finance/capital/contributions/<int:cid>/delete', methods=['POST'])
@login_required
@admin_required
def capital_contribution_delete(cid):
    c = CapitalContribution.query.filter_by(id=cid).first_or_404()
    log_audit('DELETE', 'capital_contributions', cid, f'Deleted capital contribution from {c.contributor}')
    c.deleted_at = datetime.now(timezone.utc)
    touch_sync_fields(c)
    db.session.commit()
    flash('Capital contribution deleted.', 'warning')
    return redirect(url_for('capital_list'))


@app.route('/finance/capital/drawings/add', methods=['GET', 'POST'])
@login_required
@permission_required('finance')
@handle_form_errors
def owner_drawing_add():
    if request.method == 'POST':
        d = OwnerDrawing(
            amount=form_float(request.form, 'amount', min_value=0),
            drawing_date=parse_date(request.form['drawing_date']),
            notes=request.form.get('notes', '').strip(),
            created_by=current_user.id,
        )
        db.session.add(d)
        db.session.flush()
        log_audit('CREATE', 'owner_drawings', d.id, f'Owner drawing: {d.amount}')
        touch_sync_fields(d)
        db.session.commit()
        flash('Owner drawing recorded.', 'success')
        return redirect(url_for('capital_list'))
    return render_template('finance/owner_drawing_form.html', today=date.today().strftime('%Y-%m-%d'))


@app.route('/finance/capital/drawings/<int:did>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
@handle_form_errors
def owner_drawing_edit(did):
    d = OwnerDrawing.query.filter_by(id=did).first_or_404()
    if request.method == 'POST':
        d.amount = form_float(request.form, 'amount', min_value=0)
        d.drawing_date = parse_date(request.form['drawing_date'])
        d.notes = request.form.get('notes', '').strip()
        log_audit('UPDATE', 'owner_drawings', d.id, f'Updated owner drawing of {d.amount}')
        touch_sync_fields(d)
        db.session.commit()
        flash('Owner drawing updated.', 'success')
        return redirect(url_for('capital_list'))
    return render_template('finance/owner_drawing_form.html', drawing=d,
                           today=d.drawing_date.strftime('%Y-%m-%d'))


@app.route('/finance/capital/drawings/<int:did>/delete', methods=['POST'])
@login_required
@admin_required
def owner_drawing_delete(did):
    d = OwnerDrawing.query.filter_by(id=did).first_or_404()
    log_audit('DELETE', 'owner_drawings', did, f'Deleted owner drawing of {d.amount}')
    d.deleted_at = datetime.now(timezone.utc)
    touch_sync_fields(d)
    db.session.commit()
    flash('Owner drawing deleted.', 'warning')
    return redirect(url_for('capital_list'))


# ─────────────────────────────────────────────────────────────
# Finance Ledger: Expense Categories & Expenses
# ─────────────────────────────────────────────────────────────
@app.route('/finance/expenses')
@login_required
@permission_required('finance')
def expenses_list():
    page = request.args.get('page', 1, type=int)
    vehicle_id = request.args.get('vehicle_id', '')
    q = request.args.get('q', '').strip()

    query = Expense.query
    if vehicle_id:
        query = query.filter(Expense.vehicle_id == vehicle_id)
    if q:
        like = f'%{q}%'
        query = query.join(Expense.vehicle, isouter=True).filter(
            db.or_(Expense.description.ilike(like), Vehicle.registration.ilike(like)))

    expenses = query.order_by(Expense.expense_date.desc()).paginate(page=page, per_page=20)
    headings = ExpenseCategory.query.filter_by(parent_id=None).order_by(ExpenseCategory.name).all()
    all_vehicles = Vehicle.query.order_by(Vehicle.registration).all()
    return render_template('finance/expenses.html', expenses=expenses, headings=headings,
                           vehicles=all_vehicles, vehicle_id=vehicle_id, q=q)


@app.route('/finance/expenses/export')
@login_required
@permission_required('finance')
def expenses_export():
    vehicle_id = request.args.get('vehicle_id', '')
    q = request.args.get('q', '').strip()

    query = Expense.query
    if vehicle_id:
        query = query.filter(Expense.vehicle_id == vehicle_id)
    if q:
        like = f'%{q}%'
        query = query.join(Expense.vehicle, isouter=True).filter(
            db.or_(Expense.description.ilike(like), Vehicle.registration.ilike(like)))

    all_expenses = query.order_by(Expense.expense_date.desc()).all()
    rows = [[e.expense_date, e.category.display_name, e.vehicle.registration if e.vehicle else '',
             e.driver.name if e.driver else '', e.description or '', f'{e.amount:.2f}'] for e in all_expenses]
    header = ['Date', 'Category', 'Vehicle', 'Driver', 'Description', 'Amount']
    if request.args.get('format') == 'pdf':
        return _table_pdf_response(f'expenses_{date.today()}.pdf', 'Expenses', f'Generated {date.today()}', header, rows)
    return csv_export_response(f'expenses_{date.today()}.csv', header, rows)


@app.route('/finance/expenses/add', methods=['GET', 'POST'])
@login_required
@permission_required('finance')
@handle_form_errors
def expense_add():
    if request.method == 'POST':
        client_id = request.form.get('_client_id')
        if already_synced(client_id):
            flash('Already recorded.', 'info')
            return redirect(url_for('expenses_list'))
        e = Expense(
            category_id=form_int(request.form, 'category_id'),
            vehicle_id=form_int(request.form, 'vehicle_id', required=False),
            driver_id=form_int(request.form, 'driver_id', required=False),
            expense_date=parse_date(request.form['expense_date']),
            description=request.form.get('description', '').strip(),
            amount=form_float(request.form, 'amount', min_value=0),
            created_by=current_user.id,
        )
        db.session.add(e)
        db.session.flush()
        log_audit('CREATE', 'expenses', e.id, f'Expense of {e.amount}')
        record_offline_sync(client_id, 'expense_add')
        touch_sync_fields(e)
        db.session.commit()
        flash('Expense recorded.', 'success')
        return redirect(url_for('expenses_list'))
    headings = ExpenseCategory.query.filter_by(parent_id=None).order_by(ExpenseCategory.name).all()
    all_vehicles = Vehicle.query.order_by(Vehicle.registration).all()
    all_drivers = Driver.query.filter_by(status='active').order_by(Driver.name).all()
    selected_category_id = request.args.get('new_category_id', '')
    return render_template('finance/expense_form.html', headings=headings, vehicles=all_vehicles,
                           drivers=all_drivers,
                           selected_category_id=selected_category_id,
                           today=date.today().strftime('%Y-%m-%d'))


@app.route('/finance/expenses/<int:eid>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
@handle_form_errors
def expense_edit(eid):
    e = Expense.query.filter_by(id=eid).first_or_404()
    if request.method == 'POST':
        e.category_id = form_int(request.form, 'category_id')
        e.vehicle_id = form_int(request.form, 'vehicle_id', required=False)
        e.driver_id = form_int(request.form, 'driver_id', required=False)
        e.expense_date = parse_date(request.form['expense_date'])
        e.description = request.form.get('description', '').strip()
        e.amount = form_float(request.form, 'amount', min_value=0)
        log_audit('UPDATE', 'expenses', e.id, f'Updated expense of {e.amount}')
        touch_sync_fields(e)
        db.session.commit()
        flash('Expense updated.', 'success')
        return redirect(url_for('expenses_list'))
    headings = ExpenseCategory.query.filter_by(parent_id=None).order_by(ExpenseCategory.name).all()
    all_vehicles = Vehicle.query.order_by(Vehicle.registration).all()
    all_drivers = Driver.query.filter_by(status='active').order_by(Driver.name).all()
    return render_template('finance/expense_form.html', expense=e, headings=headings, vehicles=all_vehicles,
                           drivers=all_drivers,
                           selected_category_id=e.category_id,
                           today=e.expense_date.strftime('%Y-%m-%d'))


@app.route('/finance/expenses/<int:eid>/delete', methods=['POST'])
@login_required
@admin_required
def expense_delete(eid):
    e = Expense.query.filter_by(id=eid).first_or_404()
    log_audit('DELETE', 'expenses', eid, f'Deleted expense of {e.amount}')
    e.deleted_at = datetime.now(timezone.utc)
    touch_sync_fields(e)
    db.session.commit()
    flash('Expense deleted.', 'warning')
    return redirect(url_for('expenses_list'))


@app.route('/finance/expense-categories/add', methods=['POST'])
@login_required
@permission_required('finance')
@handle_form_errors
def expense_category_add():
    name = request.form.get('name', '').strip()
    if not name:
        raise ValueError('Category name is required.')
    parent_id = form_int(request.form, 'parent_id', required=False)
    parent = ExpenseCategory.query.filter_by(id=parent_id).first() if parent_id else None
    if parent_id and not parent:
        raise ValueError('Selected heading does not exist.')

    redirect_to = request.form.get('redirect_to') or request.referrer or url_for('expenses_list')

    default_amount = form_float(request.form, 'default_amount', label='Default amount', required=False, min_value=0)

    existing = ExpenseCategory.query.filter_by(name=name, parent_id=parent_id).first()
    if existing:
        flash(f'"{name}" already exists under {parent.name if parent else "top-level headings"}.', 'warning')
    else:
        new_cat = ExpenseCategory(name=name, parent_id=parent_id, default_amount=default_amount)
        db.session.add(new_cat)
        db.session.flush()
        log_audit('CREATE', 'expense_categories', new_cat.id, f'Added expense category {new_cat.display_name}')
        touch_sync_fields(new_cat)
        db.session.commit()
        flash(f'"{new_cat.display_name}" added.', 'success')
        sep = '&' if '?' in redirect_to else '?'
        redirect_to = f'{redirect_to}{sep}new_category_id={new_cat.id}'
    return redirect(redirect_to)


@app.route('/finance/expense-categories/<int:cid>/edit', methods=['GET', 'POST'])
@login_required
@permission_required('finance')
@handle_form_errors
def expense_category_edit(cid):
    cat = ExpenseCategory.query.filter_by(id=cid).first_or_404()
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        if not name:
            raise ValueError('Category name is required.')
        parent_id = form_int(request.form, 'parent_id', required=False)
        if parent_id == cat.id:
            raise ValueError('A category cannot be its own heading.')
        if parent_id and any(child.id == parent_id for child in cat.children):
            raise ValueError('Cannot move a heading under one of its own sub-headings.')
        parent = ExpenseCategory.query.filter_by(id=parent_id).first() if parent_id else None
        if parent_id and not parent:
            raise ValueError('Selected heading does not exist.')
        if parent_id and cat.children:
            raise ValueError('This category has sub-headings under it, so it cannot also become a sub-heading itself.')

        old_label = cat.display_name
        cat.name = name
        cat.parent_id = parent_id
        cat.default_amount = form_float(request.form, 'default_amount', label='Default amount',
                                        required=False, min_value=0)
        log_audit('UPDATE', 'expense_categories', cid, f'Renamed expense category {old_label} to {cat.display_name}')
        touch_sync_fields(cat)
        db.session.commit()
        flash(f'"{cat.display_name}" updated.', 'success')
        return redirect(url_for('expenses_list'))

    headings = ExpenseCategory.query.filter_by(parent_id=None).order_by(ExpenseCategory.name).all()
    return render_template('finance/expense_category_form.html', category=cat, headings=headings)


@app.route('/finance/expense-categories/<int:cid>/delete', methods=['POST'])
@login_required
@permission_required('finance')
def expense_category_delete(cid):
    cat = ExpenseCategory.query.filter_by(id=cid).first_or_404()
    if cat.children:
        flash(f'Cannot delete "{cat.name}" — it has sub-headings under it. Delete those first.', 'danger')
        return redirect(request.referrer or url_for('expenses_list'))
    if Expense.query.filter_by(category_id=cid).first():
        flash(f'Cannot delete "{cat.display_name}" — expenses are recorded against it.', 'danger')
        return redirect(request.referrer or url_for('expenses_list'))
    name = cat.display_name
    log_audit('DELETE', 'expense_categories', cid, f'Deleted expense category {name}')
    cat.deleted_at = datetime.now(timezone.utc)
    touch_sync_fields(cat)
    db.session.commit()
    flash(f'"{name}" deleted.', 'warning')
    return redirect(request.referrer or url_for('expenses_list'))


# ─────────────────────────────────────────────────────────────
# Franchise Income — daily and weekly franchise fee collections, kept as
# two fully independent entities (see FranchiseDailyIncome/
# FranchiseWeeklyIncome above), each reconciled against its own income,
# expenditure and cash deposited.
# ─────────────────────────────────────────────────────────────
FRANCHISE_INCOME_EXPENSE_FIELDS = (
    ('exp_traffic_fines', 'Traffic Fines'),
    ('exp_facilitation_fees', 'Facilitation Fees'),
    ('exp_workshop', 'Workshop'),
    ('exp_wages', 'Wages'),
)


def _franchise_income_period_rows(model_cls, date_field, vehicle, df, dt):
    q = model_cls.query.filter(model_cls.vehicle_id == (vehicle.id if vehicle else None))
    if df:
        col = getattr(model_cls, date_field)
        q = q.filter(col >= df, col <= dt)
    return q.order_by(getattr(model_cls, date_field).desc()).all()


def _franchise_fleet_income_total(model_cls, date_field, df, dt):
    """Sum of income across every vehicle (not the shared, vehicle_id-less
    expenditure rows) for the period — the fleet-wide income half of the
    franchise P&L, since income is recorded per vehicle while expenditure is
    recorded once, shared across the whole franchise (see
    _franchise_income_period_rows(..., vehicle=None, ...) for that side)."""
    q = model_cls.query.filter(model_cls.vehicle_id.isnot(None))
    if df:
        col = getattr(model_cls, date_field)
        q = q.filter(col >= df, col <= dt)
    return sum(e.income for e in q.all())


def _franchise_income_by_vehicle_on(model_cls, date_field, target_date, vehicles, end_date=None):
    """The other cut of the same data as _franchise_income_period_rows: one
    row per vehicle for a single date/week, instead of one row per date/week
    for a single vehicle — so all vehicles can be scanned side by side for a
    given day/week (see the "By Date"/"By Week" view on the Daily/Weekly
    Income pages). A vehicle can have more than one entry on the same
    week_start (partial payments — see FranchiseWeeklyIncome), so entries
    are summed per vehicle rather than assumed unique.

    end_date, when passed, widens the match to a range [target_date,
    end_date] instead of an exact date — used for the Weekly Income "By
    Week" view, since week_start now holds each entry's own date rather
    than being normalized to that week's Monday, so a week's entries can
    fall on any day within it."""
    col = getattr(model_cls, date_field)
    cond = col == target_date if end_date is None else col.between(target_date, end_date)
    entries = model_cls.query.filter(cond, model_cls.vehicle_id.isnot(None)).all()
    by_vehicle = {}
    for e in entries:
        by_vehicle.setdefault(e.vehicle_id, []).append(e)
    rows = []
    for v in vehicles:
        v_entries = by_vehicle.get(v.id, [])
        rows.append(dict(vehicle=v, entries=v_entries, income=sum(e.income for e in v_entries),
                          description='; '.join(e.description for e in v_entries if e.description)))
    return rows


@app.route('/franchise/daily-income')
@login_required
@permission_required('franchise')
def franchise_daily_income_list():
    today = date.today()
    period, df, dt = resolve_ledger_period(request.args.get('period', 'month'), today)

    all_vehicles = FranchiseVehicle.query.filter_by(status='active').order_by(FranchiseVehicle.number_plate).all()
    vehicle_id = request.args.get('vehicle_id', type=int)
    vehicle = FranchiseVehicle.query.filter_by(id=vehicle_id).first() if vehicle_id else None
    # A first-ever visit (no vehicle_id in the URL at all) defaults to the
    # first vehicle tab rather than landing on an empty income panel. But
    # 'vehicle_id' present-and-blank is a deliberate, explicit choice — the
    # "Whole Franchise" option in the vehicle picker (see the template) —
    # and must stay that way rather than being silently forced back onto a
    # vehicle: that's the only reachable path to the bulk many-vehicles-at-
    # once CSV importer (see franchise_income_import_preview).
    if not vehicle and 'vehicle_id' not in request.args and all_vehicles:
        vehicle = all_vehicles[0]

    income_entries = _franchise_income_period_rows(FranchiseDailyIncome, 'entry_date', vehicle, df, dt) if vehicle else []
    expenditure_entries = _franchise_income_period_rows(FranchiseDailyIncome, 'entry_date', None, df, dt)
    fleet_income = _franchise_fleet_income_total(FranchiseDailyIncome, 'entry_date', df, dt)
    total_expenditure = sum(e.total_expenditure for e in expenditure_entries)
    # Cash Deposited is one lump sum per date covering every vehicle's
    # collections combined, not split per vehicle — so it's recorded on the
    # shared (vehicle-less) expenditure row, and the fleet variance compares
    # it against the whole fleet's net income for the period, not any one
    # vehicle's.
    fleet_deposited = sum(e.deposited for e in expenditure_entries)
    fleet_variance = _franchise_fleet_variance(FranchiseDailyIncome, 'entry_date', 'daily', df, dt)
    # Net Income is deposited cash minus whatever variance is still
    # outstanding, not the raw income-minus-expenditure book figure — once a
    # variance is cleared through the Suspense Account, that discrepancy is
    # considered explained/settled, so it should stop separating Net Income
    # from Total Cash Deposited. If every variance in the period is cleared,
    # fleet_variance is 0 and the two KPI cards read identically.
    net_income = fleet_deposited - fleet_variance

    # "By Date" view — every vehicle's income for one calendar date side by
    # side, complementing the "By Vehicle" view above which is one vehicle
    # across many dates. See _franchise_income_by_vehicle_on.
    view = request.args.get('view', 'vehicle')
    on_date = parse_date(request.args.get('on_date')) or today
    by_date_rows = _franchise_income_by_vehicle_on(FranchiseDailyIncome, 'entry_date', on_date, all_vehicles) \
        if view == 'date' else None
    by_date_total = sum(r['income'] for r in by_date_rows) if by_date_rows is not None else 0
    by_date_missing = sum(1 for r in by_date_rows if not r['entries']) if by_date_rows is not None else 0

    return render_template('franchise/daily_income_list.html', vehicles=all_vehicles, vehicle=vehicle,
        income_entries=income_entries, expenditure_entries=expenditure_entries,
        period=period, today=today.strftime('%Y-%m-%d'),
        vehicle_income_total=sum(e.income for e in income_entries),
        fleet_income=fleet_income, total_expenditure=total_expenditure,
        net_income=net_income, fleet_deposited=fleet_deposited, fleet_variance=fleet_variance,
        view=view, on_date=on_date.strftime('%Y-%m-%d'), by_date_rows=by_date_rows, by_date_total=by_date_total,
        by_date_missing=by_date_missing)


@app.route('/franchise/daily-income/add', methods=['POST'])
@login_required
@permission_required_any('franchise', 'franchise_entry')
def franchise_daily_income_add():
    # POST-only, no GET counterpart at this URL (see driver_ledger_add) —
    # errors are handled locally here and always redirect to the GET list page.
    # A franchise_entry-only user (see franchise_my_collections) is a
    # restricted clerk: collections against a specific vehicle only, no
    # shared/vehicle-less expenditure row and no expense fields — those
    # stay at their model default (0) below rather than being read from
    # the form, and success takes them back to their own entry page
    # instead of the full list they can't otherwise reach.
    entry_only = not current_user.has_permission('franchise')
    period = request.form.get('period', 'month')
    vehicle_id = form_int(request.form, 'vehicle_id', required=False)
    client_id = request.form.get('_client_id')

    def back():
        if entry_only:
            return redirect(url_for('franchise_my_collections'))
        return redirect(url_for('franchise_daily_income_list', vehicle_id=vehicle_id, period=period))

    if already_synced(client_id):
        flash('Already recorded.', 'info')
        return back()
    try:
        vehicle = FranchiseVehicle.query.filter_by(id=vehicle_id).first() if vehicle_id else None
        if vehicle_id and not vehicle:
            raise ValueError('Select a valid franchise vehicle.')
        if entry_only and not vehicle:
            raise ValueError('Select a vehicle for this collection.')
        label = vehicle.number_plate if vehicle else 'the franchise\'s shared expenditure'
        entry_date = parse_date(request.form['entry_date'])

        # The (entry_date, vehicle_id) uniqueness is enforced at the DB
        # level, and a soft-deleted row still occupies that slot — restore
        # it in place instead of inserting a fresh row, or the INSERT
        # below would fail with an IntegrityError the very next time
        # someone re-enters a date they'd previously deleted.
        entry = (FranchiseDailyIncome.query.execution_options(include_deleted=True)
                .filter_by(entry_date=entry_date, vehicle_id=vehicle.id if vehicle else None).first())
        if entry and entry.deleted_at is None:
            raise ValueError(f'A daily income entry for {label} on {entry_date} already exists — delete it first to re-enter.')
        if entry:
            entry.deleted_at = None
        else:
            entry = FranchiseDailyIncome(entry_date=entry_date, vehicle_id=vehicle.id if vehicle else None)
            db.session.add(entry)
        # Defaults to this vehicle's agreed Daily Fee (set on the vehicle
        # record) when left blank, rather than 0 — the form pre-fills it
        # too, but a blank submission (e.g. a stale offline form) should
        # still land on the agreed amount, not silently record nothing.
        income_default = vehicle.daily_fee if vehicle and vehicle.daily_fee is not None else 0
        entry.income = form_float(request.form, 'income', required=False, default=income_default)
        # An entry_only clerk's form has no expense/deposit fields at all —
        # those are left at whatever the entry already had (0 for a new
        # row), never read from the POST body, even if a crafted request
        # included them.
        if not entry_only:
            entry.other_expenditure = form_float(request.form, 'other_expenditure', required=False, default=0)
        elif entry.other_expenditure is None:
            entry.other_expenditure = 0
        # Cash Deposited is one lump sum per date covering every vehicle
        # combined, entered only on the shared (vehicle-less) row — see
        # franchise_daily_income_list. Admin-only (see
        # franchise_daily_income_deposit) — a non-admin or vehicle-scoped
        # submission is silently ignored rather than trusted, since the
        # field is hidden from those forms but a crafted POST could still
        # include it. Preserves whatever an admin already recorded on a
        # restored soft-deleted row instead of zeroing it back out.
        if current_user.role == 'admin' and not vehicle:
            entry.deposited = form_float(request.form, 'deposited', required=False, default=0)
        elif entry.deposited is None:
            entry.deposited = 0
        entry.description = request.form.get('description', '').strip()
        entry.created_by = current_user.id
        if not entry_only:
            for f, lbl in FRANCHISE_INCOME_EXPENSE_FIELDS:
                setattr(entry, f, form_float(request.form, f, label=lbl, required=False, default=0))
        else:
            for f, _lbl in FRANCHISE_INCOME_EXPENSE_FIELDS:
                if getattr(entry, f) is None:
                    setattr(entry, f, 0)
        db.session.flush()
        log_audit('CREATE', 'franchise_daily_income', entry.id,
                  f'Daily franchise income for {label} on {entry_date}: income {entry.income}, '
                  f'expenditure {entry.total_expenditure}, deposited {entry.deposited}')
        record_offline_sync(client_id, 'franchise_daily_income_add')
        touch_sync_fields(entry)
        db.session.commit()
        flash('Daily franchise income recorded.', 'success')
        if entry_only:
            notify_admins_whatsapp(
                f'Franchise collection entered by {current_user.username}: '
                f'{label} on {entry_date}, income {entry.income:.2f}.'
            )
    except KeyError as e:
        db.session.rollback()
        flash(f'Missing required field: {e}', 'danger')
    except ValueError as e:
        db.session.rollback()
        flash(str(e), 'danger')

    return back()


@app.route('/franchise/my-collections')
@login_required
@permission_required_any('franchise', 'franchise_entry')
def franchise_my_collections():
    """Restricted home page for a franchise_entry-only clerk (see PERMISSIONS)
    — record a collection against a vehicle (income only, no expenses/
    deposits — see franchise_daily_income_add/franchise_weekly_income_add),
    register a new vehicle (franchise_vehicle_quick_add), see their own
    recent entries, and see every clerk's recent entries (team_entries,
    below) so they can tell who's already recorded what. Nothing else in
    Franchise is reachable for this permission level — no reports,
    reconciliation, suspense account, or other vehicles' fees. A full
    franchise-permission holder can also reach this page, but has no reason
    to since the full Daily/Weekly Income pages cover everything here and
    more."""
    vehicles = FranchiseVehicle.query.filter_by(status='active').order_by(FranchiseVehicle.number_plate).all()
    my_daily = FranchiseDailyIncome.query.filter_by(created_by=current_user.id).order_by(
        FranchiseDailyIncome.entry_date.desc()).limit(50).all()
    my_weekly = FranchiseWeeklyIncome.query.filter_by(created_by=current_user.id).order_by(
        FranchiseWeeklyIncome.week_start.desc()).limit(50).all()
    my_entries = sorted(
        [dict(kind='Daily', period=e.entry_date, vehicle=e.vehicle, income=e.income) for e in my_daily] +
        [dict(kind='Weekly', period=e.week_start, vehicle=e.vehicle, income=e.income) for e in my_weekly],
        key=lambda r: r['period'], reverse=True)[:50]

    # Every clerk's recorded collections, not just this one's — so a clerk
    # can see what colleagues have recorded (who recorded what, and how
    # much), same data as my_entries above but across all users. created_by
    # is already populated on every entry (see franchise_daily_income_add /
    # franchise_daily_income_bulk_fill), so this needs no schema change.
    team_daily = FranchiseDailyIncome.query.filter(FranchiseDailyIncome.vehicle_id.isnot(None)).order_by(
        FranchiseDailyIncome.entry_date.desc()).limit(50).all()
    team_weekly = FranchiseWeeklyIncome.query.filter(FranchiseWeeklyIncome.vehicle_id.isnot(None)).order_by(
        FranchiseWeeklyIncome.week_start.desc()).limit(50).all()
    team_entries = sorted(
        [dict(kind='Daily', period=e.entry_date, vehicle=e.vehicle, income=e.income,
              recorded_by=e.creator.username if e.creator else '—') for e in team_daily] +
        [dict(kind='Weekly', period=e.week_start, vehicle=e.vehicle, income=e.income,
              recorded_by=e.creator.username if e.creator else '—') for e in team_weekly],
        key=lambda r: r['period'], reverse=True)[:50]

    return render_template('franchise/my_collections.html', vehicles=vehicles, my_entries=my_entries,
                           team_entries=team_entries, today=date.today().strftime('%Y-%m-%d'))


@app.route('/franchise/confirm-payments')
@login_required
@permission_required_any('franchise', 'franchise_entry')
def franchise_confirm_payments():
    """"Who's paid?" view for a franchise clerk — every active vehicle for
    one date (or week) side by side, with a one-click Confirm for any
    vehicle that hasn't paid yet. This is the franchise_entry-reachable
    counterpart of the admin's "By Date"/"By Week" tab on
    franchise_daily_income_list/franchise_weekly_income_list (same
    _franchise_income_by_vehicle_on data), trimmed down to just
    confirm/record — no fleet KPIs, shared expenditure, or delete, since a
    franchise_entry-only clerk only ever enters/confirms income (see
    PERMISSIONS). A full 'franchise' permission holder can reach this too
    as a shortcut, but the By Date/By Week tab on the full Income pages
    covers the same ground plus more."""
    today = date.today()
    all_vehicles = FranchiseVehicle.query.filter_by(status='active').order_by(FranchiseVehicle.number_plate).all()
    kind = request.args.get('kind')
    if kind not in ('daily', 'weekly'):
        kind = 'daily'

    if kind == 'weekly':
        on_week_raw = parse_date(request.args.get('on_week')) or today
        on_period = on_week_raw - timedelta(days=on_week_raw.weekday())
        rows = _franchise_income_by_vehicle_on(FranchiseWeeklyIncome, 'week_start', on_period, all_vehicles,
                                                end_date=on_period + timedelta(days=6))
    else:
        on_period = parse_date(request.args.get('on_date')) or today
        rows = _franchise_income_by_vehicle_on(FranchiseDailyIncome, 'entry_date', on_period, all_vehicles)

    total = sum(r['income'] for r in rows)
    missing = sum(1 for r in rows if not r['entries'])
    return render_template('franchise/confirm_payments.html', kind=kind,
                           on_period=on_period.strftime('%Y-%m-%d'), rows=rows, total=total, missing=missing)


@app.route('/franchise/daily-income/bulk-fill', methods=['POST'])
@login_required
@permission_required_any('franchise', 'franchise_entry')
def franchise_daily_income_bulk_fill():
    """One-click alternative to picking each vehicle from the dropdown and
    submitting the Add Income form 100+ times over — most franchise vehicles
    charge the same flat daily fee, so on the "By Date" view an admin (or,
    via franchise_confirm_payments, a franchise_entry-only clerk confirming
    who's paid) can fill many vehicles at once instead. Only fills vehicles
    explicitly checked in vehicle_ids — the "By Date" table lists every
    vehicle without an entry pre-checked, so an admin can uncheck the ones
    that didn't run/pay that day (rather than everyone being forced to a
    flat fee regardless of whether they were actually out that day). A
    vehicle with its own agreed Daily Fee (see FranchiseVehicle.daily_fee)
    still gets that fee rather than the typed amount — this only stands in
    for vehicles without one, it never overrides an already-agreed rate.
    Vehicles that already have a (non-deleted) entry for the date are
    skipped even if somehow checked, same as the single-entry Add form's
    own duplicate guard."""
    entry_only = not current_user.has_permission('franchise')

    def back(**kw):
        if entry_only:
            return redirect(url_for('franchise_confirm_payments', **kw))
        return redirect(url_for('franchise_daily_income_list', view='date', **kw))

    period = request.form.get('period', 'month')
    try:
        on_date = parse_date(request.form['on_date'])
        amount = form_float(request.form, 'amount', label='Income amount', required=True, min_value=0)
    except KeyError as e:
        flash(f'Missing required field: {e}', 'danger')
        return back(period=period)
    except ValueError as e:
        flash(str(e), 'danger')
        return back(period=period)

    selected_ids = set(request.form.getlist('vehicle_ids', type=int))
    if not selected_ids:
        flash('No vehicles selected — check the vehicles to fill, or uncheck the ones that didn\'t pay '
              'that day, before submitting.', 'warning')
        return back(on_date=on_date.strftime('%Y-%m-%d'), period=period)

    vehicles = FranchiseVehicle.query.filter(FranchiseVehicle.id.in_(selected_ids), FranchiseVehicle.status == 'active').all()
    already_recorded = {e.vehicle_id for e in FranchiseDailyIncome.query
                         .filter_by(entry_date=on_date).filter(FranchiseDailyIncome.vehicle_id.isnot(None)).all()}
    # Soft-deleted rows at this (date, vehicle) slot must be restored in
    # place rather than left alone — a fresh INSERT would collide with them
    # on the unique constraint, same as franchise_daily_income_add.
    deleted_by_vehicle = {e.vehicle_id: e for e in FranchiseDailyIncome.query.execution_options(include_deleted=True)
                           .filter_by(entry_date=on_date).filter(FranchiseDailyIncome.vehicle_id.isnot(None),
                                                                  FranchiseDailyIncome.deleted_at.isnot(None)).all()}

    filled = 0
    filled_labels = []
    for v in vehicles:
        if v.id in already_recorded:
            continue
        entry = deleted_by_vehicle.get(v.id)
        if entry:
            entry.deleted_at = None
        else:
            entry = FranchiseDailyIncome(entry_date=on_date, vehicle_id=v.id)
            db.session.add(entry)
        entry.income = v.daily_fee if v.daily_fee is not None else amount
        entry.created_by = current_user.id
        touch_sync_fields(entry)
        filled += 1
        filled_labels.append(v.number_plate)

    if filled:
        db.session.flush()
        log_audit('CREATE', 'franchise_daily_income', None,
                  f'Bulk-filled daily income (default {amount}) for {filled} vehicle(s) without an entry on {on_date}')
        db.session.commit()
        flash(f'Recorded daily income for {filled} vehicle(s) on {on_date}.', 'success')
        if entry_only:
            notify_admins_whatsapp(
                f'Franchise payment confirmed by {current_user.username}: '
                f'{", ".join(filled_labels)} on {on_date}.'
            )
    else:
        db.session.rollback()
        flash(f'No entries created — the selected vehicle(s) already have a daily income entry for {on_date}.', 'info')

    return back(on_date=on_date.strftime('%Y-%m-%d'), period=period)


@app.route('/franchise/daily-income/bulk-delete', methods=['POST'])
@login_required
@admin_required
def franchise_daily_income_bulk_delete():
    """Companion to bulk-fill, for the other half of the same mistake: a
    flat Fill assumes every vehicle ran that day, so any vehicle that was
    actually absent/didn't pay ends up with a wrong entry that needs
    clearing again. Lets an admin review the whole day's entries on the "By
    Date" view, check the specific vehicles whose entry was wrong, and
    remove them all in one submission instead of hunting each one down on
    its own per-vehicle tab. Same soft-delete as the single-entry Delete
    button (see franchise_daily_income_delete)."""
    period = request.form.get('period', 'month')
    try:
        on_date = parse_date(request.form['on_date'])
    except KeyError as e:
        flash(f'Missing required field: {e}', 'danger')
        return redirect(url_for('franchise_daily_income_list', view='date', period=period))
    except ValueError as e:
        flash(str(e), 'danger')
        return redirect(url_for('franchise_daily_income_list', view='date', period=period))

    selected_ids = set(request.form.getlist('vehicle_ids', type=int))
    if not selected_ids:
        flash('No vehicles selected — check the vehicles whose entry you want to remove.', 'warning')
        return redirect(url_for('franchise_daily_income_list', view='date',
                                 on_date=on_date.strftime('%Y-%m-%d'), period=period))

    entries = FranchiseDailyIncome.query.filter(
        FranchiseDailyIncome.entry_date == on_date, FranchiseDailyIncome.vehicle_id.in_(selected_ids)).all()
    deleted = 0
    for entry in entries:
        log_audit('DELETE', 'franchise_daily_income', entry.id,
                  f'Bulk-deleted daily franchise income entry for {entry.vehicle.number_plate} on {entry.entry_date}')
        entry.deleted_at = datetime.now(timezone.utc)
        touch_sync_fields(entry)
        deleted += 1

    if deleted:
        db.session.commit()
        flash(f'Deleted {deleted} daily income entry(ies) for {on_date}.', 'warning')
    else:
        db.session.rollback()
        flash(f'None of the selected vehicles had a daily income entry for {on_date}.', 'info')

    return redirect(url_for('franchise_daily_income_list', view='date', on_date=on_date.strftime('%Y-%m-%d'), period=period))


@app.route('/franchise/daily-income/<int:eid>/deposit', methods=['POST'])
@login_required
@admin_required
def franchise_daily_income_deposit(eid):
    """Record/update Cash Deposited on an existing entry — the only way to
    set this field, since the Record Income form doesn't offer it to
    non-admins and entries otherwise can't be edited after creation."""
    entry = FranchiseDailyIncome.query.filter_by(id=eid).first_or_404()
    entry.deposited = form_float(request.form, 'deposited', label='Cash deposited', required=False, default=0)
    log_audit('UPDATE', 'franchise_daily_income', entry.id,
              f'Cash deposited for {entry.entry_date} set to {entry.deposited}')
    touch_sync_fields(entry)
    db.session.commit()
    flash('Cash deposited recorded.', 'success')
    return redirect(url_for('franchise_daily_income_list', vehicle_id=entry.vehicle_id))


@app.route('/franchise/daily-income/<int:eid>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
@handle_form_errors
def franchise_daily_income_edit(eid):
    """Full edit of an existing entry — income+deposited fields for a
    vehicle-scoped row, expenditure fields for the shared (vehicle_id-less)
    row, matching whichever half of the split the entry belongs to (see
    the Daily Income list page)."""
    entry = FranchiseDailyIncome.query.filter_by(id=eid).first_or_404()
    if request.method == 'POST':
        entry.entry_date = parse_date(request.form['entry_date'])
        if entry.vehicle_id:
            entry.income = form_float(request.form, 'income', required=False, default=0)
        else:
            for f, lbl in FRANCHISE_INCOME_EXPENSE_FIELDS:
                setattr(entry, f, form_float(request.form, f, label=lbl, required=False, default=0))
            entry.other_expenditure = form_float(request.form, 'other_expenditure', required=False, default=0)
            entry.deposited = form_float(request.form, 'deposited', required=False, default=0)
        entry.description = request.form.get('description', '').strip()
        log_audit('UPDATE', 'franchise_daily_income', entry.id, f'Updated daily franchise entry for {entry.entry_date}')
        touch_sync_fields(entry)
        db.session.commit()
        flash('Entry updated.', 'success')
        return redirect(url_for('franchise_daily_income_list', vehicle_id=entry.vehicle_id))
    return render_template('franchise/income_entry_edit.html', entry=entry, date_field='entry_date',
                           date_value=entry.entry_date, date_label='Date',
                           list_endpoint='franchise_daily_income_list', edit_endpoint='franchise_daily_income_edit')


@app.route('/franchise/daily-income/export')
@login_required
@permission_required('franchise')
def franchise_daily_income_export():
    vehicle_id = request.args.get('vehicle_id', type=int)
    vehicle = FranchiseVehicle.query.filter_by(id=vehicle_id).first() if vehicle_id else None
    if vehicle_id and not vehicle:
        flash('Select a valid franchise vehicle to export.', 'danger')
        return redirect(url_for('franchise_daily_income_list'))

    period, df, dt = resolve_ledger_period(request.args.get('period', 'month'), date.today())
    entries = list(reversed(_franchise_income_period_rows(FranchiseDailyIncome, 'entry_date', vehicle, df, dt)))

    header = ['Date', 'Income', 'Traffic Fines', 'Facilitation Fees', 'Workshop', 'Wages',
              'Other Expenditure', 'Total Expenditure', 'Cash Deposited', 'Variance', 'Description']
    rows = [[e.entry_date, f'{e.income:.2f}', f'{e.exp_traffic_fines:.2f}', f'{e.exp_facilitation_fees:.2f}',
             f'{e.exp_workshop:.2f}', f'{e.exp_wages:.2f}', f'{e.other_expenditure:.2f}',
             f'{e.total_expenditure:.2f}', f'{e.deposited:.2f}', f'{e.variance:.2f}', e.description or '']
            for e in entries]

    label = vehicle.number_plate.replace(' ', '_') if vehicle else 'shared_expenditure'
    if request.args.get('format') == 'pdf':
        return _table_pdf_response(f'{label}_daily_income_{period}_{date.today()}.pdf',
            'Franchise Daily Income', f'Period: {df} to {dt}', header, rows)
    return csv_export_response(f'{label}_daily_income_{period}_{date.today()}.csv', header, rows)


@app.route('/franchise/daily-income/<int:eid>/delete', methods=['POST'])
@login_required
@admin_required
def franchise_daily_income_delete(eid):
    entry = FranchiseDailyIncome.query.filter_by(id=eid).first_or_404()
    log_audit('DELETE', 'franchise_daily_income', eid, f'Deleted daily franchise income entry for {entry.entry_date}')
    entry.deleted_at = datetime.now(timezone.utc)
    touch_sync_fields(entry)
    db.session.commit()
    flash('Daily franchise income entry deleted.', 'warning')
    return redirect(url_for('franchise_daily_income_list', vehicle_id=entry.vehicle_id))


@app.route('/franchise/weekly-income')
@login_required
@permission_required('franchise')
def franchise_weekly_income_list():
    today = date.today()
    period, df, dt = resolve_ledger_period(request.args.get('period', 'month'), today)

    all_vehicles = FranchiseVehicle.query.filter_by(status='active').order_by(FranchiseVehicle.number_plate).all()
    vehicle_id = request.args.get('vehicle_id', type=int)
    vehicle = FranchiseVehicle.query.filter_by(id=vehicle_id).first() if vehicle_id else None
    # See franchise_daily_income_list — 'vehicle_id' present-and-blank
    # (the "Whole Franchise" picker option) is a deliberate choice, not
    # the same as it being absent entirely.
    if not vehicle and 'vehicle_id' not in request.args and all_vehicles:
        vehicle = all_vehicles[0]

    income_entries = _franchise_income_period_rows(FranchiseWeeklyIncome, 'week_start', vehicle, df, dt) if vehicle else []
    expenditure_entries = _franchise_income_period_rows(FranchiseWeeklyIncome, 'week_start', None, df, dt)
    fleet_income = _franchise_fleet_income_total(FranchiseWeeklyIncome, 'week_start', df, dt)
    total_expenditure = sum(e.total_expenditure for e in expenditure_entries)
    # See franchise_daily_income_list — Cash Deposited is one lump sum per
    # week covering every vehicle combined, recorded on the shared row.
    fleet_deposited = sum(e.deposited for e in expenditure_entries)
    fleet_variance = _franchise_fleet_variance(FranchiseWeeklyIncome, 'week_start', 'weekly', df, dt)
    # See franchise_daily_income_list — Net Income reconciles to Total Cash
    # Deposited once suspense-cleared variance is excluded, rather than
    # staying at the raw income-minus-expenditure figure.
    net_income = fleet_deposited - fleet_variance

    # "By Week" view — every vehicle's income for one week side by side. See
    # franchise_daily_income_list's "By Date" view / _franchise_income_by_vehicle_on.
    view = request.args.get('view', 'vehicle')
    on_week_raw = parse_date(request.args.get('on_week')) or today
    on_week = on_week_raw - timedelta(days=on_week_raw.weekday())
    by_date_rows = _franchise_income_by_vehicle_on(FranchiseWeeklyIncome, 'week_start', on_week, all_vehicles,
        end_date=on_week + timedelta(days=6)) if view == 'date' else None
    by_date_total = sum(r['income'] for r in by_date_rows) if by_date_rows is not None else 0
    by_date_missing = sum(1 for r in by_date_rows if not r['entries']) if by_date_rows is not None else 0

    return render_template('franchise/weekly_income_list.html', vehicles=all_vehicles, vehicle=vehicle,
        income_entries=income_entries, expenditure_entries=expenditure_entries,
        period=period, today=today.strftime('%Y-%m-%d'),
        vehicle_income_total=sum(e.income for e in income_entries),
        fleet_income=fleet_income, total_expenditure=total_expenditure,
        net_income=net_income, fleet_deposited=fleet_deposited, fleet_variance=fleet_variance,
        view=view, on_week=on_week.strftime('%Y-%m-%d'), by_date_rows=by_date_rows, by_date_total=by_date_total,
        by_date_missing=by_date_missing)


@app.route('/franchise/weekly-income/add', methods=['POST'])
@login_required
@permission_required_any('franchise', 'franchise_entry')
def franchise_weekly_income_add():
    # See franchise_daily_income_add — same entry_only clerk restrictions
    # (vehicle required, no expense/deposit fields, notify admins, redirect
    # to their own entry page instead of the full list).
    entry_only = not current_user.has_permission('franchise')
    period = request.form.get('period', 'month')
    vehicle_id = form_int(request.form, 'vehicle_id', required=False)
    client_id = request.form.get('_client_id')

    def back():
        if entry_only:
            return redirect(url_for('franchise_my_collections'))
        return redirect(url_for('franchise_weekly_income_list', vehicle_id=vehicle_id, period=period))

    if already_synced(client_id):
        flash('Already recorded.', 'info')
        return back()
    try:
        vehicle = FranchiseVehicle.query.filter_by(id=vehicle_id).first() if vehicle_id else None
        if vehicle_id and not vehicle:
            raise ValueError('Select a valid franchise vehicle.')
        if entry_only and not vehicle:
            raise ValueError('Select a vehicle for this collection.')
        label = vehicle.number_plate if vehicle else 'the franchise\'s shared expenditure'
        week_start = parse_date(request.form['week_start'])  # the entry's own date, not normalized to Monday

        # Unlike franchise_daily_income_add, this always inserts a fresh row
        # rather than looking for one to restore/reuse at the same
        # (week_start, vehicle_id) — a vehicle can have more than one weekly
        # income entry for the same week (e.g. several partial payments), so
        # there's no single existing slot to find.
        entry = FranchiseWeeklyIncome(week_start=week_start, vehicle_id=vehicle.id if vehicle else None)
        db.session.add(entry)
        # Defaults to this vehicle's agreed Weekly Fee when left blank — see
        # franchise_daily_income_add.
        income_default = vehicle.weekly_fee if vehicle and vehicle.weekly_fee is not None else 0
        entry.income = form_float(request.form, 'income', required=False, default=income_default)
        # See franchise_daily_income_add — an entry_only clerk's form has
        # no expense/deposit fields, so those are never read from the POST.
        entry.other_expenditure = form_float(request.form, 'other_expenditure', required=False, default=0) \
            if not entry_only else 0
        # Cash Deposited is one lump sum per week, entered only on the
        # shared row — see franchise_daily_income_add /
        # franchise_weekly_income_deposit.
        if current_user.role == 'admin' and not vehicle:
            entry.deposited = form_float(request.form, 'deposited', required=False, default=0)
        elif entry.deposited is None:
            entry.deposited = 0
        entry.description = request.form.get('description', '').strip()
        entry.created_by = current_user.id
        if not entry_only:
            for f, lbl in FRANCHISE_INCOME_EXPENSE_FIELDS:
                setattr(entry, f, form_float(request.form, f, label=lbl, required=False, default=0))
        db.session.flush()
        log_audit('CREATE', 'franchise_weekly_income', entry.id,
                  f'Weekly franchise income for {label} for week of {week_start}: income {entry.income}, '
                  f'expenditure {entry.total_expenditure}, deposited {entry.deposited}')
        record_offline_sync(client_id, 'franchise_weekly_income_add')
        touch_sync_fields(entry)
        db.session.commit()
        flash('Weekly franchise income recorded.', 'success')
        if entry_only:
            notify_admins_whatsapp(
                f'Franchise collection entered by {current_user.username}: '
                f'{label} for week of {week_start}, income {entry.income:.2f}.'
            )
    except KeyError as e:
        db.session.rollback()
        flash(f'Missing required field: {e}', 'danger')
    except ValueError as e:
        db.session.rollback()
        flash(str(e), 'danger')

    return back()


@app.route('/franchise/weekly-income/bulk-fill', methods=['POST'])
@login_required
@permission_required_any('franchise', 'franchise_entry')
def franchise_weekly_income_bulk_fill():
    """Bulk-fill weekly income for checked vehicles that don't already
    have an entry in the target week. Mirrors the daily bulk-fill's
    behaviour (including the franchise_entry-only clerk path via
    franchise_confirm_payments) but expands the week to
    [on_week, on_week+6] when checking for existing entries so any entry
    inside the week counts as present."""
    entry_only = not current_user.has_permission('franchise')

    def back(**kw):
        if entry_only:
            return redirect(url_for('franchise_confirm_payments', kind='weekly', **kw))
        return redirect(url_for('franchise_weekly_income_list', view='date', **kw))

    period = request.form.get('period', 'month')
    try:
        on_week = parse_date(request.form['on_week'])
        amount = form_float(request.form, 'amount', label='Income amount', required=True, min_value=0)
    except KeyError as e:
        flash(f'Missing required field: {e}', 'danger')
        return back(period=period)
    except ValueError as e:
        flash(str(e), 'danger')
        return back(period=period)

    selected_ids = set(request.form.getlist('vehicle_ids', type=int))
    if not selected_ids:
        flash('No vehicles selected — check the vehicles to fill before submitting.', 'warning')
        return back(on_week=on_week.strftime('%Y-%m-%d'), period=period)

    vehicles = FranchiseVehicle.query.filter(FranchiseVehicle.id.in_(selected_ids), FranchiseVehicle.status == 'active').all()
    # Any entry for that vehicle within the week counts as already recorded
    week_end = on_week + timedelta(days=6)
    already_recorded = {e.vehicle_id for e in FranchiseWeeklyIncome.query
                         .filter(FranchiseWeeklyIncome.week_start.between(on_week, week_end))
                         .filter(FranchiseWeeklyIncome.vehicle_id.isnot(None)).all()}

    filled = 0
    filled_labels = []
    for v in vehicles:
        if v.id in already_recorded:
            continue
        entry = FranchiseWeeklyIncome(week_start=on_week, vehicle_id=v.id)
        entry.income = v.weekly_fee if v.weekly_fee is not None else amount
        entry.created_by = current_user.id
        touch_sync_fields(entry)
        db.session.add(entry)
        filled += 1
        filled_labels.append(v.number_plate)

    if filled:
        db.session.flush()
        log_audit('CREATE', 'franchise_weekly_income', None,
                  f'Bulk-filled weekly income (default {amount}) for {filled} vehicle(s) for week of {on_week}')
        db.session.commit()
        flash(f'Recorded weekly income for {filled} vehicle(s) for week of {on_week}.', 'success')
        if entry_only:
            notify_admins_whatsapp(
                f'Franchise payment confirmed by {current_user.username}: '
                f'{", ".join(filled_labels)} for week of {on_week}.'
            )
    else:
        db.session.rollback()
        flash(f'No entries created — the selected vehicle(s) already have a weekly income entry in that week.', 'info')

    return back(on_week=on_week.strftime('%Y-%m-%d'), period=period)


@app.route('/franchise/weekly-income/bulk-delete', methods=['POST'])
@login_required
@admin_required
def franchise_weekly_income_bulk_delete():
    """Bulk-delete weekly income entries for the selected vehicles in
    the given week. Soft-deletes rows like single delete."""
    period = request.form.get('period', 'month')
    try:
        on_week = parse_date(request.form['on_week'])
    except KeyError as e:
        flash(f'Missing required field: {e}', 'danger')
        return redirect(url_for('franchise_weekly_income_list', view='date', period=period))
    except ValueError as e:
        flash(str(e), 'danger')
        return redirect(url_for('franchise_weekly_income_list', view='date', period=period))

    selected_ids = set(request.form.getlist('vehicle_ids', type=int))
    if not selected_ids:
        flash('No vehicles selected — check the vehicles whose entry you want to remove.', 'warning')
        return redirect(url_for('franchise_weekly_income_list', view='date', on_week=on_week.strftime('%Y-%m-%d'), period=period))

    week_end = on_week + timedelta(days=6)
    entries = FranchiseWeeklyIncome.query.filter(
        FranchiseWeeklyIncome.week_start.between(on_week, week_end), FranchiseWeeklyIncome.vehicle_id.in_(selected_ids)).all()
    deleted = 0
    for entry in entries:
        log_audit('DELETE', 'franchise_weekly_income', entry.id,
                  f'Bulk-deleted weekly franchise income entry for {entry.vehicle.number_plate} in week of {entry.week_start}')
        entry.deleted_at = datetime.now(timezone.utc)
        touch_sync_fields(entry)
        deleted += 1

    if deleted:
        db.session.commit()
        flash(f'Deleted {deleted} weekly income entry(ies) for week of {on_week}.', 'warning')
    else:
        db.session.rollback()
        flash(f'None of the selected vehicles had a weekly income entry in that week.', 'info')

    return redirect(url_for('franchise_weekly_income_list', view='date', on_week=on_week.strftime('%Y-%m-%d'), period=period))


@app.route('/franchise/weekly-income/<int:eid>/deposit', methods=['POST'])
@login_required
@admin_required
def franchise_weekly_income_deposit(eid):
    """Record/update Cash Deposited on an existing entry — see
    franchise_daily_income_deposit."""
    entry = FranchiseWeeklyIncome.query.filter_by(id=eid).first_or_404()
    entry.deposited = form_float(request.form, 'deposited', label='Cash deposited', required=False, default=0)
    log_audit('UPDATE', 'franchise_weekly_income', entry.id,
              f'Cash deposited for week of {entry.week_start} set to {entry.deposited}')
    touch_sync_fields(entry)
    db.session.commit()
    flash('Cash deposited recorded.', 'success')
    return redirect(url_for('franchise_weekly_income_list', vehicle_id=entry.vehicle_id))


@app.route('/franchise/weekly-income/<int:eid>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
@handle_form_errors
def franchise_weekly_income_edit(eid):
    """Full edit of an existing entry — see franchise_daily_income_edit."""
    entry = FranchiseWeeklyIncome.query.filter_by(id=eid).first_or_404()
    if request.method == 'POST':
        entry.week_start = parse_date(request.form['week_start'])
        if entry.vehicle_id:
            entry.income = form_float(request.form, 'income', required=False, default=0)
        else:
            for f, lbl in FRANCHISE_INCOME_EXPENSE_FIELDS:
                setattr(entry, f, form_float(request.form, f, label=lbl, required=False, default=0))
            entry.other_expenditure = form_float(request.form, 'other_expenditure', required=False, default=0)
            entry.deposited = form_float(request.form, 'deposited', required=False, default=0)
        entry.description = request.form.get('description', '').strip()
        log_audit('UPDATE', 'franchise_weekly_income', entry.id, f'Updated weekly franchise entry for week of {entry.week_start}')
        touch_sync_fields(entry)
        db.session.commit()
        flash('Entry updated.', 'success')
        return redirect(url_for('franchise_weekly_income_list', vehicle_id=entry.vehicle_id))
    return render_template('franchise/income_entry_edit.html', entry=entry, date_field='week_start',
                           date_value=entry.week_start, date_label='Date',
                           list_endpoint='franchise_weekly_income_list', edit_endpoint='franchise_weekly_income_edit')


@app.route('/franchise/weekly-income/export')
@login_required
@permission_required('franchise')
def franchise_weekly_income_export():
    vehicle_id = request.args.get('vehicle_id', type=int)
    vehicle = FranchiseVehicle.query.filter_by(id=vehicle_id).first() if vehicle_id else None
    if vehicle_id and not vehicle:
        flash('Select a valid franchise vehicle to export.', 'danger')
        return redirect(url_for('franchise_weekly_income_list'))

    period, df, dt = resolve_ledger_period(request.args.get('period', 'month'), date.today())
    entries = list(reversed(_franchise_income_period_rows(FranchiseWeeklyIncome, 'week_start', vehicle, df, dt)))

    header = ['Week Of', 'Income', 'Traffic Fines', 'Facilitation Fees', 'Workshop', 'Wages',
              'Other Expenditure', 'Total Expenditure', 'Cash Deposited', 'Variance', 'Description']
    rows = [[e.week_start, f'{e.income:.2f}', f'{e.exp_traffic_fines:.2f}', f'{e.exp_facilitation_fees:.2f}',
             f'{e.exp_workshop:.2f}', f'{e.exp_wages:.2f}', f'{e.other_expenditure:.2f}',
             f'{e.total_expenditure:.2f}', f'{e.deposited:.2f}', f'{e.variance:.2f}', e.description or '']
            for e in entries]

    label = vehicle.number_plate.replace(' ', '_') if vehicle else 'shared_expenditure'
    if request.args.get('format') == 'pdf':
        return _table_pdf_response(f'{label}_weekly_income_{period}_{date.today()}.pdf',
            'Franchise Weekly Income', f'Period: {df} to {dt}', header, rows)
    return csv_export_response(f'{label}_weekly_income_{period}_{date.today()}.csv', header, rows)


@app.route('/franchise/weekly-income/<int:eid>/delete', methods=['POST'])
@login_required
@admin_required
def franchise_weekly_income_delete(eid):
    entry = FranchiseWeeklyIncome.query.filter_by(id=eid).first_or_404()
    log_audit('DELETE', 'franchise_weekly_income', eid, f'Deleted weekly franchise income entry for week of {entry.week_start}')
    entry.deleted_at = datetime.now(timezone.utc)
    touch_sync_fields(entry)
    db.session.commit()
    flash('Weekly franchise income entry deleted.', 'warning')
    return redirect(url_for('franchise_weekly_income_list', vehicle_id=entry.vehicle_id))


# ─────────────────────────────────────────────────────────────
# Franchise Operational Expenses — franchise-wide costs (rent, admin
# salaries, etc.) under a single 'Operational Expenses' heading with
# manageable sub-headings (FranchiseExpenseCategory). Feed only into the
# Consolidated P&L's Net Profit (see report_franchise_consolidated) — never
# the Daily/Weekly Income lists or Cash Reconciliation, since no vehicle
# handles this cash. See FranchiseOperationalExpense's docstring.
# ─────────────────────────────────────────────────────────────
@app.route('/franchise/operational-expenses')
@login_required
@permission_required('franchise')
def franchise_operational_expenses_list():
    df, dt = query_date_range()
    expenses = FranchiseOperationalExpense.query.filter(
        FranchiseOperationalExpense.expense_date.between(df, dt)
    ).order_by(FranchiseOperationalExpense.expense_date.desc()).all()
    categories = FranchiseExpenseCategory.query.order_by(FranchiseExpenseCategory.name).all()
    total = sum(e.amount for e in expenses)
    return render_template('franchise/operational_expenses.html', expenses=expenses, categories=categories,
                           total=total, date_from=df.strftime('%Y-%m-%d'), date_to=dt.strftime('%Y-%m-%d'),
                           today=date.today().strftime('%Y-%m-%d'))


@app.route('/franchise/operational-expenses/export')
@login_required
@permission_required('franchise')
def franchise_operational_expenses_export():
    df, dt = query_date_range()
    expenses = FranchiseOperationalExpense.query.filter(
        FranchiseOperationalExpense.expense_date.between(df, dt)
    ).order_by(FranchiseOperationalExpense.expense_date.asc()).all()
    rows = [[e.expense_date, e.category.name, e.description or '', f'{e.amount:.2f}'] for e in expenses]
    header = ['Date', 'Sub-heading', 'Description', 'Amount']
    if request.args.get('format') == 'pdf':
        return _table_pdf_response(f'franchise_operational_expenses_{df}_to_{dt}.pdf',
            'Franchise Operational Expenses', f'Period: {df} to {dt}', header, rows)
    return csv_export_response(f'franchise_operational_expenses_{df}_to_{dt}.csv', header, rows)


@app.route('/franchise/operational-expenses/add', methods=['POST'])
@login_required
@permission_required('franchise')
@handle_form_errors
def franchise_operational_expense_add():
    client_id = request.form.get('_client_id')
    if already_synced(client_id):
        flash('Already recorded.', 'info')
        return redirect(url_for('franchise_operational_expenses_list'))
    e = FranchiseOperationalExpense(
        expense_date=parse_date(request.form['expense_date']),
        category_id=form_int(request.form, 'category_id'),
        amount=form_float(request.form, 'amount', min_value=0),
        description=request.form.get('description', '').strip(),
        created_by=current_user.id,
    )
    db.session.add(e)
    db.session.flush()
    log_audit('CREATE', 'franchise_operational_expenses', e.id, f'Operational expense of {e.amount}')
    record_offline_sync(client_id, 'franchise_operational_expense_add')
    touch_sync_fields(e)
    db.session.commit()
    flash('Operational expense recorded.', 'success')
    return redirect(url_for('franchise_operational_expenses_list'))


@app.route('/franchise/operational-expenses/<int:eid>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
@handle_form_errors
def franchise_operational_expense_edit(eid):
    e = FranchiseOperationalExpense.query.filter_by(id=eid).first_or_404()
    if request.method == 'POST':
        e.expense_date = parse_date(request.form['expense_date'])
        e.category_id = form_int(request.form, 'category_id')
        e.amount = form_float(request.form, 'amount', min_value=0)
        e.description = request.form.get('description', '').strip()
        log_audit('UPDATE', 'franchise_operational_expenses', e.id, f'Updated operational expense of {e.amount}')
        touch_sync_fields(e)
        db.session.commit()
        flash('Operational expense updated.', 'success')
        return redirect(url_for('franchise_operational_expenses_list'))
    categories = FranchiseExpenseCategory.query.order_by(FranchiseExpenseCategory.name).all()
    return render_template('franchise/operational_expense_edit.html', expense=e, categories=categories)


@app.route('/franchise/operational-expenses/<int:eid>/delete', methods=['POST'])
@login_required
@admin_required
def franchise_operational_expense_delete(eid):
    e = FranchiseOperationalExpense.query.filter_by(id=eid).first_or_404()
    log_audit('DELETE', 'franchise_operational_expenses', eid, f'Deleted operational expense of {e.amount}')
    e.deleted_at = datetime.now(timezone.utc)
    touch_sync_fields(e)
    db.session.commit()
    flash('Operational expense deleted.', 'warning')
    return redirect(url_for('franchise_operational_expenses_list'))


@app.route('/franchise/operational-expense-categories/add', methods=['POST'])
@login_required
@permission_required('franchise')
@handle_form_errors
def franchise_expense_category_add():
    name = request.form.get('name', '').strip()
    if not name:
        raise ValueError('Sub-heading name is required.')
    existing = FranchiseExpenseCategory.query.filter_by(name=name).first()
    if existing:
        flash(f'"{name}" already exists.', 'warning')
    else:
        cat = FranchiseExpenseCategory(name=name)
        db.session.add(cat)
        db.session.flush()
        log_audit('CREATE', 'franchise_expense_categories', cat.id, f'Added operational expense sub-heading {cat.name}')
        touch_sync_fields(cat)
        db.session.commit()
        flash(f'"{cat.name}" added.', 'success')
    return redirect(url_for('franchise_operational_expenses_list'))


@app.route('/franchise/operational-expense-categories/<int:cid>/delete', methods=['POST'])
@login_required
@admin_required
def franchise_expense_category_delete(cid):
    cat = FranchiseExpenseCategory.query.filter_by(id=cid).first_or_404()
    if FranchiseOperationalExpense.query.filter_by(category_id=cid).first():
        flash(f'Cannot delete "{cat.name}" — operational expenses are recorded against it.', 'danger')
        return redirect(url_for('franchise_operational_expenses_list'))
    name = cat.name
    log_audit('DELETE', 'franchise_expense_categories', cid, f'Deleted operational expense sub-heading {name}')
    cat.deleted_at = datetime.now(timezone.utc)
    touch_sync_fields(cat)
    db.session.commit()
    flash(f'"{name}" deleted.', 'warning')
    return redirect(url_for('franchise_operational_expenses_list'))


def _franchise_income_kind(kind):
    kind = kind if kind in ('daily', 'weekly') else 'daily'
    list_endpoint = 'franchise_daily_income_list' if kind == 'daily' else 'franchise_weekly_income_list'
    model_cls = FranchiseDailyIncome if kind == 'daily' else FranchiseWeeklyIncome
    date_field = 'entry_date' if kind == 'daily' else 'week_start'
    return kind, list_endpoint, model_cls, date_field


@app.route('/franchise/income/import/preview', methods=['POST'])
@login_required
@permission_required('franchise')
def franchise_income_import_preview():
    """Two-step confirmed-mapping import for the Daily/Weekly Income pages,
    same flow as the Daily Transactions ledger import. On a specific vehicle
    tab, that vehicle is implied (no Vehicle column to map), mirroring the
    ledger's single-vehicle import. On the Whole Franchise tab there's no
    single vehicle to imply, so the full field list (including Vehicle /
    Franchisee) is offered instead — a flat file with many vehicles' rows in
    it can be imported in one pass, each row picking its own vehicle (or
    staying whole-franchise when left blank). 'kind' picks which page this
    import targets."""
    kind, list_endpoint, _model_cls, _date_field = _franchise_income_kind(request.form.get('kind', 'daily'))
    period = request.form.get('period', 'month')
    vehicle_id = form_int(request.form, 'vehicle_id', required=False)
    vehicle = FranchiseVehicle.query.filter_by(id=vehicle_id).first() if vehicle_id else None
    if vehicle_id and not vehicle:
        flash('Select a valid franchise vehicle before importing.', 'danger')
        return redirect(url_for(list_endpoint, period=period))
    fields = CANONICAL_FRANCHISE_INCOME_FIELDS_SCOPED if vehicle else CANONICAL_FRANCHISE_INCOME_FIELDS

    file = request.files.get('file')
    if file and file.filename:
        filename = file.filename
        try:
            headers, raw_rows = read_uploaded_table(file)
        except ValueError as e:
            flash(str(e), 'danger')
            return redirect(url_for(list_endpoint, vehicle_id=vehicle_id, period=period))
        mapping = auto_map_columns(headers, fields=fields)
    else:
        # Re-preview after the user adjusted the mapping — the file itself
        # isn't resubmitted, the previously parsed rows travel via raw_data.
        try:
            filename = request.form.get('filename', 'uploaded file')
            payload = json.loads(request.form.get('raw_data') or '{}')
            headers, raw_rows = payload.get('headers') or [], payload.get('rows') or []
            if not headers or not raw_rows:
                raise ValueError('Choose a CSV or Excel file to import.')
            mapping = {field_key: (request.form.get(f'map_{field_key}') or None)
                       for field_key, _label, _syn in fields}
        except (ValueError, json.JSONDecodeError, TypeError):
            flash('Choose a CSV or Excel file to import — the previous preview session expired.', 'danger')
            return redirect(url_for(list_endpoint, vehicle_id=vehicle_id, period=period))

    if not raw_rows:
        flash('That file has no data rows to import — it only has a header row. '
              'Add rows with a Date and Income/expense figures, then re-import.', 'warning')
        return redirect(url_for(list_endpoint, vehicle_id=vehicle_id, period=period))

    preview_rows = apply_column_mapping(headers, raw_rows[:10], mapping, row_key_map=FRANCHISE_INCOME_ROW_KEY_MAP)
    return render_template('franchise/income_import_preview.html',
                           kind=kind, vehicle=vehicle, vehicle_id=vehicle_id, period=period, filename=filename,
                           headers=headers, mapping=mapping, fields=fields,
                           preview_rows=preview_rows, row_count=len(raw_rows),
                           raw_data=json.dumps({'headers': headers, 'rows': raw_rows}))


@app.route('/franchise/income/import/confirm', methods=['POST'])
@login_required
@permission_required('franchise')
def franchise_income_import_confirm():
    kind, list_endpoint, model_cls, date_field = _franchise_income_kind(request.form.get('kind', 'daily'))
    period = request.form.get('period', 'month')
    vehicle_id = form_int(request.form, 'vehicle_id', required=False)
    vehicle = FranchiseVehicle.query.filter_by(id=vehicle_id).first() if vehicle_id else None
    if vehicle_id and not vehicle:
        flash('Select a valid franchise vehicle before importing.', 'danger')
        return redirect(url_for(list_endpoint, period=period))
    fields = CANONICAL_FRANCHISE_INCOME_FIELDS_SCOPED if vehicle else CANONICAL_FRANCHISE_INCOME_FIELDS

    filename = request.form.get('filename', 'uploaded file')
    try:
        payload = json.loads(request.form.get('raw_data') or '{}')
        headers, raw_rows = payload.get('headers') or [], payload.get('rows') or []
        if not raw_rows:
            raise ValueError('empty')
    except (ValueError, json.JSONDecodeError, TypeError):
        flash('That preview session expired — please choose the file again.', 'danger')
        return redirect(url_for(list_endpoint, vehicle_id=vehicle_id, period=period))

    mapping = {field_key: (request.form.get(f'map_{field_key}') or None)
               for field_key, _label, _syn in fields}
    file_rows = apply_column_mapping(headers, raw_rows, mapping, row_key_map=FRANCHISE_INCOME_ROW_KEY_MAP)
    imported, errors, error_rows, _created_vehicles, created_records = import_franchise_income_rows(
        file_rows, model_cls, date_field, week_normalize=(kind == 'weekly'), forced_vehicle=vehicle)

    label = vehicle.number_plate if vehicle else 'the whole franchise'
    if imported or error_rows:
        # Commit even when imported == 0: a batch made only of failed rows
        # still needs to persist so its quarantine CSV can be downloaded.
        save_import_batch(model_cls.__tablename__, filename, len(raw_rows), imported, error_rows, created_records)
        if imported:
            log_audit('CREATE', model_cls.__tablename__, None,
                      f'Imported {imported} {kind} income row(s) for {label} from {filename}')
        db.session.commit()
    else:
        db.session.rollback()

    if imported:
        flash(f'Imported {imported} {kind} income row(s) for {label}.', 'success')
    if errors:
        shown = errors[:10]
        more = f' (+{len(errors) - 10} more)' if len(errors) > 10 else ''
        flash('Skipped rows — ' + '; '.join(shown) + more, 'warning')
    if not imported and not errors:
        flash('No rows found to import.', 'warning')

    return redirect(url_for(list_endpoint, vehicle_id=vehicle_id, period=period))


@app.route('/franchise/income/import/bulk', methods=['POST'])
@login_required
@permission_required('franchise')
def franchise_income_import_bulk():
    """Import a franchise-wide workbook with one sheet per franchise
    vehicle (sheet name matched against each vehicle's number plate) in a
    single pass — the Franchise Income counterpart of driver_ledger_import_bulk.
    'kind' picks Daily vs Weekly Income."""
    kind, list_endpoint, model_cls, date_field = _franchise_income_kind(request.form.get('kind', 'daily'))
    period = request.form.get('period', 'month')

    file = request.files.get('file')
    if not file or not file.filename:
        flash('Choose an Excel workbook to import.', 'danger')
        return redirect(url_for(list_endpoint, period=period))

    try:
        sheets = read_uploaded_workbook_sheets(file)
    except ValueError as e:
        flash(str(e), 'danger')
        return redirect(url_for(list_endpoint, period=period))

    if not sheets:
        flash('That workbook has no sheets with data rows to import.', 'warning')
        return redirect(url_for(list_endpoint, period=period))

    vehicles_by_plate = {_normalize_registration(v.number_plate): v for v in FranchiseVehicle.query.all()}

    results = []
    for sheet_name, (headers, rows) in sheets.items():
        savepoint = db.session.begin_nested()
        try:
            mapping = auto_map_columns(headers, fields=CANONICAL_FRANCHISE_INCOME_FIELDS_SCOPED)
            vehicle = vehicles_by_plate.get(_normalize_registration(sheet_name))

            # A real per-vehicle income sheet always has both a Date and an
            # Income column — that combination is what tells this apart from
            # an unrelated sheet.
            looks_like_income_sheet = bool(mapping.get('date')) and bool(mapping.get('income'))

            if not vehicle:
                # Vehicles are never auto-registered from an import — register
                # this plate under Franchise Vehicles first, then re-import.
                reason = ('No registered franchise vehicle matches this sheet name — register it under '
                          'Franchise Vehicles first.' if looks_like_income_sheet else
                          "Doesn't look like a per-vehicle income sheet.")
                results.append({'sheet': sheet_name, 'vehicle': None, 'mapping': {},
                                 'imported': 0, 'errors': [], 'skip_reason': reason})
                savepoint.commit()
                continue

            if not mapping.get('date'):
                results.append({'sheet': sheet_name, 'vehicle': vehicle, 'mapping': mapping,
                                 'imported': 0, 'errors': [], 'skip_reason': 'No Date column detected.'})
                savepoint.commit()
                continue

            mapped_rows = apply_column_mapping(headers, rows, mapping, row_key_map=FRANCHISE_INCOME_ROW_KEY_MAP)
            imported, errors, _error_rows, _created_vehicles, _created_records = import_franchise_income_rows(
                mapped_rows, model_cls, date_field, week_normalize=(kind == 'weekly'), forced_vehicle=vehicle)
            if imported:
                log_audit('CREATE', model_cls.__tablename__, None,
                          f'Imported {imported} {kind} income row(s) for {vehicle.number_plate} '
                          f'from {file.filename} (sheet "{sheet_name}")')
            results.append({'sheet': sheet_name, 'vehicle': vehicle, 'mapping': mapping,
                             'imported': imported, 'errors': errors, 'skip_reason': None})
            savepoint.commit()
        except Exception as e:
            savepoint.rollback()
            results.append({'sheet': sheet_name, 'vehicle': None, 'mapping': {},
                             'imported': 0, 'errors': [], 'skip_reason': f'Unexpected error: {e}'})

    total_imported = sum(r['imported'] for r in results)
    if total_imported:
        db.session.commit()
    else:
        db.session.rollback()

    return render_template('franchise/income_bulk_import_result.html',
                           filename=file.filename, results=results, kind=kind, list_endpoint=list_endpoint,
                           total_imported=total_imported, period=period,
                           fields=CANONICAL_FRANCHISE_INCOME_FIELDS_SCOPED)


@app.route('/franchise/income/import/reconciliation', methods=['POST'])
@login_required
@permission_required('franchise')
def franchise_income_import_reconciliation():
    """Import a franchise monthly logbook workbook — one sheet per month,
    each holding a 'FRANCHISE COLLECTION RECONCILIATION SCHEDULE' block
    (whole-franchise daily income/expenses/deposits) alongside other,
    unrelated tables on the same sheet. Unlike the generic column-mapping
    importer above, this format is located and parsed automatically (see
    _find_franchise_reconciliation_block) since its shape is fixed and
    known, just shifted in position — there's no column mapping to confirm.
    Always Daily Income, whole-franchise: the schedule is inherently one row
    per date, not per vehicle or per week."""
    period = request.form.get('period', 'month')
    file = request.files.get('file')
    if not file or not file.filename:
        flash('Choose an Excel workbook to import.', 'danger')
        return redirect(url_for('franchise_daily_income_list', period=period))

    try:
        sheets = read_uploaded_workbook_raw_sheets(file)
    except ValueError as e:
        flash(str(e), 'danger')
        return redirect(url_for('franchise_daily_income_list', period=period))

    if not sheets:
        flash('That workbook has no sheets with data to import.', 'warning')
        return redirect(url_for('franchise_daily_income_list', period=period))

    results = []
    for sheet_name, raw_rows in sheets.items():
        savepoint = db.session.begin_nested()
        try:
            found_rows = _find_franchise_reconciliation_block(raw_rows)
            if found_rows is None:
                results.append({'sheet': sheet_name, 'rows_found': 0, 'imported': 0,
                                 'errors': [], 'skip_reason': 'No reconciliation schedule found on this sheet.'})
                savepoint.commit()
                continue
            if not found_rows:
                results.append({'sheet': sheet_name, 'rows_found': 0, 'imported': 0,
                                 'errors': [], 'skip_reason': 'Schedule found, but it has no dated rows.'})
                savepoint.commit()
                continue

            imported, errors, error_rows, _created_vehicles, created_records = import_franchise_income_rows(
                found_rows, FranchiseDailyIncome, 'entry_date', week_normalize=False, forced_vehicle=None)
            if imported or error_rows:
                save_import_batch('franchise_daily_income', f'{file.filename} ({sheet_name})',
                                  len(found_rows), imported, error_rows, created_records)
            if imported:
                log_audit('CREATE', 'franchise_daily_income', None,
                          f'Imported {imported} daily income row(s) for the whole franchise '
                          f'from {file.filename} (sheet "{sheet_name}")')
            results.append({'sheet': sheet_name, 'rows_found': len(found_rows), 'imported': imported,
                             'errors': errors, 'skip_reason': None})
            savepoint.commit()
        except Exception as e:
            savepoint.rollback()
            results.append({'sheet': sheet_name, 'rows_found': 0, 'imported': 0,
                             'errors': [], 'skip_reason': f'Unexpected error: {e}'})

    total_imported = sum(r['imported'] for r in results)
    if total_imported:
        db.session.commit()
    else:
        db.session.rollback()

    return render_template('franchise/income_reconciliation_import_result.html',
                           filename=file.filename, results=results, period=period,
                           total_imported=total_imported)


# ─────────────────────────────────────────────────────────────
# Franchise Vehicles — the registry of third-party vehicles paying to
# operate under the franchise, and what each one has paid.
# ─────────────────────────────────────────────────────────────
@app.route('/franchise/vehicles')
@login_required
@permission_required('franchise')
def franchise_vehicles():
    vehicles = FranchiseVehicle.query.order_by(FranchiseVehicle.franchisee_name).all()
    return render_template('franchise/vehicles.html', vehicles=vehicles)


@app.route('/franchise/vehicles/add', methods=['GET', 'POST'])
@login_required
@permission_required('franchise')
@handle_form_errors
def franchise_vehicle_add():
    if request.method == 'POST':
        number_plate = request.form.get('number_plate', '').strip().upper()
        if not number_plate:
            raise ValueError('Number plate is required.')
        check_unique(FranchiseVehicle, 'number_plate', number_plate, label='Number plate')
        vehicle = FranchiseVehicle(
            number_plate=number_plate,
            franchisee_name=request.form.get('franchisee_name', '').strip(),
            status=request.form.get('status', 'active'),
            daily_fee=form_float(request.form, 'daily_fee', label='Daily fee', required=False, min_value=0),
            weekly_fee=form_float(request.form, 'weekly_fee', label='Weekly fee', required=False, min_value=0),
            amount_owed=form_float(request.form, 'amount_owed', label='Amount owed', required=False, default=0, min_value=0),
            notes=request.form.get('notes', '').strip(),
        )
        if not vehicle.franchisee_name:
            raise ValueError('Franchisee name is required.')
        db.session.add(vehicle)
        db.session.flush()
        log_audit('CREATE', 'franchise_vehicles', vehicle.id,
                  f'Added franchise vehicle {vehicle.number_plate} ({vehicle.franchisee_name})')
        touch_sync_fields(vehicle)
        db.session.commit()
        flash('Franchise vehicle added.', 'success')
        return redirect(url_for('franchise_vehicles'))
    return render_template('franchise/vehicle_form.html', vehicle=None, action='Add')


@app.route('/franchise/vehicles/quick-add', methods=['POST'])
@login_required
@permission_required_any('franchise', 'franchise_entry')
def franchise_vehicle_quick_add():
    """Inline "+ New Vehicle" registration from the Daily/Weekly Income
    pages (both the By Vehicle picker and the By Date/Week all-vehicles
    listing) — plate + franchisee name only, no fees/status/notes, so a
    manager can register a new franchisee on the spot without leaving the
    income form. Flagged pending_review so an admin follows up with the
    full details (see franchise_vehicle_review) and alerted via the
    Vehicles nav badge, WhatsApp (if linked), and the audit log.

    POST-only, no GET counterpart at this URL (see franchise_daily_income_
    add) — errors are handled locally rather than via @handle_form_errors,
    whose generic redirect(request.url) would bounce a rejected submission
    (e.g. a plate that's already registered) back onto this same POST-only
    path and 405 instead of showing the error, taking the duplicate-plate
    message from check_unique with it."""
    # A franchise_entry-only clerk has no access to the Daily/Weekly Income
    # list pages at all (see franchise_my_collections) — quick-add reached
    # from their own page always bounces back there instead.
    entry_only = not current_user.has_permission('franchise')
    kind = request.form.get('kind') if request.form.get('kind') in ('daily', 'weekly') else 'daily'
    period = request.form.get('period', 'month')
    view = request.form.get('view', '')
    endpoint = 'franchise_weekly_income_list' if kind == 'weekly' else 'franchise_daily_income_list'

    def back_to_list(**extra):
        if entry_only:
            return redirect(url_for('franchise_my_collections'))
        params = {'period': period, **extra}
        if view == 'date':
            params['view'] = 'date'
            if kind == 'weekly':
                params['on_week'] = request.form.get('on_week', '')
            else:
                params['on_date'] = request.form.get('on_date', '')
        return redirect(url_for(endpoint, **params))

    number_plate = request.form.get('number_plate', '').strip().upper()
    franchisee_name = request.form.get('franchisee_name', '').strip()
    try:
        if not number_plate:
            raise ValueError('Number plate is required.')
        if not franchisee_name:
            raise ValueError('Franchisee name is required.')
        check_unique(FranchiseVehicle, 'number_plate', number_plate, label='Number plate')

        vehicle = FranchiseVehicle(number_plate=number_plate, franchisee_name=franchisee_name,
                                   pending_review=True)
        db.session.add(vehicle)
        db.session.flush()
        log_audit('CREATE', 'franchise_vehicles', vehicle.id,
                  f'Quick-registered franchise vehicle {vehicle.number_plate} ({vehicle.franchisee_name}) '
                  f'from {kind} income — pending admin review')
        touch_sync_fields(vehicle)
        db.session.commit()
    except ValueError as e:
        db.session.rollback()
        flash(str(e), 'danger')
        return back_to_list()

    notify_admins_whatsapp(
        f'New franchise vehicle registered by {current_user.username}: '
        f'{vehicle.number_plate} ({vehicle.franchisee_name}). '
        f'No fees/status set yet — review under Franchise > Vehicles.'
    )

    flash(f'Vehicle {vehicle.number_plate} registered — an admin will review and complete its details.', 'success')
    # From the By Date/Week listing, stay there (the new vehicle now shows
    # up in it, ready to have today's/this week's entry filled in) rather
    # than jumping to its own, currently-empty By Vehicle tab.
    if entry_only or view == 'date':
        return back_to_list()
    return redirect(url_for(endpoint, vehicle_id=vehicle.id, period=period))


@app.route('/franchise/vehicles/<int:vid>/review', methods=['POST'])
@login_required
@admin_required
def franchise_vehicle_review(vid):
    """Clears pending_review once an admin has checked/completed a
    quick-registered vehicle's details — the counterpart to
    franchise_vehicle_quick_add."""
    vehicle = FranchiseVehicle.query.filter_by(id=vid).first_or_404()
    vehicle.pending_review = False
    log_audit('UPDATE', 'franchise_vehicles', vehicle.id, f'Marked {vehicle.number_plate} as reviewed')
    touch_sync_fields(vehicle)
    db.session.commit()
    flash(f'{vehicle.number_plate} marked as reviewed.', 'success')
    return redirect(url_for('franchise_vehicles'))


@app.route('/franchise/vehicles/<int:vid>/edit', methods=['GET', 'POST'])
@login_required
@permission_required('franchise')
@handle_form_errors
def franchise_vehicle_edit(vid):
    vehicle = FranchiseVehicle.query.filter_by(id=vid).first_or_404()
    if request.method == 'POST':
        number_plate = request.form.get('number_plate', '').strip().upper()
        if not number_plate:
            raise ValueError('Number plate is required.')
        check_unique(FranchiseVehicle, 'number_plate', number_plate, label='Number plate', exclude_id=vehicle.id)
        franchisee_name = request.form.get('franchisee_name', '').strip()
        if not franchisee_name:
            raise ValueError('Franchisee name is required.')
        vehicle.number_plate = number_plate
        vehicle.franchisee_name = franchisee_name
        vehicle.status = request.form.get('status', 'active')
        vehicle.daily_fee = form_float(request.form, 'daily_fee', label='Daily fee', required=False, min_value=0)
        vehicle.weekly_fee = form_float(request.form, 'weekly_fee', label='Weekly fee', required=False, min_value=0)
        vehicle.amount_owed = form_float(request.form, 'amount_owed', label='Amount owed', required=False, default=0, min_value=0)
        vehicle.notes = request.form.get('notes', '').strip()
        log_audit('UPDATE', 'franchise_vehicles', vehicle.id, f'Updated franchise vehicle {vehicle.number_plate}')
        touch_sync_fields(vehicle)
        db.session.commit()
        flash('Franchise vehicle updated.', 'success')
        return redirect(url_for('franchise_vehicles'))
    return render_template('franchise/vehicle_form.html', vehicle=vehicle, action='Edit')


@app.route('/franchise/vehicles/<int:vid>/delete', methods=['POST'])
@login_required
@admin_required
def franchise_vehicle_delete(vid):
    vehicle = FranchiseVehicle.query.filter_by(id=vid).first_or_404()
    has_history = (FranchiseCollection.query.filter_by(vehicle_id=vid).first()
                   or FranchiseDailyIncome.query.filter_by(vehicle_id=vid).first()
                   or FranchiseWeeklyIncome.query.filter_by(vehicle_id=vid).first())
    if has_history:
        flash('Cannot delete this vehicle — it has income/collection history. Mark it Inactive instead.', 'danger')
        return redirect(url_for('franchise_vehicles'))
    log_audit('DELETE', 'franchise_vehicles', vid, f'Deleted franchise vehicle {vehicle.number_plate}')
    vehicle.deleted_at = datetime.now(timezone.utc)
    touch_sync_fields(vehicle)
    db.session.commit()
    flash('Franchise vehicle deleted.', 'warning')
    return redirect(url_for('franchise_vehicles'))


@app.route('/franchise/vehicles/import/preview', methods=['POST'])
@login_required
@permission_required('franchise')
def franchise_vehicles_import_preview():
    """Same two-step flow as the Collections/Income imports: parse the file,
    auto-map its columns onto the canonical vehicle-registration fields, and
    show a confirmation/adjustment step before writing anything."""
    file = request.files.get('file')
    if file and file.filename:
        filename = file.filename
        try:
            headers, raw_rows = read_uploaded_table(file)
        except ValueError as e:
            flash(str(e), 'danger')
            return redirect(url_for('franchise_vehicles'))
        mapping = auto_map_columns(headers, fields=CANONICAL_FRANCHISE_VEHICLE_FIELDS)
    else:
        try:
            filename = request.form.get('filename', 'uploaded file')
            payload = json.loads(request.form.get('raw_data') or '{}')
            headers, raw_rows = payload.get('headers') or [], payload.get('rows') or []
            if not headers or not raw_rows:
                raise ValueError('Choose a CSV or Excel file to import.')
            mapping = {field_key: (request.form.get(f'map_{field_key}') or None)
                       for field_key, _label, _syn in CANONICAL_FRANCHISE_VEHICLE_FIELDS}
        except (ValueError, json.JSONDecodeError, TypeError):
            flash('Choose a CSV or Excel file to import — the previous preview session expired.', 'danger')
            return redirect(url_for('franchise_vehicles'))

    if not raw_rows:
        flash('That file has no data rows to import — it only has a header row. '
              'Add rows with at least a Vehicle / Number Plate, then re-import.', 'warning')
        return redirect(url_for('franchise_vehicles'))

    preview_rows = apply_column_mapping(headers, raw_rows[:10], mapping, row_key_map=FRANCHISE_VEHICLE_ROW_KEY_MAP)
    return render_template('franchise/vehicles_import_preview.html',
                           filename=filename,
                           headers=headers, mapping=mapping, fields=CANONICAL_FRANCHISE_VEHICLE_FIELDS,
                           preview_rows=preview_rows, row_count=len(raw_rows),
                           raw_data=json.dumps({'headers': headers, 'rows': raw_rows}))


@app.route('/franchise/vehicles/import/confirm', methods=['POST'])
@login_required
@permission_required('franchise')
def franchise_vehicles_import_confirm():
    filename = request.form.get('filename', 'uploaded file')
    try:
        payload = json.loads(request.form.get('raw_data') or '{}')
        headers, raw_rows = payload.get('headers') or [], payload.get('rows') or []
        if not raw_rows:
            raise ValueError('empty')
    except (ValueError, json.JSONDecodeError, TypeError):
        flash('That preview session expired — please choose the file again.', 'danger')
        return redirect(url_for('franchise_vehicles'))

    mapping = {field_key: (request.form.get(f'map_{field_key}') or None)
               for field_key, _label, _syn in CANONICAL_FRANCHISE_VEHICLE_FIELDS}
    file_rows = apply_column_mapping(headers, raw_rows, mapping, row_key_map=FRANCHISE_VEHICLE_ROW_KEY_MAP)
    imported, errors, error_rows, created_records = import_franchise_vehicle_rows(file_rows)

    if imported or error_rows:
        save_import_batch('franchise_vehicles', filename, len(raw_rows), imported, error_rows, created_records)
        if imported:
            log_audit('CREATE', 'franchise_vehicles', None,
                      f'Registered/updated {imported} franchise vehicle(s) from {filename}')
        db.session.commit()
    else:
        db.session.rollback()

    if imported:
        flash(f'Registered/updated {imported} franchise vehicle(s).', 'success')
    if errors:
        shown = errors[:10]
        more = f' (+{len(errors) - 10} more)' if len(errors) > 10 else ''
        flash('Skipped rows — ' + '; '.join(shown) + more, 'warning')
    if not imported and not errors:
        flash('No rows found to import.', 'warning')

    return redirect(url_for('franchise_vehicles'))


def _income_entry_totals(entries):
    """Sum a list of FranchiseDailyIncome or FranchiseWeeklyIncome entries
    (identical shape) into the same fields as a single entry, so a period
    total can be rendered with the same formatting as one row."""
    return dict(
        income=sum(e.income for e in entries),
        exp_traffic_fines=sum(e.exp_traffic_fines for e in entries),
        exp_facilitation_fees=sum(e.exp_facilitation_fees for e in entries),
        exp_workshop=sum(e.exp_workshop for e in entries),
        exp_wages=sum(e.exp_wages for e in entries),
        other_expenditure=sum(e.other_expenditure for e in entries),
        total_expenditure=sum(e.total_expenditure for e in entries),
        cash_in_hand=sum(e.cash_in_hand for e in entries),
        deposited=sum(e.deposited for e in entries),
        variance=sum(e.variance for e in entries),
    )


def _group_income_by_period(entries, period_attr):
    """Group entries (FranchiseDailyIncome or FranchiseWeeklyIncome) by
    their date/week_start, summing every vehicle's figures into one row per
    period — matching the franchise's own paper reconciliation schedule,
    which records a single combined total per date/week rather than a
    per-vehicle breakdown. Returns rows sorted by period, each carrying the
    period value, how many vehicles fed into it, and the summed totals."""
    by_period = {}
    for e in entries:
        by_period.setdefault(getattr(e, period_attr), []).append(e)
    return [
        dict(period=period, vehicle_count=len(period_entries), **_income_entry_totals(period_entries))
        for period, period_entries in sorted(by_period.items())
    ]


def _apply_suspense_clearance(rows, source_type, date_from, date_to):
    """Zeroes out the variance on any row whose period has been cleared in
    the Suspense Account (see franchise_suspense_resolve) — once an admin
    has investigated and explained a discrepancy there, it's resolved and
    the reconciliation schedule should no longer report it as outstanding.
    The original figure is kept as raw_variance so cleared rows can still
    show what was cleared, and rows gain a `cleared` flag for display."""
    q = FranchiseSuspenseResolution.query.filter(FranchiseSuspenseResolution.source_type == source_type)
    if date_from:
        q = q.filter(FranchiseSuspenseResolution.period_date.between(date_from, date_to))
    resolutions = {r.period_date for r in q.all()}
    for r in rows:
        r['raw_variance'] = r['variance']
        r['cleared'] = r['period'] in resolutions
        if r['cleared']:
            r['variance'] = 0.0
    return rows


def _franchise_fleet_variance(model_cls, date_field, source_type, df, dt):
    """Fleet-wide Cash Variance for the Daily/Weekly Income pages' KPI card,
    with any period already cleared through the Suspense Account (see
    _apply_suspense_clearance) zeroed out of the total — otherwise a variance
    an admin has resolved there would keep reading as outstanding here,
    disagreeing with the Reconciliation Schedule for the same data."""
    col = getattr(model_cls, date_field)
    q = model_cls.query.filter(col.between(df, dt)) if df else model_cls.query
    rows = _apply_suspense_clearance(_group_income_by_period(q.all(), date_field), source_type, df, dt)
    return sum(r['variance'] for r in rows)


@app.route('/reports/franchise/reconciliation')
@login_required
@permission_required('franchise')
def report_franchise_reconciliation():
    """Reconciliation Schedule — daily and weekly franchise income are two
    independent streams, so this page shows two self-contained sections
    (each with its own per-date/week breakdown and subtotal) rather than
    one table with columns from both. Each row is a combined total across
    every vehicle for that date/week, matching the franchise's own paper
    schedule — per-vehicle detail lives on the Daily/Weekly Income list
    pages instead.

    Also carries a Day by Day comparison (daily income vs. weekly income vs.
    their combined total, by calendar date) — this used to be its own
    "Franchise Analysis" page, but it's the same two entities grouped a
    different way, so it's folded in here as a quick-scan summary ahead of
    the two full ledgers below it."""
    df, dt = query_date_range()
    daily_entries = FranchiseDailyIncome.query.filter(FranchiseDailyIncome.entry_date.between(df, dt)).all()
    weekly_entries = FranchiseWeeklyIncome.query.filter(FranchiseWeeklyIncome.week_start.between(df, dt)).all()

    daily_rows = _apply_suspense_clearance(_group_income_by_period(daily_entries, 'entry_date'), 'daily', df, dt)
    weekly_rows = _apply_suspense_clearance(_group_income_by_period(weekly_entries, 'week_start'), 'weekly', df, dt)

    daily_totals = _income_entry_totals(daily_entries)
    daily_totals['variance'] = sum(r['variance'] for r in daily_rows)
    weekly_totals = _income_entry_totals(weekly_entries)
    weekly_totals['variance'] = sum(r['variance'] for r in weekly_rows)
    combined_totals = {k: daily_totals[k] + weekly_totals[k] for k in daily_totals}
    combined_totals['net_profit'] = combined_totals['income'] - combined_totals['total_expenditure']

    days = {}
    for e in daily_entries:
        days.setdefault(e.entry_date, {'daily': 0.0, 'weekly': 0.0})['daily'] += e.income
    for e in weekly_entries:
        days.setdefault(e.week_start, {'daily': 0.0, 'weekly': 0.0})['weekly'] += e.income
    day_by_day_rows = [
        dict(entry_date=d, weekday=d.strftime('%A'), daily=b['daily'], weekly=b['weekly'],
             total=b['daily'] + b['weekly'])
        for d, b in sorted(days.items())
    ]
    day_by_day_totals = dict(daily=sum(r['daily'] for r in day_by_day_rows),
                             weekly=sum(r['weekly'] for r in day_by_day_rows),
                             total=sum(r['total'] for r in day_by_day_rows))

    return render_template('franchise/reconciliation.html', title='Franchise Collection Reconciliation Schedule',
                           daily_rows=daily_rows, daily_totals=daily_totals,
                           weekly_rows=weekly_rows, weekly_totals=weekly_totals,
                           combined_totals=combined_totals,
                           day_by_day_rows=day_by_day_rows, day_by_day_totals=day_by_day_totals,
                           date_from=df.strftime('%Y-%m-%d'), date_to=dt.strftime('%Y-%m-%d'))


@app.route('/reports/franchise/reconciliation/export')
@login_required
@permission_required('franchise')
def report_franchise_reconciliation_export():
    """Mirrors report_franchise_reconciliation's three sections exactly
    (Day by Day, Daily Reconciliation Schedule, Weekly Reconciliation
    Schedule) — the old version only ever exported the Day by Day summary,
    leaving out the two actual reconciliation schedules the page is named
    for."""
    df, dt = query_date_range()
    daily_entries = FranchiseDailyIncome.query.filter(FranchiseDailyIncome.entry_date.between(df, dt)).all()
    weekly_entries = FranchiseWeeklyIncome.query.filter(FranchiseWeeklyIncome.week_start.between(df, dt)).all()

    daily_rows = _apply_suspense_clearance(_group_income_by_period(daily_entries, 'entry_date'), 'daily', df, dt)
    weekly_rows = _apply_suspense_clearance(_group_income_by_period(weekly_entries, 'week_start'), 'weekly', df, dt)
    daily_totals = _income_entry_totals(daily_entries)
    daily_totals['variance'] = sum(r['variance'] for r in daily_rows)
    weekly_totals = _income_entry_totals(weekly_entries)
    weekly_totals['variance'] = sum(r['variance'] for r in weekly_rows)

    days = {}
    for e in daily_entries:
        days.setdefault(e.entry_date, {'daily': 0.0, 'weekly': 0.0})['daily'] += e.income
    for e in weekly_entries:
        days.setdefault(e.week_start, {'daily': 0.0, 'weekly': 0.0})['weekly'] += e.income
    day_by_day_rows = [[d, d.strftime('%A'), f"{b['daily']:.2f}", f"{b['weekly']:.2f}", f"{b['daily'] + b['weekly']:.2f}"]
                       for d, b in sorted(days.items())]
    day_totals = dict(daily=sum(b['daily'] for b in days.values()), weekly=sum(b['weekly'] for b in days.values()))

    schedule_header = ['Period', 'Income', 'Traffic Fines', 'Facilitation Fee', 'Workshop', 'Wages',
                       'Other', 'Net Income', 'Cash Deposited', 'Variance']

    def schedule_row(r):
        return [r['period'], f"{r['income']:.2f}", f"{r['exp_traffic_fines']:.2f}", f"{r['exp_facilitation_fees']:.2f}",
                f"{r['exp_workshop']:.2f}", f"{r['exp_wages']:.2f}", f"{r['other_expenditure']:.2f}",
                f"{r['cash_in_hand']:.2f}", f"{r['deposited']:.2f}", f"{r['variance']:.2f}"]

    def schedule_total_row(t):
        return ['TOTAL', f"{t['income']:.2f}", f"{t['exp_traffic_fines']:.2f}", f"{t['exp_facilitation_fees']:.2f}",
                f"{t['exp_workshop']:.2f}", f"{t['exp_wages']:.2f}", f"{t['other_expenditure']:.2f}",
                f"{t['cash_in_hand']:.2f}", f"{t['deposited']:.2f}", f"{t['variance']:.2f}"]

    if request.args.get('format') == 'pdf':
        styles = _pdf_styles()
        flowables = [
            Paragraph('Day by Day', styles['Heading3']), Spacer(1, 4),
            _pdf_table([['Date', 'Day', 'Daily Income', 'Weekly Income', 'Total Income']] + day_by_day_rows +
                      [['TOTAL', '', f"{day_totals['daily']:.2f}", f"{day_totals['weekly']:.2f}",
                        f"{day_totals['daily'] + day_totals['weekly']:.2f}"]]),
            Spacer(1, 14),
            Paragraph('Daily Reconciliation Schedule', styles['Heading3']), Spacer(1, 4),
            _pdf_table([schedule_header] + [schedule_row(r) for r in daily_rows] + [schedule_total_row(daily_totals)]),
            Spacer(1, 14),
            Paragraph('Weekly Reconciliation Schedule', styles['Heading3']), Spacer(1, 4),
            _pdf_table([schedule_header] + [schedule_row(r) for r in weekly_rows] + [schedule_total_row(weekly_totals)]),
        ]
        elements = _pdf_section('Franchise Collection Reconciliation Schedule', f'Period: {df} to {dt}', flowables)
        return _pdf_response(f'franchise_reconciliation_{df}_to_{dt}.pdf', elements, pagesize=landscape(A4))

    out = io.StringIO()
    w = csv.writer(out)
    w.writerow(['FRANCHISE COLLECTION RECONCILIATION SCHEDULE'])
    w.writerow([f'Period: {df} to {dt}'])
    w.writerow([f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")} by {current_user.username}'])
    w.writerow([])

    w.writerow(['DAY BY DAY'])
    w.writerow(['Date', 'Day', 'Daily Income', 'Weekly Income', 'Total Income'])
    w.writerows(day_by_day_rows)
    w.writerow(['TOTAL', '', f"{day_totals['daily']:.2f}", f"{day_totals['weekly']:.2f}",
                f"{day_totals['daily'] + day_totals['weekly']:.2f}"])
    w.writerow([])

    w.writerow(['DAILY RECONCILIATION SCHEDULE'])
    w.writerow(schedule_header)
    for r in daily_rows:
        w.writerow(schedule_row(r))
    w.writerow(schedule_total_row(daily_totals))
    w.writerow([])

    w.writerow(['WEEKLY RECONCILIATION SCHEDULE'])
    w.writerow(schedule_header)
    for r in weekly_rows:
        w.writerow(schedule_row(r))
    w.writerow(schedule_total_row(weekly_totals))

    out.seek(0)
    resp = make_response(out.getvalue())
    resp.headers['Content-Type'] = 'text/csv'
    resp.headers['Content-Disposition'] = f'attachment; filename=franchise_reconciliation_{df}_to_{dt}.csv'
    return resp


def _franchise_suspense_variance(source_type, period_date):
    """Recomputes a single date's/week's reconciliation variance the same
    way report_franchise_reconciliation does (via _group_income_by_period),
    scoped to just that one period_date — used when resolving a suspense
    item, so the amount snapshotted is always current, never stale."""
    model_cls = FranchiseDailyIncome if source_type == 'daily' else FranchiseWeeklyIncome
    date_field = 'entry_date' if source_type == 'daily' else 'week_start'
    entries = model_cls.query.filter(getattr(model_cls, date_field) == period_date).all()
    return _income_entry_totals(entries)['variance']


@app.route('/franchise/suspense-account')
@login_required
@permission_required('franchise')
def franchise_suspense_account():
    """Suspense Account — daily/weekly reconciliation variances (see
    report_franchise_reconciliation) sit here as 'open' until an admin
    manually investigates and clears them; nothing here reconciles itself
    automatically. The variance figures are always recomputed live from
    the source income entries (see _group_income_by_period) — only the
    clearance itself (FranchiseSuspenseResolution) is persisted.

    Cleared rows are kept as a permanent history independent of the live
    >0.01 variance filter: once a period is resolved it stays listed under
    Cleared for as long as the FranchiseSuspenseResolution row exists, even
    if later edits to its source entries happen to bring the live variance
    back to ~0 (which would otherwise silently drop it from both tables)."""
    df, dt = query_date_range()
    daily_entries = FranchiseDailyIncome.query.filter(FranchiseDailyIncome.entry_date.between(df, dt)).all()
    weekly_entries = FranchiseWeeklyIncome.query.filter(FranchiseWeeklyIncome.week_start.between(df, dt)).all()

    daily_by_period = {r['period']: r for r in _group_income_by_period(daily_entries, 'entry_date')}
    weekly_by_period = {r['period']: r for r in _group_income_by_period(weekly_entries, 'week_start')}

    resolutions = FranchiseSuspenseResolution.query.filter(
        FranchiseSuspenseResolution.period_date.between(df, dt)).all()
    resolutions_by_type = {'daily': {}, 'weekly': {}}
    for r in resolutions:
        resolutions_by_type[r.source_type][r.period_date] = r

    def split(rows_by_period, source_type):
        resolved = resolutions_by_type[source_type]
        open_rows = [r for period, r in rows_by_period.items()
                     if abs(r['variance']) > 0.01 and period not in resolved]
        cleared_rows = [dict(rows_by_period.get(period, dict(period=period)), resolution=resolution)
                        for period, resolution in resolved.items()]
        open_rows.sort(key=lambda r: r['period'])
        cleared_rows.sort(key=lambda r: r['period'])
        return open_rows, cleared_rows

    daily_open, daily_cleared = split(daily_by_period, 'daily')
    weekly_open, weekly_cleared = split(weekly_by_period, 'weekly')
    open_balance = sum(r['variance'] for r in daily_open) + sum(r['variance'] for r in weekly_open)

    return render_template('franchise/suspense_account.html',
                           daily_open=daily_open, daily_cleared=daily_cleared,
                           weekly_open=weekly_open, weekly_cleared=weekly_cleared,
                           open_balance=open_balance,
                           open_count=len(daily_open) + len(weekly_open),
                           cleared_count=len(daily_cleared) + len(weekly_cleared),
                           date_from=df.strftime('%Y-%m-%d'), date_to=dt.strftime('%Y-%m-%d'))


@app.route('/franchise/suspense-account/<source_type>/<period_date>/resolve', methods=['POST'])
@login_required
@admin_required
def franchise_suspense_resolve(source_type, period_date):
    if source_type not in ('daily', 'weekly'):
        abort(404)
    period_date = parse_date(period_date)
    notes = request.form.get('notes', '').strip()
    date_from = request.form.get('date_from', '')
    date_to = request.form.get('date_to', '')
    if not notes:
        flash('Explain the discrepancy before clearing it.', 'danger')
        return redirect(url_for('franchise_suspense_account', date_from=date_from, date_to=date_to))

    resolution = FranchiseSuspenseResolution.query.filter_by(
        source_type=source_type, period_date=period_date).first()
    amount = _franchise_suspense_variance(source_type, period_date)
    if resolution:
        resolution.notes = notes
        resolution.resolved_amount = amount
        resolution.resolved_by = current_user.id
        resolution.resolved_at = datetime.now(timezone.utc)
    else:
        resolution = FranchiseSuspenseResolution(
            source_type=source_type, period_date=period_date, resolved_amount=amount,
            notes=notes, resolved_by=current_user.id)
        db.session.add(resolution)
    db.session.flush()
    log_audit('UPDATE', 'franchise_suspense_resolutions', resolution.id,
              f'Cleared {source_type} suspense item for {period_date} (variance {amount:.2f}): {notes}')
    db.session.commit()
    flash('Suspense item cleared.', 'success')
    return redirect(url_for('franchise_suspense_account', date_from=date_from, date_to=date_to))


@app.route('/franchise/suspense-account/<source_type>/<period_date>/reopen', methods=['POST'])
@login_required
@admin_required
def franchise_suspense_reopen(source_type, period_date):
    if source_type not in ('daily', 'weekly'):
        abort(404)
    period_date = parse_date(period_date)
    date_from = request.form.get('date_from', '')
    date_to = request.form.get('date_to', '')
    resolution = FranchiseSuspenseResolution.query.filter_by(
        source_type=source_type, period_date=period_date).first_or_404()
    log_audit('DELETE', 'franchise_suspense_resolutions', resolution.id,
              f'Reopened {source_type} suspense item for {period_date} (was cleared: {resolution.notes})')
    db.session.delete(resolution)
    db.session.commit()
    flash('Suspense item reopened.', 'warning')
    return redirect(url_for('franchise_suspense_account', date_from=date_from, date_to=date_to))


@app.route('/reports/franchise/dual-frequency')
@login_required
@permission_required('franchise')
def report_franchise_dual_frequency():
    """Per-vehicle franchise income transaction summary for a date range,
    combining both independent income streams — Daily and Weekly — side by
    side so it's visible at a glance which vehicles paid via one frequency
    vs. both. A franchise vehicle can owe a daily fee, a weekly fee, or both
    (FranchiseVehicle.daily_fee / weekly_fee), so "both" is common enough
    that it's shown as a flag/filter rather than the only thing this page
    can show — restricting to strictly-both-in-range vehicles left the page
    empty whenever no vehicle happened to have activity on both schedules
    within the exact window chosen."""
    df, dt = query_date_range()
    only_both = request.args.get('only_both') == '1'

    daily_by_vehicle, weekly_by_vehicle = {}, {}
    for e in FranchiseDailyIncome.query.filter(FranchiseDailyIncome.entry_date.between(df, dt),
            FranchiseDailyIncome.vehicle_id.isnot(None)).all():
        daily_by_vehicle.setdefault(e.vehicle_id, []).append(e)
    for e in FranchiseWeeklyIncome.query.filter(FranchiseWeeklyIncome.week_start.between(df, dt),
            FranchiseWeeklyIncome.vehicle_id.isnot(None)).all():
        weekly_by_vehicle.setdefault(e.vehicle_id, []).append(e)

    all_ids = set(daily_by_vehicle) | set(weekly_by_vehicle)
    vehicles = {v.id: v for v in FranchiseVehicle.query.filter(FranchiseVehicle.id.in_(all_ids)).all()} if all_ids else {}

    rows = []
    for vid in all_ids:
        d_entries = daily_by_vehicle.get(vid, [])
        w_entries = weekly_by_vehicle.get(vid, [])
        both = bool(d_entries) and bool(w_entries)
        if only_both and not both:
            continue
        d_totals = _income_entry_totals(d_entries)
        w_totals = _income_entry_totals(w_entries)
        rows.append(dict(
            vehicle=vehicles.get(vid), both=both,
            daily_count=len(d_entries), daily_income=d_totals['income'],
            weekly_count=len(w_entries), weekly_income=w_totals['income'],
            total_income=d_totals['income'] + w_totals['income'],
            last_daily=max((e.entry_date for e in d_entries), default=None),
            last_weekly=max((e.week_start for e in w_entries), default=None),
        ))
    rows.sort(key=lambda r: (not r['both'], r['vehicle'].number_plate if r['vehicle'] else ''))

    totals = dict(
        both_count=sum(1 for r in rows if r['both']),
        daily_income=sum(r['daily_income'] for r in rows),
        weekly_income=sum(r['weekly_income'] for r in rows),
        total_income=sum(r['total_income'] for r in rows),
    )

    return render_template('franchise/dual_frequency.html',
                           title='Franchise Income by Vehicle — Daily & Weekly',
                           rows=rows, totals=totals, only_both=only_both,
                           date_from=df.strftime('%Y-%m-%d'), date_to=dt.strftime('%Y-%m-%d'))


@app.route('/reports/franchise/dual-frequency/export')
@login_required
@permission_required('franchise')
def report_franchise_dual_frequency_export():
    df, dt = query_date_range()
    only_both = request.args.get('only_both') == '1'

    daily_by_vehicle, weekly_by_vehicle = {}, {}
    for e in FranchiseDailyIncome.query.filter(FranchiseDailyIncome.entry_date.between(df, dt),
            FranchiseDailyIncome.vehicle_id.isnot(None)).all():
        daily_by_vehicle.setdefault(e.vehicle_id, []).append(e)
    for e in FranchiseWeeklyIncome.query.filter(FranchiseWeeklyIncome.week_start.between(df, dt),
            FranchiseWeeklyIncome.vehicle_id.isnot(None)).all():
        weekly_by_vehicle.setdefault(e.vehicle_id, []).append(e)

    all_ids = set(daily_by_vehicle) | set(weekly_by_vehicle)
    vehicles = {v.id: v for v in FranchiseVehicle.query.filter(FranchiseVehicle.id.in_(all_ids)).all()} if all_ids else {}

    rows = []
    for vid in all_ids:
        d_entries = daily_by_vehicle.get(vid, [])
        w_entries = weekly_by_vehicle.get(vid, [])
        both = bool(d_entries) and bool(w_entries)
        if only_both and not both:
            continue
        vehicle = vehicles.get(vid)
        d_totals = _income_entry_totals(d_entries)
        w_totals = _income_entry_totals(w_entries)
        rows.append([vehicle.number_plate if vehicle else '', vehicle.franchisee_name if vehicle else '',
                     'Both' if both else ('Daily only' if d_entries else 'Weekly only'),
                     len(d_entries), f"{d_totals['income']:.2f}", len(w_entries), f"{w_totals['income']:.2f}",
                     f"{d_totals['income'] + w_totals['income']:.2f}"])
    rows.sort(key=lambda r: r[0])
    header = ['Number Plate', 'Franchisee', 'Paid Via', 'Daily Entries', 'Daily Income',
              'Weekly Entries', 'Weekly Income', 'Total Income']
    if request.args.get('format') == 'pdf':
        return _table_pdf_response(f'franchise_dual_frequency_{df}_to_{dt}.pdf', 'Franchise Daily & Weekly Payers',
            f'Period: {df} to {dt}', header, rows)
    return csv_export_response(f'franchise_dual_frequency_{df}_to_{dt}.csv', header, rows)


@app.route('/reports/franchise/daily-defaulters')
@login_required
@permission_required('franchise')
def report_franchise_daily_defaulters():
    """Which franchise vehicles have NOT paid their daily fee across a date
    range — the Daily Income page's "By Date" tab is a single-date
    recording workflow (fill/confirm/delete), not a scannable report over a
    period, so this is the read-only "who's behind" view instead: one row
    per vehicle with at least one missed day in range, sorted worst-first.

    Eligibility deliberately matches _franchise_income_by_vehicle_on (every
    active vehicle, not just ones with FranchiseVehicle.daily_fee set) —
    that field is informational/optional in practice (many real vehicles
    have it unset), and the rest of the app already treats "no entry for
    this vehicle on this date" as owing regardless of it, via the same
    By Date bulk-fill workflow. Gating this report on it too would just
    make it silently show nothing for a fleet that hasn't populated fees."""
    df, dt = query_date_range()
    show_all = request.args.get('show_all') == '1'
    total_days = (dt - df).days + 1

    vehicles = FranchiseVehicle.query.filter_by(status='active').order_by(FranchiseVehicle.number_plate).all()

    income_by_vehicle_date = {}
    for e in FranchiseDailyIncome.query.filter(
            FranchiseDailyIncome.entry_date.between(df, dt),
            FranchiseDailyIncome.vehicle_id.isnot(None)).all():
        key = (e.vehicle_id, e.entry_date)
        income_by_vehicle_date[key] = income_by_vehicle_date.get(key, 0) + e.income

    rows = []
    for v in vehicles:
        fee = v.daily_fee or 0
        missed_dates, collected = [], 0.0
        d = df
        while d <= dt:
            amount = income_by_vehicle_date.get((v.id, d), 0)
            if amount > 0:
                collected += amount
            else:
                missed_dates.append(d)
            d += timedelta(days=1)
        days_missed = len(missed_dates)
        if days_missed == 0 and not show_all:
            continue
        rows.append(dict(
            vehicle=v, days_due=total_days, days_missed=days_missed,
            days_paid=total_days - days_missed,
            expected=fee * total_days, collected=collected,
            shortfall=(fee * total_days) - collected,
            last_missed=max(missed_dates) if missed_dates else None,
        ))
    rows.sort(key=lambda r: (-r['days_missed'], r['vehicle'].number_plate))

    totals = dict(
        vehicles_shown=len(rows),
        days_missed=sum(r['days_missed'] for r in rows),
        expected=sum(r['expected'] for r in rows),
        collected=sum(r['collected'] for r in rows),
        shortfall=sum(r['shortfall'] for r in rows),
    )

    return render_template('franchise/daily_defaulters.html',
                           title='Franchise Daily Fee Defaulters',
                           rows=rows, totals=totals, show_all=show_all, total_days=total_days,
                           date_from=df.strftime('%Y-%m-%d'), date_to=dt.strftime('%Y-%m-%d'))


@app.route('/reports/franchise/daily-defaulters/export')
@login_required
@permission_required('franchise')
def report_franchise_daily_defaulters_export():
    df, dt = query_date_range()
    show_all = request.args.get('show_all') == '1'
    total_days = (dt - df).days + 1

    vehicles = FranchiseVehicle.query.filter_by(status='active').order_by(FranchiseVehicle.number_plate).all()

    income_by_vehicle_date = {}
    for e in FranchiseDailyIncome.query.filter(
            FranchiseDailyIncome.entry_date.between(df, dt),
            FranchiseDailyIncome.vehicle_id.isnot(None)).all():
        key = (e.vehicle_id, e.entry_date)
        income_by_vehicle_date[key] = income_by_vehicle_date.get(key, 0) + e.income

    rows = []
    for v in vehicles:
        fee = v.daily_fee or 0
        missed_dates, collected = [], 0.0
        d = df
        while d <= dt:
            amount = income_by_vehicle_date.get((v.id, d), 0)
            if amount > 0:
                collected += amount
            else:
                missed_dates.append(d)
            d += timedelta(days=1)
        days_missed = len(missed_dates)
        if days_missed == 0 and not show_all:
            continue
        rows.append([v.number_plate, v.franchisee_name, f'{fee:.2f}', total_days,
                     total_days - days_missed, days_missed, f'{fee * total_days:.2f}',
                     f'{collected:.2f}', f'{(fee * total_days) - collected:.2f}',
                     max(missed_dates) if missed_dates else ''])
    rows.sort(key=lambda r: (-r[5], r[0]))

    header = ['Number Plate', 'Franchisee', 'Daily Fee', 'Days Due', 'Days Paid', 'Days Missed',
              'Expected', 'Collected', 'Shortfall', 'Last Missed Date']
    if request.args.get('format') == 'pdf':
        return _table_pdf_response(f'franchise_daily_defaulters_{df}_to_{dt}.pdf', 'Franchise Daily Fee Defaulters',
            f'Period: {df} to {dt}', header, rows)
    return csv_export_response(f'franchise_daily_defaulters_{df}_to_{dt}.csv', header, rows)


@app.route('/reports/franchise/weekly')
@login_required
@permission_required('franchise')
def report_franchise_weekly():
    """Weekly Analysis — for each Monday-Sunday week in range, roll up that
    week's Daily Income entries and combine them with the standalone Weekly
    Income entry for that week (if any), so daily and weekly figures — kept
    as separate entities — can still be seen combined at the week level."""
    df, dt = query_date_range()
    daily_entries = FranchiseDailyIncome.query.filter(FranchiseDailyIncome.entry_date.between(df, dt)) \
        .order_by(FranchiseDailyIncome.entry_date.asc()).all()
    weekly_entries = FranchiseWeeklyIncome.query.filter(FranchiseWeeklyIncome.week_start.between(df, dt)) \
        .order_by(FranchiseWeeklyIncome.week_start.asc()).all()
    # A week can have several weekly entries — one per franchisee — so group
    # into lists rather than assuming a single entry per week_start. Each
    # entry's week_start is its own recorded date (not normalized to
    # Monday), so it's floored to that week's Monday here to bucket it with
    # the same calendar week's daily entries.
    weekly_by_week = {}
    for e in weekly_entries:
        week_start = e.week_start - timedelta(days=e.week_start.weekday())
        weekly_by_week.setdefault(week_start, []).append(e)

    daily_by_week = {}
    for e in daily_entries:
        week_start = e.entry_date - timedelta(days=e.entry_date.weekday())
        daily_by_week.setdefault(week_start, []).append(e)

    week_starts = sorted(set(daily_by_week.keys()) | set(weekly_by_week.keys()))
    week_rows = []
    for start in week_starts:
        week_daily_entries = daily_by_week.get(start, [])
        daily_totals = _income_entry_totals(week_daily_entries)
        weekly_totals = _income_entry_totals(weekly_by_week.get(start, []))
        week_rows.append(dict(
            week_start=start, week_end=start + timedelta(days=6), days=len(week_daily_entries),
            daily_income=daily_totals['income'], daily_expenditure=daily_totals['total_expenditure'],
            weekly_income=weekly_totals['income'], weekly_expenditure=weekly_totals['total_expenditure'],
            total_income=daily_totals['income'] + weekly_totals['income'],
            total_expenditure=daily_totals['total_expenditure'] + weekly_totals['total_expenditure'],
            net_profit=(daily_totals['income'] + weekly_totals['income'])
                - (daily_totals['total_expenditure'] + weekly_totals['total_expenditure']),
        ))

    totals = dict(
        daily_income=sum(r['daily_income'] for r in week_rows),
        daily_expenditure=sum(r['daily_expenditure'] for r in week_rows),
        weekly_income=sum(r['weekly_income'] for r in week_rows),
        weekly_expenditure=sum(r['weekly_expenditure'] for r in week_rows),
        total_income=sum(r['total_income'] for r in week_rows),
        total_expenditure=sum(r['total_expenditure'] for r in week_rows),
        net_profit=sum(r['net_profit'] for r in week_rows),
    )
    return render_template('franchise/weekly_analysis.html', title='Franchise Weekly Analysis',
                           weeks=week_rows, totals=totals,
                           date_from=df.strftime('%Y-%m-%d'), date_to=dt.strftime('%Y-%m-%d'))


@app.route('/reports/franchise/weekly/export')
@login_required
@permission_required('franchise')
def report_franchise_weekly_export():
    df, dt = query_date_range()
    daily_entries = FranchiseDailyIncome.query.filter(FranchiseDailyIncome.entry_date.between(df, dt)).all()
    weekly_entries = FranchiseWeeklyIncome.query.filter(FranchiseWeeklyIncome.week_start.between(df, dt)).all()
    weekly_by_week = {}
    for e in weekly_entries:
        week_start = e.week_start - timedelta(days=e.week_start.weekday())
        weekly_by_week.setdefault(week_start, []).append(e)
    daily_by_week = {}
    for e in daily_entries:
        week_start = e.entry_date - timedelta(days=e.entry_date.weekday())
        daily_by_week.setdefault(week_start, []).append(e)

    week_starts = sorted(set(daily_by_week.keys()) | set(weekly_by_week.keys()))
    rows = []
    for start in week_starts:
        daily_totals = _income_entry_totals(daily_by_week.get(start, []))
        weekly_totals = _income_entry_totals(weekly_by_week.get(start, []))
        rows.append([start, start + timedelta(days=6), f"{daily_totals['income']:.2f}",
                     f"{weekly_totals['income']:.2f}", f"{daily_totals['income'] + weekly_totals['income']:.2f}",
                     f"{daily_totals['total_expenditure'] + weekly_totals['total_expenditure']:.2f}",
                     f"{(daily_totals['income'] + weekly_totals['income']) - (daily_totals['total_expenditure'] + weekly_totals['total_expenditure']):.2f}"])
    header = ['Week Start', 'Week End', 'Daily Income', 'Weekly Income', 'Total Income', 'Total Expenditure', 'Net Profit']
    if request.args.get('format') == 'pdf':
        return _table_pdf_response(f'franchise_weekly_analysis_{df}_to_{dt}.pdf', 'Franchise Weekly Analysis',
            f'Period: {df} to {dt}', header, rows)
    return csv_export_response(f'franchise_weekly_analysis_{df}_to_{dt}.csv', header, rows)


@app.route('/reports/franchise/consolidated')
@login_required
@permission_required('franchise')
def report_franchise_consolidated():
    """Consolidated P&L — single summary of income, expenditure by category,
    and the cash reconciliation for the whole period, combining the daily
    and weekly income entities. Operational Expenses (FranchiseOperationalExpense)
    are added here only — they reduce Net Profit but, unlike the per-entry
    exp_* categories above, never touch Cash Reconciliation, since they
    aren't cash any vehicle handled."""
    df, dt = query_date_range()
    daily_entries = FranchiseDailyIncome.query.filter(FranchiseDailyIncome.entry_date.between(df, dt)).all()
    weekly_entries = FranchiseWeeklyIncome.query.filter(FranchiseWeeklyIncome.week_start.between(df, dt)).all()
    op_expenses = FranchiseOperationalExpense.query.filter(
        FranchiseOperationalExpense.expense_date.between(df, dt)).all()

    daily_totals = _income_entry_totals(daily_entries)
    weekly_totals = _income_entry_totals(weekly_entries)
    totals = {k: daily_totals[k] + weekly_totals[k] for k in daily_totals}
    totals['income_daily'] = daily_totals['income']
    totals['income_weekly'] = weekly_totals['income']
    totals['total_income'] = totals.pop('income')

    op_by_category = {}
    for e in op_expenses:
        op_by_category[e.category.name] = op_by_category.get(e.category.name, 0) + e.amount
    totals['operational_expenses'] = sum(op_by_category.values())
    totals['operational_expenses_by_category'] = sorted(op_by_category.items())
    totals['net_profit'] = totals['total_income'] - totals['total_expenditure'] - totals['operational_expenses']

    return render_template('franchise/consolidated.html', title='Consolidated Franchise P&L Statement',
                           totals=totals, entry_count=len(daily_entries) + len(weekly_entries),
                           date_from=df.strftime('%Y-%m-%d'), date_to=dt.strftime('%Y-%m-%d'))


@app.route('/reports/franchise/consolidated/export')
@login_required
@permission_required('franchise')
def report_franchise_consolidated_export():
    df, dt = query_date_range()
    daily_entries = FranchiseDailyIncome.query.filter(FranchiseDailyIncome.entry_date.between(df, dt)).all()
    weekly_entries = FranchiseWeeklyIncome.query.filter(FranchiseWeeklyIncome.week_start.between(df, dt)).all()
    op_expenses = FranchiseOperationalExpense.query.filter(
        FranchiseOperationalExpense.expense_date.between(df, dt)).all()
    daily_totals = _income_entry_totals(daily_entries)
    weekly_totals = _income_entry_totals(weekly_entries)
    totals = {k: daily_totals[k] + weekly_totals[k] for k in daily_totals}

    op_by_category = {}
    for e in op_expenses:
        op_by_category[e.category.name] = op_by_category.get(e.category.name, 0) + e.amount
    total_operational_expenses = sum(op_by_category.values())

    rows = [
        ['Daily Income', f"{daily_totals['income']:.2f}"],
        ['Weekly Income', f"{weekly_totals['income']:.2f}"],
        ['Total Income', f"{totals['income']:.2f}"],
        ['Traffic Fines', f"{totals['exp_traffic_fines']:.2f}"],
        ['Facilitation Fees', f"{totals['exp_facilitation_fees']:.2f}"],
        ['Workshop', f"{totals['exp_workshop']:.2f}"],
        ['Wages', f"{totals['exp_wages']:.2f}"],
        ['Other Expenditure', f"{totals['other_expenditure']:.2f}"],
        ['Total Expenditure', f"{totals['total_expenditure']:.2f}"],
    ]
    for name, amount in sorted(op_by_category.items()):
        rows.append([f'Operational Expenses — {name}', f"{amount:.2f}"])
    rows.append(['Total Operational Expenses', f"{total_operational_expenses:.2f}"])
    rows.append(['Net Profit', f"{totals['income'] - totals['total_expenditure'] - total_operational_expenses:.2f}"])
    rows.append(['Cash Deposited', f"{totals['deposited']:.2f}"])
    rows.append(['Variance', f"{totals['variance']:.2f}"])
    header = ['Line Item', 'Amount']
    if request.args.get('format') == 'pdf':
        return _table_pdf_response(f'franchise_consolidated_pl_{df}_to_{dt}.pdf', 'Consolidated Franchise P&L Statement',
            f'Period: {df} to {dt}', header, rows)
    return csv_export_response(f'franchise_consolidated_pl_{df}_to_{dt}.csv', header, rows)


# ─────────────────────────────────────────────────────────────
# Fleet Reconciliation — the company's own fleet (Vehicle/Driver/DailyLog,
# distinct from the Franchise module above): net income (DailyLog.gross_revenue
# minus driver-attributed Expense rows, i.e. costs a driver incurs while away)
# reconciled against cash each driver deposits (DriverDeposit) — by driver,
# and consolidated by vehicle.
# ─────────────────────────────────────────────────────────────
def _fleet_driver_totals(df, dt):
    """Per active driver: income (DailyLog.gross_revenue), expenses (Expense
    rows attributed to that driver), net_income and variance against
    DriverDeposit. Every active driver appears even with zero activity, so a
    driver with income but no deposit shows up as a variance rather than
    being silently omitted. Returns (rows, totals) — totals has the same
    shape as one row, summed."""
    drivers = Driver.query.filter_by(role='driver', status='active').order_by(Driver.name).all()

    income_by_driver = dict(
        db.session.query(DailyLog.driver_id, func.sum(DailyLog.gross_revenue))
        .filter(DailyLog.log_date.between(df, dt), DailyLog.driver_id.isnot(None))
        .group_by(DailyLog.driver_id).all())
    expense_by_driver = dict(
        db.session.query(Expense.driver_id, func.sum(Expense.amount))
        .filter(Expense.expense_date.between(df, dt), Expense.driver_id.isnot(None))
        .group_by(Expense.driver_id).all())
    deposited_by_driver = dict(
        db.session.query(DriverDeposit.driver_id, func.sum(DriverDeposit.amount))
        .filter(DriverDeposit.deposit_date.between(df, dt))
        .group_by(DriverDeposit.driver_id).all())

    rows = []
    for d in drivers:
        income = income_by_driver.get(d.id) or 0.0
        expenses = expense_by_driver.get(d.id) or 0.0
        deposited = deposited_by_driver.get(d.id) or 0.0
        net_income = income - expenses
        rows.append(dict(driver=d, income=income, expenses=expenses,
                          net_income=net_income, deposited=deposited, variance=deposited - net_income))

    # Fleet-wide deposits (driver_id left blank on DriverDeposit) aren't
    # attributable to one driver, so they can't join the per-driver rows
    # above — surfaced as their own row instead of being silently dropped
    # from the totals.
    fleet_wide_deposited = db.session.query(func.sum(DriverDeposit.amount)).filter(
        DriverDeposit.deposit_date.between(df, dt), DriverDeposit.driver_id.is_(None)).scalar() or 0.0
    if fleet_wide_deposited:
        rows.append(dict(driver=SimpleNamespace(name='Fleet-wide (all vehicles)'), income=0.0, expenses=0.0,
                          net_income=0.0, deposited=fleet_wide_deposited, variance=fleet_wide_deposited))

    totals = dict(
        income=sum(r['income'] for r in rows), expenses=sum(r['expenses'] for r in rows),
        net_income=sum(r['net_income'] for r in rows), deposited=sum(r['deposited'] for r in rows),
        variance=sum(r['variance'] for r in rows))
    return rows, totals


def _fleet_vehicle_totals(df, dt):
    """Same shape as _fleet_driver_totals, consolidated per active vehicle
    instead of per driver. Expenses only count rows attributed to a driver
    (Expense.driver_id set) so the "expenses incurred while away" definition
    stays consistent between the by-driver and by-vehicle views."""
    vehicles = Vehicle.query.filter_by(status='active').order_by(Vehicle.registration).all()

    income_by_vehicle = dict(
        db.session.query(DailyLog.vehicle_id, func.sum(DailyLog.gross_revenue))
        .filter(DailyLog.log_date.between(df, dt))
        .group_by(DailyLog.vehicle_id).all())
    expense_by_vehicle = dict(
        db.session.query(Expense.vehicle_id, func.sum(Expense.amount))
        .filter(Expense.expense_date.between(df, dt), Expense.driver_id.isnot(None), Expense.vehicle_id.isnot(None))
        .group_by(Expense.vehicle_id).all())
    deposited_by_vehicle = dict(
        db.session.query(DriverDeposit.vehicle_id, func.sum(DriverDeposit.amount))
        .filter(DriverDeposit.deposit_date.between(df, dt), DriverDeposit.vehicle_id.isnot(None))
        .group_by(DriverDeposit.vehicle_id).all())

    rows = []
    for v in vehicles:
        income = income_by_vehicle.get(v.id) or 0.0
        expenses = expense_by_vehicle.get(v.id) or 0.0
        deposited = deposited_by_vehicle.get(v.id) or 0.0
        net_income = income - expenses
        rows.append(dict(vehicle=v, income=income, expenses=expenses,
                          net_income=net_income, deposited=deposited, variance=deposited - net_income))

    # Fleet-wide deposits (no driver, no vehicle attribution) aren't
    # attributable to one vehicle, so surface them as their own row rather
    # than dropping them from the totals silently.
    fleet_wide_deposited = db.session.query(func.sum(DriverDeposit.amount)).filter(
        DriverDeposit.deposit_date.between(df, dt), DriverDeposit.vehicle_id.is_(None)).scalar() or 0.0
    if fleet_wide_deposited:
        rows.append(dict(vehicle=SimpleNamespace(registration='Fleet-wide (all vehicles)'), income=0.0, expenses=0.0,
                          net_income=0.0, deposited=fleet_wide_deposited, variance=fleet_wide_deposited))

    totals = dict(
        income=sum(r['income'] for r in rows), expenses=sum(r['expenses'] for r in rows),
        net_income=sum(r['net_income'] for r in rows), deposited=sum(r['deposited'] for r in rows),
        variance=sum(r['variance'] for r in rows))
    return rows, totals


@app.route('/reports/fleet/reconciliation')
@login_required
@permission_required('reports')
def report_fleet_reconciliation():
    df, dt = query_date_range()
    rows, totals = _fleet_driver_totals(df, dt)
    return render_template('fleet/reconciliation.html', rows=rows, totals=totals,
                           date_from=df.strftime('%Y-%m-%d'), date_to=dt.strftime('%Y-%m-%d'))


@app.route('/reports/fleet/reconciliation/export')
@login_required
@permission_required('reports')
def report_fleet_reconciliation_export():
    df, dt = query_date_range()
    rows, totals = _fleet_driver_totals(df, dt)
    out_rows = [[r['driver'].name, f"{r['income']:.2f}", f"{r['expenses']:.2f}", f"{r['net_income']:.2f}",
                 f"{r['deposited']:.2f}", f"{r['variance']:.2f}"] for r in rows]
    out_rows.append(['TOTAL', f"{totals['income']:.2f}", f"{totals['expenses']:.2f}", f"{totals['net_income']:.2f}",
                      f"{totals['deposited']:.2f}", f"{totals['variance']:.2f}"])
    header = ['Driver', 'Income', 'Expenses', 'Net Income', 'Cash Deposited', 'Variance']
    if request.args.get('format') == 'pdf':
        return _table_pdf_response(f'fleet_reconciliation_{df}_to_{dt}.pdf', 'Fleet Reconciliation',
            f'Period: {df} to {dt}', header, out_rows)
    return csv_export_response(f'fleet_reconciliation_{df}_to_{dt}.csv', header, out_rows)


@app.route('/reports/fleet/consolidated')
@login_required
@permission_required('reports')
def report_fleet_consolidated():
    df, dt = query_date_range()
    rows, totals = _fleet_vehicle_totals(df, dt)
    return render_template('fleet/consolidated.html', rows=rows, totals=totals,
                           date_from=df.strftime('%Y-%m-%d'), date_to=dt.strftime('%Y-%m-%d'))


@app.route('/reports/fleet/consolidated/export')
@login_required
@permission_required('reports')
def report_fleet_consolidated_export():
    df, dt = query_date_range()
    rows, totals = _fleet_vehicle_totals(df, dt)
    out_rows = [[r['vehicle'].registration, f"{r['income']:.2f}", f"{r['expenses']:.2f}", f"{r['net_income']:.2f}",
                 f"{r['deposited']:.2f}", f"{r['variance']:.2f}"] for r in rows]
    out_rows.append(['TOTAL', f"{totals['income']:.2f}", f"{totals['expenses']:.2f}", f"{totals['net_income']:.2f}",
                      f"{totals['deposited']:.2f}", f"{totals['variance']:.2f}"])
    header = ['Vehicle', 'Income', 'Expenses', 'Net Income', 'Cash Deposited', 'Variance']
    if request.args.get('format') == 'pdf':
        return _table_pdf_response(f'fleet_consolidated_{df}_to_{dt}.pdf', 'Fleet Reconciliation (by Vehicle)',
            f'Period: {df} to {dt}', header, out_rows)
    return csv_export_response(f'fleet_consolidated_{df}_to_{dt}.csv', header, out_rows)


# ─────────────────────────────────────────────────────────────
# Compliance
# ─────────────────────────────────────────────────────────────
@app.route('/compliance')
@login_required
@permission_required('compliance')
def compliance():
    today = date.today()
    threshold = today + timedelta(days=30)
    expired = VehicleDocument.query.filter(
        VehicleDocument.expiry_date < today).order_by(VehicleDocument.expiry_date).all()
    expiring = VehicleDocument.query.filter(
        VehicleDocument.expiry_date.between(today, threshold)).order_by(
        VehicleDocument.expiry_date).all()
    valid = VehicleDocument.query.filter(
        VehicleDocument.expiry_date > threshold).order_by(
        VehicleDocument.expiry_date).all()

    # Insurance is a Vehicle field, not a VehicleDocument row, but belongs on
    # the same expiry tracker — wrapped in plain dicts (duck-typed the same
    # shape the template already expects: vehicle/doc_type/expiry_date/
    # days_to_expiry) so it sorts and displays alongside real documents.
    for v in Vehicle.query.filter(Vehicle.insurance_expiry.isnot(None)).all():
        doc_type = f'Insurance ({v.insurance_type})' if v.insurance_type else 'Insurance'
        entry = {'vehicle': v, 'doc_type': doc_type, 'expiry_date': v.insurance_expiry,
                  'days_to_expiry': v.insurance_days_to_expiry}
        if v.insurance_status == 'expired':
            expired.append(entry)
        elif v.insurance_status == 'warning':
            expiring.append(entry)
        else:
            valid.append(entry)
    expired.sort(key=lambda d: d['expiry_date'] if isinstance(d, dict) else d.expiry_date)
    expiring.sort(key=lambda d: d['expiry_date'] if isinstance(d, dict) else d.expiry_date)
    valid.sort(key=lambda d: d['expiry_date'] if isinstance(d, dict) else d.expiry_date)

    return render_template('compliance/index.html',
        expired=expired, expiring=expiring, valid=valid, today=today)


@app.route('/compliance/export')
@login_required
@permission_required('compliance')
def compliance_export():
    today = date.today()
    threshold = today + timedelta(days=30)
    docs = VehicleDocument.query.order_by(VehicleDocument.expiry_date).all()
    # Insurance is a Vehicle field, not a VehicleDocument row (see compliance()
    # above) — duck-typed into the same dict shape here too so both kinds of
    # entry export as one flat list instead of two separate files.
    entries = [{'vehicle': d.vehicle, 'doc_type': d.doc_type, 'reference_number': d.reference_number,
                'issue_date': d.issue_date, 'expiry_date': d.expiry_date} for d in docs]
    for v in Vehicle.query.filter(Vehicle.insurance_expiry.isnot(None)).all():
        doc_type = f'Insurance ({v.insurance_type})' if v.insurance_type else 'Insurance'
        entries.append({'vehicle': v, 'doc_type': doc_type, 'reference_number': v.insurance_policy_number,
                        'issue_date': None, 'expiry_date': v.insurance_expiry})
    entries.sort(key=lambda e: e['expiry_date'])

    def status(expiry):
        if expiry < today:
            return 'Expired'
        if expiry <= threshold:
            return 'Expiring Soon'
        return 'Valid'

    rows = [[e['vehicle'].registration, e['doc_type'], e['reference_number'] or '',
             e['issue_date'] or '', e['expiry_date'], status(e['expiry_date'])] for e in entries]
    header = ['Vehicle', 'Document Type', 'Reference Number', 'Issue Date', 'Expiry Date', 'Status']
    if request.args.get('format') == 'pdf':
        return _table_pdf_response(f'compliance_documents_{today}.pdf', 'Compliance Documents',
            f'As at {today}', header, rows)
    return csv_export_response(f'compliance_documents_{today}.csv', header, rows)


# ─────────────────────────────────────────────────────────────
# Users (Admin)
# ─────────────────────────────────────────────────────────────
@app.route('/users')
@login_required
@admin_required
def users():
    all_users = User.query.order_by(User.username).all()
    return render_template('users/index.html', users=all_users)


@app.route('/users/add', methods=['GET', 'POST'])
@login_required
@admin_required
@handle_form_errors
def user_add():
    if request.method == 'POST':
        password = request.form.get('password', '')
        if len(password) < 6:
            raise ValueError('Password must be at least 6 characters.')
        username = request.form['username'].strip().lower()
        email = request.form['email'].strip().lower()
        check_unique(User, 'username', username)
        check_unique(User, 'email', email)
        # 'role' is otherwise a free-text label (see users/form.html's
        # "Other — type a custom role…" option) — only the exact value
        # 'admin' has any special meaning to the app (User.has_permission's
        # full-access bypass, admin_required); everything else grants
        # nothing on its own and is scoped entirely by the permissions
        # list below, set on the Permissions page after creation.
        role_preset = request.form.get('role_preset', 'manager')
        role = request.form.get('role_custom', '').strip() if role_preset == '__custom__' else role_preset
        if not role:
            role = 'manager'
        if role.lower() == 'admin':
            role = 'admin'
        u = User(
            username=username,
            email=email,
            role=role,
        )
        # The Franchise Clerk preset is pre-seeded with the franchise_entry
        # permission so it works immediately without a second trip to the
        # Permissions page — every other role/preset still starts with no
        # access until an admin grants it there, same as before.
        if role_preset == 'Franchise Clerk':
            u.permissions = json.dumps(['franchise_entry'])
        u.set_password(password)
        db.session.add(u)
        db.session.flush()
        log_audit('CREATE', 'users', u.id, f'Created user {u.username}')
        db.session.commit()
        flash(f'User "{u.username}" created.', 'success')
        return redirect(url_for('users'))
    return render_template('users/form.html')


@app.route('/users/<int:uid>/toggle', methods=['POST'])
@login_required
@admin_required
def user_toggle(uid):
    u = User.query.get_or_404(uid)
    if u.id == current_user.id:
        flash('Cannot deactivate your own account.', 'danger')
        return redirect(url_for('users'))
    u.is_active = not u.is_active
    log_audit('UPDATE', 'users', uid,
              f'Set user {u.username} active={u.is_active}')
    db.session.commit()
    flash(f'User {u.username} {"activated" if u.is_active else "deactivated"}.', 'info')
    return redirect(url_for('users'))


@app.route('/users/<int:uid>/role', methods=['POST'])
@login_required
@admin_required
def user_set_role(uid):
    """Promote/demote between 'admin' (full system access, see
    User.has_permission) and 'manager' (access limited to whatever's
    granted on the Permissions page). This is the only way an existing
    user gains admin — role can't be changed after creation any other
    way, so a manager with every individual permission checked still
    can't reach admin-only tooling (Users, Audit Log, Sync Sites,
    Import History) until promoted here."""
    u = User.query.get_or_404(uid)
    new_role = request.form.get('role')
    if new_role not in ('admin', 'manager'):
        flash('Invalid role.', 'danger')
        return redirect(url_for('users'))
    if u.role == 'admin' and new_role != 'admin':
        if u.id == current_user.id:
            flash('Cannot remove your own admin access.', 'danger')
            return redirect(url_for('users'))
        if User.query.filter(User.role == 'admin', User.id != u.id).count() == 0:
            flash('Cannot remove the last admin account.', 'danger')
            return redirect(url_for('users'))
    u.role = new_role
    log_audit('UPDATE', 'users', uid, f'Set role for {u.username}: {new_role}')
    db.session.commit()
    flash(f'{u.username} is now {"an Admin" if new_role == "admin" else "a Manager"}.', 'success')
    return redirect(url_for('users'))


@app.route('/users/<int:uid>/delete', methods=['POST'])
@login_required
@admin_required
def user_delete(uid):
    """Hard-delete a login account that's no longer in use. Records this
    user created (expenses, logs, etc.) aren't touched — their created_by/
    user_id columns just point at an id that no longer resolves, same as
    disabling already leaves behind, so history/attribution stays intact.
    Use Disable instead of this if the account might come back — deletion
    also frees up its username/email/WhatsApp number for reuse."""
    u = User.query.get_or_404(uid)
    if u.id == current_user.id:
        flash('Cannot delete your own account.', 'danger')
        return redirect(url_for('users'))
    if u.role == 'admin' and User.query.filter(User.role == 'admin', User.id != u.id).count() == 0:
        flash('Cannot delete the last admin account.', 'danger')
        return redirect(url_for('users'))
    username = u.username
    log_audit('DELETE', 'users', uid, f'Deleted user {username}')
    db.session.delete(u)
    db.session.commit()
    flash(f'User "{username}" deleted.', 'info')
    return redirect(url_for('users'))


@app.route('/users/<int:uid>/permissions', methods=['GET', 'POST'])
@login_required
@admin_required
def user_permissions(uid):
    u = User.query.get_or_404(uid)
    if u.role == 'admin':
        flash('Admin users always have full access — no permission configuration needed.', 'info')
        return redirect(url_for('users'))
    if request.method == 'POST':
        granted = request.form.getlist('permissions')
        u.permissions = json.dumps(granted)
        log_audit('UPDATE', 'users', uid,
                  f'Updated permissions for {u.username}: {", ".join(granted) or "none"}')
        db.session.commit()
        flash(f'Permissions updated for {u.username}.', 'success')
        return redirect(url_for('users'))
    return render_template('users/permissions.html', u=u,
                           all_permissions=PERMISSIONS,
                           current_perms=u.get_permissions())


@app.route('/users/<int:uid>/whatsapp', methods=['GET', 'POST'])
@login_required
@admin_required
@handle_form_errors
def user_whatsapp(uid):
    u = User.query.get_or_404(uid)
    if request.method == 'POST':
        raw = request.form.get('whatsapp_phone', '').strip()
        phone = re.sub(r'\D', '', raw) or None
        if phone:
            check_unique(User, 'whatsapp_phone', phone, label='WhatsApp number', exclude_id=u.id)
        u.whatsapp_phone = phone
        log_audit('UPDATE', 'users', uid,
                  f'Set WhatsApp number for {u.username}: {phone or "(unlinked)"}')
        db.session.commit()
        flash(f'WhatsApp number updated for {u.username}.', 'success')
        return redirect(url_for('users'))
    return render_template('users/whatsapp.html', u=u)


@app.route('/users/<int:uid>/reset-password', methods=['POST'])
@login_required
@admin_required
def user_reset_password(uid):
    u = User.query.get_or_404(uid)
    new_pw = request.form.get('new_password', '').strip()
    if len(new_pw) < 6:
        flash('Password must be at least 6 characters.', 'danger')
        return redirect(url_for('users'))
    u.set_password(new_pw)
    log_audit('UPDATE', 'users', uid, f'Reset password for {u.username}')
    db.session.commit()
    flash(f'Password reset for {u.username}.', 'success')
    return redirect(url_for('users'))


# ─────────────────────────────────────────────────────────────
# Audit Log
# ─────────────────────────────────────────────────────────────
@app.route('/audit')
@login_required
@admin_required
def audit_log():
    page = request.args.get('page', 1, type=int)
    logs = AuditLog.query.order_by(AuditLog.timestamp.desc()).paginate(page=page, per_page=30)
    return render_template('audit/index.html', logs=logs)


@app.route('/audit/export')
@login_required
@admin_required
def audit_log_export():
    logs = AuditLog.query.order_by(AuditLog.timestamp.desc()).all()
    rows = [[l.timestamp.strftime('%Y-%m-%d %H:%M:%S'), l.user.username if l.user else 'System',
             l.action, l.table_name or '', l.record_id or '', l.description or '', l.ip_address or '']
            for l in logs]
    header = ['Timestamp', 'User', 'Action', 'Table', 'Record ID', 'Description', 'IP']
    if request.args.get('format') == 'pdf':
        return _table_pdf_response(f'audit_log_{date.today()}.pdf', 'Audit Log', f'Generated {date.today()}', header, rows)
    return csv_export_response(f'audit_log_{date.today()}.csv', header, rows)


# ─────────────────────────────────────────────────────────────
# Danger Zone — admin-only bulk clear per module (see DANGER_ZONE_MODULES).
# ─────────────────────────────────────────────────────────────
@app.route('/admin/danger-zone')
@login_required
@admin_required
def danger_zone():
    modules = []
    for key, label, tables in DANGER_ZONE_MODULES:
        table_counts = [(t, _danger_zone_table_label(t), SYNC_MODELS[t][0].query.count()) for t in tables]
        modules.append(dict(key=key, label=label, table_counts=table_counts,
                            total=sum(c for _, _, c in table_counts)))
    return render_template('admin/danger_zone.html', modules=modules)


@app.route('/admin/danger-zone/clear/<module_key>', methods=['POST'])
@login_required
@admin_required
def danger_zone_clear(module_key):
    if module_key not in DANGER_ZONE_MODULES_BY_KEY:
        abort(404)
    label, module_tables = DANGER_ZONE_MODULES_BY_KEY[module_key]

    # Only tables actually belonging to this module can be cleared — the
    # checkboxes already restrict this client-side, but the module boundary
    # (Franchise can't reach into Finance data) has to hold up server-side
    # even if the posted table list is tampered with.
    selected_tables = [t for t in request.form.getlist('tables') if t in module_tables]
    if not selected_tables:
        flash('Select at least one data type to clear.', 'danger')
        return redirect(url_for('danger_zone'))

    confirm_text = request.form.get('confirm_text', '').strip().upper()
    if confirm_text != module_key.upper():
        flash(f'Confirmation text did not match — nothing was cleared. Type {module_key.upper()} exactly to confirm.', 'danger')
        return redirect(url_for('danger_zone'))

    now = datetime.now(timezone.utc)
    table_totals = {}
    for table_key in selected_tables:
        model = SYNC_MODELS[table_key][0]
        rows = model.query.all()
        for row in rows:
            log_audit('DELETE', table_key, row.id,
                      f'Danger Zone clear ({label}): {_danger_zone_record_label(table_key, row)}')
            row.deleted_at = now
            touch_sync_fields(row)
        table_totals[table_key] = len(rows)
        if rows:
            log_audit('DANGER_ZONE_CLEAR', table_key, None,
                      f'{label} — cleared {len(rows)} {table_key} record(s)')

    db.session.commit()
    grand_total = sum(table_totals.values())
    if grand_total:
        summary = ', '.join(f'{count} {_danger_zone_table_label(t)}' for t, count in table_totals.items() if count)
        flash(f'Cleared {grand_total} {label} record(s): {summary}.', 'warning')
    else:
        flash('Selected data type(s) had no records to clear.', 'success')
    return redirect(url_for('danger_zone'))


# ─────────────────────────────────────────────────────────────
# Multi-site sync — conflict review. Every last-write-wins resolution
# apply_incoming_record() makes gets logged here with both full payloads
# (see SyncConflict, app.py near the sync API), so an unresolved conflict
# is never just silently overwritten and forgotten.
# ─────────────────────────────────────────────────────────────
@app.route('/sync/conflicts')
@login_required
@admin_required
def sync_conflicts():
    page = request.args.get('page', 1, type=int)
    status = request.args.get('status', 'unresolved')
    q = SyncConflict.query
    if status == 'unresolved':
        q = q.filter_by(resolved=False)
    elif status == 'resolved':
        q = q.filter_by(resolved=True)
    conflicts = q.order_by(SyncConflict.detected_at.desc()).paginate(page=page, per_page=30)
    unresolved_count = SyncConflict.query.filter_by(resolved=False).count()
    return render_template('sync/conflicts.html', conflicts=conflicts, status=status,
                           unresolved_count=unresolved_count)


@app.route('/sync/conflicts/<int:cid>/resolve', methods=['POST'])
@login_required
@admin_required
def sync_conflict_resolve(cid):
    conflict = SyncConflict.query.get_or_404(cid)
    conflict.resolved = True
    conflict.resolved_by = current_user.id
    conflict.resolved_at = datetime.now(timezone.utc)
    conflict.resolution_notes = request.form.get('resolution_notes', '').strip()
    db.session.commit()
    flash('Conflict marked resolved.', 'success')
    return redirect(url_for('sync_conflicts', status=request.form.get('return_status', 'unresolved')))


@app.route('/sync/health')
@login_required
@admin_required
def sync_health():
    """Multi-site sync status at a glance. On the hub this is the only
    place any site's last-seen time is visible at all — SyncPeerState is
    deliberately local-instance-only (see its docstring), so the hub has
    no other record of who's actually checking in versus gone stale.
    Doubles as the same page on a spoke, where it instead shows that
    instance's own peer_state against the hub."""
    stale_cutoff = datetime.now(timezone.utc) - timedelta(minutes=15)

    site_rows = []
    for site in SyncSite.query.order_by(SyncSite.display_name).all():
        last_push = site.last_push_at.replace(tzinfo=timezone.utc) if site.last_push_at else None
        last_pull = site.last_pull_at.replace(tzinfo=timezone.utc) if site.last_pull_at else None
        last_seen = max([t for t in (last_push, last_pull) if t], default=None)
        if not site.is_active:
            status = 'disabled'
        elif last_seen is None:
            status = 'never'
        elif last_seen < stale_cutoff:
            status = 'stale'
        else:
            status = 'online'
        site_rows.append({'site': site, 'last_push': last_push, 'last_pull': last_pull,
                          'last_seen': last_seen, 'status': status})

    peer_state = None
    update_state = None
    if app.config['SYNC_ENABLED']:
        peer_state = SyncPeerState.query.filter_by(peer_url=app.config['SYNC_HUB_URL']).first()
        if FROZEN:
            update_state = _get_update_state()

    unresolved_conflicts = SyncConflict.query.filter_by(resolved=False).count()
    conflicts_by_table = dict(
        db.session.query(SyncConflict.table_name, func.count(SyncConflict.id))
        .filter_by(resolved=False).group_by(SyncConflict.table_name).all())

    table_counts = []
    for table in SYNC_TABLE_ORDER:
        model = SYNC_MODELS[table][0]
        total = model.query.execution_options(include_deleted=True).count()
        pending = model.query.filter_by(pending_push=True).execution_options(include_deleted=True).count()
        table_counts.append({'table': table, 'total': total, 'pending': pending,
                             'conflicts': conflicts_by_table.get(table, 0)})

    return render_template('sync/health.html',
                           site_rows=site_rows, peer_state=peer_state,
                           unresolved_conflicts=unresolved_conflicts,
                           table_counts=table_counts,
                           site_id=app.config['SITE_ID'], sync_enabled=app.config['SYNC_ENABLED'],
                           sync_hub_url=app.config['SYNC_HUB_URL'],
                           update_state=update_state, app_version=APP_VERSION, frozen=FROZEN)


@app.route('/sync/check-update', methods=['POST'])
@login_required
@admin_required
def sync_check_update():
    """"Check for Update Now" on Sync Health — bypasses the usual
    SPOKE_UPDATE_CHECK_SECONDS throttle (default 6h) so publishing a
    release on the hub doesn't leave whoever's waiting on this spoke
    with nothing to do but sit on their hands. Only meaningful on a
    spoke's own packaged .exe (see check_for_spoke_update); a no-op
    redirect anywhere else, since the button itself is only ever shown
    there (see sync/health.html)."""
    if not (FROZEN and app.config['SYNC_ENABLED']):
        return redirect(url_for('sync_health'))
    check_for_spoke_update(force=True)
    state = _get_update_state()
    if state.last_error:
        flash(f'Update check failed: {state.last_error}', 'danger')
    elif state.staged_version:
        flash(f'Version {state.staged_version} downloaded and staged — it applies the next time this spoke restarts.', 'success')
    else:
        flash(f'Checked — already on the latest published version ({APP_VERSION}).', 'info')
    return redirect(url_for('sync_health'))


@app.route('/sync/sites')
@login_required
@admin_required
def sync_sites():
    """Onboarding a new spoke used to require shell access to whatever
    machine holds the hub's database (see provision_sync_site.py) just to
    run one INSERT. That's exactly the "tribal knowledge" the original
    sync plan wanted to avoid — this page (and sync_site_add below) does
    the same thing over the admin UI instead, so a new site PC can be
    registered by anyone with admin access, no shell required."""
    sites = SyncSite.query.order_by(SyncSite.created_at.desc()).all()
    return render_template('sync/sites.html', sites=sites)


def register_sync_site(site_id, display_name):
    """Shared by the admin-UI form (sync_site_add) and the self-service
    /api/sync/enroll endpoint — same validation and creation either way,
    just a different caller decides what site_id/display_name to pass in."""
    if not site_id:
        raise ValueError('Site ID is required.')
    if not re.match(r'^[a-z0-9][a-z0-9-]{1,48}[a-z0-9]$', site_id):
        raise ValueError('Site ID must be lowercase letters, numbers, and hyphens only (e.g. site-nairobi-01).')
    check_unique(SyncSite, 'site_id', site_id)
    api_key = secrets.token_urlsafe(32)
    site = SyncSite(
        site_id=site_id,
        api_key_hash=generate_password_hash(api_key),
        display_name=display_name or site_id,
        is_active=True,
    )
    db.session.add(site)
    db.session.flush()
    return site, api_key


@app.route('/sync/sites/add', methods=['GET', 'POST'])
@login_required
@admin_required
@handle_form_errors
def sync_site_add():
    if request.method == 'POST':
        site_id = request.form.get('site_id', '').strip().lower()
        display_name = request.form.get('display_name', '').strip()
        site, api_key = register_sync_site(site_id, display_name)
        log_audit('CREATE', 'sync_sites', site.id, f'Registered spoke site {site.site_id}')
        db.session.commit()
        # The plaintext key only ever exists in this one response — only
        # its hash is stored (see SyncSite.api_key_hash) — so it's shown
        # once here and never retrievable again, same guarantee
        # provision_sync_site.py gives on the command line.
        return render_template('sync/site_key.html', site=site, api_key=api_key,
                               sync_hub_url=request.url_root.rstrip('/'))
    return render_template('sync/site_form.html')


@app.route('/sync/sites/<int:site_id>/toggle', methods=['POST'])
@login_required
@admin_required
def sync_site_toggle(site_id):
    site = SyncSite.query.filter_by(id=site_id).first_or_404()
    site.is_active = not site.is_active
    log_audit('UPDATE', 'sync_sites', site.id,
              f'Set sync site {site.site_id} active={site.is_active}')
    db.session.commit()
    flash(f'Site "{site.site_id}" {"activated" if site.is_active else "deactivated"}.', 'info')
    return redirect(url_for('sync_sites'))


# ─────────────────────────────────────────────────────────────
# Spoke Releases — hub-side publishing of a new spoke .exe build. A spoke
# checks /api/spoke/latest-version and downloads whichever release here is
# flagged is_latest (see check_for_spoke_update below). This page is only
# meaningful on the hub: a spoke's own copy of the spoke_releases table
# stays empty, same as Sync Sites.
# ─────────────────────────────────────────────────────────────
@app.route('/sync/releases')
@login_required
@admin_required
def spoke_releases():
    releases = SpokeRelease.query.order_by(SpokeRelease.created_at.desc()).all()
    return render_template('sync/releases.html', releases=releases, app_version=app.config['APP_VERSION'])


@app.route('/sync/releases/upload', methods=['POST'])
@login_required
@admin_required
def spoke_release_upload():
    version = request.form.get('version', '').strip()
    notes = request.form.get('notes', '').strip()
    file = request.files.get('file')
    try:
        if not version or not re.match(r'^[A-Za-z0-9][A-Za-z0-9._-]{0,38}$', version):
            raise ValueError('Version must be non-empty and contain only letters, numbers, dots, dashes and underscores.')
        if SpokeRelease.query.filter_by(version=version).first():
            raise ValueError(f'Version "{version}" has already been published.')
        if not file or not file.filename:
            raise ValueError('Choose the release .zip file to upload.')
        if not file.filename.lower().endswith('.zip'):
            raise ValueError('The release file must be a .zip (zip up the whole dist\\TransportERP folder).')

        os.makedirs(app.config['SPOKE_RELEASES_DIR'], exist_ok=True)
        filename = secure_filename(f'{version}.zip')
        dest_path = os.path.join(app.config['SPOKE_RELEASES_DIR'], filename)

        hasher = hashlib.sha256()
        size = 0
        with open(dest_path, 'wb') as out:
            for chunk in iter(lambda: file.stream.read(1024 * 1024), b''):
                hasher.update(chunk)
                size += len(chunk)
                out.write(chunk)

        # New upload becomes the one every spoke pulls next — see
        # SpokeRelease.is_latest.
        SpokeRelease.query.filter_by(is_latest=True).update({'is_latest': False})
        release = SpokeRelease(version=version, filename=filename, sha256=hasher.hexdigest(),
                               file_size=size, notes=notes, is_latest=True, created_by=current_user.id)
        db.session.add(release)
        log_audit('CREATE', 'spoke_releases', None, f'Published spoke release {version} ({size:,} bytes)')
        db.session.commit()
        flash(f'Spoke release {version} published — every spoke will pick it up on its next check-in.', 'success')
    except ValueError as e:
        db.session.rollback()
        if file and file.filename:
            # Clean up a partially-written file from a validation failure
            # that happened mid/after the stream write (e.g. duplicate
            # version caught before writing is fine; this covers nothing
            # currently, but keeps the directory tidy if that check order
            # ever changes).
            stray = os.path.join(app.config['SPOKE_RELEASES_DIR'], secure_filename(f'{version}.zip'))
            if os.path.exists(stray) and not SpokeRelease.query.filter_by(version=version).first():
                os.remove(stray)
        flash(str(e), 'danger')
    return redirect(url_for('spoke_releases'))


@app.route('/sync/releases/<int:release_id>/set-latest', methods=['POST'])
@login_required
@admin_required
def spoke_release_set_latest(release_id):
    release = SpokeRelease.query.filter_by(id=release_id).first_or_404()
    SpokeRelease.query.filter_by(is_latest=True).update({'is_latest': False})
    release.is_latest = True
    log_audit('UPDATE', 'spoke_releases', release.id,
              f'Marked spoke release {release.version} as latest (rollback/republish)')
    db.session.commit()
    flash(f'Version {release.version} is now what every spoke will update to.', 'success')
    return redirect(url_for('spoke_releases'))


@app.route('/sync/releases/<int:release_id>/delete', methods=['POST'])
@login_required
@admin_required
def spoke_release_delete(release_id):
    release = SpokeRelease.query.filter_by(id=release_id).first_or_404()
    if release.is_latest:
        flash('Mark a different version as latest before deleting this one.', 'danger')
        return redirect(url_for('spoke_releases'))
    path = os.path.join(app.config['SPOKE_RELEASES_DIR'], release.filename)
    if os.path.exists(path):
        os.remove(path)
    log_audit('DELETE', 'spoke_releases', release.id, f'Deleted spoke release {release.version}')
    db.session.delete(release)
    db.session.commit()
    flash(f'Release {release.version} deleted.', 'warning')
    return redirect(url_for('spoke_releases'))


@app.context_processor
def inject_unresolved_sync_conflicts_count():
    """Powers the "Sync Conflicts" sidebar badge in base.html on every
    page — only admins see that nav section at all, and only admins can
    act on a conflict, so the query only runs for them."""
    if current_user.is_authenticated and current_user.role == 'admin':
        return {'unresolved_sync_conflicts_count': SyncConflict.query.filter_by(resolved=False).count()}
    return {}


@app.context_processor
def inject_pending_franchise_vehicles_count():
    """Powers the "Vehicles" sidebar badge under Franchise — counts
    quick-registered vehicles (see franchise_vehicle_quick_add) still
    awaiting an admin's review. Admin-only, same reasoning as the sync
    conflicts badge above."""
    if current_user.is_authenticated and current_user.role == 'admin':
        return {'pending_franchise_vehicles_count': FranchiseVehicle.query.filter_by(pending_review=True).count()}
    return {}


# ─────────────────────────────────────────────────────────────
# Import Batches — audit trail, quarantine error export, revert
# ─────────────────────────────────────────────────────────────
IMPORT_TARGET_PERMISSIONS = {
    'ledger': ('daily_logs', 'crew_portal'),
    'franchise_workbook': ('franchise',),
    'store_purchases': ('store',),
}

IMPORT_REVERT_MODELS = {
    'daily_logs': DailyLog,
    'fuel_logs': FuelLog,
    'franchise_vehicles': FranchiseVehicle,
    'franchise_collections': FranchiseCollection,
    'store_purchases': StorePurchase,
    'spare_parts': SparePart,
}


@app.route('/imports')
@login_required
@admin_required
def import_history():
    page = request.args.get('page', 1, type=int)
    batches = ImportBatch.query.order_by(ImportBatch.created_at.desc()).paginate(page=page, per_page=30)
    return render_template('imports/history.html', batches=batches)


@app.route('/imports/<int:batch_id>/errors.csv')
@login_required
def import_batch_errors_csv(batch_id):
    batch = ImportBatch.query.get_or_404(batch_id)
    perms = IMPORT_TARGET_PERMISSIONS.get(batch.target_type, ())
    if not any(current_user.has_permission(p) for p in perms):
        flash('You do not have permission to access that import.', 'danger')
        return redirect(first_permitted_url(current_user))

    rows = batch.error_row_list
    out = io.StringIO()
    if rows:
        writer = csv.DictWriter(out, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)
    resp = make_response(out.getvalue())
    resp.headers['Content-Type'] = 'text/csv'
    resp.headers['Content-Disposition'] = (
        f'attachment; filename={batch.target_type}_import_errors_{batch.id}.csv')
    return resp


@app.route('/imports/<int:batch_id>/revert', methods=['POST'])
@login_required
@admin_required
def import_batch_revert(batch_id):
    batch = ImportBatch.query.get_or_404(batch_id)
    if batch.status == 'reverted':
        flash('That import was already reverted.', 'warning')
        return redirect(url_for('import_history'))

    deleted = 0
    # store_purchases must be reverted before spare_parts: rolling back a
    # purchase reads obj.part, which the global soft-delete query filter
    # (see _exclude_soft_deleted_rows) would hide once that part itself has
    # been soft-deleted — so a part auto-created by this same import (whose
    # ImportBatchRecord can land in either order) must not be deleted first.
    revert_order = {'spare_parts': 1}
    for rec in sorted(batch.records, key=lambda r: revert_order.get(r.target_table, 0)):
        model = IMPORT_REVERT_MODELS.get(rec.target_table)
        # Deliberately .get(), not filter_by().first() — this must still
        # find a record even if it was already soft-deleted by a normal
        # user action, so reverting stays idempotent instead of silently
        # skipping it.
        obj = model.query.get(rec.record_id) if model else None
        if obj:
            # A StorePurchase rolled quantity into its part's stock on hand
            # (and cost into its weighted-average cost_price) at import time —
            # unlike the other revertible models, undoing it means more than
            # a soft-delete, or the part's stock would stay inflated forever.
            # Mirrors store_purchase_delete's own manual-delete behavior,
            # including its same caveat: historical average cost isn't
            # recomputed, only quantity.
            if rec.target_table == 'store_purchases' and obj.deleted_at is None:
                obj.part.quantity_on_hand = max(0, obj.part.quantity_on_hand - obj.quantity)
                touch_sync_fields(obj.part)
            obj.deleted_at = datetime.now(timezone.utc)
            touch_sync_fields(obj)
            deleted += 1

    batch.status = 'reverted'
    batch.reverted_at = datetime.now(timezone.utc)
    batch.reverted_by = current_user.id
    log_audit('DELETE', batch.target_type, batch.id,
              f'Reverted import batch #{batch.id} ({batch.filename}) — removed {deleted} record(s)')
    db.session.commit()
    flash(f'Reverted import #{batch.id} — removed {deleted} record(s).', 'success')
    return redirect(url_for('import_history'))


# ─────────────────────────────────────────────────────────────
# API (chart data)
# ─────────────────────────────────────────────────────────────
@app.route('/api/revenue/monthly')
@login_required
@permission_required('dashboard')
def api_revenue_monthly():
    today = date.today()
    data = []
    for i in range(5, -1, -1):
        m = today.month - i
        y = today.year
        while m <= 0:
            m += 12
            y -= 1
        start = date(y, m, 1)
        end = date(y + 1, 1, 1) if m == 12 else date(y, m + 1, 1)
        rev = db.session.query(func.sum(DailyLog.gross_revenue)).filter(
            DailyLog.log_date >= start, DailyLog.log_date < end).scalar() or 0
        maint = db.session.query(func.sum(MaintenanceLog.total_cost)).filter(
            MaintenanceLog.log_date >= start, MaintenanceLog.log_date < end).scalar() or 0
        data.append({
            'month': start.strftime('%b %Y'),
            'revenue': float(rev),
            'expenses': float(maint),
            'profit': float(rev - maint),
        })
    return jsonify(data)


@app.route('/api/vehicles/performance')
@login_required
@permission_required('dashboard')
def api_vehicle_performance():
    today = date.today()
    df, dt = query_date_range(default_from=today.replace(day=1), default_to=today)
    result = db.session.query(
        Vehicle.registration,
        func.sum(DailyLog.gross_revenue).label('revenue'),
        func.count(DailyLog.id).label('days'),
    ).join(DailyLog, Vehicle.id == DailyLog.vehicle_id).filter(
        DailyLog.log_date.between(df, dt)
    ).group_by(Vehicle.id).all()
    return jsonify([{'vehicle': r.registration, 'revenue': float(r.revenue or 0),
                     'days': r.days} for r in result])


@app.route('/api/expenses/breakdown')
@login_required
@permission_required('dashboard')
def api_expenses_breakdown():
    today = date.today()
    df, dt = query_date_range(default_from=today.replace(day=1), default_to=today)
    maint = db.session.query(func.sum(MaintenanceLog.total_cost)).filter(
        MaintenanceLog.log_date.between(df, dt)).scalar() or 0
    return jsonify({'maintenance': float(maint)})


@app.route('/api/csrf-token')
@login_required
def api_csrf_token():
    """Fresh CSRF token for offline.js — fetched right before every queued
    submission is (re)sent, since a token minted at page-load can go stale
    (default 1h) by the time a device that went offline reconnects."""
    return jsonify({'csrf_token': generate_csrf()})


@app.route('/api/ping')
def api_ping():
    """Unauthenticated connectivity probe for offline.js's topbar
    Online/Offline indicator — no @login_required, since "is the server
    reachable" needs to answer true even for a logged-out/expired session."""
    return jsonify({'status': 'ok'})


@app.route('/api/refdata')
@login_required
def api_refdata():
    """Reference data (vehicles/drivers/parts/franchise vehicles/expense
    categories) for the dropdowns on the 9 offline-capable forms. Fetched
    and cached in IndexedDB by offline.js on every page that has one of
    those forms, so a form opened from the service worker's page cache
    while offline can repopulate its dropdowns from real data instead of
    whatever happened to be baked into the HTML the last time it was
    cached. Read-only reference data (names/plates), so gated on being
    logged in at all rather than a specific module permission."""
    vehicles = Vehicle.query.filter_by(status='active').order_by(Vehicle.registration).all()
    drivers = Driver.query.filter_by(role='driver', status='active').order_by(Driver.name).all()
    parts = SparePart.query.filter_by(status='active').order_by(SparePart.name).all()
    franchise_vehicles = FranchiseVehicle.query.filter_by(status='active').order_by(
        FranchiseVehicle.franchisee_name).all()
    headings = ExpenseCategory.query.filter_by(parent_id=None).order_by(ExpenseCategory.name).all()

    def part_label(p):
        label = p.name
        if p.part_number:
            label += f' ({p.part_number})'
        label += f' — {qty_filter(p.quantity_on_hand)} {p.unit} in stock'
        return label

    expense_categories = []
    for h in headings:
        if h.children:
            expense_categories.append({'id': h.id, 'label': f'{h.name} (general)', 'default_amount': h.default_amount})
            for c in sorted(h.children, key=lambda x: x.name):
                expense_categories.append({'id': c.id, 'label': f'{h.name} — {c.name}', 'default_amount': c.default_amount})
        else:
            expense_categories.append({'id': h.id, 'label': h.name, 'default_amount': h.default_amount})

    return jsonify({
        'vehicles': [{'id': v.id, 'label': f'{v.registration} — {v.make} {v.model}'} for v in vehicles],
        'drivers': [{'id': d.id, 'label': d.name} for d in drivers],
        'parts': [{'id': p.id, 'label': part_label(p), 'selling_price': p.selling_price,
                   'quantity_on_hand': p.quantity_on_hand, 'cost_price': p.cost_price, 'unit': p.unit}
                  for p in parts],
        'franchise_vehicles': [{'id': v.id, 'label': f'{v.number_plate} — {v.franchisee_name}'}
                               for v in franchise_vehicles],
        'expense_categories': expense_categories,
    })


@app.route('/api/precache-urls')
@login_required
def api_precache_urls():
    """URLs of every main list/dashboard page this user can see (see
    PRECACHE_PAGES) — offline.js fetches this in the background while
    online and warms each one into the service worker's page cache, so
    the whole business's current data is available offline even for a
    page this browser has never actually visited. Same has_permission
    check as base.html's nav, so this never hands out a URL the user
    couldn't reach anyway."""
    urls = [url_for(endpoint) for perm, endpoint in PRECACHE_PAGES if current_user.has_permission(perm)]
    return jsonify({'urls': urls})


# ─────────────────────────────────────────────────────────────
# Multi-site sync API — Phase 2 of the local-server sync plan. A spoke
# (a local-server PC at a site) POSTs its pending changes to /api/sync/push
# and GETs everything changed since its last checkpoint from
# /api/sync/pull. Both directions share apply_incoming_record() below for
# the actual upsert, but use different conflict-detection strategies (see
# that function's docstring) — a push and a pull aren't symmetric: only
# the pusher knows what state it edited from, only the puller can know
# whether it has an unsynced local edit at risk.
#
# Scope so far: the 11 Phase-1 operational tables (fleet/spares/expenses)
# plus the 6 Phase-5 franchise/compliance tables. Loans/payables/
# receivables tables are deliberately NOT here yet — see the phased
# rollout plan. Users are never synced (see
# SYNC_MODELS below not including 'users') — accounts stay central-only.
# ─────────────────────────────────────────────────────────────

# table name -> (Model class, plain data columns to sync, {fk column: referenced table})
# Order matters: parents (Tier 0) before children (Tier 1/2) — see the FK
# map dependencies. Foreign keys travel as the referenced row's sync_uuid,
# never a local integer id, since two offline sites can independently mint
# the same integer id for different rows. created_by/updated_by are the
# exception — they reference Users, which are pull-only mirrored with the
# same id everywhere, so those integers are safe to sync as plain values.
SYNC_MODELS = {
    'vehicles': (Vehicle, (
        'registration', 'make', 'model', 'year', 'acquisition_cost', 'status',
        'fuel_type', 'daily_target', 'insurance_provider', 'insurance_policy_number',
        'insurance_type', 'insurance_expiry', 'created_at', 'deleted_at',
    ), {}),
    'routes': (Route, (
        'name', 'start_point', 'end_point', 'distance_km', 'fare_rate', 'status', 'created_at', 'deleted_at',
    ), {}),
    'spare_parts': (SparePart, (
        'name', 'part_number', 'unit', 'cost_price', 'markup_percent', 'quantity_on_hand',
        'reorder_level', 'status', 'notes', 'created_by', 'created_at', 'deleted_at',
    ), {}),
    'expense_categories': (ExpenseCategory, (
        'name', 'default_amount', 'created_at', 'deleted_at',
    ), {'parent_id': 'expense_categories'}),
    'drivers': (Driver, (
        'name', 'license_number', 'id_number', 'phone', 'role', 'commission_rate', 'status',
        'next_of_kin_name', 'next_of_kin_phone', 'next_of_kin_relationship', 'created_at', 'deleted_at',
    ), {'paired_driver_id': 'drivers', 'assigned_vehicle_id': 'vehicles'}),
    'expenses': (Expense, (
        'expense_date', 'description', 'amount', 'created_by', 'created_at', 'deleted_at',
    ), {'category_id': 'expense_categories', 'vehicle_id': 'vehicles'}),
    'store_purchases': (StorePurchase, (
        'purchase_date', 'quantity', 'unit_cost', 'total_cost', 'supplier', 'notes',
        'created_by', 'created_at', 'deleted_at',
    ), {'part_id': 'spare_parts'}),
    'daily_logs': (DailyLog, (
        'log_date', 'trips_completed', 'gross_revenue', 'garnish', 'reason_for_shortfall',
        'notes', 'created_by', 'updated_by', 'created_at', 'deleted_at',
    ), {'vehicle_id': 'vehicles', 'driver_id': 'drivers', 'conductor_id': 'drivers', 'route_id': 'routes'}),
    'driver_deposits': (DriverDeposit, (
        'deposit_date', 'amount', 'notes', 'created_by', 'created_at', 'deleted_at',
    ), {'driver_id': 'drivers', 'vehicle_id': 'vehicles'}),
    'fuel_logs': (FuelLog, (
        'log_date', 'liters', 'cost_per_liter', 'total_cost', 'odometer', 'supplier', 'notes',
        'created_by', 'created_at', 'deleted_at',
    ), {'vehicle_id': 'vehicles'}),
    'maintenance_logs': (MaintenanceLog, (
        'log_date', 'description', 'parts_cost', 'labor_cost', 'total_cost', 'mechanic', 'notes',
        'created_by', 'created_at', 'deleted_at',
    ), {'vehicle_id': 'vehicles'}),
    'store_sales': (StoreSale, (
        'sale_date', 'quantity', 'unit_cost', 'unit_price', 'total_amount', 'customer_name',
        'notes', 'created_by', 'created_at', 'deleted_at',
    ), {'part_id': 'spare_parts', 'vehicle_id': 'vehicles'}),
    'store_adjustments': (StoreAdjustment, (
        'adjustment_date', 'quantity_before', 'quantity_after', 'cost_price_before', 'cost_price_after',
        'reason', 'notes', 'created_by', 'created_at', 'deleted_at',
    ), {'part_id': 'spare_parts'}),
    # Phase 5 — franchise/compliance tables.
    'franchise_vehicles': (FranchiseVehicle, (
        'number_plate', 'franchisee_name', 'status', 'daily_fee', 'weekly_fee', 'amount_owed',
        'notes', 'created_at', 'deleted_at',
    ), {}),
    'vehicle_documents': (VehicleDocument, (
        'doc_type', 'reference_number', 'issue_date', 'expiry_date', 'notes', 'created_at', 'deleted_at',
    ), {'vehicle_id': 'vehicles'}),
    'maintenance_schedules': (MaintenanceSchedule, (
        'description', 'interval_days', 'interval_km', 'last_done_date', 'last_done_odometer',
        'next_due_date', 'next_due_odometer', 'status', 'created_at', 'deleted_at',
    ), {'vehicle_id': 'vehicles'}),
    'franchise_daily_income': (FranchiseDailyIncome, (
        'entry_date', 'income', 'exp_traffic_fines', 'exp_facilitation_fees', 'exp_workshop',
        'exp_wages', 'other_expenditure', 'deposited', 'description', 'created_by', 'created_at', 'deleted_at',
    ), {'vehicle_id': 'franchise_vehicles'}),
    'franchise_weekly_income': (FranchiseWeeklyIncome, (
        'week_start', 'income', 'exp_traffic_fines', 'exp_facilitation_fees', 'exp_workshop',
        'exp_wages', 'other_expenditure', 'deposited', 'description', 'created_by', 'created_at', 'deleted_at',
    ), {'vehicle_id': 'franchise_vehicles'}),
    'franchise_collections': (FranchiseCollection, (
        'entry_date', 'frequency', 'amount', 'expense', 'notes', 'created_by', 'created_at', 'deleted_at',
    ), {'vehicle_id': 'franchise_vehicles'}),
    'franchise_expense_categories': (FranchiseExpenseCategory, (
        'name', 'created_at', 'deleted_at',
    ), {}),
    'franchise_operational_expenses': (FranchiseOperationalExpense, (
        'expense_date', 'amount', 'description', 'created_by', 'created_at', 'deleted_at',
    ), {'category_id': 'franchise_expense_categories'}),
    # Phase 6 — financial tables. Full read/write on spokes, same as
    # every other synced table (Phase 6a's pull-only, spoke-write-blocked
    # burn-in period has ended).
    'loans': (Loan, (
        'lender', 'principal', 'interest_rate', 'start_date', 'term_months', 'status',
        'notes', 'created_by', 'created_at', 'deleted_at',
    ), {}),
    'payables': (Payable, (
        'supplier_name', 'description', 'amount', 'invoice_date', 'due_date', 'status',
        'paid_date', 'created_by', 'created_at', 'deleted_at',
    ), {}),
    'receivables': (Receivable, (
        'client_name', 'description', 'amount', 'invoice_date', 'due_date', 'status',
        'collected_date', 'created_by', 'created_at', 'deleted_at',
    ), {}),
    'capital_contributions': (CapitalContribution, (
        'contributor', 'amount', 'contribution_date', 'notes', 'created_by', 'created_at', 'deleted_at',
    ), {}),
    'owner_drawings': (OwnerDrawing, (
        'amount', 'drawing_date', 'notes', 'created_by', 'created_at', 'deleted_at',
    ), {}),
    'budgets': (Budget, (
        'category', 'month', 'amount', 'created_by', 'created_at', 'deleted_at',
    ), {}),
    'loan_payments': (LoanPayment, (
        'payment_date', 'amount', 'notes', 'created_by', 'created_at', 'deleted_at',
    ), {'loan_id': 'loans'}),
    'commission_payments': (CommissionPayment, (
        'payment_date', 'amount', 'period_start', 'period_end', 'method', 'notes',
        'created_by', 'created_at', 'deleted_at',
    ), {'driver_id': 'drivers'}),
    'payroll_deductions': (PayrollDeduction, (
        'deduction_date', 'amount', 'reason', 'created_by', 'created_at', 'deleted_at',
    ), {'driver_id': 'drivers'}),
}

# Dependency order for apply — parents before children (see FK maps above).
SYNC_TABLE_ORDER = [
    'vehicles', 'routes', 'spare_parts', 'expense_categories', 'franchise_vehicles',
    'franchise_expense_categories',
    'loans', 'payables', 'receivables', 'capital_contributions', 'owner_drawings', 'budgets',
    'drivers', 'expenses', 'store_purchases', 'vehicle_documents', 'maintenance_schedules',
    'franchise_daily_income', 'franchise_weekly_income', 'franchise_collections',
    'franchise_operational_expenses',
    'loan_payments', 'commission_payments', 'payroll_deductions',
    'daily_logs', 'driver_deposits', 'fuel_logs', 'maintenance_logs', 'store_sales', 'store_adjustments',
]

# ─────────────────────────────────────────────────────────────
# Danger Zone — admin-only bulk clear, grouped to match the sidebar's own
# module boundaries so "clear Franchise" can't reach into Finance data and
# vice versa. Every table listed here must already be a SYNC_MODELS entry:
# the clear itself reuses the same soft-delete (deleted_at) + touch_sync_fields
# path as every individual delete route, rather than a raw DELETE, so a
# module clear tombstones and propagates through multi-site sync exactly
# like any other deletion instead of looking like the rows never existed to
# a spoke that hasn't synced yet.
DANGER_ZONE_MODULES = [
    ('franchise', 'Franchise', ['franchise_daily_income', 'franchise_weekly_income', 'franchise_collections', 'franchise_vehicles', 'franchise_operational_expenses', 'franchise_expense_categories']),
    ('fleet', 'Fleet & Compliance', ['vehicles', 'drivers', 'routes', 'vehicle_documents', 'maintenance_schedules']),
    ('ledger', 'Daily Transactions (Crew Ledger)', ['daily_logs', 'driver_deposits']),
    ('operations', 'Operations (Fuel & Maintenance Logs)', ['fuel_logs', 'maintenance_logs']),
    ('finance', 'Finance Ledger', ['loans', 'loan_payments', 'payables', 'receivables', 'capital_contributions', 'owner_drawings', 'budgets', 'expenses', 'expense_categories', 'commission_payments', 'payroll_deductions']),
    ('store', 'Spares Store', ['spare_parts', 'store_purchases', 'store_sales', 'store_adjustments']),
]
DANGER_ZONE_MODULES_BY_KEY = {key: (label, tables) for key, label, tables in DANGER_ZONE_MODULES}

# Display name for each table's checkbox in the Danger Zone UI — falls back
# to a title-cased version of the raw table name for anything not listed.
DANGER_ZONE_TABLE_LABELS = {
    'franchise_daily_income': 'Daily Income', 'franchise_weekly_income': 'Weekly Income',
    'franchise_collections': 'Collections (legacy)', 'franchise_vehicles': 'Franchise Vehicles',
    'franchise_operational_expenses': 'Operational Expenses', 'franchise_expense_categories': 'Operational Expense Sub-headings',
    'vehicles': 'Vehicles', 'drivers': 'Drivers', 'routes': 'Routes',
    'vehicle_documents': 'Vehicle Documents', 'maintenance_schedules': 'Maintenance Schedules',
    'daily_logs': 'Daily Logs', 'fuel_logs': 'Fuel Logs', 'maintenance_logs': 'Maintenance Logs',
    'loans': 'Loans', 'loan_payments': 'Loan Payments', 'payables': 'Payables',
    'receivables': 'Receivables', 'capital_contributions': 'Capital Contributions',
    'owner_drawings': 'Owner Drawings', 'budgets': 'Budgets', 'expenses': 'Expenses',
    'expense_categories': 'Expense Categories', 'commission_payments': 'Commission Payments',
    'payroll_deductions': 'Payroll Deductions',
    'spare_parts': 'Spare Parts', 'store_purchases': 'Store Purchases', 'store_sales': 'Store Sales',
    'store_adjustments': 'Store Adjustments',
}


def _danger_zone_table_label(table_key):
    return DANGER_ZONE_TABLE_LABELS.get(table_key, table_key.replace('_', ' ').title())

# Which columns identify a row in its audit-log description — just enough
# to recognize the record later without re-deriving it from every model's
# full column list.
_DANGER_ZONE_LABEL_FIELDS = {
    'vehicles': ('registration',), 'routes': ('name',), 'spare_parts': ('name',),
    'expense_categories': ('name',), 'drivers': ('name',),
    'expenses': ('expense_date', 'description', 'amount'),
    'store_purchases': ('purchase_date', 'supplier', 'total_cost'),
    'daily_logs': ('log_date', 'gross_revenue'),
    'fuel_logs': ('log_date', 'total_cost'),
    'maintenance_logs': ('log_date', 'total_cost'),
    'store_sales': ('sale_date', 'customer_name', 'total_amount'),
    'store_adjustments': ('adjustment_date', 'reason', 'quantity_after'),
    'franchise_vehicles': ('number_plate', 'franchisee_name'),
    'vehicle_documents': ('doc_type', 'reference_number'),
    'maintenance_schedules': ('description',),
    'franchise_daily_income': ('entry_date', 'income'),
    'franchise_weekly_income': ('week_start', 'income'),
    'franchise_collections': ('entry_date', 'frequency', 'amount'),
    'franchise_operational_expenses': ('expense_date', 'amount'),
    'franchise_expense_categories': ('name',),
    'loans': ('lender', 'principal'),
    'payables': ('supplier_name', 'amount'),
    'receivables': ('client_name', 'amount'),
    'capital_contributions': ('contributor', 'amount'),
    'owner_drawings': ('drawing_date', 'amount'),
    'budgets': ('category', 'month', 'amount'),
    'loan_payments': ('payment_date', 'amount'),
    'commission_payments': ('payment_date', 'amount'),
    'payroll_deductions': ('deduction_date', 'amount', 'reason'),
}


def _danger_zone_record_label(table_key, row):
    parts = [f'id={row.id}']
    for field in _DANGER_ZONE_LABEL_FIELDS.get(table_key, ()):
        value = getattr(row, field, None)
        if value is not None:
            parts.append(f'{field}={value}')
    return ' '.join(parts)

# ─────────────────────────────────────────────────────────────
# Phase 7 — soft-delete / tombstone sync. Deleting a record now sets
# deleted_at instead of removing the row (see the delete routes below),
# so the deletion itself can propagate to other instances as a normal
# synced field change, the same way any other edit does — an actually
# DELETE'd row would just look like the record never existed to a spoke
# that hadn't synced it yet, with no way to tell "never existed" apart
# from "existed, then got removed."
#
# Rather than manually auditing and adding "WHERE deleted_at IS NULL" to
# every list/report query across this file (the largest, most
# error-prone mechanical change the phased plan called out — trivial to
# miss a call site, and a manual audit can't reach relationship-based
# access like vehicle.fuel_logs at all), a single global SQLAlchemy query
# filter excludes soft-deleted rows from every SELECT against a synced
# model, automatically, everywhere, including relationship loads.
#
# The sync engine's own bookkeeping queries (apply_incoming_record's
# upsert lookup, FK resolution, the push/pull selection queries) need to
# see soft-deleted rows too — that's the whole point of a tombstone — so
# those explicitly opt out via .execution_options(include_deleted=True).
_SOFT_DELETE_MODELS = tuple(model for model, _fields, _fk in SYNC_MODELS.values())
# with_loader_criteria takes one entity (or a common base) per call, not an
# arbitrary tuple of unrelated classes — build one option per model, once,
# and apply the whole set to every query rather than re-building it per call.
_SOFT_DELETE_CRITERIA_OPTIONS = tuple(
    with_loader_criteria(model, lambda cls: cls.deleted_at.is_(None), include_aliases=True)
    for model in _SOFT_DELETE_MODELS
)


@event.listens_for(db.session, 'do_orm_execute')
def _exclude_soft_deleted_rows(execute_state):
    if not execute_state.is_select:
        return
    if execute_state.execution_options.get('include_deleted', False):
        return
    execute_state.statement = execute_state.statement.options(*_SOFT_DELETE_CRITERIA_OPTIONS)


_SYNC_EPOCH = datetime(1970, 1, 1)  # naive — see _parse_sync_dt


def _parse_sync_dt(value):
    """Parse an incoming ISO datetime string to a naive UTC datetime.
    SQLite has no timezone-aware column type, so every DateTime column in
    this app round-trips as naive (implicitly UTC) once read back from the
    DB — an incoming tz-aware string (e.g. '...+00:00' from another
    instance's isoformat()) has to be normalized the same way before it's
    comparable to existing.updated_at, or every comparison raises
    TypeError: can't compare offset-naive and offset-aware datetimes."""
    if not value:
        return None
    dt = datetime.fromisoformat(value)
    if dt.tzinfo is not None:
        dt = dt.astimezone(timezone.utc).replace(tzinfo=None)
    return dt


def _sync_sortable_dt(dt):
    """None-safe comparison key — a row that's never been touched by
    touch_sync_fields() sorts as infinitely old, so any real incoming
    update always wins over it."""
    return dt if dt is not None else _SYNC_EPOCH


def _serialize_sync_value(model, field, value):
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return value


def _deserialize_sync_value(model, field, value):
    if value is None:
        return None
    py_type = model.__table__.columns[field].type.python_type
    if py_type is datetime:
        return _parse_sync_dt(value)
    if py_type is date:
        return date.fromisoformat(value)
    return value


def _resolve_sync_fk(ref_table, sync_uuid_value):
    if not sync_uuid_value:
        return None
    ref_model = SYNC_MODELS[ref_table][0]
    ref_obj = (ref_model.query.execution_options(include_deleted=True)
              .filter_by(sync_uuid=sync_uuid_value).first())
    return ref_obj.id if ref_obj else None


def serialize_record_for_sync(table, obj):
    """One row -> the (fields, fk) shape /api/sync/pull hands to a caller
    and apply_incoming_record() expects back — the same shape used in both
    directions."""
    model, fields, fk_map = SYNC_MODELS[table]
    field_values = {f: _serialize_sync_value(model, f, getattr(obj, f)) for f in fields}
    fk_values = {}
    for fk_col, ref_table in fk_map.items():
        ref_id = getattr(obj, fk_col)
        if ref_id is None:
            fk_values[fk_col] = None
        else:
            ref_model = SYNC_MODELS[ref_table][0]
            ref_obj = db.session.get(ref_model, ref_id)
            fk_values[fk_col] = ref_obj.sync_uuid if ref_obj else None
    return field_values, fk_values


def apply_incoming_record(table, item, incoming_site_id, direction):
    """Upsert one incoming sync item by sync_uuid — the idempotency key,
    same principle as the existing already_synced()/client_id pattern for
    the browser offline queue.

    Conflict detection is deliberately DIFFERENT depending on direction,
    because the two sides of a sync exchange don't have symmetric
    information:

    - direction='push' (a spoke telling the hub about its own edit): the
      pusher includes base_updated_at — what IT believed this row's
      updated_at was right before making its edit (see
      last_synced_updated_at / sync_push_to_hub). The hub compares that
      against its OWN CURRENT updated_at for the row. A mismatch means
      something else changed this row on the hub after the pusher last
      knew about it — e.g. a DIFFERENT spoke's push already landed — a
      genuine conflict regardless of the hub's own pending_push (the hub
      never has "its own" pending edit in this scenario; the race is
      between two senders, not sender-vs-receiver). Comparing against
      pending_push here would miss exactly this case: the second spoke's
      push would silently overwrite the first spoke's edit with no
      conflict ever logged, since the hub's pending_push is already
      False by the time the second push arrives.
    - direction='pull' (a spoke applying what the hub sent it): the hub
      has no meaningful "base" to offer — it just reports current truth.
      What matters instead is whether the RECEIVING spoke has its own
      unpushed local edit that the incoming pull would silently clobber —
      exactly what pending_push (on the spoke's own row) already tracks.
    """
    model, fields, fk_map = SYNC_MODELS[table]
    sync_uuid_value = item['sync_uuid']
    existing = model.query.execution_options(include_deleted=True).filter_by(sync_uuid=sync_uuid_value).first()

    resolved_fks = {}
    for fk_col, ref_table in fk_map.items():
        incoming_ref_uuid = (item.get('fk') or {}).get(fk_col)
        resolved_fks[fk_col] = _resolve_sync_fk(ref_table, incoming_ref_uuid)
        if incoming_ref_uuid and resolved_fks[fk_col] is None:
            return {'sync_uuid': sync_uuid_value, 'status': 'rejected', 'reason': 'fk_missing'}

    incoming_updated_at = _parse_sync_dt(item.get('updated_at')) or datetime.now(timezone.utc).replace(tzinfo=None)

    def _apply_fields(obj):
        for f in fields:
            setattr(obj, f, _deserialize_sync_value(model, f, item['fields'].get(f)))
        for fk_col, local_id in resolved_fks.items():
            setattr(obj, fk_col, local_id)
        obj.updated_at = incoming_updated_at
        obj.last_modified_site = incoming_site_id
        obj.pending_push = False  # arrived from elsewhere — nothing to push back for this edit
        obj.last_synced_updated_at = incoming_updated_at  # what we now know the OTHER side also has
        # NOT incoming_updated_at — this has to be "now," this instance's own
        # clock, regardless of the edit's original timestamp. Using the
        # edit's own timestamp here caused a real bug: a conflict-resolved
        # value can carry an edit time that's earlier than a watermark a
        # THIRD instance already advanced past (e.g. it arrives late, after
        # sync delay/conflict resolution) — with pull's since= filter keyed
        # off that edit time, such a change could fall behind the
        # watermark and never be seen again. server_touched_at is a
        # separate, purely-local, always-forward marker of "when did I
        # last write this row," used only for that filter.
        obj.server_touched_at = datetime.now(timezone.utc)

    if existing is None:
        obj = model(sync_uuid=sync_uuid_value)
        _apply_fields(obj)
        try:
            # SAVEPOINT-scoped, not db.session.rollback() — a whole-session
            # rollback here would silently discard every other record
            # already flushed earlier in the SAME pull/push batch (they
            # share one outer transaction, committed once at the end by
            # the caller). That was a real bug: one duplicate-constraint
            # row anywhere in a batch wiped out all its siblings with no
            # error surfaced anywhere. begin_nested() unwinds only this
            # one insert on failure, leaving the rest of the batch intact.
            with db.session.begin_nested():
                db.session.add(obj)
                db.session.flush()
        except IntegrityError:
            return {'sync_uuid': sync_uuid_value, 'status': 'rejected', 'reason': 'duplicate_constraint'}
        return {'sync_uuid': sync_uuid_value, 'status': 'applied'}

    if incoming_updated_at == _sync_sortable_dt(existing.updated_at):
        return {'sync_uuid': sync_uuid_value, 'status': 'applied'}  # already have this exact version

    if direction == 'push':
        incoming_base = _parse_sync_dt(item.get('base_updated_at'))
        genuine_conflict = (incoming_base is not None
                            and _sync_sortable_dt(incoming_base) != _sync_sortable_dt(existing.updated_at))
    else:
        genuine_conflict = existing.pending_push and existing.last_modified_site != incoming_site_id

    if not genuine_conflict:
        if incoming_updated_at > _sync_sortable_dt(existing.updated_at):
            _apply_fields(existing)
        # else: existing is already newer (e.g. our own earlier push echoed back) — no-op
        return {'sync_uuid': sync_uuid_value, 'status': 'applied'}

    # Genuine conflict: both sides changed this row independently while
    # neither knew about the other's edit. Last-write-wins by updated_at
    # (tie-break on site_id string), both full payloads logged either way.
    incoming_key = (incoming_updated_at, incoming_site_id or '')
    existing_key = (_sync_sortable_dt(existing.updated_at), existing.last_modified_site or '')
    incoming_wins = incoming_key > existing_key
    losing_payload = {f: _serialize_sync_value(model, f, getattr(existing, f)) for f in fields}
    winning_payload = item['fields']
    conflict = SyncConflict(
        table_name=table, sync_uuid=sync_uuid_value, conflict_type='lww',
        winning_site_id=incoming_site_id if incoming_wins else existing.last_modified_site,
        losing_site_id=existing.last_modified_site if incoming_wins else incoming_site_id,
        winning_updated_at=incoming_updated_at if incoming_wins else existing.updated_at,
        losing_updated_at=existing.updated_at if incoming_wins else incoming_updated_at,
        winning_payload=json.dumps(winning_payload if incoming_wins else losing_payload, default=str),
        losing_payload=json.dumps(losing_payload if incoming_wins else winning_payload, default=str),
    )
    db.session.add(conflict)
    if incoming_wins:
        _apply_fields(existing)
    return {'sync_uuid': sync_uuid_value, 'status': 'conflict_logged',
            'final': 'remote_won' if incoming_wins else 'local_won'}


def sync_auth_required(f):
    """Auth for the two /api/sync/* routes — a per-site API key, not a
    human login. A spoke has no browser session to carry, so this checks
    X-Sync-Api-Key against SyncSite.api_key_hash instead of @login_required."""
    @wraps(f)
    def decorated(*args, **kwargs):
        api_key = request.headers.get('X-Sync-Api-Key', '')
        matched_site = None
        if api_key:
            for candidate in SyncSite.query.filter_by(is_active=True).all():
                if check_password_hash(candidate.api_key_hash, api_key):
                    matched_site = candidate
                    break
        if not matched_site:
            return jsonify({'error': 'invalid or missing sync API key'}), 401
        request.sync_site = matched_site
        return f(*args, **kwargs)
    return decorated


@app.route('/api/sync/enroll', methods=['POST'])
@csrf.exempt
@limiter.limit('5 per minute')
def api_sync_enroll():
    """Self-service spoke registration — lets a brand-new spoke's
    first-run /setup wizard register itself against this hub directly,
    using an existing hub ADMIN's own login as proof it's allowed to
    join. Same effect as an admin visiting Sync Sites and clicking
    Register New Site (register_sync_site), just invoked machine-to-
    machine instead of through a browser, so nobody has to manually copy
    an API key into a spoke's .env.

    Trust boundary: this only requires credentials for an ACTIVE ADMIN
    account. That's a deliberate choice, not an oversight — it's the same
    authority that already gates the manual Sync Sites page
    (@admin_required), just reachable without a session. Rate-limited to
    slow down credential-guessing against it."""
    data = request.get_json(force=True, silent=True) or {}
    username = (data.get('username') or '').strip().lower()
    password = data.get('password') or ''
    display_name = (data.get('display_name') or '').strip()
    site_id = (data.get('site_id') or '').strip().lower()

    user = User.query.filter_by(username=username).first()
    if not user or not user.is_active or user.role != 'admin' or not user.check_password(password):
        return jsonify({'error': 'Invalid admin credentials.'}), 401

    if not site_id:
        base = re.sub(r'[^a-z0-9]+', '-', display_name.lower()).strip('-')[:40] or 'site'
        site_id = base
        suffix = 1
        while SyncSite.query.filter_by(site_id=site_id).first():
            suffix += 1
            site_id = f'{base[:38]}-{suffix}'

    try:
        site, api_key = register_sync_site(site_id, display_name)
    except ValueError as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400

    log_audit('CREATE', 'sync_sites', site.id, f'Self-enrolled spoke site {site.site_id} (via {username})')
    db.session.commit()
    return jsonify({'site_id': site.site_id, 'display_name': site.display_name, 'api_key': api_key})


@app.route('/api/sync/push', methods=['POST'])
@csrf.exempt
@sync_auth_required
def api_sync_push():
    data = request.get_json(force=True, silent=True) or {}
    site_id = data.get('site_id') or request.sync_site.site_id
    batch = data.get('batch', [])
    request.sync_site.last_push_at = datetime.now(timezone.utc)

    batch_by_table = {}
    for item in batch:
        batch_by_table.setdefault(item['table'], []).append(item)

    results = []
    accepted = 0
    conflicts = 0
    for table in SYNC_TABLE_ORDER:
        for item in batch_by_table.get(table, []):
            result = apply_incoming_record(table, item, site_id, direction='push')
            results.append(result)
            if result['status'] == 'applied':
                accepted += 1
            elif result['status'] == 'conflict_logged':
                conflicts += 1

    # Any table in the batch we don't sync (e.g. 'users') is rejected
    # explicitly rather than silently dropped, so a spoke's outbox doesn't
    # spin retrying something that will never be accepted.
    for table, items in batch_by_table.items():
        if table not in SYNC_MODELS:
            for item in items:
                results.append({'sync_uuid': item.get('sync_uuid'), 'status': 'rejected', 'reason': 'table_not_synced'})

    db.session.commit()
    return jsonify({'accepted': accepted, 'conflicts': conflicts, 'results': results})


@app.route('/api/sync/pull', methods=['GET'])
@csrf.exempt
@sync_auth_required
def api_sync_pull():
    since_str = request.args.get('since')
    since = _parse_sync_dt(since_str) or _SYNC_EPOCH
    tables_param = request.args.get('tables')
    requested_tables = tables_param.split(',') if tables_param else SYNC_TABLE_ORDER
    server_time = datetime.now(timezone.utc)
    request.sync_site.last_pull_at = server_time

    changes = []
    for table in SYNC_TABLE_ORDER:
        if table not in requested_tables:
            continue
        model = SYNC_MODELS[table][0]
        rows = model.query.execution_options(include_deleted=True).filter(model.server_touched_at > since).all()
        for row in rows:
            field_values, fk_values = serialize_record_for_sync(table, row)
            changes.append({
                'table': table,
                'sync_uuid': row.sync_uuid,
                'updated_at': row.updated_at.isoformat() if row.updated_at else None,
                'site_id': row.last_modified_site,
                'fields': field_values,
                'fk': fk_values,
            })

    db.session.commit()
    return jsonify({'changes': changes, 'server_time': server_time.isoformat()})


# ─────────────────────────────────────────────────────────────
# Spoke self-update API — a spoke's own check_for_spoke_update() (see the
# sync loop below) calls these on the same per-site key as /api/sync/*,
# so a revoked/decommissioned site loses update access exactly the way it
# loses data sync access, with no separate credential to manage.
# ─────────────────────────────────────────────────────────────
@app.route('/api/spoke/latest-version', methods=['GET'])
@csrf.exempt
@sync_auth_required
def api_spoke_latest_version():
    release = SpokeRelease.query.filter_by(is_latest=True).first()
    if not release:
        return jsonify({'version': None})
    return jsonify({'version': release.version, 'sha256': release.sha256,
                    'file_size': release.file_size, 'notes': release.notes,
                    'published_at': release.created_at.isoformat()})


@app.route('/api/spoke/download/<version>', methods=['GET'])
@csrf.exempt
@sync_auth_required
def api_spoke_download(version):
    release = SpokeRelease.query.filter_by(version=version).first_or_404()
    path = os.path.join(app.config['SPOKE_RELEASES_DIR'], release.filename)
    if not os.path.exists(path):
        return jsonify({'error': 'Release file missing on the hub — contact an admin.'}), 500
    return send_file(path, as_attachment=True, download_name=release.filename, mimetype='application/zip')


# ─────────────────────────────────────────────────────────────
# Local sync engine — Phase 3. Only runs on a spoke (SYNC_ENABLED=true,
# a local-server PC at a site), never on the hub (Render). Deliberately
# kept as functions here rather than a separate sync_engine.py module: this
# app has no other internal modules, and every one of these functions
# needs direct access to `app`, `db`, and the SYNC_MODELS/apply_incoming_
# record machinery defined just above — splitting it out would only buy a
# fragile bottom-of-file circular import for no real benefit.
#
# Pull and push are structurally the same operation — "upsert a batch of
# someone else's changes" — just running on different machines, so both
# reuse apply_incoming_record()/serialize_record_for_sync() from the API
# above rather than duplicating that logic.
#
# Must run single-process: a second worker would each spawn their own
# thread and race each other's push/pull cycles. Matches the same
# constraint already documented for Central's own gunicorn (-w 1 in
# render.yaml) — a spoke should be run as `python app.py`, not gunicorn
# with multiple workers.
# ─────────────────────────────────────────────────────────────

def _sync_headers():
    return {'X-Sync-Api-Key': app.config['SYNC_API_KEY'], 'Content-Type': 'application/json'}


def _get_peer_state():
    peer_url = app.config['SYNC_HUB_URL']
    state = SyncPeerState.query.filter_by(peer_url=peer_url).first()
    if not state:
        state = SyncPeerState(peer_url=peer_url)
        db.session.add(state)
        db.session.commit()
    return state


def sync_pull_from_hub():
    """Pull everything the hub has changed since our last successful pull
    and apply it locally. The watermark saved afterward is the hub's own
    server_time, not this machine's clock — two sites' clocks don't need
    to agree with each other for pull-since to stay correct."""
    state = _get_peer_state()
    params = {'since': state.last_pull_at.isoformat()} if state.last_pull_at else {}
    resp = requests.get(f"{app.config['SYNC_HUB_URL']}/api/sync/pull",
                        params=params, headers=_sync_headers(), timeout=30)
    resp.raise_for_status()
    body = resp.json()
    for change in body['changes']:
        apply_incoming_record(change['table'], change, change.get('site_id') or 'hub', direction='pull')
    state.last_pull_at = _parse_sync_dt(body['server_time'])
    state.last_error = None
    db.session.commit()
    return len(body['changes'])


def sync_push_to_hub():
    """Push this site's pending local changes (pending_push=True),
    dependency-ordered so a child row's foreign key always resolves
    against a parent the hub has already seen. Chunked at 200 rows/table
    per cycle — a large backlog spreads across several cycles rather than
    risking one oversized request."""
    batch = []
    for table in SYNC_TABLE_ORDER:
        model = SYNC_MODELS[table][0]
        for row in model.query.execution_options(include_deleted=True).filter_by(pending_push=True).limit(200).all():
            field_values, fk_values = serialize_record_for_sync(table, row)
            batch.append({
                'table': table, 'sync_uuid': row.sync_uuid,
                'updated_at': row.updated_at.isoformat() if row.updated_at else None,
                # What we last confirmed the hub agreed on for this row —
                # None for a brand-new local row the hub has never seen.
                # The hub compares this against its own current value to
                # detect whether some OTHER push landed in between (see
                # apply_incoming_record's direction='push' branch).
                'base_updated_at': row.last_synced_updated_at.isoformat() if row.last_synced_updated_at else None,
                'fields': field_values, 'fk': fk_values,
            })
    if not batch:
        return 0

    resp = requests.post(f"{app.config['SYNC_HUB_URL']}/api/sync/push",
                         json={'site_id': app.config['SITE_ID'], 'batch': batch},
                         headers=_sync_headers(), timeout=30)
    resp.raise_for_status()
    result = resp.json()

    # A rejected row (duplicate_constraint/fk_missing) stays pending_push —
    # retried next cycle rather than silently vanishing from the outbox.
    #
    # For everything else, pending_push clears either way, but
    # last_synced_updated_at only advances to OUR value when our edit is
    # what's now actually stored on the hub ('applied', or a conflict we
    # won). When our push LOSES a conflict, the hub kept its OWN differing
    # value — stamping last_synced_updated_at to our own updated_at would
    # falsely claim "the hub agrees with this," and the next pull would
    # then wrongly treat our now-stale local copy as newer and never
    # adopt the hub's real value. Leaving last_synced_updated_at
    # untouched here means the very next pull cycle's normal
    # newer-timestamp check correctly overwrites our losing edit with
    # what the hub actually kept.
    results_by_uuid = {r['sync_uuid']: r for r in result['results']}
    for table in SYNC_TABLE_ORDER:
        model = SYNC_MODELS[table][0]
        for row in (model.query.execution_options(include_deleted=True)
                    .filter(model.pending_push == True, model.sync_uuid.in_(results_by_uuid.keys())).all()):  # noqa: E712
            r = results_by_uuid[row.sync_uuid]
            if r['status'] == 'rejected':
                continue
            row.pending_push = False
            if r['status'] == 'applied' or r.get('final') == 'remote_won':
                row.last_synced_updated_at = row.updated_at

    state = _get_peer_state()
    state.last_push_at = datetime.now(timezone.utc)
    state.last_error = None
    db.session.commit()
    return len(batch)


def _get_update_state():
    state = SpokeUpdateState.query.get(1)
    if not state:
        state = SpokeUpdateState(id=1)
        db.session.add(state)
        db.session.commit()
    return state


def check_for_spoke_update(force=False):
    """Ask the hub whether a newer spoke build is published and, if so,
    download and unpack it into a staging folder next to this install —
    never applied here. This process cannot safely overwrite its own
    running .exe/_internal files (Windows keeps them locked while
    executing), so applying is the launcher script's job, run the next
    time this spoke actually restarts (see launcher.ps1 and
    SPOKE_SETUP.md). Throttled well below SYNC_INTERVAL_SECONDS — see
    SPOKE_UPDATE_CHECK_SECONDS — since this is a one-shot metadata check,
    not per-record sync, and an update is never urgent by design.

    force=True skips that throttle — used by the "Check for Update Now"
    button on Sync Health (sync_check_update) so an admin who just
    published a release isn't stuck waiting out the full interval
    (6h by default) before this spoke even looks for it.

    Everything here — including fetching/updating SpokeUpdateState itself,
    not just the network calls — lives inside one try/except: this used to
    fetch that state before the try started, so a bare exception there
    (e.g. a transient SQLite lock, since this can now also run on the
    request thread via sync_check_update while the background sync thread
    is mid-cycle on the same local DB) would escape uncaught. From a
    background cycle that silently kills the sync loop for good (see
    run_sync_cycle); from the manual check-update route it would 500 the
    request instead of flashing a message. Neither may happen."""
    try:
        state = _get_update_state()
        if state.staged_version == APP_VERSION:
            # Whatever was staged is now the version actually running —
            # the launcher already applied it on this restart, so it's no
            # longer "staged, pending a restart." Checked unconditionally,
            # ahead of the throttle below, so this clears on the very
            # first call after a restart rather than sitting stale for up
            # to a full SPOKE_UPDATE_CHECK_SECONDS.
            state.staged_version = None
            state.staged_dir = None
            db.session.commit()
        now = datetime.now(timezone.utc)
        if not force and state.last_checked_at:
            elapsed = (now - state.last_checked_at.replace(tzinfo=timezone.utc)).total_seconds()
            if elapsed < app.config['SPOKE_UPDATE_CHECK_SECONDS']:
                return
        state.last_checked_at = now

        resp = requests.get(f"{app.config['SYNC_HUB_URL']}/api/spoke/latest-version",
                            headers=_sync_headers(), timeout=30)
        resp.raise_for_status()
        info = resp.json()
        version = info.get('version')
        if not version or version == APP_VERSION or version == state.staged_version:
            state.last_error = None
            db.session.commit()
            return

        staging_root = os.path.join(BASE_DIR, '_update_staged')
        target_dir = os.path.join(staging_root, version)
        if os.path.exists(target_dir):
            shutil.rmtree(target_dir)
        os.makedirs(target_dir, exist_ok=True)

        zip_path = os.path.join(staging_root, f'{version}.zip')
        dl = requests.get(f"{app.config['SYNC_HUB_URL']}/api/spoke/download/{version}",
                          headers=_sync_headers(), timeout=300, stream=True)
        dl.raise_for_status()
        hasher = hashlib.sha256()
        with open(zip_path, 'wb') as f:
            for chunk in dl.iter_content(chunk_size=1024 * 1024):
                hasher.update(chunk)
                f.write(chunk)
        if hasher.hexdigest() != info.get('sha256'):
            os.remove(zip_path)
            raise ValueError(f'downloaded release {version} failed checksum verification — discarded')

        with zipfile.ZipFile(zip_path) as zf:
            zf.extractall(target_dir)
        os.remove(zip_path)

        state.staged_version = version
        state.staged_dir = target_dir
        state.last_error = None
        db.session.commit()

        # A plain marker file next to the .exe, not just the DB row above —
        # launcher.ps1 is a dependency-free script that runs BEFORE this
        # app (or any Python/SQLite) even starts, specifically so it can
        # swap in the new files while nothing has them locked yet. It
        # can't query this app's database, so this is the only way it
        # learns an update is waiting.
        marker = {'version': version, 'staged_dir': target_dir}
        with open(os.path.join(BASE_DIR, 'update_ready.json'), 'w', encoding='utf-8') as f:
            json.dump(marker, f)

        app.logger.info(f'spoke update {version} downloaded and staged — applies on next restart')
    except Exception as e:  # noqa: BLE001 — a bad check must never kill the loop (or any caller)
        db.session.rollback()
        try:
            # Best-effort — if fetching/writing the state row is itself
            # what failed above, there's nowhere left to record this.
            state = _get_update_state()
            state.last_error = str(e)
            db.session.commit()
        except Exception:  # noqa: BLE001
            db.session.rollback()
        app.logger.warning(f'spoke update check failed: {e}')


def run_sync_cycle():
    with app.app_context():
        try:
            pulled = sync_pull_from_hub()
            pushed = sync_push_to_hub()
            if pulled or pushed:
                app.logger.info(f'sync cycle: pulled {pulled}, pushed {pushed}')
        except Exception as e:  # noqa: BLE001 — a bad cycle must never kill the loop
            db.session.rollback()
            state = _get_peer_state()
            state.last_error = str(e)
            db.session.commit()
            app.logger.warning(f'sync cycle failed: {e}')
        if FROZEN:
            # Only a packaged spoke .exe can actually be replaced by the
            # launcher script — a dev machine or Render's own gunicorn
            # process has no launcher wrapping it, so skip there.
            #
            # Guarded separately from the try/except above: check_for_
            # spoke_update() fetches its own state row (_get_update_state())
            # before its internal try starts, so a bare exception there
            # (e.g. a transient SQLite lock from this thread and a request
            # thread hitting the same file at once) would otherwise escape
            # uncaught, silently killing this daemon thread for good — no
            # traceback survives that in a windowed .exe with no console
            # and no file logger. The whole point of this loop is that one
            # bad cycle can never end it, so nothing here may be allowed
            # to raise past this point either.
            try:
                check_for_spoke_update()
            except Exception as e:  # noqa: BLE001 — see above, must never kill the loop
                app.logger.warning(f'spoke update check crashed outside its own guard: {e}')


def _sync_loop():
    while True:
        run_sync_cycle()
        time.sleep(app.config['SYNC_INTERVAL_SECONDS'])


def start_sync_thread():
    threading.Thread(target=_sync_loop, daemon=True, name='sync-engine').start()


# ─────────────────────────────────────────────────────────────
# WhatsApp Report Bot — lets a linked user text plain commands (e.g.
# "income last month") to a WhatsApp number and get one of the existing
# reports back as a formatted text message. Built on Meta's WhatsApp
# Cloud API. Report figures are computed with the exact same
# compute_consolidated_overview/compute_income_statement/etc. helpers the
# web report pages use, so the bot can never drift from what's shown
# on-screen. Access is gated by the sender's number being linked to an
# active user with the 'reports' permission (see User.whatsapp_phone and
# users/whatsapp.html) — command parsing is deliberately simple
# keyword+period matching, not free-form NLU.
# ─────────────────────────────────────────────────────────────
def _wa_money(x):
    return f'${x:,.2f}'


def _wa_period(tokens):
    """Parse the words after a report command into a (from_date, to_date,
    label) tuple. Recognizes a handful of keywords plus an explicit
    'YYYY-MM-DD YYYY-MM-DD' range; defaults to the current month, same
    default as the web reports' query_date_range."""
    today = date.today()
    month_start = today.replace(day=1)
    joined = ' '.join(tokens).strip().lower()

    if not joined or joined in ('month', 'this month'):
        return month_start, today, 'this month'
    if joined == 'today':
        return today, today, 'today'
    if joined == 'yesterday':
        y = today - timedelta(days=1)
        return y, y, 'yesterday'
    if joined in ('week', 'this week'):
        start = today - timedelta(days=today.weekday())
        return start, today, 'this week'
    if joined in ('last month', 'lastmonth'):
        last_end = month_start - timedelta(days=1)
        return last_end.replace(day=1), last_end, 'last month'
    if joined in ('year', 'ytd', 'this year'):
        return today.replace(month=1, day=1), today, 'this year'

    parts = joined.split()
    if len(parts) == 2:
        try:
            df = datetime.strptime(parts[0], '%Y-%m-%d').date()
            dt = datetime.strptime(parts[1], '%Y-%m-%d').date()
            if df > dt:
                df, dt = dt, df
            return df, dt, f'{df} to {dt}'
        except ValueError:
            pass

    return month_start, today, 'this month'


def _wa_consolidated(tokens):
    df, dt, label = _wa_period(tokens)
    segments, totals = compute_consolidated_overview(df, dt)
    lines = [f'*Consolidated Overview — {label}*', f'_{df} to {dt}_', '']
    for s in segments:
        lines.append(f"*{s['name']}*  ({s['count']} entries)")
        lines.append(f"  Revenue: {_wa_money(s['revenue'])}")
        lines.append(f"  Expenses: {_wa_money(s['expenses'])}")
        lines.append(f"  Net: {_wa_money(s['net_profit'])}")
    lines.append('')
    lines.append('*TOTAL*')
    lines.append(f"  Revenue: {_wa_money(totals['revenue'])}")
    lines.append(f"  Expenses: {_wa_money(totals['expenses'])}")
    lines.append(f"  Net Profit: {_wa_money(totals['net_profit'])}")
    return '\n'.join(lines)


def _wa_income(tokens):
    df, dt, label = _wa_period(tokens)
    stmt = compute_income_statement(df, dt)
    lines = [f'*Income Statement — {label}*', f'_{df} to {dt}_', '']
    lines.append(f"Gross Revenue: {_wa_money(stmt['gross_revenue'])}")
    for name, amt in stmt['statement_expenses']:
        lines.append(f'  {name}: {_wa_money(amt)}')
    lines.append(f"Total Expenses: {_wa_money(stmt['total_expenses'])}")
    lines.append(f"*Net Profit: {_wa_money(stmt['net_profit'])}* ({stmt['profit_margin']:.1f}% margin)")
    top = stmt['vehicle_breakdown'][:5]
    if top:
        lines.append('')
        lines.append('Top vehicles:')
        for row in top:
            lines.append(f"  {row['vehicle'].registration}: {_wa_money(row['net_profit'])}")
    return '\n'.join(lines)


def _wa_payroll(tokens):
    df, dt, label = _wa_period(tokens)
    earnings, total_commissions, total_garnish, total_deductions, total_paid, total_outstanding = compute_payroll_earnings(df, dt)
    lines = [f'*Payroll — {label}*', f'_{df} to {dt}_', '']
    for e in earnings:
        lines.append(f"{e['driver'].name} ({e['driver'].role}): "
                      f"{_wa_money(e['commission'])} accrued, {_wa_money(e['outstanding'])} outstanding")
    lines.append('')
    lines.append(f'*Total accrued: {_wa_money(total_commissions)}*')
    if total_deductions:
        lines.append(f'Deductions: {_wa_money(total_deductions)}')
    lines.append(f'Paid: {_wa_money(total_paid)}  |  Outstanding: {_wa_money(total_outstanding)}')
    return '\n'.join(lines)


def _wa_cashflow(tokens):
    df, dt, label = _wa_period(tokens)
    cf = compute_cash_flow(df, dt)
    lines = [f'*Cash Flow — {label}*', f'_{df} to {dt}_', '']
    lines.append(f"Operating: {_wa_money(cf['net_operating'])}")
    lines.append(f"Investing: {_wa_money(cf['net_investing'])}")
    lines.append(f"Financing: {_wa_money(cf['net_financing'])}")
    lines.append(f"*Net Change: {_wa_money(cf['net_change'])}*")
    lines.append(f"Opening Cash: {_wa_money(cf['opening_cash'])}  |  Closing Cash: {_wa_money(cf['closing_cash'])}")
    return '\n'.join(lines)


def _wa_help(tokens=None):
    return (
        '*Transport ERP Report Bot*\n'
        'Send one of these commands, optionally followed by a period:\n\n'
        '• *consolidated* — company-wide P&L\n'
        '• *income* — income statement\n'
        '• *payroll* — crew commissions\n'
        '• *cashflow* — cash flow statement\n\n'
        'Periods: today, yesterday, week, month (default), last month, year, '
        'or an explicit range like `2026-08-01 2026-08-31`.\n'
        'Example: `income last month`'
    )


WHATSAPP_COMMANDS = {
    'consolidated': _wa_consolidated,
    'overview': _wa_consolidated,
    'income': _wa_income,
    'payroll': _wa_payroll,
    'commissions': _wa_payroll,
    'cashflow': _wa_cashflow,
    'cash': _wa_cashflow,
    'help': _wa_help,
}


def whatsapp_dispatch(text):
    """Route one inbound message body to a report and return the reply
    text. Caller (whatsapp_webhook) has already checked the sender is a
    linked, active user with the 'reports' permission."""
    tokens = (text or '').strip().split()
    if not tokens:
        return _wa_help()
    command, rest = tokens[0].lower(), tokens[1:]
    handler = WHATSAPP_COMMANDS.get(command)
    if not handler:
        return f'Unrecognized command "{tokens[0]}".\n\n' + _wa_help()
    try:
        return handler(rest)
    except Exception:
        app.logger.exception('WhatsApp report command failed: %r', text)
        return 'Sorry, something went wrong generating that report. Try again or contact your admin.'


def send_whatsapp_message(to, body):
    """POST a plain-text reply to a WhatsApp number via Meta's Cloud API.
    `to` is the digits-only sender id WhatsApp itself hands back in the
    webhook payload. No-ops (with a log line) if WHATSAPP_TOKEN/
    WHATSAPP_PHONE_NUMBER_ID aren't configured yet, so the webhook stays
    functional in dev before credentials are set up."""
    token = app.config['WHATSAPP_TOKEN']
    phone_number_id = app.config['WHATSAPP_PHONE_NUMBER_ID']
    if not token or not phone_number_id:
        app.logger.warning('WhatsApp not configured — dropped reply to %s', to)
        return
    url = f'https://graph.facebook.com/v20.0/{phone_number_id}/messages'
    payload = {
        'messaging_product': 'whatsapp',
        'to': to,
        'type': 'text',
        'text': {'body': body[:4096]},
    }
    try:
        resp = requests.post(url, json=payload,
                             headers={'Authorization': f'Bearer {token}'}, timeout=10)
        if resp.status_code >= 300:
            app.logger.error('WhatsApp send failed (%s): %s', resp.status_code, resp.text)
    except requests.RequestException:
        app.logger.exception('WhatsApp send request failed')


def notify_admins_whatsapp(body):
    """Push a WhatsApp alert to every active admin who has a number linked
    (see User.whatsapp_phone / users/whatsapp.html) — used for events an
    admin should know about right away (e.g. a quick-registered franchise
    vehicle awaiting review), on top of the in-app badge and audit log
    that already cover anyone without WhatsApp set up. Best-effort: a
    send failure to one admin (logged inside send_whatsapp_message) never
    blocks the others or the caller's own request."""
    admins = User.query.filter_by(role='admin', is_active=True).filter(User.whatsapp_phone.isnot(None)).all()
    for admin in admins:
        send_whatsapp_message(admin.whatsapp_phone, body)


def _whatsapp_signature_valid(raw_body):
    """Verify Meta's X-Hub-Signature-256 header against WHATSAPP_APP_SECRET
    so the webhook only acts on requests actually from Meta. Passes
    (unverified) if no app secret is configured yet, matching how the rest
    of first-run setup in this app degrades gracefully before secrets are
    filled in — set WHATSAPP_APP_SECRET before going live."""
    secret = app.config['WHATSAPP_APP_SECRET']
    if not secret:
        return True
    signature = request.headers.get('X-Hub-Signature-256', '')
    if not signature.startswith('sha256='):
        return False
    expected = hmac.new(secret.encode(), raw_body, hashlib.sha256).hexdigest()
    return hmac.compare_digest(signature[7:], expected)


@app.route('/api/whatsapp/webhook', methods=['GET'])
@csrf.exempt
def whatsapp_webhook_verify():
    """Meta's one-time handshake when the webhook URL is registered in the
    App Dashboard — echoes hub.challenge back only if hub.verify_token
    matches WHATSAPP_VERIFY_TOKEN, proving this endpoint is under our
    control."""
    mode = request.args.get('hub.mode')
    token = request.args.get('hub.verify_token')
    challenge = request.args.get('hub.challenge', '')
    if token and mode == 'subscribe' and token == app.config['WHATSAPP_VERIFY_TOKEN']:
        return challenge, 200
    return 'Forbidden', 403


@app.route('/api/whatsapp/webhook', methods=['POST'])
@csrf.exempt
@limiter.limit('60 per minute')
def whatsapp_webhook():
    """Inbound WhatsApp messages from Meta's Cloud API. Only text messages
    from a number linked to an active user with the 'reports' permission
    get a report reply; every other case still returns 200 (Meta retries
    non-2xx responses, and there's nothing to fix by retrying an
    unrecognized sender)."""
    if not _whatsapp_signature_valid(request.get_data()):
        return jsonify({'status': 'invalid signature'}), 403

    data = request.get_json(force=True, silent=True) or {}
    for entry in data.get('entry', []):
        for change in entry.get('changes', []):
            value = change.get('value', {})
            for msg in value.get('messages', []):
                if msg.get('type') != 'text':
                    continue
                sender = msg.get('from', '')
                body = msg.get('text', {}).get('body', '')
                if not sender:
                    continue

                user = User.query.filter_by(whatsapp_phone=sender).first()
                if not user or not user.is_active:
                    send_whatsapp_message(sender,
                        "This number isn't registered on the Transport ERP report bot. "
                        "Ask an admin to link it under Users.")
                    continue
                if not user.has_permission('reports'):
                    send_whatsapp_message(sender,
                        "Your account doesn't have Reports access — ask an admin to grant it.")
                    continue

                send_whatsapp_message(sender, whatsapp_dispatch(body))

    return jsonify({'status': 'received'})


# ─────────────────────────────────────────────────────────────
# Template filters
# ─────────────────────────────────────────────────────────────
@app.template_filter('currency')
def currency_filter(value):
    """Formats a stored amount for display, rounded to at most 3 decimal
    places — a figure entered as 15.455 (see the step="any" amount inputs
    and driver_ledger_add's duplicate check, which both preserve full
    precision) must still show as $15.455, not get silently rounded down
    to $15.46 or $15.45. Anything beyond the 3rd decimal (float noise, or
    a value with more precision than that) is rounded away. Trailing
    zeros past the 2nd decimal are trimmed so a plain whole-cents amount
    still reads as $15.00, not $15.4500000000."""
    if value is None:
        return '$0.00'
    value = float(value)
    neg = value < 0
    value = round(abs(value), 3)
    # repr() gives the shortest decimal string that round-trips back to
    # this exact float — unlike f'{value:.Nf}', it can't introduce binary-
    # float noise (e.g. printing 15.455 as 15.454999999999998).
    s = repr(value)
    if 'e' in s or 'E' in s:  # absurdly large/small — not a real money value, but don't crash
        s = f'{value:.10f}'.rstrip('0').rstrip('.')
    int_part, _, dec_part = s.partition('.')
    dec_part = dec_part.rstrip('0').ljust(2, '0')
    return f'{"-" if neg else ""}${int(int_part):,}.{dec_part}'


@app.template_filter('pct')
def pct_filter(value):
    return f'{value:.1f}%' if value is not None else '0.0%'


@app.template_filter('qty')
def qty_filter(value):
    """Formats a store quantity that may now carry a fractional part (e.g.
    2.5 litres of oil) without showing a pointless '.0' on whole numbers."""
    if value is None:
        return '0'
    value = float(value)
    if value == int(value):
        return str(int(value))
    return f'{value:.2f}'.rstrip('0').rstrip('.')


@app.template_filter('pretty_json')
def pretty_json_filter(value):
    """Used on sync/conflicts.html to render SyncConflict's stored payload
    JSON legibly instead of as one long compact line."""
    try:
        return json.dumps(json.loads(value), indent=2, default=str)
    except (TypeError, ValueError):
        return value or ''


# ─────────────────────────────────────────────────────────────
# Bootstrap DB + first admin
# ─────────────────────────────────────────────────────────────
def migrate_db():
    from sqlalchemy import inspect, text
    inspector = inspect(db.engine)
    user_cols = [c['name'] for c in inspector.get_columns('users')]
    with db.engine.connect() as conn:
        if 'permissions' not in user_cols:
            conn.execute(text("ALTER TABLE users ADD COLUMN permissions TEXT DEFAULT '[]'"))
        if 'driver_id' not in user_cols:
            conn.execute(text("ALTER TABLE users ADD COLUMN driver_id INTEGER REFERENCES drivers(id)"))
        if 'whatsapp_phone' not in user_cols:
            conn.execute(text("ALTER TABLE users ADD COLUMN whatsapp_phone VARCHAR(20)"))

        # Depots were removed — drop the leftover columns/table from any DB that has them.
        for table in ('vehicles', 'drivers', 'routes'):
            cols = [c['name'] for c in inspector.get_columns(table)]
            if 'depot_id' in cols:
                conn.execute(text(f"ALTER TABLE {table} DROP COLUMN depot_id"))

        vehicle_cols = [c['name'] for c in inspector.get_columns('vehicles')]
        if 'fuel_type' not in vehicle_cols:
            conn.execute(text("ALTER TABLE vehicles ADD COLUMN fuel_type VARCHAR(10) DEFAULT 'diesel'"))
        if 'daily_target' not in vehicle_cols:
            conn.execute(text("ALTER TABLE vehicles ADD COLUMN daily_target FLOAT"))
        if 'insurance_provider' not in vehicle_cols:
            conn.execute(text("ALTER TABLE vehicles ADD COLUMN insurance_provider VARCHAR(100)"))
        if 'insurance_policy_number' not in vehicle_cols:
            conn.execute(text("ALTER TABLE vehicles ADD COLUMN insurance_policy_number VARCHAR(100)"))
        if 'insurance_type' not in vehicle_cols:
            conn.execute(text("ALTER TABLE vehicles ADD COLUMN insurance_type VARCHAR(50)"))
        if 'insurance_expiry' not in vehicle_cols:
            conn.execute(text("ALTER TABLE vehicles ADD COLUMN insurance_expiry DATE"))
        if inspector.has_table('depots'):
            conn.execute(text("DROP TABLE depots"))

        if inspector.has_table('expense_categories'):
            exp_cat_cols = [c['name'] for c in inspector.get_columns('expense_categories')]
            if 'parent_id' not in exp_cat_cols:
                conn.execute(text(
                    "ALTER TABLE expense_categories ADD COLUMN parent_id INTEGER REFERENCES expense_categories(id)"))
            if 'default_amount' not in exp_cat_cols:
                conn.execute(text("ALTER TABLE expense_categories ADD COLUMN default_amount FLOAT"))

        if inspector.has_table('store_sales'):
            store_sale_cols = [c['name'] for c in inspector.get_columns('store_sales')]
            if 'vehicle_id' not in store_sale_cols:
                conn.execute(text(
                    "ALTER TABLE store_sales ADD COLUMN vehicle_id INTEGER REFERENCES vehicles(id)"))

        # Store quantities can now be fractional (e.g. 2.5 litres of oil sold
        # from a bulk container) instead of being forced to whole units.
        # SQLite columns are type-affinity only and already accept floats in
        # an INTEGER-declared column, so only Postgres — which enforces the
        # declared type strictly — needs an explicit ALTER here.
        if db.engine.name == 'postgresql':
            for qty_table, qty_column in (('spare_parts', 'quantity_on_hand'),
                                          ('store_purchases', 'quantity'),
                                          ('store_sales', 'quantity')):
                qty_col = next((c for c in inspector.get_columns(qty_table) if c['name'] == qty_column), None)
                if qty_col is not None and isinstance(qty_col['type'], db.Integer):
                    conn.execute(text(f"ALTER TABLE {qty_table} ALTER COLUMN {qty_column} TYPE DOUBLE PRECISION"))

        driver_cols = inspector.get_columns('drivers')
        driver_col_names = [c['name'] for c in driver_cols]
        if 'paired_driver_id' not in driver_col_names:
            conn.execute(text("ALTER TABLE drivers ADD COLUMN paired_driver_id INTEGER REFERENCES drivers(id)"))
        if 'assigned_vehicle_id' not in driver_col_names:
            conn.execute(text("ALTER TABLE drivers ADD COLUMN assigned_vehicle_id INTEGER REFERENCES vehicles(id)"))
        for col in ('next_of_kin_name', 'next_of_kin_phone', 'next_of_kin_relationship'):
            if col not in driver_col_names:
                conn.execute(text(f"ALTER TABLE drivers ADD COLUMN {col} VARCHAR(100)"))
        if 'id_number' not in driver_col_names:
            conn.execute(text("ALTER TABLE drivers ADD COLUMN id_number VARCHAR(30)"))

        license_col = next((c for c in driver_cols if c['name'] == 'license_number'), None)
        if license_col is not None and not license_col['nullable']:
            # SQLite can't relax a NOT NULL constraint with ALTER TABLE — rebuild the table.
            # Conductors don't need a license number, so this must become optional.
            conn.execute(text("""
                CREATE TABLE drivers_new (
                    id INTEGER PRIMARY KEY,
                    name VARCHAR(100) NOT NULL,
                    license_number VARCHAR(50) UNIQUE,
                    phone VARCHAR(20),
                    role VARCHAR(20),
                    commission_rate FLOAT,
                    status VARCHAR(20),
                    paired_driver_id INTEGER REFERENCES drivers(id),
                    assigned_vehicle_id INTEGER REFERENCES vehicles(id),
                    next_of_kin_name VARCHAR(100),
                    next_of_kin_phone VARCHAR(100),
                    next_of_kin_relationship VARCHAR(100),
                    created_at DATETIME
                )
            """))
            conn.execute(text("""
                INSERT INTO drivers_new (id, name, license_number, phone, role,
                    commission_rate, status, paired_driver_id, assigned_vehicle_id,
                    next_of_kin_name, next_of_kin_phone, next_of_kin_relationship, created_at)
                SELECT id, name, license_number, phone, role,
                    commission_rate, status, paired_driver_id, assigned_vehicle_id,
                    next_of_kin_name, next_of_kin_phone, next_of_kin_relationship, created_at FROM drivers
            """))
            conn.execute(text("DROP TABLE drivers"))
            conn.execute(text("ALTER TABLE drivers_new RENAME TO drivers"))

        daily_log_cols = inspector.get_columns('daily_logs')
        route_col = next((c for c in daily_log_cols if c['name'] == 'route_id'), None)
        if route_col is not None and not route_col['nullable']:
            # Same story as license_number above — the Vehicle Ledger doesn't
            # track a route per entry, so this must become optional.
            conn.execute(text("""
                CREATE TABLE daily_logs_new (
                    id INTEGER PRIMARY KEY,
                    vehicle_id INTEGER NOT NULL REFERENCES vehicles(id),
                    driver_id INTEGER NOT NULL REFERENCES drivers(id),
                    conductor_id INTEGER REFERENCES drivers(id),
                    route_id INTEGER REFERENCES routes(id),
                    log_date DATE NOT NULL,
                    trips_completed INTEGER,
                    gross_revenue FLOAT NOT NULL,
                    notes TEXT,
                    created_by INTEGER REFERENCES users(id),
                    updated_by INTEGER REFERENCES users(id),
                    created_at DATETIME,
                    updated_at DATETIME
                )
            """))
            conn.execute(text("""
                INSERT INTO daily_logs_new (id, vehicle_id, driver_id, conductor_id, route_id,
                    log_date, trips_completed, gross_revenue, notes, created_by, updated_by,
                    created_at, updated_at)
                SELECT id, vehicle_id, driver_id, conductor_id, route_id,
                    log_date, trips_completed, gross_revenue, notes, created_by, updated_by,
                    created_at, updated_at FROM daily_logs
            """))
            conn.execute(text("DROP TABLE daily_logs"))
            conn.execute(text("ALTER TABLE daily_logs_new RENAME TO daily_logs"))

        daily_log_col_names = [c['name'] for c in inspector.get_columns('daily_logs')]
        # Renamed from staff_deduction/deduction_notes to garnish/reason_for_shortfall
        # to match the "missed revenue target" concept this field represents.
        if 'garnish' not in daily_log_col_names:
            if 'staff_deduction' in daily_log_col_names:
                conn.execute(text("ALTER TABLE daily_logs RENAME COLUMN staff_deduction TO garnish"))
            else:
                conn.execute(text("ALTER TABLE daily_logs ADD COLUMN garnish FLOAT NOT NULL DEFAULT 0.0"))
        if 'reason_for_shortfall' not in daily_log_col_names:
            if 'deduction_notes' in daily_log_col_names:
                conn.execute(text("ALTER TABLE daily_logs RENAME COLUMN deduction_notes TO reason_for_shortfall"))
            else:
                conn.execute(text("ALTER TABLE daily_logs ADD COLUMN reason_for_shortfall TEXT"))

        driver_id_col = next((c for c in inspector.get_columns('daily_logs') if c['name'] == 'driver_id'), None)
        if driver_id_col is not None and not driver_id_col['nullable']:
            # SQLite can't relax a NOT NULL constraint with ALTER TABLE — rebuild the table.
            # Daily Transactions needs to record a fare/activity before a driver is assigned.
            conn.execute(text("""
                CREATE TABLE daily_logs_new2 (
                    id INTEGER PRIMARY KEY,
                    vehicle_id INTEGER NOT NULL REFERENCES vehicles(id),
                    driver_id INTEGER REFERENCES drivers(id),
                    conductor_id INTEGER REFERENCES drivers(id),
                    route_id INTEGER REFERENCES routes(id),
                    log_date DATE NOT NULL,
                    trips_completed INTEGER,
                    gross_revenue FLOAT NOT NULL,
                    garnish FLOAT NOT NULL DEFAULT 0.0,
                    reason_for_shortfall TEXT,
                    notes TEXT,
                    created_by INTEGER REFERENCES users(id),
                    updated_by INTEGER REFERENCES users(id),
                    created_at DATETIME,
                    updated_at DATETIME
                )
            """))
            conn.execute(text("""
                INSERT INTO daily_logs_new2 (id, vehicle_id, driver_id, conductor_id, route_id,
                    log_date, trips_completed, gross_revenue, garnish, reason_for_shortfall, notes,
                    created_by, updated_by, created_at, updated_at)
                SELECT id, vehicle_id, driver_id, conductor_id, route_id,
                    log_date, trips_completed, gross_revenue, garnish, reason_for_shortfall, notes,
                    created_by, updated_by, created_at, updated_at FROM daily_logs
            """))
            conn.execute(text("DROP TABLE daily_logs"))
            conn.execute(text("ALTER TABLE daily_logs_new2 RENAME TO daily_logs"))

        # franchise_income (combined daily+weekly columns on one row) and
        # franchise_weekly_expenses (the old free-form weekly deduction
        # list) were retired in favor of two independent entities,
        # franchise_daily_income and franchise_weekly_income — drop the old
        # tables and let db.create_all() (called again after this function
        # returns) create the new ones.
        if inspector.has_table('franchise_income'):
            conn.execute(text("DROP TABLE franchise_income"))
        if inspector.has_table('franchise_weekly_expenses'):
            conn.execute(text("DROP TABLE franchise_weekly_expenses"))

        # franchise_daily_income / franchise_weekly_income moved from a
        # free-text franchisee_name to a vehicle_id FK on FranchiseVehicle
        # (one entry per date/week per vehicle, since each vehicle pays its
        # own standalone fee), and vehicle_id was then relaxed to nullable
        # so a whole-franchise entry not attributable to one vehicle can
        # coexist with per-vehicle entries — drop and let db.create_all()
        # rebuild with the new column/nullability and unique constraint.
        if inspector.has_table('franchise_daily_income'):
            daily_income_cols = {c['name']: c for c in inspector.get_columns('franchise_daily_income')}
            if 'vehicle_id' not in daily_income_cols or not daily_income_cols['vehicle_id']['nullable']:
                conn.execute(text("DROP TABLE franchise_daily_income"))
        if inspector.has_table('franchise_weekly_income'):
            weekly_income_cols = {c['name']: c for c in inspector.get_columns('franchise_weekly_income')}
            if 'vehicle_id' not in weekly_income_cols or not weekly_income_cols['vehicle_id']['nullable']:
                conn.execute(text("DROP TABLE franchise_weekly_income"))

        # A vehicle can now log more than one weekly income entry for the
        # same week (e.g. several partial payments) — the old
        # one-entry-per-(week, vehicle) UNIQUE constraint blocked that with
        # an "already exists" error on the second Add. SQLite has no ALTER
        # TABLE DROP CONSTRAINT, so the table is rebuilt without it; unlike
        # the drop-and-let-create_all-rebuild above (which only ever ran
        # against tables from before real income data existed), this one
        # copies existing rows across instead of discarding them.
        if inspector.has_table('franchise_weekly_income'):
            weekly_constraints = inspector.get_unique_constraints('franchise_weekly_income')
            if any(c['name'] == 'uq_franchise_weekly_income_week_vehicle' for c in weekly_constraints):
                conn.execute(text("""
                    CREATE TABLE franchise_weekly_income_new (
                        id INTEGER NOT NULL,
                        week_start DATE NOT NULL,
                        vehicle_id INTEGER,
                        income FLOAT NOT NULL,
                        exp_traffic_fines FLOAT NOT NULL,
                        exp_facilitation_fees FLOAT NOT NULL,
                        exp_workshop FLOAT NOT NULL,
                        exp_wages FLOAT NOT NULL,
                        other_expenditure FLOAT NOT NULL,
                        deposited FLOAT NOT NULL,
                        description TEXT,
                        created_by INTEGER,
                        created_at DATETIME,
                        updated_at DATETIME,
                        sync_uuid VARCHAR(36),
                        pending_push BOOLEAN DEFAULT 0,
                        last_modified_site VARCHAR(50),
                        deleted_at DATETIME,
                        last_synced_updated_at DATETIME,
                        server_touched_at DATETIME,
                        PRIMARY KEY (id),
                        FOREIGN KEY(vehicle_id) REFERENCES franchise_vehicles (id),
                        FOREIGN KEY(created_by) REFERENCES users (id)
                    )
                """))
                conn.execute(text("INSERT INTO franchise_weekly_income_new SELECT * FROM franchise_weekly_income"))
                conn.execute(text("DROP TABLE franchise_weekly_income"))
                conn.execute(text("ALTER TABLE franchise_weekly_income_new RENAME TO franchise_weekly_income"))

        if inspector.has_table('franchise_collections'):
            collection_cols = [c['name'] for c in inspector.get_columns('franchise_collections')]
            if 'frequency' not in collection_cols:
                conn.execute(text(
                    "ALTER TABLE franchise_collections ADD COLUMN frequency VARCHAR(10) NOT NULL DEFAULT 'daily'"))
            if 'expense' not in collection_cols:
                conn.execute(text(
                    "ALTER TABLE franchise_collections ADD COLUMN expense FLOAT NOT NULL DEFAULT 0"))

        # franchise_vehicles gained amount_owed/notes/daily_fee/weekly_fee — added
        # in place (not drop-and-recreate) since this table holds real
        # registered vehicles, unlike the income tables above.
        if inspector.has_table('franchise_vehicles'):
            vehicle_cols = [c['name'] for c in inspector.get_columns('franchise_vehicles')]
            if 'amount_owed' not in vehicle_cols:
                conn.execute(text("ALTER TABLE franchise_vehicles ADD COLUMN amount_owed FLOAT NOT NULL DEFAULT 0"))
            if 'notes' not in vehicle_cols:
                conn.execute(text("ALTER TABLE franchise_vehicles ADD COLUMN notes TEXT"))
            if 'daily_fee' not in vehicle_cols:
                conn.execute(text("ALTER TABLE franchise_vehicles ADD COLUMN daily_fee FLOAT"))
            if 'weekly_fee' not in vehicle_cols:
                conn.execute(text("ALTER TABLE franchise_vehicles ADD COLUMN weekly_fee FLOAT"))
            if 'pending_review' not in vehicle_cols:
                conn.execute(text("ALTER TABLE franchise_vehicles ADD COLUMN pending_review BOOLEAN NOT NULL DEFAULT 0"))

        # expenses gained driver_id — attributes an expense to a driver
        # (see report_fleet_reconciliation). driver_deposits is a brand-new
        # table, so it needs no ALTER here; db.create_all() below builds it.
        if inspector.has_table('expenses'):
            expense_cols = [c['name'] for c in inspector.get_columns('expenses')]
            if 'driver_id' not in expense_cols:
                conn.execute(text("ALTER TABLE expenses ADD COLUMN driver_id INTEGER REFERENCES drivers(id)"))

        # driver_deposits.driver_id must become optional — a deposit can now
        # be a fleet-wide total (all vehicles, no single driver) instead of
        # always being attributed to one driver. SQLite can't relax a NOT
        # NULL constraint with ALTER TABLE — rebuild the table.
        if inspector.has_table('driver_deposits'):
            deposit_driver_col = next(
                (c for c in inspector.get_columns('driver_deposits') if c['name'] == 'driver_id'), None)
            if deposit_driver_col is not None and not deposit_driver_col['nullable']:
                conn.execute(text("""
                    CREATE TABLE driver_deposits_new (
                        id INTEGER PRIMARY KEY,
                        driver_id INTEGER REFERENCES drivers(id),
                        vehicle_id INTEGER REFERENCES vehicles(id),
                        deposit_date DATE NOT NULL,
                        amount FLOAT NOT NULL,
                        notes TEXT,
                        created_by INTEGER REFERENCES users(id),
                        created_at DATETIME,
                        updated_at DATETIME,
                        sync_uuid VARCHAR(36) UNIQUE,
                        pending_push BOOLEAN,
                        last_modified_site VARCHAR(50),
                        deleted_at DATETIME,
                        last_synced_updated_at DATETIME,
                        server_touched_at DATETIME
                    )
                """))
                conn.execute(text("""
                    INSERT INTO driver_deposits_new (id, driver_id, vehicle_id, deposit_date, amount,
                        notes, created_by, created_at, updated_at, sync_uuid, pending_push,
                        last_modified_site, deleted_at, last_synced_updated_at, server_touched_at)
                    SELECT id, driver_id, vehicle_id, deposit_date, amount,
                        notes, created_by, created_at, updated_at, sync_uuid, pending_push,
                        last_modified_site, deleted_at, last_synced_updated_at, server_touched_at
                    FROM driver_deposits
                """))
                conn.execute(text("DROP TABLE driver_deposits"))
                conn.execute(text("ALTER TABLE driver_deposits_new RENAME TO driver_deposits"))

        # Multi-site sync columns (see touch_sync_fields()) — added to every
        # syncable table. sync_uuid/deleted_at go in nullable (SQLite can't
        # add a NOT NULL column post-hoc without a full table rebuild, same
        # constraint hit above for license_number/route_id/driver_id);
        # sync_uuid is backfilled for existing rows just below instead.
        sync_tables = [
            'vehicles', 'vehicle_documents', 'drivers', 'routes', 'daily_logs',
            'fuel_logs', 'maintenance_logs', 'loans', 'loan_payments', 'payables',
            'receivables', 'commission_payments', 'capital_contributions',
            'owner_drawings', 'expense_categories', 'expenses', 'budgets',
            'franchise_daily_income', 'franchise_weekly_income', 'franchise_vehicles',
            'franchise_collections', 'maintenance_schedules', 'spare_parts',
            'store_purchases', 'store_sales',
        ]
        for table in sync_tables:
            if not inspector.has_table(table):
                continue  # brand-new DB — db.create_all() builds it with these columns already
            cols = [c['name'] for c in inspector.get_columns(table)]
            if 'updated_at' not in cols:
                conn.execute(text(f"ALTER TABLE {table} ADD COLUMN updated_at DATETIME"))
            if 'sync_uuid' not in cols:
                conn.execute(text(f"ALTER TABLE {table} ADD COLUMN sync_uuid VARCHAR(36)"))
            if 'pending_push' not in cols:
                conn.execute(text(f"ALTER TABLE {table} ADD COLUMN pending_push BOOLEAN DEFAULT 0"))
            if 'last_modified_site' not in cols:
                conn.execute(text(f"ALTER TABLE {table} ADD COLUMN last_modified_site VARCHAR(50)"))
            if 'deleted_at' not in cols:
                conn.execute(text(f"ALTER TABLE {table} ADD COLUMN deleted_at DATETIME"))
            if 'last_synced_updated_at' not in cols:
                conn.execute(text(f"ALTER TABLE {table} ADD COLUMN last_synced_updated_at DATETIME"))
            if 'server_touched_at' not in cols:
                conn.execute(text(f"ALTER TABLE {table} ADD COLUMN server_touched_at DATETIME"))
                # Backfill so existing rows are immediately pull-visible —
                # without this, every pre-sync row would have server_touched_at
                # NULL, and "NULL > since" is never true in SQL, so they'd
                # never appear in any /api/sync/pull response.
                conn.execute(text(
                    f"UPDATE {table} SET server_touched_at = COALESCE(updated_at, created_at) "
                    f"WHERE server_touched_at IS NULL"))

        if inspector.has_table('sync_sites'):
            site_cols = [c['name'] for c in inspector.get_columns('sync_sites')]
            if 'last_push_at' not in site_cols:
                conn.execute(text("ALTER TABLE sync_sites ADD COLUMN last_push_at DATETIME"))
            if 'last_pull_at' not in site_cols:
                conn.execute(text("ALTER TABLE sync_sites ADD COLUMN last_pull_at DATETIME"))

        # Every one of these tables is filtered by vehicle/driver + a date
        # range on essentially every ledger view, dashboard load and report
        # in the app, but until now the only indexed column anywhere in the
        # schema was sync_uuid — every one of those queries was a full table
        # scan. db.create_all() only creates indexes for brand-new tables, so
        # a table that already exists (i.e. every real install) needs these
        # added explicitly here. IF NOT EXISTS makes this a no-op after the
        # first run — cheap to leave in permanently rather than versioning it.
        for index_name, table, columns in (
            ('ix_daily_logs_vehicle_date', 'daily_logs', ('vehicle_id', 'log_date')),
            ('ix_daily_logs_date', 'daily_logs', ('log_date',)),
            ('ix_fuel_logs_vehicle_date', 'fuel_logs', ('vehicle_id', 'log_date')),
            ('ix_fuel_logs_date', 'fuel_logs', ('log_date',)),
            ('ix_maintenance_logs_vehicle_date', 'maintenance_logs', ('vehicle_id', 'log_date')),
            ('ix_maintenance_logs_date', 'maintenance_logs', ('log_date',)),
            ('ix_commission_payments_driver_date', 'commission_payments', ('driver_id', 'payment_date')),
            ('ix_commission_payments_date', 'commission_payments', ('payment_date',)),
            ('ix_payroll_deductions_driver_date', 'payroll_deductions', ('driver_id', 'deduction_date')),
            ('ix_payroll_deductions_date', 'payroll_deductions', ('deduction_date',)),
            ('ix_expenses_vehicle_date', 'expenses', ('vehicle_id', 'expense_date')),
            ('ix_expenses_date', 'expenses', ('expense_date',)),
            ('ix_driver_deposits_date', 'driver_deposits', ('deposit_date',)),
            ('ix_franchise_weekly_income_vehicle_week', 'franchise_weekly_income', ('vehicle_id', 'week_start')),
            ('ix_franchise_weekly_income_week', 'franchise_weekly_income', ('week_start',)),
            ('ix_store_purchases_date', 'store_purchases', ('purchase_date',)),
            ('ix_store_sales_vehicle_date', 'store_sales', ('vehicle_id', 'sale_date')),
            ('ix_store_sales_date', 'store_sales', ('sale_date',)),
            ('ix_store_adjustments_date', 'store_adjustments', ('adjustment_date',)),
        ):
            if inspector.has_table(table):
                conn.execute(text(
                    f"CREATE INDEX IF NOT EXISTS {index_name} ON {table} ({', '.join(columns)})"))

        conn.commit()

    # sync_uuid backfill needs uuid.uuid4() per row via the ORM, not raw SQL
    # (SQLite has no built-in UUID function) — runs as its own pass, after
    # the ALTER TABLE connection above has committed and closed, so every
    # existing row has a real cross-instance identity before the sync
    # engine (Phase 2/3) starts relying on it.
    sync_models = (
        Vehicle, VehicleDocument, Driver, Route, DailyLog, FuelLog, MaintenanceLog,
        Loan, LoanPayment, Payable, Receivable, CommissionPayment, CapitalContribution,
        OwnerDrawing, ExpenseCategory, Expense, Budget, DriverDeposit, FranchiseDailyIncome,
        FranchiseWeeklyIncome, FranchiseVehicle, FranchiseCollection,
        MaintenanceSchedule, SparePart, StorePurchase, StoreSale,
    )
    for model in sync_models:
        if not inspector.has_table(model.__tablename__):
            continue
        for row in model.query.execution_options(include_deleted=True).filter(model.sync_uuid.is_(None)).all():
            row.sync_uuid = uuid.uuid4().hex
    if db.session.dirty:
        db.session.commit()


def _seed_category_uuid(name, parent_name=None):
    """Deterministic (not random) sync_uuid for the fixed set of default
    expense categories every instance seeds independently at bootstrap.
    Using uuid4() here (like touch_sync_fields does for real user-created
    rows) would give the hub's "Maintenance" and a spoke's own
    independently-bootstrapped "Maintenance" two different identities —
    they'd never recognize each other as the same row and just pile up as
    duplicates once synced. A stable hash of the name means every instance
    converges on the same sync_uuid for the same default category."""
    return uuid.uuid5(uuid.NAMESPACE_DNS, f'expense_category:{parent_name or ""}:{name}').hex


def create_default_expense_categories():
    """Vehicle Expenses are classified under exactly six top-level
    headings: Maintenance, Wages, Traffic Fines, Insurance, Admin, Garage
    Fee. Garage Fee carries a default_amount ($100) since — unlike the
    other five — it's a roughly-fixed charge booked once per vehicle every
    month, so Add Expense can pre-fill it (see expense_form.html)."""
    headings = ('Maintenance', 'Wages', 'Traffic Fines', 'Insurance', 'Admin', 'Garage Fee')
    default_amounts = {'Garage Fee': 100}
    for name in headings:
        existing = ExpenseCategory.query.filter_by(name=name, parent_id=None).first()
        if not existing:
            db.session.add(ExpenseCategory(name=name, default_amount=default_amounts.get(name),
                                           sync_uuid=_seed_category_uuid(name)))
        elif name in default_amounts and existing.default_amount is None:
            # Backfills the default onto an instance that already created
            # this heading (e.g. via sync) before default_amount existed —
            # never overwrites one an admin has since edited.
            existing.default_amount = default_amounts[name]
    db.session.flush()

    # Retire the older, differently-named default headings this list
    # replaces. Each is only deleted outright if it holds no expenses and
    # no sub-categories — if it does, that data is folded into the closest
    # matching new heading first, so nothing already booked gets silently
    # dropped or orphaned.
    legacy_fold_into = {
        'Salaries & Wages': 'Wages',
        'Licensing & Permits': 'Admin',
        'Rent & Utilities': 'Admin',
        'Other Overhead': 'Admin',
        'Tax': 'Admin',
    }
    new_headings_by_name = {h.name: h for h in
        ExpenseCategory.query.filter(ExpenseCategory.parent_id.is_(None),
                                     ExpenseCategory.name.in_(headings)).all()}
    for old_name, new_name in legacy_fold_into.items():
        old = ExpenseCategory.query.filter_by(name=old_name, parent_id=None).first()
        if not old:
            continue
        has_data = Expense.query.filter_by(category_id=old.id).first() is not None or old.children
        if has_data:
            new = new_headings_by_name.get(new_name)
            if new:
                Expense.query.filter_by(category_id=old.id).update({'category_id': new.id})
                for child in list(old.children):
                    child.parent_id = new.id
        # Deliberately a real delete, not soft-delete + touch_sync_fields —
        # this is one-time bootstrap cleanup of legacy category names that
        # runs identically on every instance's own first boot, not a user
        # action, so there's nothing here that needs to propagate as a
        # tombstone to any other instance.
        db.session.delete(old)
    db.session.flush()

    maintenance = ExpenseCategory.query.filter_by(name='Maintenance', parent_id=None).first()
    if not maintenance:
        maintenance = ExpenseCategory(name='Maintenance', sync_uuid=_seed_category_uuid('Maintenance'))
        db.session.add(maintenance)
        db.session.flush()
    for name in ('Brake Pads and Tyres', 'Brake Shoes', 'Tie Rod Ends',
                 'Wheel Bearing', 'Ball Joints', 'Engine Oil'):
        if not ExpenseCategory.query.filter_by(name=name, parent_id=maintenance.id).first():
            db.session.add(ExpenseCategory(name=name, parent_id=maintenance.id,
                                           sync_uuid=_seed_category_uuid(name, 'Maintenance')))
    db.session.commit()


def create_default_admin():
    if User.query.filter_by(username='admin').first():
        return
    admin_password = os.environ.get('ADMIN_PASSWORD')
    if not admin_password:
        if FROZEN:
            # No admin yet and nothing pre-configured in .env — the
            # first-run /setup wizard creates the admin instead, with a
            # password the person running this actually chose in their
            # browser rather than one only ever visible in a console
            # window they may have already closed.
            return
        admin_password = secrets.token_urlsafe(12)
    admin = User(username='admin', email='admin@transport.local', role='admin')
    admin.set_password(admin_password)
    db.session.add(admin)
    db.session.commit()
    print(f'Default admin created username: admin  password: {admin_password}')
    print('Log in and change this password immediately.')


with app.app_context():
    db.create_all()
    migrate_db()
    db.create_all()  # recreate any tables migrate_db() dropped for a schema change
    create_default_admin()
    create_default_expense_categories()

# Only a spoke (SYNC_ENABLED=true in its .env, pointed at SYNC_HUB_URL)
# starts this loop — Central (Render) never does; it only serves
# /api/sync/push and /api/sync/pull. Note: Flask's debug-mode reloader
# re-executes this whole module in a child process, which could start a
# second thread alongside the reloader's parent watcher — harmless in
# practice since every sync cycle is idempotent (see
# apply_incoming_record), and a real spoke deployment runs without the
# reloader (FLASK_ENV=production) anyway.
if app.config['SYNC_ENABLED']:
    start_sync_thread()

def _disable_console_quick_edit():
    """A stray click into this console window (to read it, or by
    accident) triggers Windows Console "Quick Edit Mode": the whole
    process's console I/O pauses until Enter/Esc is pressed. Since every
    request thread logs its access line to this console (see
    werkzeug's log_request), one click freezes every in-flight AND
    future HTTP request — the server looks completely hung (port still
    listening, TCP connects fine, but no response ever comes back) until
    someone notices and hits a key. For an unattended spoke deployment
    (see SPOKE_SETUP.md's NSSM service-wrap) that's effectively a
    silent, click-triggered outage. Disabling Quick Edit Mode (keeping
    Insert Mode) removes the trap. Best-effort: never let a console-mode
    tweak block startup if the handle/API isn't available."""
    try:
        kernel32 = ctypes.windll.kernel32
        STD_INPUT_HANDLE = -10
        ENABLE_EXTENDED_FLAGS = 0x0080
        ENABLE_QUICK_EDIT_MODE = 0x0040
        ENABLE_INSERT_MODE = 0x0020
        handle = kernel32.GetStdHandle(STD_INPUT_HANDLE)
        mode = ctypes.c_uint32()
        if not kernel32.GetConsoleMode(handle, ctypes.byref(mode)):
            return
        new_mode = (mode.value & ~ENABLE_QUICK_EDIT_MODE) | ENABLE_EXTENDED_FLAGS | ENABLE_INSERT_MODE
        kernel32.SetConsoleMode(handle, new_mode)
    except Exception:
        pass


if __name__ == '__main__':
    debug_mode = not IS_PRODUCTION
    host = os.environ.get('HOST', '127.0.0.1' if IS_PRODUCTION else '0.0.0.0')
    port = int(os.environ.get('PORT', 5000))
    if FROZEN:
        if sys.platform == 'win32':
            _disable_console_quick_edit()
        # A double-clicked .exe has no terminal to print a URL into —
        # open the browser ourselves. Timer (not a direct call) so it
        # fires shortly after app.run() below has actually bound the
        # port, instead of racing it.
        threading.Timer(1.5, lambda: webbrowser.open(f'http://127.0.0.1:{port}')).start()
    app.run(debug=debug_mode, host=host, port=port)
